#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
進出口資料歸檔自動化工具
========================

使用說明：
---------
1. 建立訂單資料夾結構：
   python3 資料歸檔工具.py --create-folders --order SH-2026-TH-0047 --customer BEC --year 2026 --month 06

2. 自動歸檔文件（掃描目錄並按規則歸類）：
   python3 資料歸檔工具.py --organize --source ./散落的文件/ --order SH-2026-TH-0047 --customer BEC

3. 更新進度追蹤表：
   python3 資料歸檔工具.py --update-progress --order SH-2026-TH-0047 --doc 商業發票 --status 已歸檔

4. 查看訂單歸檔進度：
   python3 資料歸檔工具.py --progress SH-2026-TH-0047

5. 查看所有訂單進度匯總：
   python3 資料歸檔工具.py --summary

功能說明：
---------
- NamingEngine: 文件自動命名規則引擎
- FolderManager: 資料夾結構管理與文件歸檔
- ProgressTracker: 進度追蹤表自動更新

作者：自動生成
日期：2026-06-16
"""

import os
import sys
import shutil
import argparse
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("\033[91m錯誤：需要安裝 openpyxl 套件\033[0m")
    print("請執行：pip install openpyxl")
    sys.exit(1)


class ColorPrinter:
    """彩色終端輸出"""
    
    @staticmethod
    def success(msg: str):
        print(f"\033[92m✓ {msg}\033[0m")
    
    @staticmethod
    def error(msg: str):
        print(f"\033[91m✗ {msg}\033[0m")
    
    @staticmethod
    def warning(msg: str):
        print(f"\033[93m⚠ {msg}\033[0m")
    
    @staticmethod
    def info(msg: str):
        print(f"\033[94mℹ {msg}\033[0m")
    
    @staticmethod
    def header(msg: str):
        print(f"\n\033[1;96m{'='*60}\033[0m")
        print(f"\033[1;96m  {msg}\033[0m")
        print(f"\033[1;96m{'='*60}\033[0m\n")
    
    @staticmethod
    def table(headers: List[str], rows: List[List[str]]):
        """簡單的表格輸出"""
        col_widths = [len(h) for h in headers]
        for row in rows:
            for i, cell in enumerate(row):
                col_widths[i] = max(col_widths[i], len(str(cell)))
        
        # 表頭
        header_line = " | ".join(h.ljust(col_widths[i]) for i, h in enumerate(headers))
        separator = "-+-".join("-" * w for w in col_widths)
        
        print(f"\033[1;97m{header_line}\033[0m")
        print(separator)
        
        # 資料行
        for row in rows:
            line = " | ".join(str(cell).ljust(col_widths[i]) for i, cell in enumerate(row))
            print(line)


class NamingEngine:
    """文件命名規則引擎"""
    
    NAMING_RULES = {
        "商業發票": "{訂單號}_CI_{日期}.xlsx",
        "裝箱單": "{訂單號}_PL_{日期}.xlsx",
        "提單": "{訂單號}_BL_{日期}.pdf",
        "原產地證": "{訂單號}_CO_{日期}.pdf",
        "申報要素": "{訂單號}_DECL_{日期}.xlsx",
        "嘜頭_正": "{訂單號}_MainMark_Box{箱號}.png",
        "嘜頭_側": "{訂單號}_SideMark_Box{箱號}.png",
        "嘜頭匯總": "{訂單號}_Marks_{日期}.pdf",
        "報關委託書": "{訂單號}_Customs_{日期}.pdf",
        "HS編碼查詢": "{訂單號}_HS_{日期}.xlsx",
        "校驗報告": "{訂單號}_Validation_{日期}.xlsx",
        "清關合同": "{訂單號}_Contract_{日期}.pdf",
        "稅費單據": "{訂單號}_Tax_{日期}.pdf",
        "其他": "{訂單號}_{文件類型}_{日期}.{副檔名}",
    }
    
    # 用於自動偵測文件類型的關鍵字
    DOC_TYPE_KEYWORDS = {
        "商業發票": ["invoice", "發票", "ci", "commercial"],
        "裝箱單": ["packing", "裝箱", "pl", "packing list"],
        "提單": ["bill of lading", "bl", "提單", "b/l"],
        "原產地證": ["certificate of origin", "co", "原產地", "產地證"],
        "申報要素": ["declaration", "申報", "要素", "decl"],
        "嘜頭_正": ["main mark", "正嘜", "主嘜"],
        "嘜頭_側": ["side mark", "側嘜"],
        "嘜頭匯總": ["marks summary", "嘜頭匯總", "marks"],
        "報關委託書": ["customs", "報關", "委託"],
        "HS編碼查詢": ["hs code", "hs編碼", "海關編碼"],
        "校驗報告": ["validation", "校驗", "檢查報告"],
        "清關合同": ["contract", "合同", "清關"],
        "稅費單據": ["tax", "稅", "稅費"],
    }
    
    def __init__(self):
        self.printer = ColorPrinter()
    
    def auto_rename(self, file_path: str, order_number: str, doc_type: str, 
                    date: Optional[str] = None, box_number: Optional[str] = None) -> str:
        """根據規則自動重命名文件"""
        try:
            file_path = Path(file_path)
            if not file_path.exists():
                raise FileNotFoundError(f"文件不存在：{file_path}")
            
            # 預設日期為今天
            if date is None:
                date = datetime.now().strftime("%Y%m%d")
            
            # 取得命名規則
            if doc_type in self.NAMING_RULES:
                template = self.NAMING_RULES[doc_type]
            else:
                template = self.NAMING_RULES["其他"]
            
            # 取得副檔名
            ext = file_path.suffix.lstrip(".")
            
            # 建立新檔名
            new_name = template.format(
                訂單號=order_number,
                日期=date,
                文件類型=doc_type,
                副檔名=ext,
                箱號=box_number or "01"
            )
            
            # 執行重命名
            new_path = file_path.parent / new_name
            if new_path.exists() and new_path != file_path:
                self.printer.warning(f"目標文件已存在，跳過：{new_name}")
                return str(file_path)
            
            file_path.rename(new_path)
            self.printer.success(f"重命名：{file_path.name} -> {new_name}")
            return str(new_path)
            
        except Exception as e:
            self.printer.error(f"重命名失敗：{e}")
            return str(file_path)
    
    def batch_rename(self, files_dir: str, order_number: str, date: Optional[str] = None) -> Dict[str, str]:
        """批量重命名目錄中的文件"""
        files_dir = Path(files_dir)
        if not files_dir.exists():
            self.printer.error(f"目錄不存在：{files_dir}")
            return {}
        
        results = {}
        files = list(files_dir.iterdir())
        
        self.printer.header(f"批量重命名 - {order_number}")
        self.printer.info(f"掃描目錄：{files_dir}")
        self.printer.info(f"找到文件：{len(files)} 個")
        
        for file_path in files:
            if file_path.is_file():
                # 自動偵測文件類型
                doc_type = self.detect_doc_type(file_path)
                if doc_type:
                    new_path = self.auto_rename(str(file_path), order_number, doc_type, date)
                    results[str(file_path)] = new_path
                else:
                    self.printer.warning(f"無法識別文件類型：{file_path.name}")
        
        self.printer.success(f"完成重命名 {len(results)} 個文件")
        return results
    
    def detect_doc_type(self, file_path: str) -> Optional[str]:
        """自動偵測文件類型（根據內容或檔名）"""
        file_path = Path(file_path)
        file_name_lower = file_path.stem.lower()
        
        # 根據檔名關鍵字偵測
        for doc_type, keywords in self.DOC_TYPE_KEYWORDS.items():
            for keyword in keywords:
                if keyword.lower() in file_name_lower:
                    return doc_type
        
        # 嘗試讀取文件內容進行偵測（僅限文字文件）
        try:
            if file_path.suffix.lower() in [".txt", ".csv", ".json"]:
                with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                    content = f.read(2000).lower()
                    for doc_type, keywords in self.DOC_TYPE_KEYWORDS.items():
                        for keyword in keywords:
                            if keyword.lower() in content:
                                return doc_type
        except Exception:
            pass
        
        return None


class FolderManager:
    """資料夾結構管理"""
    
    FOLDER_STRUCTURE = {
        "01_報關單據": ["商業發票", "裝箱單", "提單", "原產地證", "申報要素"],
        "02_嘜頭包裝": ["正嘜", "側嘜", "嘜頭匯總"],
        "03_物流單據": ["訂艙確認", "提貨單"],
        "04_清關資料": ["清關合同", "稅費單據"],
        "05_校驗報告": [],
        "06_其他": [],
    }
    
    # 文件類型到資料夾的映射
    DOC_TYPE_TO_FOLDER = {
        "商業發票": ("01_報關單據", "商業發票"),
        "裝箱單": ("01_報關單據", "裝箱單"),
        "提單": ("01_報關單據", "提單"),
        "原產地證": ("01_報關單據", "原產地證"),
        "申報要素": ("01_報關單據", "申報要素"),
        "嘜頭_正": ("02_嘜頭包裝", "正嘜"),
        "嘜頭_側": ("02_嘜頭包裝", "側嘜"),
        "嘜頭匯總": ("02_嘜頭包裝", "嘜頭匯總"),
        "報關委託書": ("01_報關單據", ""),
        "HS編碼查詢": ("06_其他", ""),
        "校驗報告": ("05_校驗報告", ""),
        "清關合同": ("04_清關資料", "清關合同"),
        "稅費單據": ("04_清關資料", "稅費單據"),
    }
    
    def __init__(self):
        self.printer = ColorPrinter()
        self.naming_engine = NamingEngine()
    
    def create_order_folders(self, base_dir: str, year: str, month: str, 
                             order_number: str, customer_short: str) -> str:
        """為一筆訂單建立完整的資料夾結構"""
        try:
            base_path = Path(base_dir)
            order_folder_name = f"{order_number}_{customer_short}"
            order_path = base_path / "報關資料" / year / month / order_folder_name
            
            if order_path.exists():
                self.printer.warning(f"訂單資料夾已存在：{order_path}")
                return str(order_path)
            
            self.printer.header(f"建立訂單資料夾結構")
            self.printer.info(f"訂單編號：{order_number}")
            self.printer.info(f"客戶簡稱：{customer_short}")
            self.printer.info(f"目標路徑：{order_path}")
            
            # 建立主資料夾
            order_path.mkdir(parents=True, exist_ok=True)
            
            # 建立子資料夾結構
            for main_folder, sub_folders in self.FOLDER_STRUCTURE.items():
                main_path = order_path / main_folder
                main_path.mkdir(exist_ok=True)
                
                for sub_folder in sub_folders:
                    sub_path = main_path / sub_folder
                    sub_path.mkdir(exist_ok=True)
            
            self.printer.success(f"成功建立資料夾結構：{order_path}")
            self._print_folder_tree(order_path)
            
            return str(order_path)
            
        except Exception as e:
            self.printer.error(f"建立資料夾失敗：{e}")
            raise
    
    def _print_folder_tree(self, path: Path, prefix: str = ""):
        """打印資料夾樹狀結構"""
        items = sorted(path.iterdir(), key=lambda x: (not x.is_dir(), x.name))
        
        for i, item in enumerate(items):
            is_last = i == len(items) - 1
            connector = "└── " if is_last else "├── "
            
            if item.is_dir():
                print(f"{prefix}{connector}\033[94m{item.name}/\033[0m")
                extension = "    " if is_last else "│   "
                self._print_folder_tree(item, prefix + extension)
    
    def organize_files(self, source_dir: str, target_base: str, 
                       order_number: str, customer_short: str,
                       year: Optional[str] = None, month: Optional[str] = None) -> Dict[str, str]:
        """將散落的文件自動歸類到正確的資料夾"""
        try:
            source_path = Path(source_dir)
            if not source_path.exists():
                raise FileNotFoundError(f"來源目錄不存在：{source_dir}")
            
            # 預設年月為當前
            if year is None:
                year = datetime.now().strftime("%Y")
            if month is None:
                month = datetime.now().strftime("%m")
            
            # 確保目標資料夾存在
            order_folder_name = f"{order_number}_{customer_short}"
            target_path = Path(target_base) / "報關資料" / year / month / order_folder_name
            
            if not target_path.exists():
                self.create_order_folders(target_base, year, month, order_number, customer_short)
            
            self.printer.header(f"自動歸檔文件")
            self.printer.info(f"來源目錄：{source_path}")
            self.printer.info(f"目標目錄：{target_path}")
            
            results = {}
            files = [f for f in source_path.iterdir() if f.is_file()]
            
            self.printer.info(f"找到文件：{len(files)} 個")
            
            for file_path in files:
                # 偵測文件類型
                doc_type = self.naming_engine.detect_doc_type(file_path)
                
                if doc_type and doc_type in self.DOC_TYPE_TO_FOLDER:
                    main_folder, sub_folder = self.DOC_TYPE_TO_FOLDER[doc_type]
                    
                    # 確定目標路徑
                    if sub_folder:
                        dest_folder = target_path / main_folder / sub_folder
                    else:
                        dest_folder = target_path / main_folder
                    
                    dest_folder.mkdir(parents=True, exist_ok=True)
                    dest_path = dest_folder / file_path.name
                    
                    # 移動文件
                    if dest_path.exists():
                        self.printer.warning(f"目標文件已存在，跳過：{file_path.name}")
                    else:
                        shutil.move(str(file_path), str(dest_path))
                        results[str(file_path)] = str(dest_path)
                        self.printer.success(f"歸檔：{file_path.name} -> {main_folder}/{sub_folder}")
                else:
                    # 無法識別類型的文件放入「其他」
                    dest_folder = target_path / "06_其他"
                    dest_folder.mkdir(parents=True, exist_ok=True)
                    dest_path = dest_folder / file_path.name
                    
                    if not dest_path.exists():
                        shutil.move(str(file_path), str(dest_path))
                        results[str(file_path)] = str(dest_path)
                        self.printer.warning(f"無法識別類型，放入「其他」：{file_path.name}")
            
            self.printer.success(f"完成歸檔 {len(results)} 個文件")
            return results
            
        except Exception as e:
            self.printer.error(f"歸檔失敗：{e}")
            raise


class ProgressTracker:
    """進度追蹤表自動更新"""
    
    REQUIRED_DOCS = [
        "商業發票", "裝箱單", "提單", "原產地證", "申報要素",
        "嘜頭", "校驗報告", "HS查詢", "報關委託書", "清關合同", "稅費單據"
    ]
    
    COLUMNS = [
        "訂單編號", "客戶", "國家", "方向", "訂單日期",
        "商業發票", "裝箱單", "提單", "原產地證", "申報要素",
        "嘜頭", "校驗報告", "HS查詢", "報關委託書", "清關合同", "稅費單據",
        "歸檔完成度%", "最後更新", "備註"
    ]
    
    STATUS_VALUES = ["已歸檔", "未歸檔", "不需要"]
    
    def __init__(self, progress_file: str):
        self.progress_file = Path(progress_file)
        self.printer = ColorPrinter()
        self.wb = None
        self.ws = None
        self._load_or_create()
    
    @staticmethod
    def _normalize_completion(value) -> float:
        """正規化完成度為 0-1 範圍"""
        if not isinstance(value, (int, float)):
            return 0.0
        # 如果值大於 1，假設是百分比整數（如 73），轉換為 0-1
        if value > 1:
            return value / 100.0
        return float(value)
    
    def _load_or_create(self):
        """載入或建立進度追蹤 Excel"""
        try:
            if self.progress_file.exists():
                self.wb = load_workbook(str(self.progress_file))
                if "歸檔進度" in self.wb.sheetnames:
                    self.ws = self.wb["歸檔進度"]
                else:
                    self.ws = self.wb.active
                    self.ws.title = "歸檔進度"
            else:
                self.wb = Workbook()
                self.ws = self.wb.active
                self.ws.title = "歸檔進度"
                self._create_header()
                self._add_sample_data()
            
            self._save()
            
        except Exception as e:
            self.printer.error(f"載入進度表失敗：{e}")
            raise
    
    def _create_header(self):
        """建立表頭"""
        # 寫入欄位名稱
        for col_idx, col_name in enumerate(self.COLUMNS, 1):
            cell = self.ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="014725", end_color="014725", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # 設定欄寬
        column_widths = [15, 10, 8, 8, 12] + [10] * 11 + [12, 12, 20]
        for col_idx, width in enumerate(column_widths, 1):
            self.ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # 凍結首行
        self.ws.freeze_panes = "A2"
    
    def _add_sample_data(self):
        """添加範例數據"""
        sample_data = [
            {
                "訂單編號": "SH-2026-TH-0045",
                "客戶": "BEC Company",
                "國家": "泰國",
                "方向": "出口",
                "訂單日期": "2026-05-15",
                "商業發票": "已歸檔",
                "裝箱單": "已歸檔",
                "提單": "已歸檔",
                "原產地證": "已歸檔",
                "申報要素": "已歸檔",
                "嘜頭": "已歸檔",
                "校驗報告": "已歸檔",
                "HS查詢": "已歸檔",
                "報關委託書": "已歸檔",
                "清關合同": "不需要",
                "稅費單據": "不需要",
                "歸檔完成度%": 1.0,
                "最後更新": "2026-06-10",
                "備註": "已完成全部歸檔"
            },
            {
                "訂單編號": "SH-2026-VN-0046",
                "客戶": "Vina Tech",
                "國家": "越南",
                "方向": "出口",
                "訂單日期": "2026-05-20",
                "商業發票": "已歸檔",
                "裝箱單": "已歸檔",
                "提單": "已歸檔",
                "原產地證": "已歸檔",
                "申報要素": "未歸檔",
                "嘜頭": "已歸檔",
                "校驗報告": "未歸檔",
                "HS查詢": "已歸檔",
                "報關委託書": "已歸檔",
                "清關合同": "不需要",
                "稅費單據": "不需要",
                "歸檔完成度%": 0.73,
                "最後更新": "2026-06-12",
                "備註": "等待申報要素和校驗報告"
            },
            {
                "訂單編號": "SH-2026-MY-0047",
                "客戶": "MY Electronics",
                "國家": "馬來西亞",
                "方向": "進口",
                "訂單日期": "2026-06-01",
                "商業發票": "已歸檔",
                "裝箱單": "已歸檔",
                "提單": "未歸檔",
                "原產地證": "未歸檔",
                "申報要素": "未歸檔",
                "嘜頭": "未歸檔",
                "校驗報告": "未歸檔",
                "HS查詢": "已歸檔",
                "報關委託書": "已歸檔",
                "清關合同": "已歸檔",
                "稅費單據": "未歸檔",
                "歸檔完成度%": 0.36,
                "最後更新": "2026-06-14",
                "備註": "進口清關中"
            },
            {
                "訂單編號": "SH-2026-ID-0048",
                "客戶": "Jakarta Import",
                "國家": "印尼",
                "方向": "進口",
                "訂單日期": "2026-06-05",
                "商業發票": "已歸檔",
                "裝箱單": "未歸檔",
                "提單": "未歸檔",
                "原產地證": "未歸檔",
                "申報要素": "未歸檔",
                "嘜頭": "未歸檔",
                "校驗報告": "未歸檔",
                "HS查詢": "未歸檔",
                "報關委託書": "未歸檔",
                "清關合同": "未歸檔",
                "稅費單據": "未歸檔",
                "歸檔完成度%": 0.09,
                "最後更新": "2026-06-15",
                "備註": "剛開始處理"
            },
            {
                "訂單編號": "SH-2026-SG-0049",
                "客戶": "SG Logistics",
                "國家": "新加坡",
                "方向": "出口",
                "訂單日期": "2026-06-10",
                "商業發票": "已歸檔",
                "裝箱單": "已歸檔",
                "提單": "已歸檔",
                "原產地證": "不需要",
                "申報要素": "已歸檔",
                "嘜頭": "已歸檔",
                "校驗報告": "已歸檔",
                "HS查詢": "已歸檔",
                "報關委託書": "已歸檔",
                "清關合同": "不需要",
                "稅費單據": "不需要",
                "歸檔完成度%": 1.0,
                "最後更新": "2026-06-16",
                "備註": "新加坡免原產地證"
            }
        ]
        
        for row_idx, data in enumerate(sample_data, 2):
            for col_idx, col_name in enumerate(self.COLUMNS, 1):
                value = data.get(col_name, "")
                cell = self.ws.cell(row=row_idx, column=col_idx, value=value)
                
                # 狀態欄位上色
                if col_name in self.REQUIRED_DOCS:
                    if value == "已歸檔":
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        cell.font = Font(color="006100")
                    elif value == "未歸檔":
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        cell.font = Font(color="9C0006")
                    elif value == "不需要":
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        cell.font = Font(color="9C5700")
                
                # 完成度欄位
                if col_name == "歸檔完成度%":
                    cell.number_format = "0%"
                    if isinstance(value, (int, float)):
                        norm_val = value if value <= 1 else value / 100.0
                        if norm_val >= 0.8:
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif norm_val >= 0.5:
                            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        else:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # 添加邊框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in self.ws.iter_rows(min_row=1, max_row=len(sample_data)+1, 
                                      min_col=1, max_col=len(self.COLUMNS)):
            for cell in row:
                cell.border = thin_border
    
    def _save(self):
        """儲存 Excel"""
        self.progress_file.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(str(self.progress_file))
    
    def _find_order_row(self, order_number: str) -> Optional[int]:
        """查找訂單所在行"""
        for row_idx in range(2, self.ws.max_row + 1):
            if self.ws.cell(row=row_idx, column=1).value == order_number:
                return row_idx
        return None
    
    def update_status(self, order_number: str, doc_type: str, status: str, 
                      file_path: Optional[str] = None, customer: Optional[str] = None,
                      country: Optional[str] = None, direction: Optional[str] = None,
                      order_date: Optional[str] = None, remark: Optional[str] = None):
        """更新某訂單某文件的狀態"""
        try:
            # 驗證狀態值
            if status not in self.STATUS_VALUES:
                raise ValueError(f"無效的狀態值：{status}，必須是 {self.STATUS_VALUES} 之一")
            
            # 查找訂單行
            row_idx = self._find_order_row(order_number)
            
            # 如果訂單不存在，建立新行
            if row_idx is None:
                row_idx = self.ws.max_row + 1
                self.ws.cell(row=row_idx, column=1, value=order_number)
                self.ws.cell(row=row_idx, column=2, value=customer or "")
                self.ws.cell(row=row_idx, column=3, value=country or "")
                self.ws.cell(row=row_idx, column=4, value=direction or "")
                self.ws.cell(row=row_idx, column=5, value=order_date or datetime.now().strftime("%Y-%m-%d"))
                
                # 預設所有文件為未歸檔
                for col_idx, col_name in enumerate(self.COLUMNS, 1):
                    if col_name in self.REQUIRED_DOCS:
                        self.ws.cell(row=row_idx, column=col_idx, value="未歸檔")
            
            # 找到文件類型對應的列
            col_idx = None
            for i, col_name in enumerate(self.COLUMNS, 1):
                if col_name == doc_type:
                    col_idx = i
                    break
            
            if col_idx is None:
                raise ValueError(f"未知的文件類型：{doc_type}")
            
            # 更新狀態
            cell = self.ws.cell(row=row_idx, column=col_idx, value=status)
            
            # 上色
            if status == "已歸檔":
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                cell.font = Font(color="006100")
            elif status == "未歸檔":
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                cell.font = Font(color="9C0006")
            elif status == "不需要":
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                cell.font = Font(color="9C5700")
            
            # 更新完成度
            self._update_completion_rate(row_idx)
            
            # 更新最後更新時間
            date_col = self.COLUMNS.index("最後更新") + 1
            self.ws.cell(row=row_idx, column=date_col, value=datetime.now().strftime("%Y-%m-%d"))
            
            # 更新備註
            if remark:
                remark_col = self.COLUMNS.index("備註") + 1
                self.ws.cell(row=row_idx, column=remark_col, value=remark)
            
            # 添加邊框
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            for cell in self.ws[row_idx]:
                cell.border = thin_border
            
            self._save()
            
            self.printer.success(f"更新成功：{order_number} - {doc_type} -> {status}")
            
        except Exception as e:
            self.printer.error(f"更新失敗：{e}")
            raise
    
    def _update_completion_rate(self, row_idx: int):
        """更新歸檔完成度"""
        total = 0
        completed = 0
        
        for doc_type in self.REQUIRED_DOCS:
            col_idx = self.COLUMNS.index(doc_type) + 1
            status = self.ws.cell(row=row_idx, column=col_idx).value
            
            if status and status != "不需要":
                total += 1
                if status == "已歸檔":
                    completed += 1
        
        if total > 0:
            rate = completed / total
        else:
            rate = 1.0
        
        rate_col = self.COLUMNS.index("歸檔完成度%") + 1
        self.ws.cell(row=row_idx, column=rate_col, value=rate)
    
    def get_order_progress(self, order_number: str) -> Optional[Dict]:
        """獲取某訂單的文件歸檔進度"""
        row_idx = self._find_order_row(order_number)
        
        if row_idx is None:
            self.printer.error(f"找不到訂單：{order_number}")
            return None
        
        progress = {}
        for col_idx, col_name in enumerate(self.COLUMNS, 1):
            progress[col_name] = self.ws.cell(row=row_idx, column=col_idx).value
        
        return progress
    
    def generate_summary(self, output_path: Optional[str] = None) -> List[Dict]:
        """生成所有訂單的歸檔進度匯總"""
        self.printer.header("訂單歸檔進度匯總")
        
        summaries = []
        headers = ["訂單編號", "客戶", "國家", "方向", "完成度", "狀態"]
        rows = []
        
        for row_idx in range(2, self.ws.max_row + 1):
            order_number = self.ws.cell(row=row_idx, column=1).value
            if not order_number:
                continue
            
            customer = self.ws.cell(row=row_idx, column=2).value or ""
            country = self.ws.cell(row=row_idx, column=3).value or ""
            direction = self.ws.cell(row=row_idx, column=4).value or ""
            raw_completion = self.ws.cell(row=row_idx, column=self.COLUMNS.index("歸檔完成度%") + 1).value or 0
            
            # 正規化完成度
            completion = self._normalize_completion(raw_completion)
            completion_pct = f"{completion*100:.0f}%"
            
            if completion >= 0.8:
                status_str = "\033[92m✓ 接近完成\033[0m"
            elif completion >= 0.5:
                status_str = "\033[93m⚠ 進行中\033[0m"
            else:
                status_str = "\033[91m✗ 待處理\033[0m"
            
            rows.append([order_number, customer, country, direction, completion_pct, status_str])
            
            summaries.append({
                "訂單編號": order_number,
                "客戶": customer,
                "國家": country,
                "方向": direction,
                "完成度": completion,
            })
        
        ColorPrinter.table(headers, rows)
        
        # 統計資訊
        total_orders = len(summaries)
        completed_orders = sum(1 for s in summaries if s["完成度"] >= 0.99)
        avg_completion = sum(s["完成度"] for s in summaries) / total_orders if total_orders > 0 else 0
        
        print(f"\n\033[1;96m統計資訊：\033[0m")
        print(f"  總訂單數：{total_orders}")
        print(f"  已完成訂單：{completed_orders}")
        print(f"  平均完成度：{avg_completion*100:.1f}%")
        
        # 如果指定輸出路径，保存匯總報告
        if output_path:
            self._save_summary_report(output_path, summaries)
        
        return summaries
    
    def _save_summary_report(self, output_path: str, summaries: List[Dict]):
        """保存匯總報告到 Excel"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "進度匯總"
            
            # 寫入標題
            ws.cell(row=1, column=1, value="進出口資料歸檔進度匯總報告")
            ws.cell(row=1, column=1).font = Font(bold=True, size=14)
            ws.merge_cells("A1:F1")
            
            ws.cell(row=2, column=1, value=f"生成時間：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            
            # 寫入數據
            headers = ["訂單編號", "客戶", "國家", "方向", "完成度"]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="014725", end_color="014725", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            
            for row_idx, summary in enumerate(summaries, 5):
                ws.cell(row=row_idx, column=1, value=summary["訂單編號"])
                ws.cell(row=row_idx, column=2, value=summary["客戶"])
                ws.cell(row=row_idx, column=3, value=summary["國家"])
                ws.cell(row=row_idx, column=4, value=summary["方向"])
                ws.cell(row=row_idx, column=5, value=summary["完成度"])
                ws.cell(row=row_idx, column=5).number_format = "0%"
            
            wb.save(output_path)
            self.printer.success(f"匯總報告已保存：{output_path}")
            
        except Exception as e:
            self.printer.error(f"保存匯總報告失敗：{e}")


def main():
    """主程式入口"""
    parser = argparse.ArgumentParser(
        description="進出口資料歸檔自動化工具",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用範例：
  python3 資料歸檔工具.py --create-folders --order SH-2026-TH-0047 --customer BEC --year 2026 --month 06
  python3 資料歸檔工具.py --organize --source ./散落文件/ --order SH-2026-TH-0047 --customer BEC
  python3 資料歸檔工具.py --update-progress --order SH-2026-TH-0047 --doc 商業發票 --status 已歸檔
  python3 資料歸檔工具.py --progress SH-2026-TH-0047
  python3 資料歸檔工具.py --summary
        """
    )
    
    # 互斥操作組
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--create-folders", action="store_true", help="建立訂單資料夾結構")
    group.add_argument("--organize", action="store_true", help="自動歸檔文件")
    group.add_argument("--update-progress", action="store_true", help="更新進度追蹤表")
    group.add_argument("--progress", metavar="ORDER", help="查看訂單歸檔進度")
    group.add_argument("--summary", action="store_true", help="查看所有訂單進度匯總")
    group.add_argument("--batch-rename", action="store_true", help="批量重命名文件")
    
    # 共用參數
    parser.add_argument("--order", help="訂單編號")
    parser.add_argument("--customer", help="客戶簡稱")
    parser.add_argument("--year", default=datetime.now().strftime("%Y"), help="年份")
    parser.add_argument("--month", default=datetime.now().strftime("%m"), help="月份")
    parser.add_argument("--base-dir", default="./輸出", help="基礎目錄路徑")
    
    # 歸檔參數
    parser.add_argument("--source", help="來源目錄路徑")
    
    # 進度更新參數
    parser.add_argument("--doc", help="文件類型")
    parser.add_argument("--status", help="狀態（已歸檔/未歸檔/不需要）")
    parser.add_argument("--remark", help="備註")
    
    # 批量重命名參數
    parser.add_argument("--dir", help="批量重命名目錄")
    parser.add_argument("--date", help="日期（YYYYMMDD 格式）")
    
    # 進度表路徑
    parser.add_argument("--progress-file", default="./輸出/歸檔進度追蹤表.xlsx", help="進度追蹤表路徑")
    
    args = parser.parse_args()
    
    printer = ColorPrinter()
    
    try:
        if args.create_folders:
            # 建立資料夾結構
            if not args.order or not args.customer:
                parser.error("--create-folders 需要 --order 和 --customer 參數")
            
            fm = FolderManager()
            path = fm.create_order_folders(
                args.base_dir, args.year, args.month, 
                args.order, args.customer
            )
            printer.success(f"資料夾已建立：{path}")
        
        elif args.organize:
            # 自動歸檔文件
            if not args.source or not args.order or not args.customer:
                parser.error("--organize 需要 --source、--order 和 --customer 參數")
            
            fm = FolderManager()
            results = fm.organize_files(
                args.source, args.base_dir, 
                args.order, args.customer,
                args.year, args.month
            )
            printer.success(f"完成歸檔 {len(results)} 個文件")
        
        elif args.update_progress:
            # 更新進度
            if not args.order or not args.doc or not args.status:
                parser.error("--update-progress 需要 --order、--doc 和 --status 參數")
            
            tracker = ProgressTracker(args.progress_file)
            tracker.update_status(
                args.order, args.doc, args.status,
                remark=args.remark
            )
        
        elif args.progress:
            # 查看訂單進度
            tracker = ProgressTracker(args.progress_file)
            progress = tracker.get_order_progress(args.progress)
            
            if progress:
                printer.header(f"訂單進度：{args.progress}")
                
                print(f"\033[1;97m基本資訊：\033[0m")
                print(f"  客戶：{progress.get('客戶', 'N/A')}")
                print(f"  國家：{progress.get('國家', 'N/A')}")
                print(f"  方向：{progress.get('方向', 'N/A')}")
                print(f"  訂單日期：{progress.get('訂單日期', 'N/A')}")
                
                print(f"\n\033[1;97m文件歸檔狀態：\033[0m")
                for doc in tracker.REQUIRED_DOCS:
                    status = progress.get(doc, "未歸檔")
                    if status == "已歸檔":
                        icon = "\033[92m✓\033[0m"
                    elif status == "不需要":
                        icon = "\033[93m-\033[0m"
                    else:
                        icon = "\033[91m✗\033[0m"
                    print(f"  {icon} {doc}: {status}")
                
                raw_completion = progress.get("歸檔完成度%", 0)
                completion = tracker._normalize_completion(raw_completion)
                print(f"\n\033[1;97m歸檔完成度：{completion*100:.1f}%\033[0m")
                print(f"最後更新：{progress.get('最後更新', 'N/A')}")
                print(f"備註：{progress.get('備註', '')}")
        
        elif args.summary:
            # 查看所有訂單匯總
            tracker = ProgressTracker(args.progress_file)
            tracker.generate_summary()
        
        elif args.batch_rename:
            # 批量重命名
            if not args.dir or not args.order:
                parser.error("--batch-rename 需要 --dir 和 --order 參數")
            
            engine = NamingEngine()
            results = engine.batch_rename(args.dir, args.order, args.date)
            printer.success(f"完成重命名 {len(results)} 個文件")
    
    except Exception as e:
        printer.error(f"執行失敗：{e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
