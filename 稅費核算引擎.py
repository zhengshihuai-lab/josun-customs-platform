#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
稅費自動核算引擎 v1.0
Thailand & Indonesia Customs Import Duty Calculation Engine

支援功能:
  - 泰國/印尼進口稅費完整計算
  - FTA優惠稅率比較與最优方案建議
  - Excel報表產出 (openpyxl)
  - CLI命令列操作 (argparse)
  - Decimal精確計算 (無浮點誤差)
"""

import argparse
import json
import os
import sys
from datetime import datetime, date
from decimal import Decimal, ROUND_HALF_UP, getcontext
from typing import Any, Dict, List, Optional, Tuple

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print("錯誤: 請先安裝 openpyxl: pip install openpyxl")
    sys.exit(1)

# 設定 Decimal 精度
getcontext().prec = 28

# ─────────────────────────────────────────────────────────
# 常量與格式定義
# ─────────────────────────────────────────────────────────

D = Decimal
ZERO = D("0")
HUNDRED = D("100")

DEEP_BLUE = "1F4E79"
WHITE = "FFFFFF"
LIGHT_BLUE = "D6E4F0"
LIGHT_GREEN = "E2EFDA"
LIGHT_YELLOW = "FFF2CC"

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

BOX_H = "─"
BOX_V = "│"
BOX_TL = "┌"
BOX_TR = "┐"
BOX_BL = "└"
BOX_BR = "┘"
BOX_LT = "├"
BOX_RT = "┤"
BOX_DH = "═"
BOX_DV = "║"
BOX_DTL = "╔"
BOX_DTR = "╗"
BOX_DBL = "╚"
BOX_DBR = "╝"


def fmt_money(val: Decimal) -> str:
    """格式化金額，含千分位逗號，小數兩位"""
    if not isinstance(val, Decimal):
        val = D(str(val))
    q = val.quantize(D("0.01"), rounding=ROUND_HALF_UP)
    sign = "-" if q < 0 else ""
    q_abs = abs(q)
    integer_part = int(q_abs)
    decimal_part = q_abs - integer_part
    dec_str = str(decimal_part.quantize(D("0.01"), rounding=ROUND_HALF_UP))[2:]
    int_str = f"{integer_part:,}"
    return f"{sign}{int_str}.{dec_str}"


def fmt_pct(val: Decimal) -> str:
    """格式化百分比"""
    if not isinstance(val, Decimal):
        val = D(str(val))
    return str(val.quantize(D("0.01"), rounding=ROUND_HALF_UP)) + "%"


def draw_box(title: str, lines: List[str], width: int = 70) -> str:
    """以 box-drawing 字元繪製框線"""
    result = []
    inner_w = width - 2
    result.append(BOX_TL + BOX_H * inner_w + BOX_TR)
    title_pad = inner_w - len(title) - 2
    left_pad = title_pad // 2
    right_pad = title_pad - left_pad
    result.append(BOX_V + " " * left_pad + " " + title + " " + " " * right_pad + BOX_V)
    result.append(BOX_LT + BOX_H * inner_w + BOX_RT)
    for line in lines:
        # 計算顯示寬度 (CJK字元佔2格)
        display_len = 0
        for ch in line:
            if '\u4e00' <= ch <= '\u9fff' or '\u3000' <= ch <= '\u303f' or '\uff00' <= ch <= '\uffef':
                display_len += 2
            else:
                display_len += 1
        padding = inner_w - display_len
        if padding < 0:
            padding = 0
        result.append(BOX_V + " " + line + " " * max(padding - 1, 0) + BOX_V)
    result.append(BOX_BL + BOX_H * inner_w + BOX_BR)
    return "\n".join(result)


def draw_double_box(title: str, lines: List[str], width: int = 70) -> str:
    """雙線框"""
    result = []
    inner_w = width - 2
    result.append(BOX_DTL + BOX_DH * inner_w + BOX_DTR)
    title_pad = inner_w - len(title) - 2
    left_pad = title_pad // 2
    right_pad = title_pad - left_pad
    result.append(BOX_DV + " " * left_pad + " " + title + " " + " " * right_pad + BOX_DV)
    result.append(BOX_LT + BOX_H * inner_w + BOX_RT)
    for line in lines:
        display_len = 0
        for ch in line:
            if '\u4e00' <= ch <= '\u9fff' or '\u3000' <= ch <= '\u303f' or '\uff00' <= ch <= '\uffef':
                display_len += 2
            else:
                display_len += 1
        padding = inner_w - display_len
        if padding < 0:
            padding = 0
        result.append(BOX_V + " " + line + " " * max(padding - 1, 0) + BOX_V)
    result.append(BOX_DBL + BOX_DH * inner_w + BOX_DBR)
    return "\n".join(result)


# ─────────────────────────────────────────────────────────
# 1. TaxRateDatabase - 稅率資料庫
# ─────────────────────────────────────────────────────────

class TaxRateDatabase:
    """綜合稅率資料庫，涵蓋泰國與印尼主要HS編碼"""

    def __init__(self):
        self._thailand_rates: Dict[str, Dict[str, Any]] = {}
        self._indonesia_rates: Dict[str, Dict[str, Any]] = {}
        self._fta_agreements: List[Dict[str, str]] = []
        self._load_thailand_rates()
        self._load_indonesia_rates()
        self._load_fta_agreements()

    def _load_thailand_rates(self):
        """載入泰國稅率資料 (30+筆)"""
        data = [
            # Chapter 84 - Machinery
            {"hs_code": "8471.30", "description": "可攜式自動資料處理機(筆記型電腦)",
             "duty_rate": D("5"), "fta_rate": D("0"), "special_tax": D("0"),
             "vat": D("7"), "withholding_tax": D("1"), "fta_agreement": "ACFTA/ATIGA/JTEPA",
             "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            {"hs_code": "8471.41", "description": "其他自動資料處理機(桌上型)",
             "duty_rate": D("5"), "fta_rate": D("0"), "special_tax": D("0"),
             "vat": D("7"), "withholding_tax": D("1"), "fta_agreement": "ACFTA/ATIGA",
             "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            {"hs_code": "8471.49", "description": "其他自動資料處理機組件",
             "duty_rate": D("5"), "fta_rate": D("0"), "special_tax": D("0"),
             "vat": D("7"), "withholding_tax": D("1"), "fta_agreement": "ACFTA/ATIGA",
             "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            {"hs_code": "8473.30", "description": "電腦零件及附件",
             "duty_rate": D("5"), "fta_rate": D("0"), "special_tax": D("0"),
             "vat": D("7"), "withholding_tax": D("1"), "fta_agreement": "ACFTA/ATIGA/ITA",
             "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            {"hs_code": "8418.10", "description": "冷藏/冷凍組合櫃",
             "duty_rate": D("10"), "fta_rate": D("0"), "special_tax": D("0"),
             "vat": D("7"), "withholding_tax": D("1"), "fta_agreement": "ATIGA/ACFTA",
             "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            {"hs_code": "8415.10", "description": "空氣調節器(分離式)",
             "duty_rate": D("10"), "fta_rate": D("0"), "special_tax": D("0"),
             "vat": D("7"), "withholding_tax": D("1"), "fta_agreement": "ATIGA/ACFTA",
             "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            # Chapter 85 - Electrical
            {"hs_code": "8544.42", "description": "電纜及導體(裝配接頭)",
             "duty_rate": D("5"), "fta_rate": D("0"), "special_tax": D("0"),
             "vat": D("7"), "withholding_tax": D("1"), "fta_agreement": "ACFTA/ATIGA",
             "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            {"hs_code": "8544.49", "description": "其他電導體",
             "duty_rate": D("5"), "fta_rate": D("0"), "special_tax": D("0"),
             "vat": D("7"), "withholding_tax": D("1"), "fta_agreement": "ACFTA/ATIGA",
             "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            {"hs_code": "8517.12", "description": "電話機及通訊設備",
             "duty_rate": D("5"), "fta_rate": D("0"), "special_tax": D("0"),
             "vat": D("7"), "withholding_tax": D("1"), "fta_agreement": "ACFTA/ATIGA/ITA",
             "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            {"hs_code": "8517.62", "description": "數據通訊設備(路由器/交換器)",
             "duty_rate": D("5"), "fta_rate": D("0"), "special_tax": D("0"),
             "vat": D("7"), "withholding_tax": D("1"), "fta_agreement": "ACFTA/ATIGA",
             "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            {"hs_code": "8528.72", "description": "彩色電視接收機",
             "duty_rate": D("10"), "fta_rate": D("0"), "special_tax": D("1.5"),
             "vat": D("7"), "withholding_tax": D("1"), "fta_agreement": "ATIGA/ACFTA",
             "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            {"hs_code": "8536.50", "description": "電路開關及保險絲",
             "duty_rate": D("5"), "fta_rate": D("0"), "special_tax": D("0"),
             "vat": D("7"), "withholding_tax": D("1"), "fta_agreement": "ACFTA/ATIGA",
             "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            {"hs_code": "8541.40", "description": "半導體器件(二極體/電晶體)",
             "duty_rate": D("0"), "fta_rate": D("0"), "special_tax": D("0"),
             "vat": D("7"), "withholding_tax": D("1"), "fta_agreement": "ITA/ACFTA",
             "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            # Chapter 73 - Steel
            {"hs_code": "7308.90", "description": "鋼鐵結構體及其零件",
             "duty_rate": D("10"), "fta_rate": D("0"), "special_tax": D("0"),
             "vat": D("7"), "withholding_tax": D("1"), "fta_agreement": "ATIGA/ACFTA",
             "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            {"hs_code": "7306.30", "description": "鋼鐵管(焊接圓管)",
             "duty_rate": D("5"), "fta_rate": D("0"), "special_tax": D("0"),
             "vat": D("7"), "withholding_tax": D("1"), "fta_agreement": "ATIGA/ACFTA",
             "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            {"hs_code": "7318.15", "description": "螺釘及螺栓",
             "duty_rate": D("10"), "fta_rate": D("0"), "special_tax": D("0"),
             "vat": D("7"), "withholding_tax": D("1"), "fta_agreement": "ATIGA/ACFTA",
             "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            # Chapter 39 - Plastics
            {"hs_code": "3923.21", "description": "聚乙烯塑膠袋",
             "duty_rate": D("5"), "fta_rate": D("0"), "special_tax": D("0"),
             "vat": D("7"), "withholding_tax": D("1"), "fta_agreement": "ATIGA/ACFTA",
             "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            {"hs_code": "3923.30", "description": "塑膠瓶及容器",
             "duty_rate": D("5"), "fta_rate": D("0"), "special_tax": D("0"),
             "vat": D("7"), "withholding_tax": D("1"), "fta_agreement": "ATIGA/ACFTA",
             "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            {"hs_code": "3926.90", "description": "其他塑膠製品",
             "duty_rate": D("5"), "fta_rate": D("0"), "special_tax": D("0"),
             "vat": D("7"), "withholding_tax": D("1"), "fta_agreement": "ATIGA/ACFTA",
             "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            # Chapter 40 - Rubber
            {"hs_code": "4011.10", "description": "新充氣橡膠輪胎(汽車用)",
             "duty_rate": D("10"), "fta_rate": D("0"), "special_tax": D("0"),
             "vat": D("7"), "withholding_tax": D("1"), "fta_agreement": "ATIGA/ACFTA",
             "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            {"hs_code": "4016.93", "description": "橡膠墊圈及密封件",
             "duty_rate": D("5"), "fta_rate": D("0"), "special_tax": D("0"),
             "vat": D("7"), "withholding_tax": D("1"), "fta_agreement": "ATIGA/ACFTA",
             "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            # Chapter 87 - Vehicles
            {"hs_code": "8708.99", "description": "汽車零件及附件",
             "duty_rate": D("10"), "fta_rate": D("0"), "special_tax": D("0"),
             "vat": D("7"), "withholding_tax": D("1"), "fta_agreement": "ATIGA/JTEPA",
             "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            {"hs_code": "8708.29", "description": "車身零件",
             "duty_rate": D("10"), "fta_rate": D("0"), "special_tax": D("0"),
             "vat": D("7"), "withholding_tax": D("1"), "fta_agreement": "ATIGA/JTEPA",
             "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            # Chapter 90 - Instruments
            {"hs_code": "9018.11", "description": "心電圖記錄儀",
             "duty_rate": D("5"), "fta_rate": D("0"), "special_tax": D("0"),
             "vat": D("7"), "withholding_tax": D("1"), "fta_agreement": "ACFTA/ATIGA",
             "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            {"hs_code": "9027.80", "description": "分析儀器及設備",
             "duty_rate": D("0"), "fta_rate": D("0"), "special_tax": D("0"),
             "vat": D("7"), "withholding_tax": D("1"), "fta_agreement": "ITA/ACFTA",
             "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            {"hs_code": "9031.80", "description": "計量檢驗儀器",
             "duty_rate": D("3"), "fta_rate": D("0"), "special_tax": D("0"),
             "vat": D("7"), "withholding_tax": D("1"), "fta_agreement": "ACFTA/ATIGA",
             "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            # Chapter 62 - Textiles
            {"hs_code": "6204.62", "description": "女褲(棉質)",
             "duty_rate": D("30"), "fta_rate": D("0"), "special_tax": D("0"),
             "vat": D("7"), "withholding_tax": D("1"), "fta_agreement": "ATIGA",
             "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            {"hs_code": "6203.42", "description": "男褲(棉質)",
             "duty_rate": D("30"), "fta_rate": D("0"), "special_tax": D("0"),
             "vat": D("7"), "withholding_tax": D("1"), "fta_agreement": "ATIGA",
             "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            # Chapter 94 - Furniture
            {"hs_code": "9403.20", "description": "金屬家具",
             "duty_rate": D("10"), "fta_rate": D("0"), "special_tax": D("0"),
             "vat": D("7"), "withholding_tax": D("1"), "fta_agreement": "ATIGA/ACFTA",
             "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            {"hs_code": "9403.60", "description": "木製家具",
             "duty_rate": D("10"), "fta_rate": D("0"), "special_tax": D("0"),
             "vat": D("7"), "withholding_tax": D("1"), "fta_agreement": "ATIGA/ACFTA",
             "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            # Chapter 76 - Aluminum
            {"hs_code": "7606.12", "description": "鋁合金板及片",
             "duty_rate": D("5"), "fta_rate": D("0"), "special_tax": D("0"),
             "vat": D("7"), "withholding_tax": D("1"), "fta_agreement": "ATIGA/ACFTA",
             "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            {"hs_code": "7604.10", "description": "鋁合金棒及條",
             "duty_rate": D("5"), "fta_rate": D("0"), "special_tax": D("0"),
             "vat": D("7"), "withholding_tax": D("1"), "fta_agreement": "ATIGA/ACFTA",
             "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
        ]
        for item in data:
            self._thailand_rates[item["hs_code"]] = item

    def _load_indonesia_rates(self):
        """載入印尼稅率資料 (30+筆)"""
        data = [
            # Chapter 84 - Machinery
            {"hs_code": "8471.30", "description": "可攜式自動資料處理機(筆記型電腦)",
             "mfn_rate": D("0"), "atiga_rate": D("0"), "acfta_rate": D("0"),
             "ppn": D("11"), "pph": D("2.5"), "excise": D("0"), "pdri": D("2.5"),
             "fta_agreement": "ITA/ATIGA/ACFTA", "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            {"hs_code": "8471.41", "description": "其他自動資料處理機(桌上型)",
             "mfn_rate": D("0"), "atiga_rate": D("0"), "acfta_rate": D("0"),
             "ppn": D("11"), "pph": D("2.5"), "excise": D("0"), "pdri": D("2.5"),
             "fta_agreement": "ITA/ATIGA/ACFTA", "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            {"hs_code": "8471.49", "description": "其他自動資料處理機組件",
             "mfn_rate": D("0"), "atiga_rate": D("0"), "acfta_rate": D("0"),
             "ppn": D("11"), "pph": D("2.5"), "excise": D("0"), "pdri": D("2.5"),
             "fta_agreement": "ITA/ATIGA/ACFTA", "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            {"hs_code": "8473.30", "description": "電腦零件及附件",
             "mfn_rate": D("0"), "atiga_rate": D("0"), "acfta_rate": D("0"),
             "ppn": D("11"), "pph": D("2.5"), "excise": D("0"), "pdri": D("2.5"),
             "fta_agreement": "ITA/ATIGA/ACFTA", "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            {"hs_code": "8418.10", "description": "冷藏/冷凍組合櫃",
             "mfn_rate": D("5"), "atiga_rate": D("0"), "acfta_rate": D("0"),
             "ppn": D("11"), "pph": D("2.5"), "excise": D("0"), "pdri": D("2.5"),
             "fta_agreement": "ATIGA/ACFTA", "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            {"hs_code": "8415.10", "description": "空氣調節器(分離式)",
             "mfn_rate": D("5"), "atiga_rate": D("0"), "acfta_rate": D("0"),
             "ppn": D("11"), "pph": D("2.5"), "excise": D("0"), "pdri": D("2.5"),
             "fta_agreement": "ATIGA/ACFTA", "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            # Chapter 85 - Electrical
            {"hs_code": "8544.42", "description": "電纜及導體(裝配接頭)",
             "mfn_rate": D("5"), "atiga_rate": D("0"), "acfta_rate": D("0"),
             "ppn": D("11"), "pph": D("2.5"), "excise": D("0"), "pdri": D("2.5"),
             "fta_agreement": "ATIGA/ACFTA", "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            {"hs_code": "8544.49", "description": "其他電導體",
             "mfn_rate": D("5"), "atiga_rate": D("0"), "acfta_rate": D("0"),
             "ppn": D("11"), "pph": D("2.5"), "excise": D("0"), "pdri": D("2.5"),
             "fta_agreement": "ATIGA/ACFTA", "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            {"hs_code": "8517.12", "description": "電話機及通訊設備",
             "mfn_rate": D("0"), "atiga_rate": D("0"), "acfta_rate": D("0"),
             "ppn": D("11"), "pph": D("2.5"), "excise": D("0"), "pdri": D("2.5"),
             "fta_agreement": "ITA/ATIGA/ACFTA", "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            {"hs_code": "8517.62", "description": "數據通訊設備(路由器/交換器)",
             "mfn_rate": D("0"), "atiga_rate": D("0"), "acfta_rate": D("0"),
             "ppn": D("11"), "pph": D("2.5"), "excise": D("0"), "pdri": D("2.5"),
             "fta_agreement": "ITA/ATIGA/ACFTA", "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            {"hs_code": "8528.72", "description": "彩色電視接收機",
             "mfn_rate": D("10"), "atiga_rate": D("0"), "acfta_rate": D("5"),
             "ppn": D("11"), "pph": D("2.5"), "excise": D("0"), "pdri": D("2.5"),
             "fta_agreement": "ATIGA/ACFTA", "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            {"hs_code": "8536.50", "description": "電路開關及保險絲",
             "mfn_rate": D("5"), "atiga_rate": D("0"), "acfta_rate": D("0"),
             "ppn": D("11"), "pph": D("2.5"), "excise": D("0"), "pdri": D("2.5"),
             "fta_agreement": "ATIGA/ACFTA", "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            {"hs_code": "8541.40", "description": "半導體器件(二極體/電晶體)",
             "mfn_rate": D("0"), "atiga_rate": D("0"), "acfta_rate": D("0"),
             "ppn": D("11"), "pph": D("2.5"), "excise": D("0"), "pdri": D("2.5"),
             "fta_agreement": "ITA/ATIGA/ACFTA", "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            # Chapter 73 - Steel
            {"hs_code": "7308.90", "description": "鋼鐵結構體及其零件",
             "mfn_rate": D("10"), "atiga_rate": D("0"), "acfta_rate": D("0"),
             "ppn": D("11"), "pph": D("2.5"), "excise": D("0"), "pdri": D("2.5"),
             "fta_agreement": "ATIGA/ACFTA", "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            {"hs_code": "7306.30", "description": "鋼鐵管(焊接圓管)",
             "mfn_rate": D("10"), "atiga_rate": D("0"), "acfta_rate": D("0"),
             "ppn": D("11"), "pph": D("2.5"), "excise": D("0"), "pdri": D("2.5"),
             "fta_agreement": "ATIGA/ACFTA", "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            {"hs_code": "7318.15", "description": "螺釘及螺栓",
             "mfn_rate": D("5"), "atiga_rate": D("0"), "acfta_rate": D("0"),
             "ppn": D("11"), "pph": D("2.5"), "excise": D("0"), "pdri": D("2.5"),
             "fta_agreement": "ATIGA/ACFTA", "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            # Chapter 39 - Plastics
            {"hs_code": "3923.21", "description": "聚乙烯塑膠袋",
             "mfn_rate": D("5"), "atiga_rate": D("0"), "acfta_rate": D("0"),
             "ppn": D("11"), "pph": D("2.5"), "excise": D("0"), "pdri": D("2.5"),
             "fta_agreement": "ATIGA/ACFTA", "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            {"hs_code": "3923.30", "description": "塑膠瓶及容器",
             "mfn_rate": D("5"), "atiga_rate": D("0"), "acfta_rate": D("0"),
             "ppn": D("11"), "pph": D("2.5"), "excise": D("0"), "pdri": D("2.5"),
             "fta_agreement": "ATIGA/ACFTA", "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            {"hs_code": "3926.90", "description": "其他塑膠製品",
             "mfn_rate": D("5"), "atiga_rate": D("0"), "acfta_rate": D("0"),
             "ppn": D("11"), "pph": D("2.5"), "excise": D("0"), "pdri": D("2.5"),
             "fta_agreement": "ATIGA/ACFTA", "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            # Chapter 40 - Rubber
            {"hs_code": "4011.10", "description": "新充氣橡膠輪胎(汽車用)",
             "mfn_rate": D("15"), "atiga_rate": D("0"), "acfta_rate": D("5"),
             "ppn": D("11"), "pph": D("2.5"), "excise": D("0"), "pdri": D("2.5"),
             "fta_agreement": "ATIGA/ACFTA", "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            {"hs_code": "4016.93", "description": "橡膠墊圈及密封件",
             "mfn_rate": D("5"), "atiga_rate": D("0"), "acfta_rate": D("0"),
             "ppn": D("11"), "pph": D("2.5"), "excise": D("0"), "pdri": D("2.5"),
             "fta_agreement": "ATIGA/ACFTA", "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            # Chapter 87 - Vehicles
            {"hs_code": "8708.99", "description": "汽車零件及附件",
             "mfn_rate": D("0"), "atiga_rate": D("0"), "acfta_rate": D("0"),
             "ppn": D("11"), "pph": D("2.5"), "excise": D("0"), "pdri": D("2.5"),
             "fta_agreement": "ATIGA/ACFTA", "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            {"hs_code": "8708.29", "description": "車身零件",
             "mfn_rate": D("0"), "atiga_rate": D("0"), "acfta_rate": D("0"),
             "ppn": D("11"), "pph": D("2.5"), "excise": D("0"), "pdri": D("2.5"),
             "fta_agreement": "ATIGA/ACFTA", "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            # Chapter 90 - Instruments
            {"hs_code": "9018.11", "description": "心電圖記錄儀",
             "mfn_rate": D("5"), "atiga_rate": D("0"), "acfta_rate": D("0"),
             "ppn": D("11"), "pph": D("2.5"), "excise": D("0"), "pdri": D("2.5"),
             "fta_agreement": "ATIGA/ACFTA", "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            {"hs_code": "9027.80", "description": "分析儀器及設備",
             "mfn_rate": D("0"), "atiga_rate": D("0"), "acfta_rate": D("0"),
             "ppn": D("11"), "pph": D("2.5"), "excise": D("0"), "pdri": D("2.5"),
             "fta_agreement": "ITA/ATIGA/ACFTA", "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            {"hs_code": "9031.80", "description": "計量檢驗儀器",
             "mfn_rate": D("5"), "atiga_rate": D("0"), "acfta_rate": D("0"),
             "ppn": D("11"), "pph": D("2.5"), "excise": D("0"), "pdri": D("2.5"),
             "fta_agreement": "ATIGA/ACFTA", "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            # Chapter 62 - Textiles
            {"hs_code": "6204.62", "description": "女褲(棉質)",
             "mfn_rate": D("25"), "atiga_rate": D("0"), "acfta_rate": D("10"),
             "ppn": D("11"), "pph": D("2.5"), "excise": D("0"), "pdri": D("2.5"),
             "fta_agreement": "ATIGA", "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            {"hs_code": "6203.42", "description": "男褲(棉質)",
             "mfn_rate": D("25"), "atiga_rate": D("0"), "acfta_rate": D("10"),
             "ppn": D("11"), "pph": D("2.5"), "excise": D("0"), "pdri": D("2.5"),
             "fta_agreement": "ATIGA", "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            # Chapter 94 - Furniture
            {"hs_code": "9403.20", "description": "金屬家具",
             "mfn_rate": D("10"), "atiga_rate": D("0"), "acfta_rate": D("0"),
             "ppn": D("11"), "pph": D("2.5"), "excise": D("0"), "pdri": D("2.5"),
             "fta_agreement": "ATIGA/ACFTA", "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            {"hs_code": "9403.60", "description": "木製家具",
             "mfn_rate": D("10"), "atiga_rate": D("0"), "acfta_rate": D("0"),
             "ppn": D("11"), "pph": D("2.5"), "excise": D("0"), "pdri": D("2.5"),
             "fta_agreement": "ATIGA/ACFTA", "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            # Chapter 76 - Aluminum
            {"hs_code": "7606.12", "description": "鋁合金板及片",
             "mfn_rate": D("5"), "atiga_rate": D("0"), "acfta_rate": D("0"),
             "ppn": D("11"), "pph": D("2.5"), "excise": D("0"), "pdri": D("2.5"),
             "fta_agreement": "ATIGA/ACFTA", "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
            {"hs_code": "7604.10", "description": "鋁合金棒及條",
             "mfn_rate": D("5"), "atiga_rate": D("0"), "acfta_rate": D("0"),
             "ppn": D("11"), "pph": D("2.5"), "excise": D("0"), "pdri": D("2.5"),
             "fta_agreement": "ATIGA/ACFTA", "valid_from": "2024-01-01", "valid_to": "2026-12-31"},
        ]
        for item in data:
            self._indonesia_rates[item["hs_code"]] = item

    def _load_fta_agreements(self):
        """載入FTA協定參考資料"""
        self._fta_agreements = [
            {"code": "ATIGA", "name": "ASEAN Trade in Goods Agreement (東協貨物貿易協定)",
             "members": "ASEAN 10國", "condition": "原產地證明 Form D",
             "general_rate": "0%", "notes": "泰國/印尼互為ASEAN成員"},
            {"code": "ACFTA", "name": "ASEAN-China Free Trade Area (東協-中國自貿區)",
             "members": "ASEAN 10國 + 中國", "condition": "原產地證明 Form E",
             "general_rate": "0-5%", "notes": "自中國進口適用"},
            {"code": "JTEPA", "name": "Japan-Thailand Economic Partnership Agreement",
             "members": "日本 + 泰國", "condition": "原產地證明 Form JTEPA",
             "general_rate": "0-10%", "notes": "僅適用於泰國"},
            {"code": "AJCEP", "name": "ASEAN-Japan Comprehensive Economic Partnership",
             "members": "ASEAN + 日本", "condition": "原產地證明 Form AJ",
             "general_rate": "0-5%", "notes": "泰國/印尼均適用"},
            {"code": "AKFTA", "name": "ASEAN-Korea Free Trade Area",
             "members": "ASEAN + 韓國", "condition": "原產地證明 Form AK",
             "general_rate": "0-5%", "notes": "泰國/印尼均適用"},
            {"code": "AANZFTA", "name": "ASEAN-Australia-New Zealand FTA",
             "members": "ASEAN + 澳洲 + 紐西蘭", "condition": "原產地證明 Form AANZ",
             "general_rate": "0-5%", "notes": "泰國/印尼均適用"},
            {"code": "RCEP", "name": "Regional Comprehensive Economic Partnership (區域全面經濟夥伴協定)",
             "members": "ASEAN+中日韓澳紐", "condition": "原產地聲明或證明",
             "general_rate": "0-5%", "notes": "2022年起陸續生效"},
            {"code": "ITA", "name": "Information Technology Agreement (資訊科技協定)",
             "members": "WTO成員", "condition": "符合ITA產品範圍",
             "general_rate": "0%", "notes": "資訊科技產品免關稅"},
        ]

    # ── 查詢方法 ──

    def lookup_rate(self, hs_code: str, country: str) -> Optional[Dict[str, Any]]:
        """查詢指定HS編碼與國家的稅率"""
        country = country.lower().strip()
        if country in ("thailand", "th", "泰國"):
            result = self._thailand_rates.get(hs_code)
        elif country in ("indonesia", "id", "印尼"):
            result = self._indonesia_rates.get(hs_code)
        else:
            result = None

        if result is None:
            # 嘗試模糊匹配 (前4碼)
            prefix = hs_code.replace(".", "")[:4]
            db = self._thailand_rates if country in ("thailand", "th", "泰國") else self._indonesia_rates
            closest = None
            for code in db:
                if code.replace(".", "").startswith(prefix):
                    closest = db[code]
                    break
            if closest:
                return {"_closest_match": closest, "_original_query": hs_code}
        return result

    def search_by_keyword(self, keyword: str, country: str = "all") -> List[Dict[str, Any]]:
        """以關鍵字搜尋品名"""
        results = []
        keyword = keyword.lower()
        if country in ("all", "both"):
            dbs = [
                ("泰國", self._thailand_rates),
                ("印尼", self._indonesia_rates),
            ]
        elif country.lower() in ("thailand", "th", "泰國"):
            dbs = [("泰國", self._thailand_rates)]
        else:
            dbs = [("印尼", self._indonesia_rates)]

        for country_name, db in dbs:
            for hs_code, item in db.items():
                if keyword in item["description"].lower():
                    results.append({**item, "country": country_name})
        return results

    def get_fta_options(self, hs_code: str, country: str) -> List[Dict[str, str]]:
        """取得可用FTA協定選項"""
        rate_info = self.lookup_rate(hs_code, country)
        if not rate_info:
            return []
        fta_str = rate_info.get("fta_agreement", "")
        fta_codes = fta_str.split("/")
        options = []
        for fta in self._fta_agreements:
            if fta["code"] in fta_codes:
                options.append(fta)
        return options

    def update_rate(self, hs_code: str, country: str, new_rate: Dict[str, Any]) -> bool:
        """更新稅率"""
        country = country.lower().strip()
        if country in ("thailand", "th", "泰國"):
            if hs_code in self._thailand_rates:
                self._thailand_rates[hs_code].update(new_rate)
                return True
        elif country in ("indonesia", "id", "印尼"):
            if hs_code in self._indonesia_rates:
                self._indonesia_rates[hs_code].update(new_rate)
                return True
        return False

    def get_all_thailand(self) -> Dict[str, Dict[str, Any]]:
        return dict(self._thailand_rates)

    def get_all_indonesia(self) -> Dict[str, Dict[str, Any]]:
        return dict(self._indonesia_rates)

    def get_fta_agreements(self) -> List[Dict[str, str]]:
        return list(self._fta_agreements)


# ─────────────────────────────────────────────────────────
# 2. TaxCalculator - 核心計算引擎
# ─────────────────────────────────────────────────────────

class TaxCalculator:
    """稅費計算引擎"""

    def __init__(self, db: Optional[TaxRateDatabase] = None):
        self.db = db or TaxRateDatabase()

    @staticmethod
    def _pct(value: Decimal, rate: Decimal) -> Decimal:
        """百分比計算"""
        return (value * rate / HUNDRED).quantize(D("0.01"), rounding=ROUND_HALF_UP)

    def calculate_thailand(
        self,
        hs_code: str,
        cif_value: Decimal,
        use_fta: bool = False,
    ) -> Dict[str, Any]:
        """
        泰國進口稅費計算
        CIF = FOB + freight + insurance (若直接給CIF則直接使用)
        duty = CIF x duty_rate
        special_tax = (CIF + duty) x special_tax_rate
        VAT = (CIF + duty + special_tax) x 7%
        withholding_tax = CIF x withholding_rate
        total = duty + special_tax + VAT + withholding_tax
        """
        rate_info = self.db.lookup_rate(hs_code, "thailand")
        if rate_info is None:
            raise ValueError(f"找不到HS編碼 '{hs_code}' 的泰國稅率資料")
        if "_closest_match" in rate_info:
            closest = rate_info["_closest_match"]
            raise ValueError(
                f"找不到HS編碼 '{hs_code}' 的泰國稅率，最接近的編碼為: "
                f"{closest['hs_code']} ({closest['description']})"
            )

        cif = cif_value if isinstance(cif_value, Decimal) else D(str(cif_value))

        # 決定關稅率
        if use_fta:
            duty_rate = rate_info["fta_rate"]
            applied_rate_label = "FTA優惠稅率"
        else:
            duty_rate = rate_info["duty_rate"]
            applied_rate_label = "基本關稅率"

        # 計算各項稅費
        duty = self._pct(cif, duty_rate)
        special_tax_rate = rate_info["special_tax"]
        special_tax = self._pct(cif + duty, special_tax_rate)
        vat_rate = rate_info["vat"]
        vat = self._pct(cif + duty + special_tax, vat_rate)
        wht_rate = rate_info["withholding_tax"]
        withholding_tax = self._pct(cif, wht_rate)
        total_tax = duty + special_tax + vat + withholding_tax
        total_cost = cif + total_tax

        return {
            "country": "泰國",
            "hs_code": hs_code,
            "description": rate_info["description"],
            "cif": cif,
            "duty_rate": duty_rate,
            "duty_rate_label": applied_rate_label,
            "duty": duty,
            "special_tax_rate": special_tax_rate,
            "special_tax": special_tax,
            "vat_rate": vat_rate,
            "vat": vat,
            "withholding_tax_rate": wht_rate,
            "withholding_tax": withholding_tax,
            "total_tax": total_tax,
            "total_cost": total_cost,
            "use_fta": use_fta,
            "fta_agreement": rate_info["fta_agreement"],
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }

    def calculate_indonesia(
        self,
        hs_code: str,
        cif_value: Decimal,
        has_api: bool = True,
        use_fta: str = "MFN",
    ) -> Dict[str, Any]:
        """
        印尼進口稅費計算
        CIF = FOB + freight + insurance
        Bea_Masuk(duty) = CIF x duty_rate
        PPN(VAT) = (CIF + duty) x 11%
        PPh22 = (CIF + duty) x pph_rate (有API: 2.5%, 無API: 7.5%)
        excise = CIF x excise_rate (若適用)
        PDRI = (CIF + duty) x pdri_rate
        total = duty + PPN + PPh22 + excise + PDRI
        """
        rate_info = self.db.lookup_rate(hs_code, "indonesia")
        if rate_info is None:
            raise ValueError(f"找不到HS編碼 '{hs_code}' 的印尼稅率資料")
        if "_closest_match" in rate_info:
            closest = rate_info["_closest_match"]
            raise ValueError(
                f"找不到HS編碼 '{hs_code}' 的印尼稅率，最接近的編碼為: "
                f"{closest['hs_code']} ({closest['description']})"
            )

        cif = cif_value if isinstance(cif_value, Decimal) else D(str(cif_value))

        # 決定關稅率
        fta_upper = use_fta.upper().strip()
        if fta_upper in ("ACFTA", "中國-東協"):
            duty_rate = rate_info["acfta_rate"]
            applied_label = "ACFTA稅率"
        elif fta_upper in ("ATIGA", "東協"):
            duty_rate = rate_info["atiga_rate"]
            applied_label = "ATIGA稅率"
        else:
            duty_rate = rate_info["mfn_rate"]
            applied_label = "MFN關稅率"

        # 計算各項稅費
        duty = self._pct(cif, duty_rate)

        ppn_rate = rate_info["ppn"]
        ppn = self._pct(cif + duty, ppn_rate)

        # PPh22: 有API 2.5%, 無API 7.5%
        if has_api:
            pph_rate = rate_info["pph"]  # 通常 2.5%
        else:
            pph_rate = D("7.5")
        pph22 = self._pct(cif + duty, pph_rate)

        excise_rate = rate_info["excise"]
        excise = self._pct(cif, excise_rate)

        pdri_rate = rate_info["pdri"]
        pdri = self._pct(cif + duty, pdri_rate)

        total_tax = duty + ppn + pph22 + excise + pdri
        total_cost = cif + total_tax

        return {
            "country": "印尼",
            "hs_code": hs_code,
            "description": rate_info["description"],
            "cif": cif,
            "duty_rate": duty_rate,
            "duty_rate_label": applied_label,
            "duty": duty,
            "ppn_rate": ppn_rate,
            "ppn": ppn,
            "pph_rate": pph_rate,
            "pph22": pph22,
            "has_api": has_api,
            "excise_rate": excise_rate,
            "excise": excise,
            "pdri_rate": pdri_rate,
            "pdri": pdri,
            "total_tax": total_tax,
            "total_cost": total_cost,
            "use_fta": use_fta,
            "fta_agreement": rate_info["fta_agreement"],
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }

    def calculate_from_breakdown(
        self,
        fob: Decimal,
        freight: Decimal,
        insurance: Decimal,
        hs_code: str,
        country: str,
        **kwargs,
    ) -> Dict[str, Any]:
        """從FOB+運費+保險費計算CIF後進行完整計算"""
        fob = fob if isinstance(fob, Decimal) else D(str(fob))
        freight = freight if isinstance(freight, Decimal) else D(str(freight))
        insurance = insurance if isinstance(insurance, Decimal) else D(str(insurance))
        cif = fob + freight + insurance

        country_lower = country.lower().strip()
        result_base = {
            "fob": fob,
            "freight": freight,
            "insurance": insurance,
            "cif": cif,
        }

        if country_lower in ("thailand", "th", "泰國"):
            calc = self.calculate_thailand(hs_code, cif, **kwargs)
        else:
            calc = self.calculate_indonesia(hs_code, cif, **kwargs)

        calc.update(result_base)
        return calc

    def compare_scenarios(
        self,
        hs_code: str,
        cif_value: Decimal,
        country: str,
    ) -> Dict[str, Any]:
        """比較不同FTA情境下的稅費"""
        cif = cif_value if isinstance(cif_value, Decimal) else D(str(cif_value))
        country_lower = country.lower().strip()
        scenarios = []

        if country_lower in ("thailand", "th", "泰國"):
            # MFN (標準)
            s_mfn = self.calculate_thailand(hs_code, cif, use_fta=False)
            s_mfn["scenario_name"] = "MFN標準稅率"
            scenarios.append(s_mfn)

            # FTA
            s_fta = self.calculate_thailand(hs_code, cif, use_fta=True)
            s_fta["scenario_name"] = "FTA優惠稅率"
            scenarios.append(s_fta)
        else:
            # MFN
            s_mfn = self.calculate_indonesia(hs_code, cif, has_api=True, use_fta="MFN")
            s_mfn["scenario_name"] = "MFN標準稅率(有API)"
            scenarios.append(s_mfn)

            # MFN no API
            s_noapi = self.calculate_indonesia(hs_code, cif, has_api=False, use_fta="MFN")
            s_noapi["scenario_name"] = "MFN標準稅率(無API)"
            scenarios.append(s_noapi)

            # ATIGA
            s_atiga = self.calculate_indonesia(hs_code, cif, has_api=True, use_fta="ATIGA")
            s_atiga["scenario_name"] = "ATIGA優惠稅率"
            scenarios.append(s_atiga)

            # ACFTA
            s_acfta = self.calculate_indonesia(hs_code, cif, has_api=True, use_fta="ACFTA")
            s_acfta["scenario_name"] = "ACFTA優惠稅率"
            scenarios.append(s_acfta)

        # 找出最佳方案
        best = min(scenarios, key=lambda s: s["total_tax"])
        savings_vs_worst = max(s["total_tax"] for s in scenarios) - best["total_tax"]

        return {
            "hs_code": hs_code,
            "cif": cif,
            "country": "泰國" if country_lower in ("thailand", "th", "泰國") else "印尼",
            "scenarios": scenarios,
            "best_scenario": best["scenario_name"],
            "max_savings": savings_vs_worst,
        }

    def suggest_optimal(
        self,
        hs_code: str,
        cif_value: Decimal,
        country: str,
    ) -> Dict[str, Any]:
        """建議最佳稅務策略"""
        comparison = self.compare_scenarios(hs_code, cif_value, country)
        best = None
        for s in comparison["scenarios"]:
            if best is None or s["total_tax"] < best["total_tax"]:
                best = s

        suggestions = []
        if best["total_tax"] < comparison["scenarios"][0]["total_tax"]:
            saving = comparison["scenarios"][0]["total_tax"] - best["total_tax"]
            saving_pct = (saving / comparison["scenarios"][0]["total_tax"] * HUNDRED) if comparison["scenarios"][0]["total_tax"] > 0 else ZERO
            suggestions.append(f"建議使用 {best.get('scenario_name', best.get('duty_rate_label', 'N/A'))} 方案")
            suggestions.append(f"可節省 USD {fmt_money(saving)} (約 {fmt_pct(saving_pct)})")
        else:
            suggestions.append("所有方案稅額相同，使用標準稅率即可")

        # FTA申請建議
        fta_opts = self.db.get_fta_options(hs_code, country)
        if fta_opts:
            for fta in fta_opts:
                suggestions.append(f"可用FTA: {fta['code']} - {fta['name']}")
                suggestions.append(f"  條件: {fta['condition']}")

        return {
            "comparison": comparison,
            "optimal": best,
            "suggestions": suggestions,
            "fta_options": fta_opts,
        }


# ─────────────────────────────────────────────────────────
# 3. TaxReportGenerator - Excel報表產出
# ─────────────────────────────────────────────────────────

class TaxReportGenerator:
    """Excel報表產生器"""

    def __init__(self, db: Optional[TaxRateDatabase] = None):
        self.db = db or TaxRateDatabase()
        self.header_fill = PatternFill(start_color=DEEP_BLUE, end_color=DEEP_BLUE, fill_type="solid")
        self.header_font = Font(name="Microsoft JhengHei", size=11, bold=True, color=WHITE)
        self.data_font = Font(name="Microsoft JhengHei", size=10)
        self.title_font = Font(name="Microsoft JhengHei", size=14, bold=True, color=DEEP_BLUE)
        self.center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        self.right_align = Alignment(horizontal="right", vertical="center")

    def _apply_header_style(self, ws, row: int, cols: int):
        """套用表頭樣式"""
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_align
            cell.border = THIN_BORDER

    def _apply_data_style(self, ws, row: int, cols: int, money_cols: Optional[List[int]] = None):
        """套用資料列樣式"""
        money_cols = money_cols or []
        alt_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = self.data_font
            cell.border = THIN_BORDER
            if col in money_cols:
                cell.alignment = self.right_align
                if isinstance(cell.value, (int, float, Decimal)):
                    cell.number_format = '#,##0.00'
            else:
                cell.alignment = self.center_align
            if row % 2 == 0:
                cell.fill = alt_fill

    def generate_calculation_report(
        self,
        calc_result: Dict[str, Any],
        order_info: Optional[Dict[str, str]] = None,
        filename: str = "稅費計算報告.xlsx",
    ) -> str:
        """產生單筆計算Excel報表"""
        wb = Workbook()
        ws = wb.active
        ws.title = "稅費計算明細"

        # 標題
        ws.merge_cells("A1:F1")
        title_cell = ws["A1"]
        country_label = calc_result.get("country", "")
        title_cell.value = f"進口稅費計算明細 - {country_label}"
        title_cell.font = self.title_font
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 35

        # 訂單資訊
        row = 3
        if order_info:
            for key, val in order_info.items():
                ws.cell(row=row, column=1, value=key).font = Font(name="Microsoft JhengHei", size=10, bold=True)
                ws.cell(row=row, column=2, value=val).font = self.data_font
                row += 1
            row += 1

        # 基本資訊
        info_headers = ["項目", "內容"]
        for c, h in enumerate(info_headers, 1):
            ws.cell(row=row, column=c, value=h)
        self._apply_header_style(ws, row, 2)
        row += 1

        info_data = [
            ("HS編碼", calc_result.get("hs_code", "")),
            ("品名描述", calc_result.get("description", "")),
            ("CIF價值 (USD)", calc_result.get("cif", "")),
            ("適用稅率方案", calc_result.get("duty_rate_label", "")),
            ("可用FTA協定", calc_result.get("fta_agreement", "")),
            ("計算時間", calc_result.get("timestamp", "")),
        ]
        for label, val in info_data:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=str(val) if not isinstance(val, Decimal) else float(val))
            self._apply_data_style(ws, row, 2)
            row += 1

        row += 1
        # 稅費明細表頭
        tax_headers = ["稅費項目", "稅率(%)", "金額(USD)"]
        for c, h in enumerate(tax_headers, 1):
            ws.cell(row=row, column=c, value=h)
        self._apply_header_style(ws, row, 3)
        row += 1

        # 稅費項目
        tax_items = []
        tax_items.append(("關稅 (Duty)", calc_result.get("duty_rate", ZERO), calc_result.get("duty", ZERO)))

        if calc_result.get("country") == "泰國":
            tax_items.append(("特別稅 (Special Tax)", calc_result.get("special_tax_rate", ZERO), calc_result.get("special_tax", ZERO)))
            tax_items.append(("增值稅 VAT", calc_result.get("vat_rate", ZERO), calc_result.get("vat", ZERO)))
            tax_items.append(("預扣稅 (WHT)", calc_result.get("withholding_tax_rate", ZERO), calc_result.get("withholding_tax", ZERO)))
        else:
            tax_items.append(("增值稅 PPN", calc_result.get("ppn_rate", ZERO), calc_result.get("ppn", ZERO)))
            tax_items.append(("所得稅 PPh22", calc_result.get("pph_rate", ZERO), calc_result.get("pph22", ZERO)))
            tax_items.append(("消費稅 (Excise)", calc_result.get("excise_rate", ZERO), calc_result.get("excise", ZERO)))
            tax_items.append(("PDRI", calc_result.get("pdri_rate", ZERO), calc_result.get("pdri", ZERO)))

        for label, rate, amount in tax_items:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=float(rate))
            ws.cell(row=row, column=3, value=float(amount))
            self._apply_data_style(ws, row, 3, money_cols=[3])
            row += 1

        # 合計
        total_fill = PatternFill(start_color=DEEP_BLUE, end_color=DEEP_BLUE, fill_type="solid")
        total_font = Font(name="Microsoft JhengHei", size=11, bold=True, color=WHITE)
        for c, val in enumerate(["稅費合計", "", float(calc_result.get("total_tax", ZERO))], 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.fill = total_fill
            cell.font = total_font
            cell.border = THIN_BORDER
            if c == 3:
                cell.number_format = '#,##0.00'
                cell.alignment = self.right_align
        row += 1

        for c, val in enumerate(["CIF + 稅費總計", "", float(calc_result.get("total_cost", ZERO))], 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.fill = total_fill
            cell.font = total_font
            cell.border = THIN_BORDER
            if c == 3:
                cell.number_format = '#,##0.00'
                cell.alignment = self.right_align

        # 欄寬
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 20

        # 凍結
        ws.freeze_panes = "A3"

        wb.save(filename)
        return filename

    def generate_comparison_report(
        self,
        comparison: Dict[str, Any],
        filename: str = "稅費比較報告.xlsx",
    ) -> str:
        """產生情境比較Excel報表"""
        wb = Workbook()
        ws = wb.active
        ws.title = "稅費情境比較"

        scenarios = comparison.get("scenarios", [])
        if not scenarios:
            return filename

        # 標題
        num_cols = 2 + len(scenarios)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
        title_cell = ws["A1"]
        title_cell.value = f"進口稅費情境比較 - {comparison.get('country', '')} HS {comparison.get('hs_code', '')}"
        title_cell.font = self.title_font
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 35

        # 表頭
        row = 3
        headers = ["稅費項目", "稅率(%)"]
        for s in scenarios:
            headers.append(s.get("scenario_name", "情境"))
        for c, h in enumerate(headers, 1):
            ws.cell(row=row, column=c, value=h)
        self._apply_header_style(ws, row, len(headers))
        row += 1

        # 資料列
        is_thailand = comparison.get("country", "") == "泰國"
        if is_thailand:
            line_items = [
                ("關稅 (Duty)", "duty_rate", "duty"),
                ("特別稅", "special_tax_rate", "special_tax"),
                ("增值稅 VAT", "vat_rate", "vat"),
                ("預扣稅 WHT", "withholding_tax_rate", "withholding_tax"),
            ]
        else:
            line_items = [
                ("關稅 Bea Masuk", "duty_rate", "duty"),
                ("增值稅 PPN", "ppn_rate", "ppn"),
                ("所得稅 PPh22", "pph_rate", "pph22"),
                ("消費稅 Excise", "excise_rate", "excise"),
                ("PDRI", "pdri_rate", "pdri"),
            ]

        money_cols = list(range(3, len(headers) + 1))
        for label, rate_key, val_key in line_items:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=float(scenarios[0].get(rate_key, ZERO)))
            for i, s in enumerate(scenarios):
                ws.cell(row=row, column=3 + i, value=float(s.get(val_key, ZERO)))
            self._apply_data_style(ws, row, len(headers), money_cols=money_cols)
            row += 1

        # 合計列
        for c, val in enumerate(["稅費合計", ""], 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = THIN_BORDER
        for i, s in enumerate(scenarios):
            cell = ws.cell(row=row, column=3 + i, value=float(s.get("total_tax", ZERO)))
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = THIN_BORDER
            cell.number_format = '#,##0.00'
            cell.alignment = self.right_align
        row += 1

        # CIF + 總計
        for c, val in enumerate(["CIF + 稅費總計", ""], 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = THIN_BORDER
        for i, s in enumerate(scenarios):
            cell = ws.cell(row=row, column=3 + i, value=float(s.get("total_cost", ZERO)))
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = THIN_BORDER
            cell.number_format = '#,##0.00'
            cell.alignment = self.right_align
        row += 2

        # 最佳方案
        ws.cell(row=row, column=1, value="最佳方案:").font = Font(name="Microsoft JhengHei", size=11, bold=True, color="228B22")
        ws.cell(row=row, column=2, value=comparison.get("best_scenario", "")).font = Font(name="Microsoft JhengHei", size=11, bold=True, color="228B22")
        row += 1
        ws.cell(row=row, column=1, value="最大節省:").font = Font(name="Microsoft JhengHei", size=11, bold=True)
        ws.cell(row=row, column=2, value=float(comparison.get("max_savings", ZERO))).number_format = '#,##0.00'

        # 欄寬
        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 14
        for i in range(len(scenarios)):
            ws.column_dimensions[get_column_letter(3 + i)].width = 22

        ws.freeze_panes = "A4"

        wb.save(filename)
        return filename

    def generate_monthly_summary(
        self,
        calculations_list: List[Dict[str, Any]],
        filename: str = "月度稅費彙總.xlsx",
    ) -> str:
        """產生月度稅費彙總報表"""
        wb = Workbook()
        ws = wb.active
        ws.title = "月度稅費彙總"

        # 標題
        ws.merge_cells("A1:I1")
        ws["A1"].value = "月度進口稅費彙總報表"
        ws["A1"].font = self.title_font
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 35

        # 表頭
        row = 3
        headers = ["序號", "國家", "HS編碼", "品名描述", "CIF(USD)", "關稅(USD)",
                    "其他稅費(USD)", "稅費合計(USD)", "CIF+稅費總計(USD)"]
        for c, h in enumerate(headers, 1):
            ws.cell(row=row, column=c, value=h)
        self._apply_header_style(ws, row, len(headers))
        row += 1

        money_cols = [5, 6, 7, 8, 9]
        total_cif = ZERO
        total_duty = ZERO
        total_other = ZERO
        total_tax = ZERO
        total_cost = ZERO

        for idx, calc in enumerate(calculations_list, 1):
            duty = calc.get("duty", ZERO)
            total_tax_val = calc.get("total_tax", ZERO)
            other_tax = total_tax_val - duty
            cif = calc.get("cif", ZERO)

            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=calc.get("country", ""))
            ws.cell(row=row, column=3, value=calc.get("hs_code", ""))
            ws.cell(row=row, column=4, value=calc.get("description", ""))
            ws.cell(row=row, column=5, value=float(cif))
            ws.cell(row=row, column=6, value=float(duty))
            ws.cell(row=row, column=7, value=float(other_tax))
            ws.cell(row=row, column=8, value=float(total_tax_val))
            ws.cell(row=row, column=9, value=float(calc.get("total_cost", ZERO)))
            self._apply_data_style(ws, row, len(headers), money_cols=money_cols)

            total_cif += cif
            total_duty += duty
            total_other += other_tax
            total_tax += total_tax_val
            total_cost += calc.get("total_cost", ZERO)
            row += 1

        # 合計列
        total_data = ["", "合計", "", "", float(total_cif), float(total_duty),
                      float(total_other), float(total_tax), float(total_cost)]
        for c, val in enumerate(total_data, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = THIN_BORDER
            if c >= 5:
                cell.number_format = '#,##0.00'
                cell.alignment = self.right_align

        # 欄寬
        widths = [6, 8, 12, 30, 16, 16, 16, 16, 18]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = "A4"
        wb.save(filename)
        return filename

    def generate_fta_analysis(
        self,
        hs_codes: List[str],
        filename: str = "FTA分析報告.xlsx",
    ) -> str:
        """產生FTA利用率分析報表"""
        wb = Workbook()
        ws = wb.active
        ws.title = "FTA分析"

        # 標題
        ws.merge_cells("A1:H1")
        ws["A1"].value = "FTA優惠稅率利用分析"
        ws["A1"].font = self.title_font
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 35

        # 表頭
        row = 3
        headers = ["HS編碼", "品名描述", "國家", "基本/MFN稅率(%)", "FTA最優稅率(%)",
                    "可用FTA協定", "節稅潛力(每萬USD)", "建議"]
        for c, h in enumerate(headers, 1):
            ws.cell(row=row, column=c, value=h)
        self._apply_header_style(ws, row, len(headers))
        row += 1

        money_cols = [7]
        for hs_code in hs_codes:
            for country in ("thailand", "indonesia"):
                rate_info = self.db.lookup_rate(hs_code, country)
                if rate_info is None or "_closest_match" in rate_info:
                    continue

                country_name = "泰國" if country == "thailand" else "印尼"
                if country == "thailand":
                    base_rate = rate_info["duty_rate"]
                    fta_rate = rate_info["fta_rate"]
                else:
                    base_rate = rate_info["mfn_rate"]
                    fta_rate = min(rate_info["atiga_rate"], rate_info["acfta_rate"])

                saving_per_10k = (D("10000") * (base_rate - fta_rate) / HUNDRED).quantize(D("0.01"), rounding=ROUND_HALF_UP)
                fta_opts = self.db.get_fta_options(hs_code, country)
                fta_names = "/".join([f["code"] for f in fta_opts])

                suggestion = "建議申請FTA原產地證明" if base_rate > fta_rate else "無需FTA即可享優惠"

                ws.cell(row=row, column=1, value=hs_code)
                ws.cell(row=row, column=2, value=rate_info["description"])
                ws.cell(row=row, column=3, value=country_name)
                ws.cell(row=row, column=4, value=float(base_rate))
                ws.cell(row=row, column=5, value=float(fta_rate))
                ws.cell(row=row, column=6, value=fta_names)
                ws.cell(row=row, column=7, value=float(saving_per_10k))
                ws.cell(row=row, column=8, value=suggestion)
                self._apply_data_style(ws, row, len(headers), money_cols=money_cols)
                row += 1

        # 欄寬
        widths = [12, 30, 8, 16, 16, 24, 18, 24]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = "A4"
        wb.save(filename)
        return filename

    def generate_database_workbook(self, filename: str = "稅率數據庫.xlsx") -> str:
        """產生完整稅率資料庫Excel"""
        wb = Workbook()

        # ── Sheet 1: 泰國稅率 ──
        ws_th = wb.active
        ws_th.title = "泰國稅率數據庫"
        th_headers = ["HS編碼", "品名描述", "基本關稅率(%)", "FTA優惠稅率(%)",
                       "特別稅(%)", "VAT(%)", "預扣稅(%)", "適用FTA協定", "有效期"]
        for c, h in enumerate(th_headers, 1):
            ws_th.cell(row=1, column=c, value=h)
        self._apply_header_style(ws_th, 1, len(th_headers))

        row = 2
        for hs_code, item in sorted(self.db.get_all_thailand().items()):
            ws_th.cell(row=row, column=1, value=item["hs_code"])
            ws_th.cell(row=row, column=2, value=item["description"])
            ws_th.cell(row=row, column=3, value=float(item["duty_rate"]))
            ws_th.cell(row=row, column=4, value=float(item["fta_rate"]))
            ws_th.cell(row=row, column=5, value=float(item["special_tax"]))
            ws_th.cell(row=row, column=6, value=float(item["vat"]))
            ws_th.cell(row=row, column=7, value=float(item["withholding_tax"]))
            ws_th.cell(row=row, column=8, value=item["fta_agreement"])
            ws_th.cell(row=row, column=9, value=f"{item['valid_from']} ~ {item['valid_to']}")
            self._apply_data_style(ws_th, row, len(th_headers))
            row += 1

        th_widths = [12, 32, 14, 14, 12, 10, 12, 24, 24]
        for i, w in enumerate(th_widths, 1):
            ws_th.column_dimensions[get_column_letter(i)].width = w
        ws_th.freeze_panes = "A2"
        ws_th.auto_filter.ref = f"A1:{get_column_letter(len(th_headers))}{row - 1}"

        # ── Sheet 2: 印尼稅率 ──
        ws_id = wb.create_sheet("印尼稅率數據庫")
        id_headers = ["HS編碼", "品名描述", "MFN關稅率(%)", "ATIGA稅率(%)",
                       "ACFTA稅率(%)", "PPN增值稅(%)", "PPh所得稅(%)", "消費稅(%)",
                       "PDRI(%)", "適用協定", "有效期"]
        for c, h in enumerate(id_headers, 1):
            ws_id.cell(row=1, column=c, value=h)
        self._apply_header_style(ws_id, 1, len(id_headers))

        row = 2
        for hs_code, item in sorted(self.db.get_all_indonesia().items()):
            ws_id.cell(row=row, column=1, value=item["hs_code"])
            ws_id.cell(row=row, column=2, value=item["description"])
            ws_id.cell(row=row, column=3, value=float(item["mfn_rate"]))
            ws_id.cell(row=row, column=4, value=float(item["atiga_rate"]))
            ws_id.cell(row=row, column=5, value=float(item["acfta_rate"]))
            ws_id.cell(row=row, column=6, value=float(item["ppn"]))
            ws_id.cell(row=row, column=7, value=float(item["pph"]))
            ws_id.cell(row=row, column=8, value=float(item["excise"]))
            ws_id.cell(row=row, column=9, value=float(item["pdri"]))
            ws_id.cell(row=row, column=10, value=item["fta_agreement"])
            ws_id.cell(row=row, column=11, value=f"{item['valid_from']} ~ {item['valid_to']}")
            self._apply_data_style(ws_id, row, len(id_headers))
            row += 1

        id_widths = [12, 32, 14, 14, 14, 14, 14, 12, 10, 24, 24]
        for i, w in enumerate(id_widths, 1):
            ws_id.column_dimensions[get_column_letter(i)].width = w
        ws_id.freeze_panes = "A2"
        ws_id.auto_filter.ref = f"A1:{get_column_letter(len(id_headers))}{row - 1}"

        # ── Sheet 3: FTA協定參考 ──
        ws_fta = wb.create_sheet("FTA協定參考")
        fta_headers = ["協定代碼", "協定名稱", "成員國", "原產地證明條件",
                        "一般優惠稅率", "備註說明"]
        for c, h in enumerate(fta_headers, 1):
            ws_fta.cell(row=1, column=c, value=h)
        self._apply_header_style(ws_fta, 1, len(fta_headers))

        row = 2
        for fta in self.db.get_fta_agreements():
            ws_fta.cell(row=row, column=1, value=fta["code"])
            ws_fta.cell(row=row, column=2, value=fta["name"])
            ws_fta.cell(row=row, column=3, value=fta["members"])
            ws_fta.cell(row=row, column=4, value=fta["condition"])
            ws_fta.cell(row=row, column=5, value=fta["general_rate"])
            ws_fta.cell(row=row, column=6, value=fta["notes"])
            self._apply_data_style(ws_fta, row, len(fta_headers))
            row += 1

        fta_widths = [14, 50, 24, 24, 14, 30]
        for i, w in enumerate(fta_widths, 1):
            ws_fta.column_dimensions[get_column_letter(i)].width = w
        ws_fta.freeze_panes = "A2"
        ws_fta.auto_filter.ref = f"A1:{get_column_letter(len(fta_headers))}{row - 1}"

        wb.save(filename)
        return filename


# ─────────────────────────────────────────────────────────
# 4. TaxQueryTool - CLI命令列工具
# ─────────────────────────────────────────────────────────

class TaxQueryTool:
    """CLI命令列查詢工具"""

    def __init__(self):
        self.db = TaxRateDatabase()
        self.calc = TaxCalculator(self.db)
        self.report = TaxReportGenerator(self.db)

    def _print_calc_result(self, result: Dict[str, Any]):
        """格式化輸出計算結果"""
        country = result.get("country", "")
        lines = []
        lines.append(f"HS編碼: {result['hs_code']}  {result['description']}")
        lines.append(f"適用地區: {country}")
        lines.append(f"適用方案: {result.get('duty_rate_label', 'N/A')}")
        lines.append(f"可用FTA:  {result.get('fta_agreement', 'N/A')}")
        lines.append("─" * 66)

        if "fob" in result:
            lines.append(f"FOB價值:          USD {fmt_money(result['fob']):>16}")
            lines.append(f"運費 (Freight):   USD {fmt_money(result['freight']):>16}")
            lines.append(f"保險費:           USD {fmt_money(result['insurance']):>16}")
            lines.append(f"CIF價值:          USD {fmt_money(result['cif']):>16}")
        else:
            lines.append(f"CIF價值:          USD {fmt_money(result['cif']):>16}")

        lines.append("─" * 66)

        duty_rate = result.get("duty_rate", ZERO)
        duty = result.get("duty", ZERO)
        lines.append(f"關稅 ({fmt_pct(duty_rate):>6}):      USD {fmt_money(duty):>16}")

        if country == "泰國":
            st_rate = result.get("special_tax_rate", ZERO)
            st = result.get("special_tax", ZERO)
            lines.append(f"特別稅 ({fmt_pct(st_rate):>6}):  USD {fmt_money(st):>16}")

            vat_rate = result.get("vat_rate", ZERO)
            vat = result.get("vat", ZERO)
            lines.append(f"VAT ({fmt_pct(vat_rate):>6}):       USD {fmt_money(vat):>16}")

            wht_rate = result.get("withholding_tax_rate", ZERO)
            wht = result.get("withholding_tax", ZERO)
            lines.append(f"預扣稅 ({fmt_pct(wht_rate):>6}):   USD {fmt_money(wht):>16}")
        else:
            ppn_rate = result.get("ppn_rate", ZERO)
            ppn = result.get("ppn", ZERO)
            lines.append(f"PPN增值稅 ({fmt_pct(ppn_rate):>6}): USD {fmt_money(ppn):>16}")

            pph_rate = result.get("pph_rate", ZERO)
            pph = result.get("pph22", ZERO)
            api_label = "有API" if result.get("has_api", True) else "無API"
            lines.append(f"PPh22 ({fmt_pct(pph_rate):>6}, {api_label}): USD {fmt_money(pph):>16}")

            ex_rate = result.get("excise_rate", ZERO)
            ex = result.get("excise", ZERO)
            lines.append(f"消費稅 ({fmt_pct(ex_rate):>6}):  USD {fmt_money(ex):>16}")

            pdri_rate = result.get("pdri_rate", ZERO)
            pdri = result.get("pdri", ZERO)
            lines.append(f"PDRI ({fmt_pct(pdri_rate):>6}):      USD {fmt_money(pdri):>16}")

        lines.append("═" * 66)
        total_tax = result.get("total_tax", ZERO)
        total_cost = result.get("total_cost", ZERO)
        lines.append(f"稅費合計:         USD {fmt_money(total_tax):>16}")
        lines.append(f"CIF + 稅費總計:  USD {fmt_money(total_cost):>16}")

        title = f" {country}進口稅費計算結果 "
        print(draw_box(title, lines, width=70))

    def _print_comparison(self, comparison: Dict[str, Any]):
        """格式化輸出比較結果"""
        lines = []
        lines.append(f"HS編碼: {comparison['hs_code']}   CIF: USD {fmt_money(comparison['cif'])}")
        lines.append(f"國家:   {comparison['country']}")
        lines.append("─" * 66)

        for s in comparison["scenarios"]:
            lines.append(f"  [{s['scenario_name']}]")
            lines.append(f"    關稅:   USD {fmt_money(s['duty']):>12}")
            if s.get("country") == "泰國":
                lines.append(f"    特別稅: USD {fmt_money(s.get('special_tax', ZERO)):>12}")
                lines.append(f"    VAT:    USD {fmt_money(s.get('vat', ZERO)):>12}")
                lines.append(f"    預扣稅: USD {fmt_money(s.get('withholding_tax', ZERO)):>12}")
            else:
                lines.append(f"    PPN:    USD {fmt_money(s.get('ppn', ZERO)):>12}")
                lines.append(f"    PPh22:  USD {fmt_money(s.get('pph22', ZERO)):>12}")
                lines.append(f"    消費稅: USD {fmt_money(s.get('excise', ZERO)):>12}")
                lines.append(f"    PDRI:   USD {fmt_money(s.get('pdri', ZERO)):>12}")
            lines.append(f"    稅費合計: USD {fmt_money(s['total_tax']):>10}")
            lines.append("")

        lines.append("═" * 66)
        lines.append(f"  最佳方案: {comparison['best_scenario']}")
        lines.append(f"  最大節省: USD {fmt_money(comparison['max_savings'])}")

        title = f" 稅費情境比較 - {comparison['country']} "
        print(draw_double_box(title, lines, width=70))

    def _print_suggest(self, suggestion: Dict[str, Any]):
        """輸出最佳建議"""
        lines = []
        lines.append("稅務策略建議:")
        lines.append("")
        for s in suggestion["suggestions"]:
            lines.append(f"  {s}")
        if suggestion.get("fta_options"):
            lines.append("")
            lines.append("可用FTA協定:")
            for fta in suggestion["fta_options"]:
                lines.append(f"  {fta['code']}: {fta['name']}")
                lines.append(f"    成員: {fta['members']}")
                lines.append(f"    條件: {fta['condition']}")

        title = " 最佳稅務策略建議 "
        print(draw_double_box(title, lines, width=70))

    def do_calculate(self, args):
        """執行計算"""
        hs_code = args.hs_code
        cif = D(str(args.cif))
        country = args.country.lower()

        try:
            if country in ("thailand", "th"):
                use_fta = getattr(args, "fta", False)
                result = self.calc.calculate_thailand(hs_code, cif, use_fta=use_fta)
            else:
                fta_type = getattr(args, "fta", "MFN") or "MFN"
                has_api = not getattr(args, "no_api", False)
                result = self.calc.calculate_indonesia(hs_code, cif, has_api=has_api, use_fta=fta_type)

            self._print_calc_result(result)

            if getattr(args, "report", False):
                report_file = self.report.generate_calculation_report(result)
                print(f"\n已產生Excel報表: {report_file}")
        except ValueError as e:
            print(f"錯誤: {e}")

    def do_compare(self, args):
        """執行比較"""
        hs_code = args.hs_code
        cif = D(str(args.cif))
        country = args.country.lower()

        try:
            comparison = self.calc.compare_scenarios(hs_code, cif, country)
            self._print_comparison(comparison)

            if getattr(args, "report", False):
                report_file = self.report.generate_comparison_report(comparison)
                print(f"\n已產生Excel報表: {report_file}")
        except ValueError as e:
            print(f"錯誤: {e}")

    def do_lookup(self, args):
        """查詢稅率"""
        hs_code = args.hs_code
        country = args.country.lower()
        rate = self.db.lookup_rate(hs_code, country)

        if rate is None:
            print(f"找不到 HS編碼 '{hs_code}' 在 {country} 的稅率資料")
            return
        if "_closest_match" in rate:
            closest = rate["_closest_match"]
            print(f"找不到 '{hs_code}'，最接近的編碼為: {closest['hs_code']} ({closest['description']})")
            rate = closest

        country_name = "泰國" if country in ("thailand", "th") else "印尼"
        lines = []
        lines.append(f"HS編碼: {rate['hs_code']}")
        lines.append(f"品名:   {rate['description']}")
        lines.append(f"國家:   {country_name}")
        lines.append(f"適用FTA: {rate['fta_agreement']}")
        lines.append(f"有效期: {rate['valid_from']} ~ {rate['valid_to']}")
        lines.append("─" * 60)

        if country_name == "泰國":
            lines.append(f"基本關稅率: {fmt_pct(rate['duty_rate'])}")
            lines.append(f"FTA優惠稅率: {fmt_pct(rate['fta_rate'])}")
            lines.append(f"特別稅: {fmt_pct(rate['special_tax'])}")
            lines.append(f"VAT: {fmt_pct(rate['vat'])}")
            lines.append(f"預扣稅: {fmt_pct(rate['withholding_tax'])}")
        else:
            lines.append(f"MFN關稅率: {fmt_pct(rate['mfn_rate'])}")
            lines.append(f"ATIGA稅率: {fmt_pct(rate['atiga_rate'])}")
            lines.append(f"ACFTA稅率: {fmt_pct(rate['acfta_rate'])}")
            lines.append(f"PPN增值稅: {fmt_pct(rate['ppn'])}")
            lines.append(f"PPh所得稅(有API): {fmt_pct(rate['pph'])}")
            lines.append(f"消費稅: {fmt_pct(rate['excise'])}")
            lines.append(f"PDRI: {fmt_pct(rate['pdri'])}")

        title = " 稅率查詢結果 "
        print(draw_box(title, lines, width=64))

    def do_search(self, args):
        """關鍵字搜尋"""
        keyword = args.keyword
        country = getattr(args, "country", "all") or "all"
        results = self.db.search_by_keyword(keyword, country)

        if not results:
            print(f"未找到包含 '{keyword}' 的品項")
            return

        lines = [f"搜尋關鍵字: '{keyword}'  找到 {len(results)} 筆結果"]
        lines.append("─" * 60)
        for r in results:
            lines.append(f"[{r['country']}] {r['hs_code']} - {r['description']}")

        title = " 品名搜尋結果 "
        print(draw_box(title, lines, width=64))

    def do_suggest(self, args):
        """最佳策略建議"""
        hs_code = args.hs_code
        cif = D(str(args.cif))
        country = args.country.lower()

        try:
            suggestion = self.calc.suggest_optimal(hs_code, cif, country)
            self._print_suggest(suggestion)
        except ValueError as e:
            print(f"錯誤: {e}")

    def do_report(self, args):
        """從計算結果產生報表"""
        # 預設計算範例並產出報表
        calc_results = []
        demo_items = [
            ("8471.30", D("50000"), "thailand", True),
            ("8544.42", D("12000"), "thailand", False),
            ("7308.90", D("85000"), "thailand", False),
            ("8471.30", D("50000"), "indonesia", "ACFTA"),
            ("3923.21", D("30000"), "indonesia", "MFN"),
        ]
        for hs, cif, country, fta_opt in demo_items:
            if country == "thailand":
                r = self.calc.calculate_thailand(hs, cif, use_fta=fta_opt)
            else:
                has_api = True
                r = self.calc.calculate_indonesia(hs, cif, has_api=has_api, use_fta=fta_opt)
            calc_results.append(r)

        filename = self.report.generate_monthly_summary(calc_results, args.report_file)
        print(f"已產生月度彙總報表: {filename}")

        # 也產生 FTA 分析
        hs_codes = [item[0] for item in demo_items]
        fta_file = self.report.generate_fta_analysis(hs_codes, "FTA分析報告.xlsx")
        print(f"已產生FTA分析報表: {fta_file}")

    def do_batch(self, args):
        """批次計算"""
        filepath = args.batch_file
        if not os.path.exists(filepath):
            print(f"錯誤: 找不到檔案 '{filepath}'")
            return

        try:
            wb = load_workbook(filepath)
            ws = wb.active
            results = []
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                hs = str(row_dict.get("HS編碼", row_dict.get("hs_code", "")))
                cif = D(str(row_dict.get("CIF", row_dict.get("cif", "0"))))
                country = str(row_dict.get("國家", row_dict.get("country", "thailand"))).lower()
                if country in ("thailand", "th", "泰國"):
                    r = self.calc.calculate_thailand(hs, cif)
                else:
                    r = self.calc.calculate_indonesia(hs, cif)
                results.append(r)

            filename = self.report.generate_monthly_summary(results, "批次計算結果.xlsx")
            print(f"批次計算完成，共 {len(results)} 筆")
            print(f"結果已輸出至: {filename}")
        except Exception as e:
            print(f"批次處理錯誤: {e}")

    def do_demo(self, args):
        """執行5個示範情境"""
        print()
        print("=" * 70)
        print("       稅費自動核算引擎 v1.0 - 示範情境")
        print("       Thailand & Indonesia Import Duty Calculator")
        print("=" * 70)

        # ── 情境 1: 泰國進口筆記型電腦 ACFTA ──
        print("\n\n【情境 1】泰國進口 HS 8471.30 筆記型電腦 | CIF USD 50,000 | ACFTA優惠稅率")
        print()
        r1 = self.calc.calculate_thailand("8471.30", D("50000"), use_fta=True)
        self._print_calc_result(r1)

        # 與標準稅率比較
        r1_mfn = self.calc.calculate_thailand("8471.30", D("50000"), use_fta=False)
        saving1 = r1_mfn["total_tax"] - r1["total_tax"]
        if saving1 > 0:
            print(f"\n  >> ACFTA方案較標準稅率節省: USD {fmt_money(saving1)}")
        else:
            print(f"\n  >> ACFTA方案與標準稅率稅額相同 (已為最優)")

        # ── 情境 2: 泰國進口電纜 標準稅率 ──
        print("\n\n【情境 2】泰國進口 HS 8544.42 電纜及導體 | CIF USD 12,000 | 標準稅率")
        print()
        r2 = self.calc.calculate_thailand("8544.42", D("12000"), use_fta=False)
        self._print_calc_result(r2)

        # ── 情境 3: 泰國進口鋼鐵結構體 比較所有FTA ──
        print("\n\n【情境 3】泰國進口 HS 7308.90 鋼鐵結構體 | CIF USD 85,000 | 比較所有FTA方案")
        print()
        c3 = self.calc.compare_scenarios("7308.90", D("85000"), "thailand")
        self._print_comparison(c3)

        # ── 情境 4: 印尼進口筆記型電腦 有API MFN vs ACFTA ──
        print("\n\n【情境 4】印尼進口 HS 8471.30 筆記型電腦 | CIF USD 50,000 | 有API | MFN vs ACFTA")
        print()
        c4 = self.calc.compare_scenarios("8471.30", D("50000"), "indonesia")
        self._print_comparison(c4)

        # ── 情境 5: 印尼進口塑膠袋 無API ──
        print("\n\n【情境 5】印尼進口 HS 3923.21 聚乙烯塑膠袋 | CIF USD 30,000 | 無API")
        print()
        r5 = self.calc.calculate_indonesia("3923.21", D("30000"), has_api=False, use_fta="MFN")
        self._print_calc_result(r5)

        # 最佳建議
        print("\n\n【最佳策略建議】HS 3923.21 印尼進口")
        s5 = self.calc.suggest_optimal("3923.21", D("30000"), "indonesia")
        self._print_suggest(s5)

        # ── 產生 Excel 報表 ──
        print("\n\n" + "=" * 70)
        print("  正在產生 Excel 報表...")
        print("=" * 70)

        # 1. 單筆計算報表
        r1_report = self.report.generate_calculation_report(
            r1,
            order_info={"訂單編號": "PO-2026-001", "供應商": "深圳科技有限公司", "品名": "筆記型電腦"},
            filename="demo_泰國_筆記型電腦_稅費計算.xlsx",
        )
        print(f"  [1] {r1_report}")

        # 2. 比較報表
        c3_report = self.report.generate_comparison_report(c3, filename="demo_泰國_鋼鐵結構體_比較.xlsx")
        print(f"  [2] {c3_report}")

        # 3. 月度彙總
        all_results = [r1, r2, c3["scenarios"][0], c4["scenarios"][0], r5]
        m_report = self.report.generate_monthly_summary(all_results, filename="demo_月度稅費彙總.xlsx")
        print(f"  [3] {m_report}")

        # 4. FTA分析
        fta_report = self.report.generate_fta_analysis(
            ["8471.30", "8544.42", "7308.90", "3923.21"],
            filename="demo_FTA分析報告.xlsx",
        )
        print(f"  [4] {fta_report}")

        # 5. 完整稅率資料庫
        db_report = self.report.generate_database_workbook(filename="稅率數據庫.xlsx")
        print(f"  [5] {db_report}")

        print("\n  所有報表已成功產生!")
        print("=" * 70)


# ─────────────────────────────────────────────────────────
# CLI 入口
# ─────────────────────────────────────────────────────────

def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="稅費自動核算引擎 v1.0 - 泰國/印尼進口關稅計算",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用範例:
  python3 稅費核算引擎.py --demo
  python3 稅費核算引擎.py --calculate --hs-code 8471.30 --cif 50000 --country thailand --fta
  python3 稅費核算引擎.py --compare --hs-code 7308.90 --cif 85000 --country thailand
  python3 稅費核算引擎.py --lookup --hs-code 8471.30 --country indonesia
  python3 稅費核算引擎.py --search --keyword 電腦 --country thailand
  python3 稅費核算引擎.py --suggest --hs-code 8471.30 --cif 50000 --country thailand
  python3 稅費核算引擎.py --report 月度彙總.xlsx
  python3 稅費核算引擎.py --batch input.xlsx
        """,
    )

    parser.add_argument("--calculate", action="store_true", help="互動式稅費計算")
    parser.add_argument("--compare", action="store_true", help="比較多個FTA情境")
    parser.add_argument("--lookup", action="store_true", help="查詢HS編碼稅率")
    parser.add_argument("--search", action="store_true", help="以品名關鍵字搜尋")
    parser.add_argument("--suggest", action="store_true", help="建議最佳稅務策略")
    parser.add_argument("--report", nargs="?", const="月度彙總.xlsx", help="產生Excel報表")
    parser.add_argument("--batch", nargs="?", const="batch_input.xlsx", help="批次計算")
    parser.add_argument("--demo", action="store_true", help="執行5個示範情境")

    parser.add_argument("--hs-code", type=str, help="HS編碼 (如 8471.30)")
    parser.add_argument("--keyword", type=str, help="品名關鍵字 (用於 --search)")
    parser.add_argument("--cif", type=str, help="CIF價值 (USD)")
    parser.add_argument("--country", type=str, help="國家 (thailand/indonesia)")
    parser.add_argument("--fta", nargs="?", const="ACFTA", default=None, help="FTA類型 (ACFTA/ATIGA/MFN)")
    parser.add_argument("--no-api", action="store_true", help="無API (印尼)")
    parser.add_argument("--report-file", type=str, default="月度彙總.xlsx", help="報表輸出檔名")

    return parser


def main():
    parser = build_parser()
    args = parser.parse_args()
    tool = TaxQueryTool()

    if args.demo:
        tool.do_demo(args)
    elif args.calculate:
        if not args.hs_code or not args.cif or not args.country:
            parser.error("--calculate 需要 --hs-code, --cif, --country")
        tool.do_calculate(args)
    elif args.compare:
        if not args.hs_code or not args.cif or not args.country:
            parser.error("--compare 需要 --hs-code, --cif, --country")
        tool.do_compare(args)
    elif args.lookup:
        if not args.hs_code or not args.country:
            parser.error("--lookup 需要 --hs-code, --country")
        tool.do_lookup(args)
    elif args.search:
        if not args.keyword:
            parser.error("--search 需要 --keyword")
        tool.do_search(args)
    elif args.suggest:
        if not args.hs_code or not args.cif or not args.country:
            parser.error("--suggest 需要 --hs-code, --cif, --country")
        tool.do_suggest(args)
    elif args.report:
        args.report_file = args.report if args.report != "月度彙總.xlsx" else "月度彙總.xlsx"
        tool.do_report(args)
    elif args.batch:
        args.batch_file = args.batch
        tool.do_batch(args)
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
