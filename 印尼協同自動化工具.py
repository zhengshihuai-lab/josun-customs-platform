#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
印尼市場協同自動化工具
Indonesia Market Collaboration Automation Tool
===============================================
用於泰國/印尼進出口海關部門的協同作業自動化工具。
功能包括：訂單進度追蹤、雙語訊息範本、術語翻譯、週報產生。

Author: Qoder Auto-Generated
Version: 1.0.0
"""

import argparse
import json
import os
import sys
from datetime import datetime, timedelta
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference, PieChart
except ImportError:
    print("[錯誤] 缺少 openpyxl 套件，請執行: pip install openpyxl")
    sys.exit(1)

# ============================================================
# 全域常數
# ============================================================

STATUS_PIPELINE = [
    "生產中",
    "已出貨",
    "海上運輸中",
    "已到港",
    "清關中",
    "已放行",
    "已送達",
]

STATUS_EN = {
    "生產中": "In Production",
    "已出貨": "Shipped",
    "海上運輸中": "In Transit (Sea)",
    "已到港": "Arrived at Port",
    "清關中": "Customs Clearance",
    "已放行": "Released",
    "已送達": "Delivered",
}

STATUS_ID = {
    "生產中": "Sedang Diproduksi",
    "已出貨": "Telah Dikirim",
    "海上運輸中": "Dalam Pengiriman Laut",
    "已到港": "Tiba di Pelabuhan",
    "清關中": "Proses Bea Cukai",
    "已放行": "Telah Dilepaskan",
    "已送達": "Telah Diterima",
}

# Status overdue thresholds (days)
STATUS_OVERDUE_DAYS = {
    "生產中": 14,
    "已出貨": 5,
    "海上運輸中": 21,
    "已到港": 5,
    "清關中": 5,
    "已放行": 3,
    "已送達": 999,
}

# Deep blue header style
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Microsoft JhengHei", size=11, bold=True, color="FFFFFF")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBHEADER_FONT = Font(name="Microsoft JhengHei", size=10, bold=True, color="1F4E79")
NORMAL_FONT = Font(name="Microsoft JhengHei", size=10)
ALERT_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ALERT_FONT = Font(name="Microsoft JhengHei", size=10, color="9C0006")
OK_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
OK_FONT = Font(name="Microsoft JhengHei", size=10, color="006100")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# ============================================================
# Indonesia / Thailand Holiday Calendar 2026
# ============================================================

HOLIDAYS_2026 = {
    # Indonesia
    "2026-01-01": ("元旦", "Tahun Baru Masehi", "ID"),
    "2026-01-27": ("先知穆罕默德升天日", "Isra Mi'raj Nabi Muhammad", "ID"),
    "2026-02-08": ("農曆新年", "Tahun Baru Imlek", "ID"),
    "2026-03-19": ("靜居日", "Hari Raya Nyepi", "ID"),
    "2026-03-20": ("開齋節", "Hari Raya Idul Fitri", "ID"),
    "2026-03-21": ("開齋節次日", "Hari Raya Idul Fitri (Hari ke-2)", "ID"),
    "2026-04-03": ("耶穌受難日", "Wafat Isa Al Masih", "ID"),
    "2026-05-01": ("勞動節", "Hari Buruh Internasional", "ID"),
    "2026-05-14": ("佛陀誕辰", "Hari Raya Waisak", "ID"),
    "2026-05-27": ("先知穆罕默德誕辰", "Maulid Nabi Muhammad", "ID"),
    "2026-06-01": ("潘查希拉日", "Hari Lahir Pancasila", "ID"),
    "2026-05-28": ("耶穌升天日", "Kenaikan Isa Al Masih", "ID"),
    "2026-06-08": ("宰牲節", "Hari Raya Idul Adha", "ID"),
    "2026-06-29": ("伊斯蘭新年", "Tahun Baru Islam", "ID"),
    "2026-08-17": ("印尼獨立日", "Hari Kemerdekaan RI", "ID"),
    "2026-09-05": ("先知穆罕默德升天日", "Isra Mi'raj Nabi Muhammad", "ID"),
    "2026-12-25": ("聖誕節", "Hari Raya Natal", "ID"),
    # Thailand
    "2026-01-01b": ("元旦", "วันขึ้นปีใหม่", "TH"),
    "2026-02-12": ("萬佛節", "วันมาฆบูชา", "TH"),
    "2026-04-06": ("卻克里王朝紀念日", "วันจักรี", "TH"),
    "2026-04-13": ("潑水節", "วันสงกรานต์", "TH"),
    "2026-04-14": ("潑水節次日", "วันสงกรานต์ (วันที่ 2)", "TH"),
    "2026-04-15": ("潑水節第三天", "วันสงกรานต์ (วันที่ 3)", "TH"),
    "2026-05-04": ("泰王加冕日", "วันฉัตรมงคล", "TH"),
    "2026-05-11": ("皇家犁耕節", "วันพืชมงคล", "TH"),
    "2026-05-31": ("衛塞節", "วันวิสาขบูชา", "TH"),
    "2026-07-10": ("三寶佛節", "วันอาสาฬหบูชา", "TH"),
    "2026-07-11": ("守夏節", "วันเข้าพรรษา", "TH"),
    "2026-07-28": ("泰王誕辰", "วันเฉลิมพระชนมพรรษา", "TH"),
    "2026-08-12": ("王太后誕辰/母親節", "วันเฉลิมพระชนมพรรษาพระบรมราชินีนาถ", "TH"),
    "2026-10-13": ("已故泰王紀念日", "วันคล้ายวันสวรรคต", "TH"),
    "2026-10-23": ("朱拉隆功大帝紀念日", "วันปิยมหาราช", "TH"),
    "2026-12-05": ("泰王父親節", "วันคล้ายวันพระบรมราชสมภพ", "TH"),
    "2026-12-10": ("憲法日", "วันรัฐธรรมนูญ", "TH"),
    "2026-12-31": ("除夕", "วันสิ้นปี", "TH"),
}


# ============================================================
# Class 1: OrderProgressTracker
# ============================================================

class OrderProgressTracker:
    """共享訂單進度追蹤器"""

    def __init__(self, filepath="印尼協同追蹤表.xlsx"):
        self.filepath = filepath
        self.orders = {}

    # ------ helpers ------

    def _apply_header_style(self, ws, row=1, max_col=None):
        if max_col is None:
            max_col = ws.max_column
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER

    def _auto_width(self, ws):
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                try:
                    val = str(cell.value or "")
                    # Rough CJK width estimation
                    cjk_count = sum(1 for c in val if '\u4e00' <= c <= '\u9fff' or '\u0e00' <= c <= '\u0e7f')
                    ascii_count = len(val) - cjk_count
                    est_width = cjk_count * 2.2 + ascii_count * 1.1
                    if est_width > max_len:
                        max_len = est_width
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    def _save(self, wb):
        wb.save(self.filepath)

    # ------ public API ------

    def create_tracker(self, sample_orders=None):
        """建立新的追蹤 Excel 檔案"""
        wb = Workbook()

        # === Sheet 1: 訂單進度追蹤 ===
        ws1 = wb.active
        ws1.title = "訂單進度追蹤"
        headers1 = [
            "訂單編號", "客戶名稱", "產品描述", "數量",
            "出發地", "目的地", "提單號(B/L)", "船舶名稱",
            "預計出貨日", "預計到港日(ETA)", "當前狀態",
            "狀態更新時間", "備註", "負責人", "異常標記"
        ]
        ws1.append(headers1)
        self._apply_header_style(ws1, row=1, max_col=len(headers1))

        if sample_orders:
            for order in sample_orders:
                ws1.append([
                    order.get("order_id", ""),
                    order.get("customer", ""),
                    order.get("product", ""),
                    order.get("quantity", ""),
                    order.get("origin", ""),
                    order.get("destination", ""),
                    order.get("bl_number", ""),
                    order.get("vessel", ""),
                    order.get("ship_date", ""),
                    order.get("eta", ""),
                    order.get("status", ""),
                    order.get("update_time", ""),
                    order.get("notes", ""),
                    order.get("owner", ""),
                    order.get("anomaly", ""),
                ])
                # Color-code anomaly rows
                if order.get("anomaly") and "異常" in str(order.get("anomaly", "")):
                    row_idx = ws1.max_row
                    for col in range(1, len(headers1) + 1):
                        cell = ws1.cell(row=row_idx, column=col)
                        cell.fill = ALERT_FILL
                        cell.font = ALERT_FONT
                elif order.get("status") == "已送達":
                    row_idx = ws1.max_row
                    for col in range(1, len(headers1) + 1):
                        cell = ws1.cell(row=row_idx, column=col)
                        cell.fill = OK_FILL
                        cell.font = OK_FONT
                # Style normal rows
                else:
                    row_idx = ws1.max_row
                    for col in range(1, len(headers1) + 1):
                        cell = ws1.cell(row=row_idx, column=col)
                        cell.font = NORMAL_FONT
                        cell.border = THIN_BORDER
                        cell.alignment = Alignment(vertical="center", wrap_text=True)

        self._auto_width(ws1)

        # === Sheet 2: 溝通記錄 ===
        ws2 = wb.create_sheet("溝通記錄")
        headers2 = [
            "日期", "訂單編號", "溝通類型", "發送方", "接收方",
            "訊息內容(中文)", "訊息內容(印尼文)", "回覆狀態", "回覆內容", "備註"
        ]
        ws2.append(headers2)
        self._apply_header_style(ws2, row=1, max_col=len(headers2))

        sample_comms = [
            ["2026-06-10", "ORD-2026-062", "出貨通知", "台灣總部", "雅加達辦事處",
             "訂單ORD-2026-062已出貨，提單號BKKT2606123", "Pesanan ORD-2026-062 telah dikirim, nomor B/L BKKT2606123",
             "已回覆", "收到，準備清關文件", ""],
            ["2026-06-12", "ORD-2026-064", "到港通知", "雅加達辦事處", "報關行",
             "貨物已到港，請安排清關", "Barang sudah tiba di pelabuhan, mohon atur bea cukai",
             "已回覆", "清關文件已提交", ""],
            ["2026-06-14", "ORD-2026-065", "清關文件確認", "報關行", "雅加達辦事處",
             "需要補充原產地證明", "Perlu tambahan sertifikat asal",
             "待回覆", "", "緊急"],
            ["2026-06-08", "ORD-2026-068", "異常回報", "雅加達辦事處", "台灣總部",
             "清關延誤，已超過8天", "Keterlambatan bea cukai, sudah lebih dari 8 hari",
             "已回覆", "請跟進海關官員", "高層關注"],
            ["2026-06-15", "ORD-2026-063", "到港通知", "船公司", "雅加達辦事處",
             "船舶MV Pacific Star預計6月20日抵達", "Kapal MV Pacific Star diperkirakan tiba 20 Juni",
             "未回覆", "", "追蹤中"],
        ]
        for comm in sample_comms:
            ws2.append(comm)
            row_idx = ws2.max_row
            for col in range(1, len(headers2) + 1):
                cell = ws2.cell(row=row_idx, column=col)
                cell.font = NORMAL_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        self._auto_width(ws2)

        # === Sheet 3: 文件到期追蹤 ===
        ws3 = wb.create_sheet("文件到期追蹤")
        headers3 = [
            "文件名稱", "文件類型", "持有人/單位", "到期日",
            "剩餘天數", "續期狀態", "續期步驟", "負責人", "備註"
        ]
        ws3.append(headers3)
        self._apply_header_style(ws3, row=1, max_col=len(headers3))

        today = datetime.now()
        sample_docs = [
            ["印尼進口許可證(API)", "進口許可", "雅加達辦事處",
             (today + timedelta(days=45)).strftime("%Y-%m-%d"),
             45, "正常", "到期前30天向貿易部申請續期", "張經理", ""],
            ["海關登記證(NIB)", "海關登記", "雅加達辦事處",
             (today + timedelta(days=15)).strftime("%Y-%m-%d"),
             15, "需續期", "立即聯繫海關辦公室辦理", "李專員", "緊急"],
            ["原產地證明(CO)", "貿易文件", "台灣總部",
             (today + timedelta(days=90)).strftime("%Y-%m-%d"),
             90, "正常", "向經濟部國貿署申請", "王主任", ""],
            ["熏蒸證書", "檢疫文件", "泰國辦事處",
             (today + timedelta(days=7)).strftime("%Y-%m-%d"),
             7, "即將到期", "聯繫泰國農業部辦理新證", "陳專員", "需立即處理"],
            ["危險品運輸許可", "特許文件", "台灣總部",
             (today + timedelta(days=120)).strftime("%Y-%m-%d"),
             120, "正常", "向交通部申請", "趙經理", ""],
        ]
        for doc in sample_docs:
            ws3.append(doc)
            row_idx = ws3.max_row
            for col in range(1, len(headers3) + 1):
                cell = ws3.cell(row=row_idx, column=col)
                cell.font = NORMAL_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            # Highlight expiring soon
            if doc[4] <= 15:
                for col in range(1, len(headers3) + 1):
                    cell = ws3.cell(row=row_idx, column=col)
                    cell.fill = ALERT_FILL
                    cell.font = ALERT_FONT
        self._auto_width(ws3)

        self._save(wb)
        # Store orders internally
        if sample_orders:
            for o in sample_orders:
                self.orders[o["order_id"]] = o
        return self.filepath

    def update_status(self, order_id, new_status, notes=""):
        """更新訂單狀態"""
        if order_id not in self.orders:
            # Try loading from file
            self._load_orders_from_file()
        if order_id in self.orders:
            old_status = self.orders[order_id]["status"]
            self.orders[order_id]["status"] = new_status
            self.orders[order_id]["update_time"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            if notes:
                self.orders[order_id]["notes"] = notes
            # Clear anomaly if moving forward
            if STATUS_PIPELINE.index(new_status) > STATUS_PIPELINE.index(old_status):
                self.orders[order_id]["anomaly"] = ""
            # Update the Excel file
            self._update_excel_row(order_id)
            return {
                "order_id": order_id,
                "old_status": old_status,
                "new_status": new_status,
                "timestamp": self.orders[order_id]["update_time"],
                "notes": notes,
            }
        return {"error": f"找不到訂單 {order_id}"}

    def _load_orders_from_file(self):
        """從 Excel 載入訂單"""
        if not os.path.exists(self.filepath):
            return
        wb = load_workbook(self.filepath)
        ws = wb["訂單進度追蹤"]
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                order = dict(zip(headers, row))
                self.orders[order["訂單編號"]] = {
                    "order_id": order["訂單編號"],
                    "customer": order.get("客戶名稱", ""),
                    "product": order.get("產品描述", ""),
                    "quantity": order.get("數量", ""),
                    "origin": order.get("出發地", ""),
                    "destination": order.get("目的地", ""),
                    "bl_number": order.get("提單號(B/L)", ""),
                    "vessel": order.get("船舶名稱", ""),
                    "ship_date": str(order.get("預計出貨日", "")),
                    "eta": str(order.get("預計到港日(ETA)", "")),
                    "status": order.get("當前狀態", ""),
                    "update_time": str(order.get("狀態更新時間", "")),
                    "notes": order.get("備註", ""),
                    "owner": order.get("負責人", ""),
                    "anomaly": order.get("異常標記", ""),
                }

    def _update_excel_row(self, order_id):
        """更新 Excel 中的指定行"""
        if not os.path.exists(self.filepath):
            return
        wb = load_workbook(self.filepath)
        ws = wb["訂單進度追蹤"]
        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row=row_idx, column=1).value == order_id:
                order = self.orders[order_id]
                ws.cell(row=row_idx, column=11).value = order["status"]
                ws.cell(row=row_idx, column=12).value = order["update_time"]
                ws.cell(row=row_idx, column=13).value = order.get("notes", "")
                ws.cell(row=row_idx, column=15).value = order.get("anomaly", "")
                break
        self._save(wb)

    def get_daily_report(self):
        """產生每日報告"""
        if not self.orders:
            self._load_orders_from_file()

        today_str = datetime.now().strftime("%Y-%m-%d")
        report = {
            "report_date": today_str,
            "report_title": "每日進度報告",
            "today_arrivals": [],
            "in_customs": [],
            "attention_needed": [],
            "summary": {},
        }

        for oid, order in self.orders.items():
            eta = str(order.get("eta", ""))
            status = order.get("status", "")

            # Today arrivals
            if eta and today_str in eta:
                report["today_arrivals"].append(order)

            # In customs
            if status == "清關中":
                report["in_customs"].append(order)

            # Attention: overdue or anomaly
            if order.get("anomaly"):
                report["attention_needed"].append(order)

        # Summary stats
        status_counts = {}
        for s in STATUS_PIPELINE:
            status_counts[s] = 0
        for order in self.orders.values():
            s = order.get("status", "")
            if s in status_counts:
                status_counts[s] += 1
        report["summary"] = status_counts
        report["total_orders"] = len(self.orders)

        return report

    def get_weekly_report(self):
        """產生週報告"""
        if not self.orders:
            self._load_orders_from_file()

        now = datetime.now()
        week_start = now - timedelta(days=now.weekday())
        week_end = week_start + timedelta(days=6)

        report = {
            "report_title": "週報摘要",
            "date_range": f"{week_start.strftime('%Y-%m-%d')} ~ {week_end.strftime('%Y-%m-%d')}",
            "shipment_stats": {
                "total": len(self.orders),
                "in_production": 0,
                "in_transit": 0,
                "in_customs": 0,
                "delivered": 0,
            },
            "status_breakdown": {},
            "anomalies": [],
            "upcoming_deadlines": [],
            "action_items": [],
            "next_week_forecast": [],
        }

        for s in STATUS_PIPELINE:
            report["status_breakdown"][s] = 0

        for oid, order in self.orders.items():
            s = order.get("status", "")
            if s in report["status_breakdown"]:
                report["status_breakdown"][s] += 1

            if s == "生產中":
                report["shipment_stats"]["in_production"] += 1
            elif s in ("已出貨", "海上運輸中"):
                report["shipment_stats"]["in_transit"] += 1
            elif s in ("清關中", "已到港"):
                report["shipment_stats"]["in_customs"] += 1
            elif s in ("已送達", "已放行"):
                report["shipment_stats"]["delivered"] += 1

            if order.get("anomaly"):
                report["anomalies"].append(order)

            eta = str(order.get("eta", ""))
            if eta:
                try:
                    eta_date = datetime.strptime(eta[:10], "%Y-%m-%d")
                    next_week_end = week_end + timedelta(days=7)
                    if week_start <= eta_date <= next_week_end:
                        report["upcoming_deadlines"].append(order)
                        report["next_week_forecast"].append(order)
                except ValueError:
                    pass

            if s == "清關中":
                report["action_items"].append({
                    "order_id": oid,
                    "action": "跟進清關進度",
                    "priority": "高" if order.get("anomaly") else "中",
                })

        return report

    def get_overdue_items(self):
        """取得逾期/卡住的訂單"""
        if not self.orders:
            self._load_orders_from_file()

        overdue = []
        now = datetime.now()

        for oid, order in self.orders.items():
            status = order.get("status", "")
            if status == "已送達":
                continue

            update_time = order.get("update_time", "")
            if not update_time:
                continue

            try:
                if " " in str(update_time):
                    last_update = datetime.strptime(str(update_time)[:19], "%Y-%m-%d %H:%M:%S")
                else:
                    last_update = datetime.strptime(str(update_time)[:10], "%Y-%m-%d")
                days_stuck = (now - last_update).days
            except (ValueError, TypeError):
                days_stuck = 0

            threshold = STATUS_OVERDUE_DAYS.get(status, 7)
            if days_stuck >= threshold:
                overdue.append({
                    "order_id": oid,
                    "status": status,
                    "days_stuck": days_stuck,
                    "threshold": threshold,
                    "severity": "嚴重" if days_stuck >= threshold * 2 else "警告",
                    "last_update": str(update_time),
                    "customer": order.get("customer", ""),
                    "notes": order.get("notes", ""),
                })

        overdue.sort(key=lambda x: x["days_stuck"], reverse=True)
        return overdue

    def calculate_eta(self, order_id):
        """計算預計到達時間"""
        if not self.orders:
            self._load_orders_from_file()

        if order_id not in self.orders:
            return {"error": f"找不到訂單 {order_id}"}

        order = self.orders[order_id]
        status = order.get("status", "")
        eta = order.get("eta", "")

        result = {
            "order_id": order_id,
            "current_status": status,
            "stated_eta": eta,
        }

        # Simple ETA calculation based on status
        typical_days = {
            "生產中": {"sea": 14, "customs": 5, "delivery": 2},
            "已出貨": {"sea": 12, "customs": 5, "delivery": 2},
            "海上運輸中": {"sea": 7, "customs": 5, "delivery": 2},
            "已到港": {"customs": 5, "delivery": 2},
            "清關中": {"customs": 3, "delivery": 2},
            "已放行": {"delivery": 2},
            "已送達": {"delivery": 0},
        }

        remaining = typical_days.get(status, {})
        total_remaining = sum(remaining.values())
        if eta:
            try:
                eta_date = datetime.strptime(str(eta)[:10], "%Y-%m-%d")
                result["calculated_arrival"] = eta_date.strftime("%Y-%m-%d")
                result["days_remaining"] = max(0, (eta_date - datetime.now()).days)
            except ValueError:
                result["calculated_arrival"] = (datetime.now() + timedelta(days=total_remaining)).strftime("%Y-%m-%d")
                result["days_remaining"] = total_remaining
        else:
            result["calculated_arrival"] = (datetime.now() + timedelta(days=total_remaining)).strftime("%Y-%m-%d")
            result["days_remaining"] = total_remaining

        result["breakdown"] = remaining
        return result


# ============================================================
# Class 2: MessageTemplateEngine
# ============================================================

class MessageTemplateEngine:
    """雙語訊息範本引擎 (中文/印尼文)"""

    TEMPLATES = {
        "出貨通知": {
            "cn": (
                "【出貨通知】\n"
                "訂單編號：{order_id}\n"
                "出貨日期：{ship_date}\n"
                "提單號碼：{bl_number}\n"
                "預計到達：{eta}\n"
                "目的港口：{destination}\n"
                "船舶名稱：{vessel}\n"
                "貨物描述：{product}\n"
                "數量：{quantity}\n\n"
                "請提前準備清關文件，如有問題請隨時聯繫。"
            ),
            "id": (
                "[Pemberitahuan Pengiriman]\n"
                "Nomor Pesanan: {order_id}\n"
                "Tanggal Pengiriman: {ship_date}\n"
                "Nomor B/L: {bl_number}\n"
                "Perkiraan Tiba: {eta}\n"
                "Pelabuhan Tujuan: {destination}\n"
                "Nama Kapal: {vessel}\n"
                "Deskripsi Barang: {product}\n"
                "Jumlah: {quantity}\n\n"
                "Mohon siapkan dokumen bea cukai terlebih dahulu. Hubungi kami jika ada pertanyaan."
            ),
        },
        "清關文件確認": {
            "cn": (
                "【清關文件確認】\n"
                "訂單編號：{order_id}\n"
                "以下文件請確認是否齊全：\n"
                "☐ 商業發票 (Commercial Invoice)\n"
                "☐ 裝箱單 (Packing List)\n"
                "☐ 提單 (Bill of Lading)\n"
                "☐ 原產地證明 (Certificate of Origin)\n"
                "☐ 熏蒸證書 (Fumigation Certificate)\n"
                "☐ 進口許可證 (Import License/API)\n"
                "☐ 保險單 (Insurance Policy)\n\n"
                "請於 {deadline} 前回覆確認。"
            ),
            "id": (
                "[Konfirmasi Dokumen Bea Cukai]\n"
                "Nomor Pesanan: {order_id}\n"
                "Mohon konfirmasi kelengkapan dokumen berikut:\n"
                "☐ Faktur Komersial (Commercial Invoice)\n"
                "☐ Daftar Kemasan (Packing List)\n"
                "☐ Konosemen (Bill of Lading)\n"
                "☐ Sertifikat Asal (Certificate of Origin)\n"
                "☐ Sertifikat Fumigasi (Fumigation Certificate)\n"
                "☐ Izin Impor (API)\n"
                "☐ Polis Asuransi (Insurance Policy)\n\n"
                "Mohon balas konfirmasi sebelum {deadline}."
            ),
        },
        "到港通知": {
            "cn": (
                "【到港通知】\n"
                "訂單編號：{order_id}\n"
                "船舶名稱：{vessel}\n"
                "預計到港：{eta}\n"
                "目的港口：{destination}\n"
                "提單號碼：{bl_number}\n\n"
                "請注意：\n"
                "1. 請確認清關文件已準備就緒\n"
                "2. 請聯繫報關行安排清關\n"
                "3. 如需倉儲安排請提前告知\n"
                "4. 預計清關時間：到港後3-5個工作天"
            ),
            "id": (
                "[Pemberitahuan Kedatangan]\n"
                "Nomor Pesanan: {order_id}\n"
                "Nama Kapal: {vessel}\n"
                "Perkiraan Tiba: {eta}\n"
                "Pelabuhan Tujuan: {destination}\n"
                "Nomor B/L: {bl_number}\n\n"
                "Perhatian:\n"
                "1. Pastikan dokumen bea cukai sudah siap\n"
                "2. Hubungi agen bea cukai untuk mengatur proses\n"
                "3. Beri tahu kami jika perlu pengaturan gudang\n"
                "4. Perkiraan waktu bea cukai: 3-5 hari kerja setelah tiba"
            ),
        },
        "異常回報": {
            "cn": (
                "【異常回報】\n"
                "訂單編號：{order_id}\n"
                "異常類型：{issue_type}\n"
                "異常描述：{description}\n"
                "當前狀態：{status}\n"
                "持續天數：{days} 天\n\n"
                "建議處理方案：{resolution}\n"
                "處理期限：{deadline}\n"
                "嚴重程度：{severity}\n\n"
                "請相關人員立即跟進處理。"
            ),
            "id": (
                "[Laporan Anomali]\n"
                "Nomor Pesanan: {order_id}\n"
                "Jenis Anomali: {issue_type}\n"
                "Deskripsi: {description}\n"
                "Status Saat Ini: {status}\n"
                "Durasi: {days} hari\n\n"
                "Solusi yang Disarankan: {resolution}\n"
                "Batas Waktu: {deadline}\n"
                "Tingkat Keparahan: {severity}\n\n"
                "Mohon personel terkait segera menindaklanjuti."
            ),
        },
        "文件到期提醒": {
            "cn": (
                "【文件到期提醒】\n"
                "文件名稱：{doc_name}\n"
                "文件類型：{doc_type}\n"
                "到期日期：{expiry_date}\n"
                "剩餘天數：{days_remaining} 天\n\n"
                "續期步驟：\n"
                "{renewal_steps}\n\n"
                "負責人：{owner}\n"
                "請盡快處理續期事宜，避免影響正常業務。"
            ),
            "id": (
                "[Pengingat Kedaluwarsa Dokumen]\n"
                "Nama Dokumen: {doc_name}\n"
                "Jenis Dokumen: {doc_type}\n"
                "Tanggal Kedaluwarsa: {expiry_date}\n"
                "Sisa Hari: {days_remaining} hari\n\n"
                "Langkah Perpanjangan:\n"
                "{renewal_steps}\n\n"
                "Penanggung Jawab: {owner}\n"
                "Mohon segera proses perpanjangan agar tidak mengganggu operasional."
            ),
        },
        "詢價請求": {
            "cn": (
                "【詢價請求】\n"
                "產品名稱：{product}\n"
                "規格要求：{specs}\n"
                "需求數量：{quantity}\n"
                "交貨地點：{destination}\n"
                "回覆期限：{deadline}\n\n"
                "請提供以下資訊：\n"
                "1. FOB/CIF 報價\n"
                "2. 最短交貨期\n"
                "3. 付款條件\n"
                "4. 報價有效期"
            ),
            "id": (
                "[Permintaan Penawaran Harga]\n"
                "Nama Produk: {product}\n"
                "Spesifikasi: {specs}\n"
                "Jumlah Dibutuhkan: {quantity}\n"
                "Tempat Pengiriman: {destination}\n"
                "Batas Waktu Balasan: {deadline}\n\n"
                "Mohon berikan informasi berikut:\n"
                "1. Harga FOB/CIF\n"
                "2. Waktu pengiriman tercepat\n"
                "3. Syarat pembayaran\n"
                "4. Masa berlaku penawaran"
            ),
        },
        "付款確認": {
            "cn": (
                "【付款確認】\n"
                "訂單編號：{order_id}\n"
                "付款金額：{currency} {amount}\n"
                "付款參考號：{reference}\n"
                "付款日期：{payment_date}\n"
                "付款方式：{payment_method}\n\n"
                "請確認收款並安排後續出貨事宜。"
            ),
            "id": (
                "[Konfirmasi Pembayaran]\n"
                "Nomor Pesanan: {order_id}\n"
                "Jumlah Pembayaran: {currency} {amount}\n"
                "Referensi Pembayaran: {reference}\n"
                "Tanggal Pembayaran: {payment_date}\n"
                "Metode Pembayaran: {payment_method}\n\n"
                "Mohon konfirmasi penerimaan dan atur pengiriman selanjutnya."
            ),
        },
        "週報摘要": {
            "cn": (
                "【週報摘要】{date_range}\n"
                "━━━━━━━━━━━━━━━━━━━━━\n"
                "📊 本週統計：\n"
                "  訂單總數：{total}\n"
                "  生產中：{in_production}\n"
                "  運輸中：{in_transit}\n"
                "  清關中：{in_customs}\n"
                "  已完成：{delivered}\n\n"
                "⚠️ 異常項目：{anomaly_count} 件\n"
                "{anomaly_details}\n\n"
                "📋 待辦事項：\n"
                "{action_items}\n\n"
                "📅 下週預計：\n"
                "{forecast}"
            ),
            "id": (
                "[Ringkasan Mingguan] {date_range}\n"
                "━━━━━━━━━━━━━━━━━━━━━\n"
                "Statistik Minggu Ini:\n"
                "  Total Pesanan: {total}\n"
                "  Dalam Produksi: {in_production}\n"
                "  Dalam Pengiriman: {in_transit}\n"
                "  Proses Bea Cukai: {in_customs}\n"
                "  Selesai: {delivered}\n\n"
                "Anomali: {anomaly_count} item\n"
                "{anomaly_details}\n\n"
                "Tindakan yang Diperlukan:\n"
                "{action_items}\n\n"
                "Perkiraan Minggu Depan:\n"
                "{forecast}"
            ),
        },
        "催促回覆": {
            "cn": (
                "【催促回覆】\n"
                "原始請求：{original_request}\n"
                "訂單編號：{order_id}\n"
                "等待天數：{days_pending} 天\n"
                "緊急程度：{urgency}\n\n"
                "此事項影響後續作業進度，請盡快回覆。\n"
                "如需更多資訊請告知，謝謝。"
            ),
            "id": (
                "[Tindak Lanjut]\n"
                "Permintaan Awal: {original_request}\n"
                "Nomor Pesanan: {order_id}\n"
                "Hari Menunggu: {days_pending} hari\n"
                "Tingkat Urgensi: {urgency}\n\n"
                "Hal ini memengaruhi proses selanjutnya, mohon balas sesegera mungkin.\n"
                "Jika membutuhkan informasi lebih lanjut, silakan beri tahu. Terima kasih."
            ),
        },
        "節日問候": {
            "cn": (
                "【節日問候】\n"
                "{holiday_cn}快樂！\n\n"
                "值此 {holiday_cn} 佳節，謹向 貴公司致以最誠摯的祝福。\n"
                "祝 業務蒸蒸日上，合作愉快！\n\n"
                "備註：假日期間 ({holiday_date}) 可能影響作業時程，\n"
                "請提前安排相關事宜。"
            ),
            "id": (
                "[Ucapan Hari Raya]\n"
                "Selamat {holiday_id}!\n\n"
                "Dalam rangka {holiday_id}, kami mengucapkan selamat dan salam hangat.\n"
                "Semoga bisnis semakin sukses dan kerja sama semakin baik!\n\n"
                "Catatan: Selama hari libur ({holiday_date}), jadwal operasional mungkin terpengaruh.\n"
                "Mohon atur segala sesuatunya terlebih dahulu."
            ),
        },
    }

    def fill_template(self, template_name, data_dict):
        """填入範本資料，回傳中英文訊息"""
        if template_name not in self.TEMPLATES:
            return {"error": f"找不到範本: {template_name}", "available": list(self.TEMPLATES.keys())}

        template = self.TEMPLATES[template_name]
        try:
            cn_msg = template["cn"].format(**data_dict)
            id_msg = template["id"].format(**data_dict)
            return {
                "template": template_name,
                "chinese": cn_msg,
                "indonesian": id_msg,
            }
        except KeyError as e:
            return {"error": f"缺少參數: {e}"}

    def generate_wecom_message(self, template_name, data_dict, recipients=None):
        """產生 WeCom 格式訊息"""
        filled = self.fill_template(template_name, data_dict)
        if "error" in filled:
            return filled

        wecom_msg = {
            "msgtype": "text",
            "text": {
                "content": filled["chinese"] + "\n\n---\n\n" + filled["indonesian"]
            },
            "safe": 0,
        }
        if recipients:
            wecom_msg["touser"] = "|".join(recipients)

        return {
            "template": template_name,
            "wecom_payload": wecom_msg,
            "chinese": filled["chinese"],
            "indonesian": filled["indonesian"],
        }

    def list_templates(self):
        """列出所有可用範本"""
        return list(self.TEMPLATES.keys())

    def get_holiday_greeting(self, date_str=None):
        """根據日期取得節日問候"""
        if date_str is None:
            date_str = datetime.now().strftime("%Y-%m-%d")

        for key, (cn_name, id_name, country) in HOLIDAYS_2026.items():
            key_date = key.rstrip("b")
            if date_str in key_date or key_date in date_str:
                return self.fill_template("節日問候", {
                    "holiday_cn": cn_name,
                    "holiday_id": id_name,
                    "holiday_date": date_str,
                })
        return {"info": "指定日期無對應節日"}


# ============================================================
# Class 3: TranslationHelper
# ============================================================

class TranslationHelper:
    """海關/貿易術語翻譯助手 (中文 ↔ 印尼文 ↔ 英文)"""

    GLOSSARY = {
        # --- 基本貿易術語 ---
        "提單": {"id": "Konosemen / Bill of Lading", "en": "Bill of Lading (B/L)"},
        "商業發票": {"id": "Faktur Komersial", "en": "Commercial Invoice"},
        "裝箱單": {"id": "Daftar Kemasan", "en": "Packing List"},
        "原產地證明": {"id": "Sertifikat Asal", "en": "Certificate of Origin"},
        "熏蒸證書": {"id": "Sertifikat Fumigasi", "en": "Fumigation Certificate"},
        "植物檢疫證書": {"id": "Sertifikat Fitosanitari", "en": "Phytosanitary Certificate"},
        "進口許可證": {"id": "Izin Impor (API)", "en": "Import License (API)"},
        "出口許可證": {"id": "Izin Ekspor", "en": "Export License"},
        "海關申報單": {"id": "Pemberitahuan Impor Barang (PIB)", "en": "Customs Declaration"},
        "報關行": {"id": "Agen Bea Cukai / PPJK", "en": "Customs Broker"},
        "關稅": {"id": "Bea Masuk", "en": "Import Duty"},
        "增值稅": {"id": "Pajak Pertambahan Nilai (PPN)", "en": "Value Added Tax (VAT)"},
        "所得稅": {"id": "Pajak Penghasilan (PPh)", "en": "Income Tax"},
        "奢侈品稅": {"id": "Pajak Penjualan atas Barang Mewah (PPnBM)", "en": "Luxury Goods Sales Tax"},
        "保稅區": {"id": "Kawasan Berikat", "en": "Bonded Zone"},
        "自由貿易區": {"id": "Kawasan Perdagangan Bebas", "en": "Free Trade Zone"},
        "海關編碼": {"id": "Kode HS (Harmonized System)", "en": "HS Code"},
        "清關": {"id": "Proses Bea Cukai / Clearance", "en": "Customs Clearance"},
        "放行": {"id": "Pengeluaran Barang", "en": "Release of Goods"},
        "查驗": {"id": "Pemeriksaan Fisik", "en": "Physical Inspection"},
        "X光檢查": {"id": "Pemeriksaan Sinar-X", "en": "X-ray Inspection"},
        "抽樣檢驗": {"id": "Pemeriksaan Sampel", "en": "Sampling Inspection"},
        # --- 運輸術語 ---
        "貨櫃": {"id": "Kontainer / Peti Kemas", "en": "Container"},
        "整櫃": {"id": "FCL (Full Container Load)", "en": "Full Container Load (FCL)"},
        "拼櫃": {"id": "LCL (Less Container Load)", "en": "Less Container Load (LCL)"},
        "20呎貨櫃": {"id": "Kontainer 20 kaki", "en": "20ft Container (TEU)"},
        "40呎貨櫃": {"id": "Kontainer 40 kaki", "en": "40ft Container (FEU)"},
        "散貨": {"id": "Kargo Curah", "en": "Bulk Cargo"},
        "船舶": {"id": "Kapal", "en": "Vessel / Ship"},
        "航次": {"id": "Pelayaran / Voy.", "en": "Voyage"},
        "裝貨港": {"id": "Pelabuhan Muat", "en": "Port of Loading"},
        "卸貨港": {"id": "Pelabuhan Bongkar", "en": "Port of Discharge"},
        "轉運港": {"id": "Pelabuhan Transit", "en": "Transshipment Port"},
        "到港通知": {"id": "Pemberitahuan Kedatangan", "en": "Arrival Notice"},
        "滯港費": {"id": "Biaya Demurrage", "en": "Demurrage Charge"},
        "滯箱費": {"id": "Biaya Detention", "en": "Detention Charge"},
        "倉儲費": {"id": "Biaya Pergudangan", "en": "Storage / Warehousing Fee"},
        "運費": {"id": "Biaya Pengiriman / Freight", "en": "Freight Charge"},
        "海運": {"id": "Pengiriman Laut", "en": "Sea Freight"},
        "空運": {"id": "Pengiriman Udara", "en": "Air Freight"},
        "內陸運輸": {"id": "Transportasi Darat", "en": "Inland Transportation"},
        # --- 貿易條款 ---
        "離岸價": {"id": "FOB (Free on Board)", "en": "FOB - Free on Board"},
        "到岸價": {"id": "CIF (Cost, Insurance, Freight)", "en": "CIF - Cost, Insurance & Freight"},
        "工廠交貨": {"id": "EXW (Ex Works)", "en": "EXW - Ex Works"},
        "完稅交貨": {"id": "DDP (Delivered Duty Paid)", "en": "DDP - Delivered Duty Paid"},
        "信用狀": {"id": "Letter of Credit (L/C)", "en": "Letter of Credit (L/C)"},
        "電匯": {"id": "Transfer Telegraphic (T/T)", "en": "Telegraphic Transfer (T/T)"},
        "付款交單": {"id": "D/P (Documents against Payment)", "en": "Documents against Payment (D/P)"},
        "承兌交單": {"id": "D/A (Documents against Acceptance)", "en": "Documents against Acceptance (D/A)"},
        "訂金": {"id": "Uang Muka / Down Payment", "en": "Down Payment / Deposit"},
        "尾款": {"id": "Sisa Pembayaran", "en": "Balance Payment"},
        # --- 產品/貨物相關 ---
        "淨重": {"id": "Berat Bersih", "en": "Net Weight"},
        "毛重": {"id": "Berat Kotor", "en": "Gross Weight"},
        "體積": {"id": "Volume / Kubikasi", "en": "Volume / CBM"},
        "立方米": {"id": "Meter Kubik (CBM)", "en": "Cubic Meter (CBM)"},
        "嘜頭": {"id": "Tanda Pengiriman / Shipping Mark", "en": "Shipping Mark"},
        "危險品": {"id": "Barang Berbahaya (DG)", "en": "Dangerous Goods (DG)"},
        "易腐品": {"id": "Barang Mudah Rusak", "en": "Perishable Goods"},
        "原產地": {"id": "Negara Asal", "en": "Country of Origin"},
        "製造日期": {"id": "Tanggal Produksi", "en": "Production Date"},
        "有效期": {"id": "Masa Berlaku", "en": "Validity Period"},
        "材質證明": {"id": "Sertifikat Material", "en": "Material Certificate"},
        "品質檢驗報告": {"id": "Laporan Inspeksi Kualitas", "en": "Quality Inspection Report"},
        # --- 印尼海關專用 ---
        "印尼國家標準": {"id": "Standar Nasional Indonesia (SNI)", "en": "Indonesian National Standard (SNI)"},
        "食品藥品監督局": {"id": "Badan Pengawas Obat dan Makanan (BPOM)", "en": "Food and Drug Agency (BPOM)"},
        "印尼海關總署": {"id": "Direktorat Jenderal Bea dan Cukai (DJBC)", "en": "Directorate General of Customs and Excise"},
        "進口商識別號": {"id": "Nomor Pokok Wajib Pajak (NPWP)", "en": "Taxpayer ID Number (NPWP)"},
        "企業登記號": {"id": "Nomor Induk Berusaha (NIB)", "en": "Business ID Number (NIB)"},
        "印尼港口代碼": {"id": "Kode Pelabuhan Indonesia", "en": "Indonesian Port Code"},
        "雅加達港": {"id": "Pelabuhan Tanjung Priok, Jakarta", "en": "Tanjung Priok Port, Jakarta"},
        "泗水港": {"id": "Pelabuhan Tanjung Perak, Surabaya", "en": "Tanjung Perak Port, Surabaya"},
        "棉蘭港": {"id": "Pelabuhan Belawan, Medan", "en": "Belawan Port, Medan"},
        "紅色通道": {"id": "Jalur Merah (Pemeriksaan)", "en": "Red Channel (Inspection Required)"},
        "綠色通道": {"id": "Jalur Hijau (Tanpa Pemeriksaan)", "en": "Green Channel (No Inspection)"},
        "黃色通道": {"id": "Jalur Kuning (Pemeriksaan Dokumen)", "en": "Yellow Channel (Document Check)"},
        "印尼盾": {"id": "Rupiah Indonesia (IDR)", "en": "Indonesian Rupiah (IDR)"},
        "匯率": {"id": "Nilai Tukar / Kurs", "en": "Exchange Rate"},
        # --- 泰國相關 ---
        "泰國海關": {"id": "Bea Cukai Thailand", "en": "Thai Customs Department"},
        "泰國銖": {"id": "Baht Thailand (THB)", "en": "Thai Baht (THB)"},
        "林查班港": {"id": "Pelabuhan Laem Chabang", "en": "Laem Chabang Port"},
        "曼谷港": {"id": "Pelabuhan Bangkok (PAT)", "en": "Bangkok Port (PAT)"},
        "泰國原產地證明": {"id": "Sertifikat Asal Thailand (Form D)", "en": "Thai Certificate of Origin (Form D)"},
        "東協自貿區": {"id": "Kawasan Perdagangan Bebas ASEAN (AFTA)", "en": "ASEAN Free Trade Area (AFTA)"},
        "RCEP原產地證明": {"id": "Sertifikat Asal RCEP", "en": "RCEP Certificate of Origin"},
        # --- 作業流程 ---
        "生產中": {"id": "Sedang Diproduksi", "en": "In Production"},
        "已出貨": {"id": "Telah Dikirim", "en": "Shipped"},
        "海上運輸中": {"id": "Dalam Pengiriman Laut", "en": "In Transit (Sea)"},
        "已到港": {"id": "Tiba di Pelabuhan", "en": "Arrived at Port"},
        "清關中": {"id": "Proses Bea Cukai", "en": "Customs Clearance"},
        "已放行": {"id": "Telah Dilepaskan", "en": "Released"},
        "已送達": {"id": "Telah Diterima", "en": "Delivered"},
        "異常": {"id": "Anomali / Masalah", "en": "Anomaly / Issue"},
        "逾期": {"id": "Terlambat / Overdue", "en": "Overdue"},
        "催促": {"id": "Tindak Lanjut / Follow-up", "en": "Follow-up"},
        "訂單": {"id": "Pesanan", "en": "Order"},
        "出貨": {"id": "Pengiriman", "en": "Shipment"},
        "倉庫": {"id": "Gudang", "en": "Warehouse"},
        "客戶": {"id": "Pelanggan", "en": "Customer"},
        "供應商": {"id": "Pemasok / Supplier", "en": "Supplier"},
        "報價": {"id": "Penawaran Harga", "en": "Quotation"},
        "合約": {"id": "Kontrak", "en": "Contract"},
        "保險": {"id": "Asuransi", "en": "Insurance"},
        "索賠": {"id": "Klaim", "en": "Claim"},
        "退運": {"id": "Pengembalian Barang", "en": "Return Shipment"},
        "轉運": {"id": "Transhipmen", "en": "Transshipment"},
        "預配": {"id": "Pra-alokasi Kontainer", "en": "Container Pre-allocation"},
        "報檢": {"id": "Pemberitahuan Inspeksi", "en": "Inspection Declaration"},
        "核銷": {"id": "Verifikasi Pembayaran", "en": "Payment Verification"},
        "退稅": {"id": "Pengembalian Pajak (Restitusi)", "en": "Tax Refund / Restitution"},
        "免稅": {"id": "Bebas Pajak / Pembebasan", "en": "Tax Exemption"},
        "減稅": {"id": "Pengurangan Pajak", "en": "Tax Reduction"},
    }

    def translate_term(self, term, from_lang="zh", to_lang="id"):
        """翻譯單一術語"""
        # Direct lookup
        if term in self.GLOSSARY:
            if to_lang == "id":
                return {"term": term, "translation": self.GLOSSARY[term]["id"], "from": "zh", "to": "id"}
            elif to_lang == "en":
                return {"term": term, "translation": self.GLOSSARY[term]["en"], "from": "zh", "to": "en"}
            elif to_lang == "zh":
                return {"term": term, "source": "zh"}

        # Reverse lookup
        for zh_term, translations in self.GLOSSARY.items():
            if from_lang == "id" and to_lang == "zh":
                if term.lower() in translations["id"].lower():
                    return {"term": term, "translation": zh_term, "from": "id", "to": "zh"}
            elif from_lang == "en" and to_lang == "zh":
                if term.lower() in translations["en"].lower():
                    return {"term": term, "translation": zh_term, "from": "en", "to": "zh"}
            elif from_lang == "id" and to_lang == "en":
                if term.lower() in translations["id"].lower():
                    return {"term": term, "translation": translations["en"], "from": "id", "to": "en"}
            elif from_lang == "en" and to_lang == "id":
                if term.lower() in translations["en"].lower():
                    return {"term": term, "translation": translations["id"], "from": "en", "to": "id"}

        return {"term": term, "translation": f"[未找到翻譯: {term}]", "from": from_lang, "to": to_lang}

    def translate_message(self, message, target_lang="id"):
        """簡單的術語替換式翻譯"""
        result = message
        replacements = []

        for zh_term, translations in self.GLOSSARY.items():
            if zh_term in result:
                if target_lang == "id":
                    result = result.replace(zh_term, translations["id"])
                    replacements.append(f"{zh_term} → {translations['id']}")
                elif target_lang == "en":
                    result = result.replace(zh_term, translations["en"])
                    replacements.append(f"{zh_term} → {translations['en']}")

        return {
            "original": message,
            "translated": result,
            "replacements": replacements,
            "note": "此為術語替換式翻譯，僅供參考，正式文件請使用專業翻譯服務。",
        }

    def get_glossary(self):
        """匯出完整術語表"""
        rows = []
        for zh, trans in self.GLOSSARY.items():
            rows.append({
                "中文": zh,
                "印尼文": trans["id"],
                "英文": trans["en"],
            })
        return rows


# ============================================================
# Class 4: WeeklyReportGenerator
# ============================================================

class WeeklyReportGenerator:
    """週報自動產生器"""

    def __init__(self, output_dir="."):
        self.output_dir = output_dir

    def generate_report(self, tracker_data, date_range=None):
        """產生格式化週報"""
        now = datetime.now()
        if date_range is None:
            week_start = now - timedelta(days=now.weekday())
            week_end = week_start + timedelta(days=6)
            date_range = f"{week_start.strftime('%Y-%m-%d')} ~ {week_end.strftime('%Y-%m-%d')}"

        stats = tracker_data.get("shipment_stats", {})
        status_bd = tracker_data.get("status_breakdown", {})
        anomalies = tracker_data.get("anomalies", [])
        deadlines = tracker_data.get("upcoming_deadlines", [])
        actions = tracker_data.get("action_items", [])
        forecast = tracker_data.get("next_week_forecast", [])

        # Build text report
        lines = []
        lines.append("=" * 60)
        lines.append(f"  印尼市場協同作業 — 週報摘要")
        lines.append(f"  報告期間：{date_range}")
        lines.append(f"  產生時間：{now.strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append("=" * 60)
        lines.append("")

        lines.append("一、出貨統計")
        lines.append(f"  訂單總數：{stats.get('total', 0)}")
        lines.append(f"  生產中：{stats.get('in_production', 0)} 筆")
        lines.append(f"  運輸中：{stats.get('in_transit', 0)} 筆")
        lines.append(f"  清關中：{stats.get('in_customs', 0)} 筆")
        lines.append(f"  已完成：{stats.get('delivered', 0)} 筆")
        lines.append("")

        lines.append("二、狀態分布")
        for s in STATUS_PIPELINE:
            count = status_bd.get(s, 0)
            bar = "█" * count + "░" * (10 - count)
            lines.append(f"  {s:<8} {bar} {count} 筆")
        lines.append("")

        lines.append("三、異常項目")
        if anomalies:
            for a in anomalies:
                lines.append(f"  ⚠ {a.get('order_id', 'N/A')} — {a.get('anomaly', '異常')} (客戶: {a.get('customer', 'N/A')})")
        else:
            lines.append("  ✓ 本週無異常項目")
        lines.append("")

        lines.append("四、即將到期的截止日")
        if deadlines:
            for d in deadlines:
                lines.append(f"  📅 {d.get('order_id', 'N/A')} — ETA: {d.get('eta', 'N/A')} ({d.get('status', '')})")
        else:
            lines.append("  ✓ 無即將到期的截止日")
        lines.append("")

        lines.append("五、待辦事項")
        if actions:
            for a in actions:
                priority_mark = "🔴" if a.get("priority") == "高" else "🟡"
                lines.append(f"  {priority_mark} [{a.get('priority', '')}] {a.get('order_id', 'N/A')} — {a.get('action', '')}")
        else:
            lines.append("  ✓ 無待辦事項")
        lines.append("")

        lines.append("六、下週預測")
        if forecast:
            for f in forecast:
                lines.append(f"  → {f.get('order_id', 'N/A')} 預計到港: {f.get('eta', 'N/A')}")
        else:
            lines.append("  無預計到港的訂單")
        lines.append("")

        lines.append("=" * 60)
        lines.append("  報告結束 — 印尼市場協同自動化工具")
        lines.append("=" * 60)

        report_text = "\n".join(lines)
        return report_text

    def export_to_excel(self, tracker_data, filepath=None):
        """匯出週報至 Excel"""
        if filepath is None:
            filepath = os.path.join(self.output_dir, f"週報_{datetime.now().strftime('%Y%m%d')}.xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "週報統計"

        # Title
        ws.merge_cells("A1:F1")
        ws["A1"] = f"印尼市場協同作業週報 — {tracker_data.get('date_range', '')}"
        ws["A1"].font = Font(name="Microsoft JhengHei", size=14, bold=True, color="1F4E79")
        ws["A1"].alignment = Alignment(horizontal="center")

        # Stats table
        ws["A3"] = "狀態"
        ws["B3"] = "數量"
        ws["A3"].fill = HEADER_FILL
        ws["B3"].fill = HEADER_FILL
        ws["A3"].font = HEADER_FONT
        ws["B3"].font = HEADER_FONT

        status_bd = tracker_data.get("status_breakdown", {})
        row = 4
        for s in STATUS_PIPELINE:
            ws.cell(row=row, column=1, value=s).font = NORMAL_FONT
            ws.cell(row=row, column=2, value=status_bd.get(s, 0)).font = NORMAL_FONT
            row += 1

        # Bar chart
        chart = BarChart()
        chart.title = "訂單狀態分布"
        chart.y_axis.title = "數量"
        chart.x_axis.title = "狀態"
        chart.style = 10
        data_ref = Reference(ws, min_col=2, min_row=3, max_row=3 + len(STATUS_PIPELINE))
        cats_ref = Reference(ws, min_col=1, min_row=4, max_row=3 + len(STATUS_PIPELINE))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.shape = 4
        chart.width = 18
        chart.height = 10
        ws.add_chart(chart, "D3")

        # Anomalies section
        anomaly_row = row + 2
        ws.cell(row=anomaly_row, column=1, value="異常項目").font = SUBHEADER_FONT
        anomaly_row += 1
        for header_idx, h in enumerate(["訂單編號", "狀態", "異常描述", "客戶"], 1):
            c = ws.cell(row=anomaly_row, column=header_idx, value=h)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
        anomaly_row += 1
        for a in tracker_data.get("anomalies", []):
            ws.cell(row=anomaly_row, column=1, value=a.get("order_id", "")).font = NORMAL_FONT
            ws.cell(row=anomaly_row, column=2, value=a.get("status", "")).font = NORMAL_FONT
            ws.cell(row=anomaly_row, column=3, value=a.get("anomaly", "")).font = NORMAL_FONT
            ws.cell(row=anomaly_row, column=4, value=a.get("customer", "")).font = NORMAL_FONT
            anomaly_row += 1

        wb.save(filepath)
        return filepath

    def generate_wecom_summary(self, tracker_data):
        """產生 WeCom 相容的摘要文字"""
        stats = tracker_data.get("shipment_stats", {})
        anomalies = tracker_data.get("anomalies", [])
        actions = tracker_data.get("action_items", [])

        summary = f"📊 印尼市場週報 {tracker_data.get('date_range', '')}\n"
        summary += f"訂單總數: {stats.get('total', 0)}\n"
        summary += f"生產中: {stats.get('in_production', 0)} | "
        summary += f"運輸中: {stats.get('in_transit', 0)} | "
        summary += f"清關中: {stats.get('in_customs', 0)} | "
        summary += f"已完成: {stats.get('delivered', 0)}\n"

        if anomalies:
            summary += f"\n⚠️ 異常 ({len(anomalies)} 件):\n"
            for a in anomalies:
                summary += f"  - {a.get('order_id', 'N/A')}: {a.get('anomaly', '')}\n"

        if actions:
            summary += f"\n📋 待辦 ({len(actions)} 件):\n"
            for a in actions:
                summary += f"  - [{a.get('priority', '')}] {a.get('order_id', 'N/A')}: {a.get('action', '')}\n"

        return summary


# ============================================================
# Class 5: IndonesiaCollabTool (Main CLI Orchestrator)
# ============================================================

class IndonesiaCollabTool:
    """主 CLI 協調器"""

    def __init__(self, output_dir="."):
        self.output_dir = output_dir
        self.tracker = OrderProgressTracker(
            filepath=os.path.join(output_dir, "印尼協同追蹤表.xlsx")
        )
        self.msg_engine = MessageTemplateEngine()
        self.translator = TranslationHelper()
        self.report_gen = WeeklyReportGenerator(output_dir=output_dir)

    def _get_sample_orders(self):
        """取得範例訂單資料"""
        now = datetime.now()
        return [
            {
                "order_id": "ORD-2026-061",
                "customer": "PT Jaya Makmur (雅加達)",
                "product": "不鏽鋼管件 DN100",
                "quantity": "500 PCS",
                "origin": "台灣高雄",
                "destination": "雅加達 Tanjung Priok",
                "bl_number": "",
                "vessel": "",
                "ship_date": "2026-06-20",
                "eta": "2026-07-05",
                "status": "生產中",
                "update_time": now.strftime("%Y-%m-%d %H:%M:%S"),
                "notes": "預計6月20日出貨",
                "owner": "張經理",
                "anomaly": "",
            },
            {
                "order_id": "ORD-2026-062",
                "customer": "CV Sentosa Teknik (泗水)",
                "product": "工業閥門 DN80",
                "quantity": "200 PCS",
                "origin": "台灣高雄",
                "destination": "泗水 Tanjung Perak",
                "bl_number": "BKKT2606123",
                "vessel": "MV Pacific Star",
                "ship_date": "2026-06-12",
                "eta": "2026-06-22",
                "status": "已出貨",
                "update_time": "2026-06-12 09:30:00",
                "notes": "已裝船，提單已簽發",
                "owner": "李專員",
                "anomaly": "",
            },
            {
                "order_id": "ORD-2026-063",
                "customer": "PT Indo Steel (雅加達)",
                "product": "碳鋼法蘭 DN150",
                "quantity": "300 PCS",
                "origin": "台灣高雄",
                "destination": "雅加達 Tanjung Priok",
                "bl_number": "BKKT2606101",
                "vessel": "MV Ocean Grace",
                "ship_date": "2026-06-10",
                "eta": "2026-06-20",
                "status": "海上運輸中",
                "update_time": "2026-06-10 14:00:00",
                "notes": "6月10日離開高雄港",
                "owner": "李專員",
                "anomaly": "",
            },
            {
                "order_id": "ORD-2026-064",
                "customer": "PT Abadi Logam (棉蘭)",
                "product": "鍍鋅鋼管 DN50",
                "quantity": "1000 PCS",
                "origin": "台灣高雄",
                "destination": "棉蘭 Belawan",
                "bl_number": "BKKT2606085",
                "vessel": "MV Nusantara Jaya",
                "ship_date": "2026-06-05",
                "eta": "2026-06-15",
                "status": "已到港",
                "update_time": "2026-06-15 08:00:00",
                "notes": "6月15日到港，等待清關安排",
                "owner": "陳專員",
                "anomaly": "",
            },
            {
                "order_id": "ORD-2026-065",
                "customer": "PT Karya Bersama (雅加達)",
                "product": "精密軸承 SKF 6205",
                "quantity": "800 PCS",
                "origin": "台灣高雄",
                "destination": "雅加達 Tanjung Priok",
                "bl_number": "BKKT2606090",
                "vessel": "MV Harmony Sea",
                "ship_date": "2026-06-06",
                "eta": "2026-06-16",
                "status": "清關中",
                "update_time": "2026-06-14 10:00:00",
                "notes": "清關文件已提交，等待查驗結果",
                "owner": "陳專員",
                "anomaly": "清關已3天，需關注",
            },
            {
                "order_id": "ORD-2026-066",
                "customer": "PT Sumber Rejeki (泗水)",
                "product": "液壓軟管 SAE 100R2",
                "quantity": "2000 M",
                "origin": "泰國林查班",
                "destination": "泗水 Tanjung Perak",
                "bl_number": "LCHB2606001",
                "vessel": "MV Thai Fortune",
                "ship_date": "2026-06-01",
                "eta": "2026-06-10",
                "status": "已放行",
                "update_time": "2026-06-13 16:00:00",
                "notes": "海關已放行，等待配送",
                "owner": "張經理",
                "anomaly": "",
            },
            {
                "order_id": "ORD-2026-067",
                "customer": "PT Maju Jaya (雅加達)",
                "product": "氣動接頭 KQ2系列",
                "quantity": "5000 PCS",
                "origin": "台灣高雄",
                "destination": "雅加達 Tanjung Priok",
                "bl_number": "BKKT2605250",
                "vessel": "MV Golden Bridge",
                "ship_date": "2026-05-25",
                "eta": "2026-06-05",
                "status": "已送達",
                "update_time": "2026-06-08 14:00:00",
                "notes": "已完成配送，客戶簽收",
                "owner": "李專員",
                "anomaly": "",
            },
            {
                "order_id": "ORD-2026-068",
                "customer": "PT Nusantara Prima (雅加達)",
                "product": "特殊合金管件 Inconel 625",
                "quantity": "100 PCS",
                "origin": "台灣高雄",
                "destination": "雅加達 Tanjung Priok",
                "bl_number": "BKKT2606020",
                "vessel": "MV Celebes Trader",
                "ship_date": "2026-06-02",
                "eta": "2026-06-12",
                "status": "清關中",
                "update_time": "2026-06-08 09:00:00",
                "notes": "特殊合金需額外檢驗，海關要求補充材質證明",
                "owner": "張經理",
                "anomaly": "異常：清關已8天，嚴重逾期",
            },
            {
                "order_id": "ORD-2026-069",
                "customer": "CV Teknik Mandiri (泗水)",
                "product": "電焊條 AWS E7018",
                "quantity": "3000 KG",
                "origin": "泰國林查班",
                "destination": "泗水 Tanjung Perak",
                "bl_number": "LCHB2606010",
                "vessel": "MV Siam Pioneer",
                "ship_date": "2026-06-08",
                "eta": "2026-06-18",
                "status": "海上運輸中",
                "update_time": "2026-06-08 11:00:00",
                "notes": "泰國出貨，運輸中",
                "owner": "陳專員",
                "anomaly": "",
            },
            {
                "order_id": "ORD-2026-070",
                "customer": "PT Global Indo (棉蘭)",
                "product": "橡膠密封件 O-Ring",
                "quantity": "10000 PCS",
                "origin": "台灣高雄",
                "destination": "棉蘭 Belawan",
                "bl_number": "",
                "vessel": "",
                "ship_date": "2026-06-25",
                "eta": "2026-07-10",
                "status": "生產中",
                "update_time": now.strftime("%Y-%m-%d %H:%M:%S"),
                "notes": "新訂單，生產排程中",
                "owner": "王主任",
                "anomaly": "",
            },
        ]

    def run_demo(self):
        """執行完整示範"""
        separator = "=" * 65
        sub_sep = "-" * 65

        print(separator)
        print("  印尼市場協同自動化工具 — 完整示範 (Demo)")
        print("  Indonesia Market Collaboration Automation — Full Demo")
        print(separator)
        print()

        # --------------------------------------------------
        # Step 1: 建立追蹤表
        # --------------------------------------------------
        print(f"[Step 1] 建立訂單進度追蹤表...")
        print(sub_sep)
        orders = self._get_sample_orders()
        filepath = self.tracker.create_tracker(orders)
        print(f"  ✓ 已建立: {filepath}")
        print(f"  ✓ 包含 {len(orders)} 筆範例訂單")
        print(f"  ✓ 包含 3 個工作表: 訂單進度追蹤、溝通記錄、文件到期追蹤")
        print()

        # --------------------------------------------------
        # Step 2: 每日報告
        # --------------------------------------------------
        print(f"[Step 2] 產生每日報告...")
        print(sub_sep)
        daily = self.tracker.get_daily_report()
        print(f"  報告日期: {daily['report_date']}")
        print(f"  訂單總數: {daily['total_orders']}")
        print()

        print("  今日預計到港:")
        if daily["today_arrivals"]:
            for o in daily["today_arrivals"]:
                print(f"    → {o['order_id']} ({o['customer']}) — {o['product']}")
        else:
            print("    (無)")
        print()

        print("  清關中訂單:")
        for o in daily["in_customs"]:
            print(f"    ⚠ {o['order_id']} ({o['customer']}) — {o.get('anomaly', '正常')}")
        print()

        print("  需關注項目:")
        for o in daily["attention_needed"]:
            print(f"    🔴 {o['order_id']}: {o.get('anomaly', '')}")
        print()

        print("  狀態統計:")
        for s in STATUS_PIPELINE:
            count = daily["summary"].get(s, 0)
            bar = "█" * count + "░" * (10 - count)
            print(f"    {s:<8} {bar} {count}")
        print()

        # --------------------------------------------------
        # Step 3: 更新訂單狀態
        # --------------------------------------------------
        print(f"[Step 3] 更新訂單狀態...")
        print(sub_sep)
        result = self.tracker.update_status(
            "ORD-2026-064",
            "清關中",
            notes="已提交清關文件，報關行處理中"
        )
        print(f"  訂單: {result['order_id']}")
        print(f"  狀態變更: {result['old_status']} → {result['new_status']}")
        print(f"  更新時間: {result['timestamp']}")
        print(f"  備註: {result['notes']}")
        print()

        # --------------------------------------------------
        # Step 4: 產生訊息範本
        # --------------------------------------------------
        print(f"[Step 4] 產生雙語訊息範本...")
        print(sub_sep)

        # 出貨通知
        print("  【出貨通知範本】")
        print(sub_sep)
        shipment_msg = self.msg_engine.fill_template("出貨通知", {
            "order_id": "ORD-2026-062",
            "ship_date": "2026-06-12",
            "bl_number": "BKKT2606123",
            "eta": "2026-06-22",
            "destination": "泗水 Tanjung Perak",
            "vessel": "MV Pacific Star",
            "product": "工業閥門 DN80",
            "quantity": "200 PCS",
        })
        print(shipment_msg["chinese"])
        print()
        print(shipment_msg["indonesian"])
        print()

        # 異常回報
        print("  【異常回報範本】")
        print(sub_sep)
        anomaly_msg = self.msg_engine.fill_template("異常回報", {
            "order_id": "ORD-2026-068",
            "issue_type": "清關延誤",
            "description": "特殊合金管件 Inconel 625 清關已超過8天，海關要求補充材質證明文件",
            "status": "清關中",
            "days": "8",
            "resolution": "1. 聯繫台灣總部補發材質證明 2. 委託報關行與海關溝通 3. 申請綠色通道快速通關",
            "deadline": "2026-06-18",
            "severity": "嚴重",
        })
        print(anomaly_msg["chinese"])
        print()
        print(anomaly_msg["indonesian"])
        print()

        # WeCom 格式
        print("  【WeCom 訊息格式範例】")
        print(sub_sep)
        wecom_msg = self.msg_engine.generate_wecom_message(
            "清關文件確認",
            {
                "order_id": "ORD-2026-065",
                "deadline": "2026-06-18",
            },
            recipients=["zhangsan", "lisi", "chenwu"]
        )
        if "wecom_payload" in wecom_msg:
            print(f"  收件人: {wecom_msg['wecom_payload'].get('touser', 'N/A')}")
            print(f"  訊息長度: {len(wecom_msg['wecom_payload']['text']['content'])} 字元")
            print(f"  訊息類型: {wecom_msg['wecom_payload']['msgtype']}")
        print()

        # --------------------------------------------------
        # Step 5: 翻譯術語
        # --------------------------------------------------
        print(f"[Step 5] 海關術語翻譯 (中文 → 印尼文)...")
        print(sub_sep)
        test_terms = ["提單", "原產地證明", "清關", "滯港費", "紅色通道"]
        for term in test_terms:
            result = self.translator.translate_term(term, from_lang="zh", to_lang="id")
            print(f"  {term:<12} → {result['translation']}")
        print()

        # 反向翻譯示範
        print("  反向翻譯 (印尼文 → 中文):")
        reverse_result = self.translator.translate_term("Bea Masuk", from_lang="id", to_lang="zh")
        print(f"  Bea Masuk → {reverse_result['translation']}")
        print()

        # --------------------------------------------------
        # Step 6: 週報告
        # --------------------------------------------------
        print(f"[Step 6] 產生週報告...")
        print(sub_sep)
        weekly = self.tracker.get_weekly_report()
        report_text = self.report_gen.generate_report(weekly)
        print(report_text)
        print()

        # WeCom summary
        print("  【WeCom 週報摘要】")
        print(sub_sep)
        wecom_summary = self.report_gen.generate_wecom_summary(weekly)
        print(wecom_summary)
        print()

        # Export weekly report Excel
        weekly_excel = self.report_gen.export_to_excel(
            weekly,
            filepath=os.path.join(self.output_dir, f"週報_{datetime.now().strftime('%Y%m%d')}.xlsx")
        )
        print(f"  ✓ 週報 Excel 已匯出: {weekly_excel}")
        print()

        # --------------------------------------------------
        # Step 7: 逾期項目
        # --------------------------------------------------
        print(f"[Step 7] 逾期/卡住的訂單...")
        print(sub_sep)
        overdue = self.tracker.get_overdue_items()
        if overdue:
            for item in overdue:
                severity_icon = "🔴" if item["severity"] == "嚴重" else "🟡"
                print(f"  {severity_icon} {item['order_id']}")
                print(f"     狀態: {item['status']} (已停留 {item['days_stuck']} 天)")
                print(f"     嚴重程度: {item['severity']} (閾值: {item['threshold']} 天)")
                print(f"     客戶: {item['customer']}")
                print()
        else:
            print("  ✓ 無逾期訂單")
            print()

        # --------------------------------------------------
        # Step 8: 術語表
        # --------------------------------------------------
        print(f"[Step 8] 術語表匯出摘要...")
        print(sub_sep)
        glossary = self.translator.get_glossary()
        print(f"  術語表共 {len(glossary)} 筆")
        print("  前10筆預覽:")
        for i, entry in enumerate(glossary[:10]):
            print(f"    {i+1:2d}. {entry['中文']:<14} | {entry['印尼文'][:30]:<32} | {entry['英文'][:30]}")
        print()

        # --------------------------------------------------
        # Summary
        # --------------------------------------------------
        print(separator)
        print("  示範執行完成！(Demo Complete!)")
        print(separator)
        print()
        print("  已產生以下檔案:")
        print(f"    1. {self.tracker.filepath}")
        print(f"    2. {weekly_excel}")
        print()
        print("  可用命令列選項:")
        print("    python3 印尼協同自動化工具.py --create-tracker")
        print("    python3 印尼協同自動化工具.py --update ORD-2026-062 已到港")
        print("    python3 印尼協同自動化工具.py --daily-report")
        print("    python3 印尼協同自動化工具.py --weekly-report")
        print("    python3 印尼協同自動化工具.py --template 出貨通知")
        print("    python3 印尼協同自動化工具.py --translate 提單")
        print("    python3 印尼協同自動化工具.py --glossary")
        print("    python3 印尼協同自動化工具.py --overdue")
        print("    python3 印尼協同自動化工具.py --demo")
        print()
        print(separator)


# ============================================================
# CLI Entry Point
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description="印尼市場協同自動化工具 — Indonesia Market Collaboration Automation",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
範例用法:
  python3 印尼協同自動化工具.py --demo                  # 執行完整示範
  python3 印尼協同自動化工具.py --create-tracker        # 建立訂單追蹤表
  python3 印尼協同自動化工具.py --update ORD-2026-062 已到港
  python3 印尼協同自動化工具.py --daily-report          # 每日報告
  python3 印尼協同自動化工具.py --weekly-report         # 週報告
  python3 印尼協同自動化工具.py --template 出貨通知     # 產生訊息範本
  python3 印尼協同自動化工具.py --translate 提單        # 翻譯術語
  python3 印尼協同自動化工具.py --glossary              # 匯出術語表
  python3 印尼協同自動化工具.py --overdue               # 顯示逾期訂單
        """
    )

    parser.add_argument("--create-tracker", action="store_true",
                        help="建立新的訂單進度追蹤表 (含範例資料)")
    parser.add_argument("--update", nargs=2, metavar=("ORDER_ID", "STATUS"),
                        help="更新訂單狀態 (例: --update ORD-2026-062 已到港)")
    parser.add_argument("--daily-report", action="store_true",
                        help="產生每日報告")
    parser.add_argument("--weekly-report", action="store_true",
                        help="產生週報告")
    parser.add_argument("--template", type=str,
                        help="產生訊息範本 (可用: 出貨通知, 清關文件確認, 到港通知, 異常回報, 文件到期提醒, 詢價請求, 付款確認, 週報摘要, 催促回覆, 節日問候)")
    parser.add_argument("--translate", type=str,
                        help="翻譯海關/貿易術語 (中文→印尼文)")
    parser.add_argument("--glossary", action="store_true",
                        help="匯出完整 CN/ID/EN 術語表")
    parser.add_argument("--overdue", action="store_true",
                        help="顯示逾期/卡住的訂單")
    parser.add_argument("--demo", action="store_true",
                        help="執行完整示範")
    parser.add_argument("--output-dir", type=str, default=".",
                        help="輸出目錄 (預設: 當前目錄)")

    args = parser.parse_args()
    tool = IndonesiaCollabTool(output_dir=args.output_dir)

    if args.demo:
        tool.run_demo()
        return

    if args.create_tracker:
        orders = tool._get_sample_orders()
        filepath = tool.tracker.create_tracker(orders)
        print(f"✓ 訂單追蹤表已建立: {filepath}")
        print(f"  包含 {len(orders)} 筆範例訂單")
        return

    if args.update:
        order_id, new_status = args.update
        if new_status not in STATUS_PIPELINE:
            print(f"✗ 無效狀態: {new_status}")
            print(f"  可用狀態: {', '.join(STATUS_PIPELINE)}")
            return
        result = tool.tracker.update_status(order_id, new_status)
        if "error" in result:
            print(f"✗ {result['error']}")
        else:
            print(f"✓ 訂單 {result['order_id']} 狀態已更新:")
            print(f"  {result['old_status']} → {result['new_status']}")
            print(f"  更新時間: {result['timestamp']}")
        return

    if args.daily_report:
        daily = tool.tracker.get_daily_report()
        print(f"每日報告 — {daily['report_date']}")
        print(f"訂單總數: {daily['total_orders']}")
        print(f"\n今日到港: {len(daily['today_arrivals'])} 筆")
        for o in daily["today_arrivals"]:
            print(f"  → {o['order_id']} ({o['customer']})")
        print(f"\n清關中: {len(daily['in_customs'])} 筆")
        for o in daily["in_customs"]:
            print(f"  ⚠ {o['order_id']} ({o['customer']}) — {o.get('anomaly', '正常')}")
        print(f"\n需關注: {len(daily['attention_needed'])} 筆")
        for o in daily["attention_needed"]:
            print(f"  🔴 {o['order_id']}: {o.get('anomaly', '')}")
        print(f"\n狀態統計:")
        for s in STATUS_PIPELINE:
            print(f"  {s}: {daily['summary'].get(s, 0)}")
        return

    if args.weekly_report:
        weekly = tool.tracker.get_weekly_report()
        report_text = tool.report_gen.generate_report(weekly)
        print(report_text)
        excel_path = tool.report_gen.export_to_excel(weekly)
        print(f"\n✓ 週報 Excel 已匯出: {excel_path}")
        wecom_summary = tool.report_gen.generate_wecom_summary(weekly)
        print(f"\nWeCom 摘要:\n{wecom_summary}")
        return

    if args.template:
        template_name = args.template
        available = tool.msg_engine.list_templates()
        if template_name not in available:
            print(f"✗ 找不到範本: {template_name}")
            print(f"  可用範本: {', '.join(available)}")
            return

        # Use sample data for the template
        sample_data_map = {
            "出貨通知": {
                "order_id": "ORD-2026-062", "ship_date": "2026-06-12",
                "bl_number": "BKKT2606123", "eta": "2026-06-22",
                "destination": "泗水 Tanjung Perak", "vessel": "MV Pacific Star",
                "product": "工業閥門 DN80", "quantity": "200 PCS",
            },
            "清關文件確認": {"order_id": "ORD-2026-065", "deadline": "2026-06-18"},
            "到港通知": {
                "order_id": "ORD-2026-064", "vessel": "MV Nusantara Jaya",
                "eta": "2026-06-15", "destination": "棉蘭 Belawan",
                "bl_number": "BKKT2606085",
            },
            "異常回報": {
                "order_id": "ORD-2026-068", "issue_type": "清關延誤",
                "description": "清關已超過8天，需補充材質證明", "status": "清關中",
                "days": "8", "resolution": "聯繫總部補發文件", "deadline": "2026-06-18",
                "severity": "嚴重",
            },
            "文件到期提醒": {
                "doc_name": "海關登記證(NIB)", "doc_type": "海關登記",
                "expiry_date": "2026-07-01", "days_remaining": "15",
                "renewal_steps": "聯繫海關辦公室辦理續期", "owner": "李專員",
            },
            "詢價請求": {
                "product": "不鏽鋼管件 DN200", "specs": "ASTM A312 TP304",
                "quantity": "300 PCS", "destination": "雅加達",
                "deadline": "2026-06-20",
            },
            "付款確認": {
                "order_id": "ORD-2026-062", "currency": "USD", "amount": "45,000",
                "reference": "WIRE-20260612-001", "payment_date": "2026-06-12",
                "payment_method": "電匯 T/T",
            },
            "週報摘要": {
                "date_range": "2026-06-10 ~ 2026-06-16", "total": "10",
                "in_production": "2", "in_transit": "3", "in_customs": "2",
                "delivered": "3", "anomaly_count": "2",
                "anomaly_details": "- ORD-2026-068: 清關逾期\n- ORD-2026-065: 清關關注",
                "action_items": "- 跟進 ORD-2026-068 材質證明\n- 確認 ORD-2026-065 查驗結果",
                "forecast": "- ORD-2026-063 預計 6/20 到港\n- ORD-2026-062 預計 6/22 到港",
            },
            "催促回覆": {
                "original_request": "清關文件確認", "order_id": "ORD-2026-065",
                "days_pending": "3", "urgency": "高",
            },
            "節日問候": {
                "holiday_cn": "印尼獨立日", "holiday_id": "Hari Kemerdekaan RI",
                "holiday_date": "2026-08-17",
            },
        }
        data = sample_data_map.get(template_name, {})
        result = tool.msg_engine.fill_template(template_name, data)
        if "error" in result:
            print(f"✗ {result['error']}")
        else:
            print(f"【{template_name}】\n")
            print("── 中文版 ──")
            print(result["chinese"])
            print()
            print("── Versi Indonesia ──")
            print(result["indonesian"])
        return

    if args.translate:
        term = args.translate
        result = tool.translator.translate_term(term, from_lang="zh", to_lang="id")
        print(f"術語翻譯:")
        print(f"  中文: {term}")
        print(f"  印尼文: {result['translation']}")
        result_en = tool.translator.translate_term(term, from_lang="zh", to_lang="en")
        print(f"  英文: {result_en['translation']}")
        return

    if args.glossary:
        glossary = tool.translator.get_glossary()
        print(f"海關/貿易術語表 (共 {len(glossary)} 筆)")
        print("=" * 80)
        print(f"{'中文':<16} {'印尼文':<36} {'英文'}")
        print("-" * 80)
        for entry in glossary:
            print(f"{entry['中文']:<16} {entry['印尼文'][:34]:<36} {entry['英文'][:30]}")
        return

    if args.overdue:
        overdue = tool.tracker.get_overdue_items()
        if overdue:
            print(f"逾期/卡住的訂單 (共 {len(overdue)} 筆)")
            print("-" * 60)
            for item in overdue:
                severity_icon = "🔴" if item["severity"] == "嚴重" else "🟡"
                print(f"{severity_icon} {item['order_id']} — {item['status']}")
                print(f"  已停留 {item['days_stuck']} 天 (閾值: {item['threshold']} 天)")
                print(f"  嚴重程度: {item['severity']}")
                print(f"  客戶: {item['customer']}")
                print()
        else:
            print("✓ 無逾期訂單")
        return

    # No arguments
    parser.print_help()


if __name__ == "__main__":
    main()
