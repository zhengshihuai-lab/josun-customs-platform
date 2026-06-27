#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""建立範例輸入檔案：產品主數據庫 + 採購訂單"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os

BASE = os.path.dirname(os.path.abspath(__file__))

# ============================================================
# 1. 產品主數據庫
# ============================================================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Product Master"

headers = [
    "產品內部編碼", "中文品名", "英文品名", "型號規格", "HS編碼",
    "原產地", "單位", "淨重kg", "毛重kg", "長cm", "寬cm", "高cm",
    "單價USD", "申報要素描述_中文", "申報要素描述_英文",
    "品牌", "材質", "用途", "功能描述_中文", "功能描述_英文"
]

products = [
    [
        "YD-EL-001", "工業級壓力感測器", "Industrial Pressure Sensor",
        "YD-PS-200A", "9031809090", "中國", "PCS",
        0.35, 0.50, 12, 8, 6,
        128.00,
        "用途:工業壓力測量|功能:將壓力信號轉換為4-20mA電信號|品牌:御大|型號:YD-PS-200A|材質:316L不鏽鋼",
        "Usage:Industrial pressure measurement|Function:Converts pressure signal to 4-20mA electrical signal|Brand:YUDA|Model:YD-PS-200A|Material:316L Stainless Steel",
        "YUDA", "316L不鏽鋼", "工業壓力測量",
        "將壓力信號轉換為4-20mA標準電信號輸出，量程0-10MPa，精度±0.25%FS",
        "Converts pressure signal to 4-20mA standard output, range 0-10MPa, accuracy ±0.25%FS"
    ],
    [
        "YD-EL-002", "數位溫度控制器", "Digital Temperature Controller",
        "YD-TC-500B", "9032100000", "中國", "PCS",
        0.80, 1.10, 20, 15, 10,
        245.00,
        "用途:工業溫度控制|功能:PID自動調節控制|品牌:御大|型號:YD-TC-500B|材質:鋁合金外殼",
        "Usage:Industrial temperature control|Function:PID automatic regulation control|Brand:YUDA|Model:YD-TC-500B|Material:Aluminum alloy housing",
        "YUDA", "鋁合金", "工業溫度控制",
        "PID智能溫控，範圍-50~400°C，精度±0.5°C，RS485通訊接口",
        "PID intelligent temperature control, range -50~400°C, accuracy ±0.5°C, RS485 communication"
    ],
    [
        "YD-EL-003", "超聲波流量計", "Ultrasonic Flow Meter",
        "YD-UF-800C", "9026100000", "中國", "SET",
        3.50, 4.80, 35, 25, 20,
        890.00,
        "用途:液體流量測量|功能:超聲波時差法測量流速|品牌:御大|型號:YD-UF-800C|材質:碳鋼/不鏽鋼",
        "Usage:Liquid flow measurement|Function:Ultrasonic transit-time flow velocity measurement|Brand:YUDA|Model:YD-UF-800C|Material:Carbon steel/Stainless steel",
        "YUDA", "碳鋼/不鏽鋼", "液體流量測量",
        "超聲波時差法測量，管徑DN50-3000，精度±1.0%，IP68防護等級",
        "Ultrasonic transit-time measurement, pipe DN50-3000, accuracy ±1.0%, IP68 protection"
    ],
    [
        "YD-MC-004", "電磁閥組件", "Solenoid Valve Assembly",
        "YD-SV-120D", "8481804090", "中國", "PCS",
        1.20, 1.60, 18, 12, 10,
        67.50,
        "用途:流體管路控制|功能:電磁驅動開關控制|品牌:御大|型號:YD-SV-120D|材質:黃銅",
        "Usage:Fluid pipeline control|Function:Electromagnetic drive on/off control|Brand:YUDA|Model:YD-SV-120D|Material:Brass",
        "YUDA", "黃銅", "流體管路控制",
        "二位二通電磁閥，口徑DN15-50，工作壓力0-1.0MPa，AC220V/DC24V可選",
        "2-way solenoid valve, bore DN15-50, working pressure 0-1.0MPa, AC220V/DC24V optional"
    ],
    [
        "YD-OP-005", "不鏽鋼法蘭接頭", "Stainless Steel Flange Connector",
        "YD-FL-300E", "7307210000", "中國", "PCS",
        2.80, 3.20, 22, 22, 5,
        35.00,
        "用途:管路連接|功能:法蘭密封連接|品牌:御大|型號:YD-FL-300E|材質:304不鏽鋼",
        "Usage:Pipeline connection|Function:Flange sealed connection|Brand:YUDA|Model:YD-FL-300E|Material:304 Stainless steel",
        "YUDA", "304不鏽鋼", "管路連接",
        "DN25-200法蘭連接件，PN10/16/25，RF密封面，符合HG/T20592標準",
        "DN25-200 flange connector, PN10/16/25, RF sealing face, HG/T20592 standard"
    ],
]

# Write headers
header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font_white = Font(bold=True, size=11, color="FFFFFF")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

for row_idx, prod in enumerate(products, 2):
    for col_idx, val in enumerate(prod, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)

# Auto-adjust column widths
for col in ws.columns:
    max_len = 0
    col_letter = col[0].column_letter
    for cell in col:
        if cell.value:
            max_len = max(max_len, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)

wb.save(os.path.join(BASE, "產品主數據庫_Product_Master.xlsx"))
print("✓ 產品主數據庫已建立")

# ============================================================
# 2. 範例採購訂單
# ============================================================
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = "採購訂單"

# Order header info (A1:B10 key-value pairs)
header_info = [
    ("訂單編號", "PO-2026-0616-001"),
    ("訂單日期", "2026-06-16"),
    ("客戶名稱", "SOUTHEAST ASIA INSTRUMENTATION SDN BHD"),
    ("客戶地址", "No. 88, Jalan Teknologi 3/5, Taman Sains Selangor, 47810 Petaling Jaya, Selangor, Malaysia"),
    ("目的港", "Port Klang, Malaysia"),
    ("貿易條款", "CIF Port Klang"),
    ("付款方式", "T/T 30% deposit, 70% against B/L copy"),
    ("合同編號", "CT-YD-2026-0158"),
    ("預計出貨日", "2026-07-15"),
    ("備註", "Please mark shipping marks as per buyer's instruction"),
]

key_font = Font(bold=True, size=11)
val_font = Font(size=11)
info_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

for row_idx, (key, val) in enumerate(header_info, 1):
    cell_k = ws2.cell(row=row_idx, column=1, value=key)
    cell_k.font = key_font
    cell_k.fill = info_fill
    cell_k.border = thin_border
    cell_k.alignment = Alignment(vertical='center')

    cell_v = ws2.cell(row=row_idx, column=2, value=val)
    cell_v.font = val_font
    cell_v.border = thin_border
    cell_v.alignment = Alignment(vertical='center', wrap_text=True)

# Detail header at row 11 (row index 10 = row 11 in sheet)
detail_start_row = 11
detail_headers = [
    "序號", "產品內部編碼", "中文品名", "英文品名", "型號規格",
    "數量", "單位", "單價USD", "金額USD", "備註"
]

for col_idx, h in enumerate(detail_headers, 1):
    cell = ws2.cell(row=detail_start_row, column=col_idx, value=h)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# Detail items
items = [
    (1, "YD-EL-001", "工業級壓力感測器", "Industrial Pressure Sensor", "YD-PS-200A", 200, "PCS", 128.00, 25600.00, ""),
    (2, "YD-EL-002", "數位溫度控制器", "Digital Temperature Controller", "YD-TC-500B", 100, "PCS", 245.00, 24500.00, "需附出廠檢驗報告"),
    (3, "YD-EL-003", "超聲波流量計", "Ultrasonic Flow Meter", "YD-UF-800C", 30, "SET", 890.00, 26700.00, ""),
    (4, "YD-MC-004", "電磁閥組件", "Solenoid Valve Assembly", "YD-SV-120D", 500, "PCS", 67.50, 33750.00, ""),
    (5, "YD-OP-005", "不鏽鋼法蘭接頭", "Stainless Steel Flange Connector", "YD-FL-300E", 300, "PCS", 35.00, 10500.00, "DN50 PN16"),
]

for row_offset, item in enumerate(items):
    row = detail_start_row + 1 + row_offset
    for col_idx, val in enumerate(item, 1):
        cell = ws2.cell(row=row, column=col_idx, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center')
        if col_idx in (8, 9) and isinstance(val, (int, float)):
            cell.number_format = '#,##0.00'

# Column widths
ws2.column_dimensions['A'].width = 16
ws2.column_dimensions['B'].width = 50
ws2.column_dimensions['C'].width = 16
ws2.column_dimensions['D'].width = 16
ws2.column_dimensions['E'].width = 16
ws2.column_dimensions['F'].width = 10
ws2.column_dimensions['G'].width = 10
ws2.column_dimensions['H'].width = 14
ws2.column_dimensions['I'].width = 16
ws2.column_dimensions['J'].width = 25

wb2.save(os.path.join(BASE, "範例採購訂單_Sample_PO.xlsx"))
print("✓ 範例採購訂單已建立")
