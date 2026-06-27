#!/usr/bin/env python3
"""
HS Code Classification Database Builder
Creates a comprehensive HS code historical classification database Excel file.
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from datetime import date, timedelta
import random

wb = openpyxl.Workbook()

# ── Color & Style Constants ──────────────────────────────────────────
DARK_BLUE = "1F4E79"
WHITE = "FFFFFF"
LIGHT_BLUE = "D6E4F0"
LIGHT_GRAY = "F2F2F2"
LIGHT_YELLOW = "FFF2CC"
LIGHT_RED = "FCE4EC"
LIGHT_GREEN = "E8F5E9"
BORDER_COLOR = "B4C6E7"

title_font = Font(name="Microsoft JhengHei", bold=True, color=WHITE, size=11)
title_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
title_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

data_font = Font(name="Microsoft JhengHei", size=10)
data_align = Alignment(vertical="center", wrap_text=True)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
number_align = Alignment(horizontal="right", vertical="center")
date_align = Alignment(horizontal="center", vertical="center")

thin_border = Border(
    left=Side(style="thin", color=BORDER_COLOR),
    right=Side(style="thin", color=BORDER_COLOR),
    top=Side(style="thin", color=BORDER_COLOR),
    bottom=Side(style="thin", color=BORDER_COLOR),
)

# ── Helper Functions ─────────────────────────────────────────────────
def style_header(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = title_font
        cell.fill = title_fill
        cell.alignment = title_align
        cell.border = thin_border

def style_data_cell(cell, row_idx, is_center=False, is_number=False, is_date=False):
    cell.font = data_font
    cell.border = thin_border
    if is_center:
        cell.alignment = center_align
    elif is_number:
        cell.alignment = number_align
    elif is_date:
        cell.alignment = date_align
        cell.number_format = "YYYY-MM-DD"
    else:
        cell.alignment = data_align
    # Alternating row colors
    if row_idx % 2 == 0:
        cell.fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")

def apply_status_fill(cell, status):
    if status == "已確認":
        cell.fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
    elif status == "海關已核准":
        cell.fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
    elif status == "待確認":
        cell.fill = PatternFill(start_color=LIGHT_YELLOW, end_color=LIGHT_YELLOW, fill_type="solid")

def apply_risk_fill(cell, risk):
    if risk == "低風險":
        cell.fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
    elif risk == "中風險":
        cell.fill = PatternFill(start_color=LIGHT_YELLOW, end_color=LIGHT_YELLOW, fill_type="solid")
    elif risk == "高風險":
        cell.fill = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid")


# =====================================================================
# SHEET 1: HS歸類記錄
# =====================================================================
ws1 = wb.active
ws1.title = "HS歸類記錄"

headers1 = [
    "記錄編號", "產品描述_中文", "產品描述_英文", "材質", "用途",
    "功能描述", "加工方式", "已歸類HS編碼", "HS編碼章節",
    "歸類依據", "中國進口關稅稅率%", "泰國進口關稅稅率%",
    "印尼進口關稅稅率%", "增值稅率%", "申報要素_中文",
    "申報要素_英文", "歸類日期", "確認狀態", "風險標記",
    "風險說明", "備註"
]
for col, h in enumerate(headers1, 1):
    ws1.cell(row=1, column=col, value=h)

style_header(ws1, len(headers1))

# ── 40 Records ───────────────────────────────────────────────────────
base_date = date(2024, 1, 15)
records = [
    # ── 電機電氣類（第85章）──
    {
        "id": "HS-001",
        "cn": "電路板模組",
        "en": "Circuit Board Module",
        "material": "FR-4玻璃纖維基材",
        "use": "工業控制設備",
        "func": "信號處理與傳輸",
        "process": "SMT貼片焊接",
        "hs": "8534.00",
        "chapter": "第85章 電機電氣設備",
        "basis": "歸類總規則一：品目8534明確列名「印刷電路」；歸類總規則六：子目8534.00",
        "cn_tariff": 0, "th_tariff": 0, "id_tariff": 0,
        "vat_cn": 13, "vat_th": 7, "vat_id": 11,
        "decl_cn": "品名|用途|材質|品牌|型號",
        "decl_en": "Product Name|Application|Material|Brand|Model",
        "date": date(2024, 1, 15), "status": "已確認", "risk": "低風險",
        "risk_note": "", "note": "PCB空板歸8534，已貼裝元器件的PCBA歸8537或對應整机品目"
    },
    {
        "id": "HS-002",
        "cn": "電力電纜 3芯16mm²",
        "en": "Power Cable 3-core 16mm²",
        "material": "銅芯導體/XLPE交聯聚乙烯絕緣/PVC護套",
        "use": "電力傳輸配電系統",
        "func": "傳輸和分配電能",
        "process": "絞合/擠包絕緣/成纜/擠包護套",
        "hs": "8544.49",
        "chapter": "第85章 電機電氣設備",
        "basis": "歸類總規則一：品目8544列名「絕緣電線電纜」；歸類總規則六：子目8544.49其他導體",
        "cn_tariff": 6.0, "th_tariff": 10.0, "id_tariff": 10.0,
        "vat_cn": 13, "vat_th": 7, "vat_id": 11,
        "decl_cn": "品名|用途|材質|額定電壓|芯數|截面積|品牌|型號",
        "decl_en": "Product Name|Application|Material|Rated Voltage|Cores|Cross-section|Brand|Model",
        "date": date(2024, 2, 8), "status": "已確認", "risk": "低風險",
        "risk_note": "", "note": "額定電壓超過1000V歸8544.60"
    },
    {
        "id": "HS-003",
        "cn": "電動馬達 550W",
        "en": "Electric Motor 550W",
        "material": "鑄鐵外殼/矽鋼片定轉子/銅線繞組",
        "use": "驅動泵、風機等旋轉設備",
        "func": "將電能轉化為旋轉機械能",
        "process": "鑄造/沖壓/繞線/組裝",
        "hs": "8501.52",
        "chapter": "第85章 電機電氣設備",
        "basis": "歸類總規則一：品目8501列名「電動機」；歸類總規則六：子目8501.52輸出功率750W以下多相交流電動機",
        "cn_tariff": 8.0, "th_tariff": 10.0, "id_tariff": 5.0,
        "vat_cn": 13, "vat_th": 7, "vat_id": 11,
        "decl_cn": "品名|用途|功率|相數|品牌|型號",
        "decl_en": "Product Name|Application|Power|Phase|Brand|Model",
        "date": date(2024, 2, 20), "status": "已確認", "risk": "低風險",
        "risk_note": "", "note": "550W屬750W以下多相交流電動機子目"
    },
    {
        "id": "HS-004",
        "cn": "工業用LED燈具",
        "en": "Industrial LED Lamp",
        "material": "鋁合金散熱器+PC透鏡",
        "use": "工廠車間照明",
        "func": "提供LED照明光源",
        "process": "壓鑄/注塑/SMT/組裝",
        "hs": "8539.50",
        "chapter": "第85章 電機電氣設備",
        "basis": "歸類總規則一：品目8539列名「白熾燈泡或放電燈管…LED燈泡」；歸類總規則六：子目8539.50發光二極體(LED)燈",
        "cn_tariff": 5.0, "th_tariff": 10.0, "id_tariff": 5.0,
        "vat_cn": 13, "vat_th": 7, "vat_id": 11,
        "decl_cn": "品名|用途|光源類型|功率|品牌|型號",
        "decl_en": "Product Name|Application|Light Source|Wattage|Brand|Model",
        "date": date(2024, 3, 5), "status": "已確認", "risk": "低風險",
        "risk_note": "", "note": "LED光源歸8539.50，LED驅動電源單獨歸8504"
    },
    {
        "id": "HS-005",
        "cn": "變頻器 7.5kW",
        "en": "Frequency Inverter 7.5kW",
        "material": "金屬外殼/IGBT模組/PCB控制板",
        "use": "交流電機變頻調速控制",
        "func": "將固定頻率交流電轉換為可調頻率交流電",
        "process": "SMT/波峰焊/組裝/調試",
        "hs": "8504.40",
        "chapter": "第85章 電機電氣設備",
        "basis": "歸類總規則一：品目8504列名「靜態變流器」；歸類總規則六：子目8504.40靜態變流器",
        "cn_tariff": 8.0, "th_tariff": 10.0, "id_tariff": 0,
        "vat_cn": 13, "vat_th": 7, "vat_id": 11,
        "decl_cn": "品名|用途|功率|品牌|型號",
        "decl_en": "Product Name|Application|Power|Brand|Model",
        "date": date(2024, 3, 18), "status": "已確認", "risk": "低風險",
        "risk_note": "", "note": "變頻器屬靜態變流器歸8504.40"
    },
    {
        "id": "HS-006",
        "cn": "溫度傳感器",
        "en": "Temperature Sensor",
        "material": "不鏽鋼探頭/熱敏電阻/矽膠線",
        "use": "工業設備溫度監測與控制",
        "func": "將溫度信號轉換為電信號輸出",
        "process": "精密焊接/封裝/校準",
        "hs": "9025.80",
        "chapter": "第90章 儀器儀表",
        "basis": "歸類總規則一：品目9025列名「溫度計、高溫計…」；歸類總規則六：子目9025.80其他儀器",
        "cn_tariff": 0, "th_tariff": 5.0, "id_tariff": 5.0,
        "vat_cn": 13, "vat_th": 7, "vat_id": 11,
        "decl_cn": "品名|用途|原理|測量範圍|品牌|型號",
        "decl_en": "Product Name|Application|Principle|Range|Brand|Model",
        "date": date(2024, 4, 2), "status": "已確認", "risk": "低風險",
        "risk_note": "", "note": "屬第90章儀器儀表，雖材質為不鏽鋼但按功能歸類"
    },
    {
        "id": "HS-007",
        "cn": "配電箱",
        "en": "Distribution Box",
        "material": "冷軋鋼板箱體/銅母線排/斷路器",
        "use": "低壓電力分配與保護",
        "func": "接收電能並分配到各用電回路",
        "process": "鈑金加工/噴塗/組裝配線",
        "hs": "8537.10",
        "chapter": "第85章 電機電氣設備",
        "basis": "歸類總規則一：品目8537列名「配電盤、配電箱…」；歸類總規則六：子目8537.10額定電壓不超過1000V",
        "cn_tariff": 8.4, "th_tariff": 10.0, "id_tariff": 10.0,
        "vat_cn": 13, "vat_th": 7, "vat_id": 11,
        "decl_cn": "品名|用途|額定電壓|品牌|型號",
        "decl_en": "Product Name|Application|Rated Voltage|Brand|Model",
        "date": date(2024, 4, 15), "status": "已確認", "risk": "低風險",
        "risk_note": "", "note": "低壓配電箱歸8537.10；高壓歸8537.20"
    },
    {
        "id": "HS-008",
        "cn": "太陽能電池板",
        "en": "Solar Panel",
        "material": "單晶矽晶圓/鋼化玻璃/EVA封裝/鋁合金框",
        "use": "太陽能光伏發電系統",
        "func": "將太陽能光輻射轉換為直流電能",
        "process": "晶片製造/電池片焊接/層壓/裝框",
        "hs": "8541.43",
        "chapter": "第85章 電機電氣設備",
        "basis": "歸類總規則一：品目8541列名「光敏半導體器件…光伏電池」；歸類總規則六：子目8541.43光伏電池組裝成模組",
        "cn_tariff": 0, "th_tariff": 0, "id_tariff": 0,
        "vat_cn": 13, "vat_th": 7, "vat_id": 11,
        "decl_cn": "品名|用途|類型|功率|品牌|型號",
        "decl_en": "Product Name|Application|Type|Power|Brand|Model",
        "date": date(2024, 5, 1), "status": "海關已核准", "risk": "低風險",
        "risk_note": "", "note": "2022版HS將光伏電池移至8541.43"
    },

    # ── 金屬製品類（第73章）──
    {
        "id": "HS-009",
        "cn": "不鏽鋼管件 DN50",
        "en": "Stainless Steel Fitting DN50",
        "material": "304不鏽鋼(0Cr18Ni9)",
        "use": "工業管道系統連接",
        "func": "管道彎頭、三通、異徑等連接",
        "process": "鍛造/機加工/拋光",
        "hs": "7307.23",
        "chapter": "第73章 鋼鐵製品",
        "basis": "歸類總規則一：品目7307列名「鋼鐵管配件」；歸類總規則六：子目7307.23不鏽鋼對焊配件",
        "cn_tariff": 8.0, "th_tariff": 10.0, "id_tariff": 10.0,
        "vat_cn": 13, "vat_th": 7, "vat_id": 11,
        "decl_cn": "品名|用途|材質|規格|加工方式|品牌",
        "decl_en": "Product Name|Application|Material|Specification|Process|Brand",
        "date": date(2024, 1, 25), "status": "已確認", "risk": "低風險",
        "risk_note": "", "note": "對焊配件歸7307.23；法蘭歸7307.21"
    },
    {
        "id": "HS-010",
        "cn": "法蘭盤 DN80 PN16",
        "en": "Flange DN80 PN16",
        "material": "304不鏽鋼鍛件",
        "use": "管道法蘭連接",
        "func": "管道間法蘭密封連接",
        "process": "鍛造/車削/鑽孔",
        "hs": "7307.21",
        "chapter": "第73章 鋼鐵製品",
        "basis": "歸類總規則一：品目7307列名「鋼鐵管配件」；歸類總規則六：子目7307.21不鏽鋼法蘭",
        "cn_tariff": 8.0, "th_tariff": 10.0, "id_tariff": 10.0,
        "vat_cn": 13, "vat_th": 7, "vat_id": 11,
        "decl_cn": "品名|用途|材質|規格|壓力等級|品牌",
        "decl_en": "Product Name|Application|Material|Size|Pressure Rating|Brand",
        "date": date(2024, 2, 12), "status": "已確認", "risk": "低風險",
        "risk_note": "", "note": "法蘭明確列名於7307.21"
    },
    {
        "id": "HS-011",
        "cn": "碳鋼無縫管",
        "en": "Carbon Steel Seamless Pipe",
        "material": "20#碳鋼",
        "use": "工業流體輸送管道系統",
        "func": "輸送水、氣體、油等流體介質",
        "process": "熱軋穿孔/冷拔/退火/矯直",
        "hs": "7304.39",
        "chapter": "第73章 鋼鐵製品",
        "basis": "歸類總規則一：品目7304列名「無縫管（鐵或非合金鋼）」；歸類總規則六：子目7304.39其他",
        "cn_tariff": 5.0, "th_tariff": 5.0, "id_tariff": 5.0,
        "vat_cn": 13, "vat_th": 7, "vat_id": 11,
        "decl_cn": "品名|用途|材質|外徑|壁厚|執行標準",
        "decl_en": "Product Name|Application|Material|OD|Wall Thickness|Standard",
        "date": date(2024, 3, 1), "status": "已確認", "risk": "低風險",
        "risk_note": "", "note": "20#碳鋼屬非合金鋼歸7304；合金鋼歸7304.51/59"
    },
    {
        "id": "HS-012",
        "cn": "工業水箱 5000L",
        "en": "Industrial Water Tank 5000L",
        "material": "304不鏽鋼板材",
        "use": "工業用水儲存",
        "func": "儲存和供應工業用水",
        "process": "切割/折彎/焊接/拋光",
        "hs": "7310.29",
        "chapter": "第73章 鋼鐵製品",
        "basis": "歸類總規則一：品目7310列名「鋼鐵製容器…容積不超過300升」；歸類總規則三(乙)：按構成特徵歸入7310.29其他",
        "cn_tariff": 10.0, "th_tariff": 20.0, "id_tariff": 15.0,
        "vat_cn": 13, "vat_th": 7, "vat_id": 11,
        "decl_cn": "品名|用途|材質|容積|是否裝有機械或熱力裝置|品牌",
        "decl_en": "Product Name|Application|Material|Volume|Mechanical/Thermal Device|Brand",
        "date": date(2024, 3, 22), "status": "待確認", "risk": "高風險",
        "risk_note": "5000L超過300L，容易誤歸7309(容積>300L儲液罐)或7310(容積≤300L容器)。需確認7309與7310的分界點：7309為容積超過300升的儲液罐；此水箱5000L應歸7309.00而非7310",
        "note": "【高風險】5000L超過300L，應歸7309.00而非7310.29。7309為容積>300升鋼鐵製儲液罐。此記錄需重新確認歸類"
    },
    {
        "id": "HS-013",
        "cn": "鋼結構件",
        "en": "Steel Structure Component",
        "material": "Q235B碳素結構鋼",
        "use": "工業廠房建築結構支撐",
        "func": "承受和傳遞建築荷載",
        "process": "切割/焊接/鑽孔/熱鍍鋅",
        "hs": "7308.90",
        "chapter": "第73章 鋼鐵製品",
        "basis": "歸類總規則一：品目7308列名「鋼鐵結構體及其部件」；歸類總規則六：子目7308.90其他",
        "cn_tariff": 8.0, "th_tariff": 10.0, "id_tariff": 10.0,
        "vat_cn": 13, "vat_th": 7, "vat_id": 11,
        "decl_cn": "品名|用途|材質|規格|表面處理",
        "decl_en": "Product Name|Application|Material|Specification|Surface Treatment",
        "date": date(2024, 4, 8), "status": "已確認", "risk": "中風險",
        "risk_note": "鋼結構件品種繁多，需確認是否為完整結構體或僅為部件。預製建築歸9406",
        "note": "僅為部件歸7308.90；完整預製建築歸9406.00"
    },

    # ── 機械設備類（第84章）──
    {
        "id": "HS-014",
        "cn": "離心泵",
        "en": "Centrifugal Pump",
        "material": "鑄鐵泵體/不鏽鋼葉輪",
        "use": "工業液體輸送循環系統",
        "func": "利用離心力將機械能轉化為液體動能和壓力能",
        "process": "鑄造/精密機加工/動平衡/組裝",
        "hs": "8413.70",
        "chapter": "第84章 機械設備",
        "basis": "歸類總規則一：品目8413列名「液體泵」；歸類總規則六：子目8413.70其他離心泵",
        "cn_tariff": 8.0, "th_tariff": 10.0, "id_tariff": 5.0,
        "vat_cn": 13, "vat_th": 7, "vat_id": 11,
        "decl_cn": "品名|用途|驅動方式|流量|揚程|品牌|型號",
        "decl_en": "Product Name|Application|Drive Type|Flow Rate|Head|Brand|Model",
        "date": date(2024, 1, 30), "status": "已確認", "risk": "低風險",
        "risk_note": "", "note": "離心泵歸8413.70；容積泵歸8413.60"
    },
    {
        "id": "HS-015",
        "cn": "閥門組件 DN100",
        "en": "Valve Assembly DN100",
        "material": "304不鏽鋼閥體/PTFE密封",
        "use": "工業管道流量調節與控制",
        "func": "調節、啟閉管道中介質流量",
        "process": "鑄造/機加工/研磨/組裝/試壓",
        "hs": "8481.80",
        "chapter": "第84章 機械設備",
        "basis": "歸類總規則一：品目8481列名「水龍頭、閥門…」；歸類總規則六：子目8481.80其他閥門",
        "cn_tariff": 6.0, "th_tariff": 10.0, "id_tariff": 5.0,
        "vat_cn": 13, "vat_th": 7, "vat_id": 11,
        "decl_cn": "品名|用途|類型|材質|口徑|壓力等級|品牌|型號",
        "decl_en": "Product Name|Application|Type|Material|Size|Pressure|Brand|Model",
        "date": date(2024, 2, 18), "status": "已確認", "risk": "低風險",
        "risk_note": "", "note": "安全閥歸8481.40；止回閥歸8481.30"
    },
    {
        "id": "HS-016",
        "cn": "空氣壓縮機",
        "en": "Air Compressor",
        "material": "鑄鐵缸體/鋼製儲氣罐/電動機",
        "use": "為氣動工具和設備提供壓縮空氣",
        "func": "將機械能轉化為氣體壓力能",
        "process": "鑄造/機加工/組裝/調試",
        "hs": "8414.80",
        "chapter": "第84章 機械設備",
        "basis": "歸類總規則一：品目8414列名「空氣泵或真空泵、空氣或其他氣體壓縮機」；歸類總規則六：子目8414.80其他",
        "cn_tariff": 8.0, "th_tariff": 10.0, "id_tariff": 5.0,
        "vat_cn": 13, "vat_th": 7, "vat_id": 11,
        "decl_cn": "品名|用途|類型|排氣量|壓力|功率|品牌|型號",
        "decl_en": "Product Name|Application|Type|Displacement|Pressure|Power|Brand|Model",
        "date": date(2024, 3, 10), "status": "已確認", "risk": "中風險",
        "risk_note": "螺桿式空壓機組（含乾燥機、過濾器）可能需要按整套設備歸類",
        "note": "單純壓縮機歸8414.80；含乾燥過濾整套設備需評估是否按功能單元歸類"
    },
    {
        "id": "HS-017",
        "cn": "工業冷水機",
        "en": "Industrial Chiller",
        "material": "鈑金外殼/銅管換熱器/壓縮機",
        "use": "工業冷卻循環系統",
        "func": "通過製冷循環將熱量從冷卻水移走",
        "process": "鈑金/焊接/管路組裝/充注冷媒/調試",
        "hs": "8418.69",
        "chapter": "第84章 機械設備",
        "basis": "歸類總規則一：品目8418列名「冷藏或冷凍設備」；歸類總規則六：子目8418.69其他",
        "cn_tariff": 10.0, "th_tariff": 10.0, "id_tariff": 10.0,
        "vat_cn": 13, "vat_th": 7, "vat_id": 11,
        "decl_cn": "品名|用途|製冷量|冷媒類型|品牌|型號",
        "decl_en": "Product Name|Application|Cooling Capacity|Refrigerant|Brand|Model",
        "date": date(2024, 4, 5), "status": "待確認", "risk": "中風險",
        "risk_note": "工業冷水機歸類有爭議：可能歸8418.69(製冷設備)或8419.89(熱交換設備)，需根據製冷量和用途判斷",
        "note": "製冷量較大(>4000大卡/h)歸8418.69；小功率精密冷卻可能歸8419.89"
    },
    {
        "id": "HS-018",
        "cn": "液壓缸",
        "en": "Hydraulic Cylinder",
        "material": "合金鋼缸筒/鍍鉻活塞桿/聚氨酯密封",
        "use": "工程機械直線運動驅動",
        "func": "將液壓能轉換為直線機械運動",
        "process": "無縫管珩磨/鍍鉻/密封組裝/試壓",
        "hs": "8412.21",
        "chapter": "第84章 機械設備",
        "basis": "歸類總規則一：品目8412列名「其他發動機及馬達」；歸類總規則六：子目8412.21液壓動力裝置直線作用(液壓缸)",
        "cn_tariff": 8.0, "th_tariff": 10.0, "id_tariff": 5.0,
        "vat_cn": 13, "vat_th": 7, "vat_id": 11,
        "decl_cn": "品名|用途|作用方式|缸徑|行程|品牌|型號",
        "decl_en": "Product Name|Application|Action|Bore|Stroke|Brand|Model",
        "date": date(2024, 5, 12), "status": "已確認", "risk": "低風險",
        "risk_note": "", "note": "直線作用液壓缸歸8412.21；旋轉液壓馬達歸8412.29"
    },

    # ── 化工類（第39/38/27章）──
    {
        "id": "HS-019",
        "cn": "化工添加劑 聚丙烯酸",
        "en": "Chemical Additive Polyacrylic Acid",
        "material": "聚丙烯酸高分子聚合物",
        "use": "工業循環水處理阻垢分散劑",
        "func": "抑制水垢沉積、分散懸浮物",
        "process": "溶液聚合/中和/稀釋",
        "hs": "3906.90",
        "chapter": "第39章 塑膠及其製品",
        "basis": "歸類總規則一：品目3906列名「丙烯或其他丙烯酸聚合物」；歸類總規則六：子目3906.90其他",
        "cn_tariff": 6.5, "th_tariff": 10.0, "id_tariff": 5.0,
        "vat_cn": 13, "vat_th": 7, "vat_id": 11,
        "decl_cn": "品名|用途|成分含量|外觀|包裝|品牌|型號",
        "decl_en": "Product Name|Application|Composition|Appearance|Packaging|Brand|Model",
        "date": date(2024, 2, 5), "status": "已確認", "risk": "低風險",
        "risk_note": "", "note": "聚丙烯酸歸3906.90；聚丙烯醯胺歸3906.90；聚甲基丙烯酸甲酯歸3906.10"
    },
    {
        "id": "HS-020",
        "cn": "工業潤滑油",
        "en": "Industrial Lubricant",
        "material": "礦物油基礎油/添加劑(抗磨、抗氧化)",
        "use": "機械設備齒輪和軸承潤滑",
        "func": "減少摩擦、冷卻、密封、防鏽",
        "process": "基礎油精煉/添加劑調合/過濾/灌裝",
        "hs": "2710.19",
        "chapter": "第27章 礦物燃料",
        "basis": "歸類總規則一：品目2710列名「石油及從瀝青礦物提取的油類…」；歸類總規則六：子目2710.19其他",
        "cn_tariff": 6.0, "th_tariff": 10.0, "id_tariff": 5.0,
        "vat_cn": 13, "vat_th": 7, "vat_id": 11,
        "decl_cn": "品名|用途|成分含量|包裝|品牌|型號",
        "decl_en": "Product Name|Application|Composition|Packaging|Brand|Model",
        "date": date(2024, 2, 25), "status": "已確認", "risk": "中風險",
        "risk_note": "潤滑油歸2710，但含合成基礎油(如PAO)的配方可能歸3403，需依成分比例判定",
        "note": "礦物油基歸2710.19；合成潤滑油歸3403.99"
    },
    {
        "id": "HS-021",
        "cn": "環氧樹脂",
        "en": "Epoxy Resin",
        "material": "雙酚A型環氧化合物",
        "use": "塗料、粘接劑、複合材料基體",
        "func": "與固化劑反應後形成高強度交聯網絡",
        "process": "縮合聚合/脫溶劑/過濾/包裝",
        "hs": "3907.30",
        "chapter": "第39章 塑膠及其製品",
        "basis": "歸類總規則一：品目3907列名「聚縮醛、其他聚醚及環氧樹脂」；歸類總規則六：子目3907.30環氧樹脂",
        "cn_tariff": 6.5, "th_tariff": 10.0, "id_tariff": 5.0,
        "vat_cn": 13, "vat_th": 7, "vat_id": 11,
        "decl_cn": "品名|用途|環氧當量|外觀|包裝|品牌|型號",
        "decl_en": "Product Name|Application|Epoxy Equivalent|Appearance|Packaging|Brand|Model",
        "date": date(2024, 3, 15), "status": "已確認", "risk": "低風險",
        "risk_note": "", "note": "環氧樹脂明確列名於3907.30"
    },
    {
        "id": "HS-022",
        "cn": "活性炭",
        "en": "Activated Carbon",
        "material": "椰殼/煤質顆粒",
        "use": "工業水處理和空氣淨化",
        "func": "通過多孔結構吸附有機物和雜質",
        "process": "炭化/活化/篩分/包裝",
        "hs": "3801.10",
        "chapter": "第38章 化學工業產品",
        "basis": "歸類總規則一：品目3801列名「活性炭」；歸類總規則六：子目3801.10活性炭",
        "cn_tariff": 6.5, "th_tariff": 5.0, "id_tariff": 5.0,
        "vat_cn": 13, "vat_th": 7, "vat_id": 11,
        "decl_cn": "品名|用途|原料|外觀|碘值|粒度|品牌",
        "decl_en": "Product Name|Application|Raw Material|Appearance|Iodine Value|Particle Size|Brand",
        "date": date(2024, 4, 1), "status": "已確認", "risk": "低風險",
        "risk_note": "", "note": "活性炭明確列名3801.10；廢活性炭歸3825"
    },

    # ── 塑膠橡膠類（第39/40章）──
    {
        "id": "HS-023",
        "cn": "PVC管材",
        "en": "PVC Pipe",
        "material": "PVC聚氯乙烯樹脂/穩定劑/填充劑",
        "use": "建築排水和工業廢水管道",
        "func": "輸送排放液體介質",
        "process": "擠出成型/冷卻定型/切割/擴口",
        "hs": "3917.32",
        "chapter": "第39章 塑膠及其製品",
        "basis": "歸類總規則一：品目3917列名「塑膠管及其配件」；歸類總規則六：子目3917.32其他管（非硬質、未加強也未與其他材料合製）",
        "cn_tariff": 6.5, "th_tariff": 10.0, "id_tariff": 10.0,
        "vat_cn": 13, "vat_th": 7, "vat_id": 11,
        "decl_cn": "品名|用途|材質|外徑|壁厚|壓力等級|品牌",
        "decl_en": "Product Name|Application|Material|OD|Wall Thickness|Pressure|Brand",
        "date": date(2024, 1, 20), "status": "已確認", "risk": "低風險",
        "risk_note": "", "note": "硬質PVC管歸3917.32；加強/複合管歸3917.31"
    },
    {
        "id": "HS-024",
        "cn": "橡膠密封墊",
        "en": "Rubber Gasket",
        "material": "NBR丁腈橡膠",
        "use": "工業管道法蘭和設備密封",
        "func": "在接合面間提供彈性密封防止洩漏",
        "process": "混煉/模壓硫化/修邊",
        "hs": "4016.93",
        "chapter": "第40章 橡膠及其製品",
        "basis": "歸類總規則一：品目4016列名「硫化橡膠(硬質橡膠除外)其他製品」；歸類總規則六：子目4016.93墊圈、密封墊及其他密封件",
        "cn_tariff": 10.0, "th_tariff": 10.0, "id_tariff": 10.0,
        "vat_cn": 13, "vat_th": 7, "vat_id": 11,
        "decl_cn": "品名|用途|材質|規格|硬度|品牌",
        "decl_en": "Product Name|Application|Material|Specification|Hardness|Brand",
        "date": date(2024, 2, 28), "status": "已確認", "risk": "低風險",
        "risk_note": "", "note": "橡膠密封墊明確列名4016.93"
    },
    {
        "id": "HS-025",
        "cn": "塑膠接頭",
        "en": "Plastic Connector",
        "material": "PP聚丙烯",
        "use": "塑膠管路系統連接",
        "func": "管路彎頭、直通、三通連接",
        "process": "注塑成型/去毛刺",
        "hs": "3917.40",
        "chapter": "第39章 塑膠及其製品",
        "basis": "歸類總規則一：品目3917列名「塑膠管及其配件」；歸類總規則六：子目3917.40配件",
        "cn_tariff": 6.5, "th_tariff": 10.0, "id_tariff": 10.0,
        "vat_cn": 13, "vat_th": 7, "vat_id": 11,
        "decl_cn": "品名|用途|材質|規格|品牌",
        "decl_en": "Product Name|Application|Material|Specification|Brand",
        "date": date(2024, 3, 20), "status": "已確認", "risk": "低風險",
        "risk_note": "", "note": "塑膠管配件歸3917.40"
    },

    # ── 建築材料類（第68-70/76章）──
    {
        "id": "HS-026",
        "cn": "瓷磚",
        "en": "Ceramic Tile",
        "material": "陶瓷(粘土/長石/石英)",
        "use": "建築地面和牆面鋪設",
        "func": "裝飾和保護建築表面",
        "process": "壓制成型/施釉/燒成/分級",
        "hs": "6907.21",
        "chapter": "第69章 陶瓷產品",
        "basis": "歸類總規則一：品目6907列名「陶瓷製鋪面磚…」；歸類總規則六：子目6907.21吸水率不超過0.5%",
        "cn_tariff": 7.0, "th_tariff": 20.0, "id_tariff": 15.0,
        "vat_cn": 13, "vat_th": 7, "vat_id": 11,
        "decl_cn": "品名|用途|材質|規格|吸水率|品牌",
        "decl_en": "Product Name|Application|Material|Size|Water Absorption|Brand",
        "date": date(2024, 4, 12), "status": "已確認", "risk": "低風險",
        "risk_note": "", "note": "吸水率≤0.5%歸6907.21；>0.5%歸6907.22"
    },
    {
        "id": "HS-027",
        "cn": "隔熱棉",
        "en": "Insulation Wool",
        "material": "玻璃纖維(離心棉)",
        "use": "建築和管道保溫隔熱",
        "func": "通過纖維間空氣層減少熱量傳遞",
        "process": "離心甩絲/噴膠/烘乾/切割",
        "hs": "7019.39",
        "chapter": "第70章 玻璃及其製品",
        "basis": "歸類總規則一：品目7019列名「玻璃纖維及其製品」；歸類總規則六：子目7019.39其他",
        "cn_tariff": 7.0, "th_tariff": 10.0, "id_tariff": 10.0,
        "vat_cn": 13, "vat_th": 7, "vat_id": 11,
        "decl_cn": "品名|用途|材質|密度|厚度|導熱係數|品牌",
        "decl_en": "Product Name|Application|Material|Density|Thickness|Thermal Conductivity|Brand",
        "date": date(2024, 5, 8), "status": "已確認", "risk": "低風險",
        "risk_note": "", "note": "玻璃纖維隔熱棉歸7019.39；岩棉歸6806.10"
    },
    {
        "id": "HS-028",
        "cn": "鋁合金型材",
        "en": "Aluminum Profile",
        "material": "6063鋁合金(T5狀態)",
        "use": "門窗幕牆框架結構",
        "func": "建築門窗、幕牆的結構框架",
        "process": "擠壓成型/陽極氧化/電泳塗裝",
        "hs": "7604.29",
        "chapter": "第76章 鋁及其製品",
        "basis": "歸類總規則一：品目7604列名「鋁條、桿、型材及異型材」；歸類總規則六：子目7604.29鋁合金製其他",
        "cn_tariff": 6.0, "th_tariff": 10.0, "id_tariff": 5.0,
        "vat_cn": 13, "vat_th": 7, "vat_id": 11,
        "decl_cn": "品名|用途|材質|牌號|規格|表面處理",
        "decl_en": "Product Name|Application|Material|Alloy|Specification|Surface Treatment",
        "date": date(2024, 5, 20), "status": "已確認", "risk": "低風險",
        "risk_note": "", "note": "鋁合金型材歸7604.29；進一步加工成門窗歸7610"
    },

    # ── 儀器儀表類（第90/85/84章）──
    {
        "id": "HS-029",
        "cn": "壓力表",
        "en": "Pressure Gauge",
        "material": "304不鏽鋼錶殼/銅合金彈簧管",
        "use": "工業管道和容器壓力測量",
        "func": "將壓力信號轉換為指針位移顯示",
        "process": "精密機加工/組裝/校準/檢定",
        "hs": "9026.20",
        "chapter": "第90章 儀器儀表",
        "basis": "歸類總規則一：品目9026列名「液體或氣體流量、液位、壓力…測量或檢驗儀器」；歸類總規則六：子目9026.20壓力測量儀器",
        "cn_tariff": 0, "th_tariff": 5.0, "id_tariff": 5.0,
        "vat_cn": 13, "vat_th": 7, "vat_id": 11,
        "decl_cn": "品名|用途|測量範圍|精度等級|接口尺寸|品牌|型號",
        "decl_en": "Product Name|Application|Range|Accuracy|Connection|Brand|Model",
        "date": date(2024, 1, 28), "status": "已確認", "risk": "低風險",
        "risk_note": "", "note": "壓力表明確列名9026.20"
    },
    {
        "id": "HS-030",
        "cn": "流量計",
        "en": "Flow Meter",
        "material": "304不鏽鋼測量管/電子信號處理模組",
        "use": "工業液體流量精確測量",
        "func": "測量管道中液體體積流量並輸出信號",
        "process": "精密機加工/電子組裝/標定校準",
        "hs": "9026.10",
        "chapter": "第90章 儀器儀表",
        "basis": "歸類總規則一：品目9026列名「液體或氣體流量…測量或檢驗儀器」；歸類總規則六：子目9026.10流量測量儀器",
        "cn_tariff": 0, "th_tariff": 5.0, "id_tariff": 5.0,
        "vat_cn": 13, "vat_th": 7, "vat_id": 11,
        "decl_cn": "品名|用途|測量原理|口徑|測量範圍|精度|品牌|型號",
        "decl_en": "Product Name|Application|Principle|Size|Range|Accuracy|Brand|Model",
        "date": date(2024, 2, 15), "status": "已確認", "risk": "低風險",
        "risk_note": "", "note": "流量計明確列名9026.10"
    },
    {
        "id": "HS-031",
        "cn": "工業相機",
        "en": "Industrial Camera",
        "material": "金屬外殼/CMOS圖像傳感器/光學鏡頭",
        "use": "生產線產品品質視覺檢測",
        "func": "將光學圖像轉換為數位信號供圖像處理分析",
        "process": "SMT/光學組裝/調試/校準",
        "hs": "8525.89",
        "chapter": "第85章 電機電氣設備",
        "basis": "歸類總規則一：品目8525列名「電視攝影機、數位相機…」；歸類總規則六：子目8525.89其他電視攝影機",
        "cn_tariff": 0, "th_tariff": 0, "id_tariff": 0,
        "vat_cn": 13, "vat_th": 7, "vat_id": 11,
        "decl_cn": "品名|用途|解析度|幀率|接口類型|品牌|型號",
        "decl_en": "Product Name|Application|Resolution|Frame Rate|Interface|Brand|Model",
        "date": date(2024, 3, 8), "status": "已確認", "risk": "中風險",
        "risk_note": "工業相機可能與機器視覺系統混淆；含圖像處理軟體的整套系統可能歸9031",
        "note": "單獨相機歸8525.89；含視覺系統整套歸9031.49"
    },
    {
        "id": "HS-032",
        "cn": "電子秤",
        "en": "Electronic Scale",
        "material": "不鏽鋼秤台/稱重傳感器/電子顯示模組",
        "use": "工業物料和成品重量稱量",
        "func": "將重量信號轉換為電信號顯示",
        "process": "傳感器標定/電子組裝/計量檢定",
        "hs": "8423.81",
        "chapter": "第84章 機械設備",
        "basis": "歸類總規則一：品目8423列名「衡器(秤)」；歸類總規則六：子目8423.81最大稱量不超過30kg",
        "cn_tariff": 10.0, "th_tariff": 10.0, "id_tariff": 5.0,
        "vat_cn": 13, "vat_th": 7, "vat_id": 11,
        "decl_cn": "品名|用途|最大稱量|精度|品牌|型號",
        "decl_en": "Product Name|Application|Max Capacity|Accuracy|Brand|Model",
        "date": date(2024, 4, 18), "status": "待確認", "risk": "中風險",
        "risk_note": "需區分最大稱量：≤30kg歸8423.81；>30kg歸8423.82；>5000kg歸8423.89",
        "note": "需確認實際最大稱量以確定正確子目"
    },

    # ── 其他常見品項 ──
    {
        "id": "HS-033",
        "cn": "木質包裝箱",
        "en": "Wooden Packing Case",
        "material": "松木/膠合板",
        "use": "工業設備和貨物出口包裝",
        "func": "保護貨物在運輸過程中不受損壞",
        "process": "鋸切/釘製/熏蒸處理",
        "hs": "4415.10",
        "chapter": "第44章 木材及其製品",
        "basis": "歸類總規則一：品目4415列名「包裝箱、盒、匣…」；歸類總規則六：子目4415.10箱、盒、匣",
        "cn_tariff": 7.5, "th_tariff": 10.0, "id_tariff": 10.0,
        "vat_cn": 13, "vat_th": 7, "vat_id": 11,
        "decl_cn": "品名|用途|材質|規格|是否熏蒸處理",
        "decl_en": "Product Name|Application|Material|Size|Fumigation Treatment",
        "date": date(2024, 5, 5), "status": "已確認", "risk": "低風險",
        "risk_note": "", "note": "木質包裝需符合ISPM-15熏蒸要求"
    },
    {
        "id": "HS-034",
        "cn": "工業膠帶",
        "en": "Industrial Tape",
        "material": "PE聚乙烯基材/丙烯酸酯壓敏膠",
        "use": "工業產品粘接固定和保護",
        "func": "提供持久粘接力連接或固定部件",
        "process": "塗布/複合/分切/卷繞",
        "hs": "3919.10",
        "chapter": "第39章 塑膠及其製品",
        "basis": "歸類總規則一：品目3919列名「自粘性塑膠板、片、膜…成卷或成條」；歸類總規則六：子目3919.10成卷寬度不超過20cm",
        "cn_tariff": 6.5, "th_tariff": 10.0, "id_tariff": 10.0,
        "vat_cn": 13, "vat_th": 7, "vat_id": 11,
        "decl_cn": "品名|用途|材質|規格(寬度×長度)|品牌|型號",
        "decl_en": "Product Name|Application|Material|Size(W×L)|Brand|Model",
        "date": date(2024, 5, 15), "status": "已確認", "risk": "低風險",
        "risk_note": "", "note": "寬≤20cm歸3919.10；寬>20cm歸3919.90"
    },
    {
        "id": "HS-035",
        "cn": "焊接材料 焊絲",
        "en": "Welding Wire",
        "material": "低碳鋼芯/鍍銅層",
        "use": "CO2氣體保護電弧焊接",
        "func": "作為填充金屬和導電電極",
        "process": "拉絲/鍍銅/層繞/包裝",
        "hs": "8311.20",
        "chapter": "第83章 雜項金屬製品",
        "basis": "歸類總規則一：品目8311列名「焊條…電弧焊用塗藥或包劑焊條、焊絲」；歸類總規則六：子目8311.20塗藥焊絲",
        "cn_tariff": 8.0, "th_tariff": 10.0, "id_tariff": 5.0,
        "vat_cn": 13, "vat_th": 7, "vat_id": 11,
        "decl_cn": "品名|用途|材質|直徑|執行標準|品牌",
        "decl_en": "Product Name|Application|Material|Diameter|Standard|Brand",
        "date": date(2024, 2, 10), "status": "已確認", "risk": "低風險",
        "risk_note": "", "note": "實芯焊絲和藥芯焊絲均歸8311.20"
    },
    {
        "id": "HS-036",
        "cn": "工業螺栓",
        "en": "Industrial Bolt",
        "material": "8.8級中碳鋼(鍍鋅)",
        "use": "機械設備和鋼結構緊固連接",
        "func": "通過螺紋連接緊固兩個或多個零件",
        "process": "冷鐓/搓絲/熱處理/鍍鋅",
        "hs": "7318.15",
        "chapter": "第73章 鋼鐵製品",
        "basis": "歸類總規則一：品目7318列名「螺釘、螺栓、螺母…」；歸類總規則六：子目7318.15其他螺釘及螺栓",
        "cn_tariff": 8.0, "th_tariff": 10.0, "id_tariff": 10.0,
        "vat_cn": 13, "vat_th": 7, "vat_id": 11,
        "decl_cn": "品名|用途|材質|規格|強度等級|表面處理",
        "decl_en": "Product Name|Application|Material|Specification|Grade|Surface Treatment",
        "date": date(2024, 3, 5), "status": "已確認", "risk": "低風險",
        "risk_note": "", "note": "不鏽鋼螺栓歸7318.15；高強度螺栓(10.9級以上)也歸7318.15"
    },
    {
        "id": "HS-037",
        "cn": "軸承",
        "en": "Bearing",
        "material": "GCr15軸承鋼(高碳鉻軸承鋼)",
        "use": "旋轉機械的軸承支撐",
        "func": "減少旋轉摩擦、承受徑向和軸向載荷",
        "process": "鍛造/車削/熱處理/磨削/超精加工/組裝",
        "hs": "8482.10",
        "chapter": "第84章 機械設備",
        "basis": "歸類總規則一：品目8482列名「滾珠或滾子軸承」；歸類總規則六：子目8482.10滾珠軸承",
        "cn_tariff": 8.0, "th_tariff": 10.0, "id_tariff": 5.0,
        "vat_cn": 13, "vat_th": 7, "vat_id": 11,
        "decl_cn": "品名|用途|類型|內徑|外徑|品牌|型號",
        "decl_en": "Product Name|Application|Type|ID|OD|Brand|Model",
        "date": date(2024, 3, 25), "status": "已確認", "risk": "低風險",
        "risk_note": "", "note": "滾珠軸承歸8482.10；滾子軸承歸8482.20-8482.50"
    },
    {
        "id": "HS-038",
        "cn": "傳動皮帶",
        "en": "Drive Belt",
        "material": "橡膠基體/芳綸纖維簾線/尼龍包布",
        "use": "工業設備動力傳遞",
        "func": "在主動輪和從動輪之間傳遞旋轉動力",
        "process": "混煉/壓延/成型/硫化/切割",
        "hs": "4010.39",
        "chapter": "第40章 橡膠及其製品",
        "basis": "歸類總規則一：品目4010列名「硫化橡膠製傳動帶或輸送帶」；歸類總規則六：子目4010.39其他傳動帶",
        "cn_tariff": 8.0, "th_tariff": 10.0, "id_tariff": 10.0,
        "vat_cn": 13, "vat_th": 7, "vat_id": 11,
        "decl_cn": "品名|用途|材質|規格(寬度×周長)|類型|品牌",
        "decl_en": "Product Name|Application|Material|Size(W×L)|Type|Brand",
        "date": date(2024, 4, 10), "status": "已確認", "risk": "低風險",
        "risk_note": "", "note": "V型帶歸4010.33；同步帶歸4010.35；其他歸4010.39"
    },
    {
        "id": "HS-039",
        "cn": "工業過濾芯",
        "en": "Industrial Filter Element",
        "material": "玻璃纖維濾材/不鏽鋼支撐網",
        "use": "液壓系統和潤滑系統液體過濾",
        "func": "去除液體中固體顆粒污染物",
        "process": "折波/組裝/焊接/密封檢測",
        "hs": "8421.99",
        "chapter": "第84章 機械設備",
        "basis": "歸類總規則一：品目8421列名「過濾或淨化機器及裝置」；歸類總規則六：子目8421.99零件",
        "cn_tariff": 0, "th_tariff": 5.0, "id_tariff": 5.0,
        "vat_cn": 13, "vat_th": 7, "vat_id": 11,
        "decl_cn": "品名|用途|過濾精度|材質|規格|品牌|型號",
        "decl_en": "Product Name|Application|Filtration Rating|Material|Specification|Brand|Model",
        "date": date(2024, 5, 1), "status": "待確認", "risk": "中風險",
        "risk_note": "過濾芯作為零件歸8421.99，但完整過濾器歸8421.29；需確認申報的是濾芯零件還是完整過濾器",
        "note": "濾芯(零件)歸8421.99；完整過濾器歸8421.29"
    },
    {
        "id": "HS-040",
        "cn": "電氣接線端子",
        "en": "Electrical Terminal Block",
        "material": "PA66尼龍絕緣體/銅導體接線柱",
        "use": "電氣控制櫃內電線連接分配",
        "func": "提供可靠的電線端接和電氣連接",
        "process": "注塑/沖壓/電鍍/組裝",
        "hs": "8536.90",
        "chapter": "第85章 電機電氣設備",
        "basis": "歸類總規則一：品目8536列名「電路開關、繼電器…接線端子」；歸類總規則六：子目8536.90其他裝置（接線端子）",
        "cn_tariff": 0, "th_tariff": 5.0, "id_tariff": 5.0,
        "vat_cn": 13, "vat_th": 7, "vat_id": 11,
        "decl_cn": "品名|用途|材質|額定電壓|額定電流|品牌|型號",
        "decl_en": "Product Name|Application|Material|Rated Voltage|Rated Current|Brand|Model",
        "date": date(2024, 5, 25), "status": "已確認", "risk": "低風險",
        "risk_note": "", "note": "接線端子歸8536.90；額定電壓>1000V歸8535"
    },
]

# Write records
for row_idx, rec in enumerate(records, 2):
    ws1.cell(row=row_idx, column=1, value=rec["id"])
    ws1.cell(row=row_idx, column=2, value=rec["cn"])
    ws1.cell(row=row_idx, column=3, value=rec["en"])
    ws1.cell(row=row_idx, column=4, value=rec["material"])
    ws1.cell(row=row_idx, column=5, value=rec["use"])
    ws1.cell(row=row_idx, column=6, value=rec["func"])
    ws1.cell(row=row_idx, column=7, value=rec["process"])
    ws1.cell(row=row_idx, column=8, value=rec["hs"])
    ws1.cell(row=row_idx, column=9, value=rec["chapter"])
    ws1.cell(row=row_idx, column=10, value=rec["basis"])
    ws1.cell(row=row_idx, column=11, value=rec["cn_tariff"])
    ws1.cell(row=row_idx, column=12, value=rec["th_tariff"])
    ws1.cell(row=row_idx, column=13, value=rec["id_tariff"])
    # VAT as combined text: CN/TH/ID
    ws1.cell(row=row_idx, column=14, value=f"CN:{rec['vat_cn']}% / TH:{rec['vat_th']}% / ID:{rec['vat_id']}%")
    ws1.cell(row=row_idx, column=15, value=rec["decl_cn"])
    ws1.cell(row=row_idx, column=16, value=rec["decl_en"])
    ws1.cell(row=row_idx, column=17, value=rec["date"])
    ws1.cell(row=row_idx, column=18, value=rec["status"])
    ws1.cell(row=row_idx, column=19, value=rec["risk"])
    ws1.cell(row=row_idx, column=20, value=rec["risk_note"])
    ws1.cell(row=row_idx, column=21, value=rec["note"])

    # Apply styles
    for col in range(1, len(headers1) + 1):
        cell = ws1.cell(row=row_idx, column=col)
        is_center = col in (1, 8, 9, 11, 12, 13, 14, 17, 18, 19)
        is_number = False
        is_date = (col == 17)
        style_data_cell(cell, row_idx, is_center=is_center, is_date=is_date)

    # Special fills for status and risk
    apply_status_fill(ws1.cell(row=row_idx, column=18), rec["status"])
    apply_risk_fill(ws1.cell(row=row_idx, column=19), rec["risk"])
    if rec["risk"] == "高風險":
        apply_risk_fill(ws1.cell(row=row_idx, column=20), rec["risk"])

# Column widths for Sheet 1
col_widths_1 = [12, 22, 30, 28, 22, 28, 24, 14, 22, 48, 16, 16, 16, 28, 42, 42, 14, 12, 10, 50, 48]
for i, w in enumerate(col_widths_1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# Freeze and filter
ws1.freeze_panes = "A2"
ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers1))}{len(records)+1}"

# Row height
ws1.row_dimensions[1].height = 32
for r in range(2, len(records) + 2):
    ws1.row_dimensions[r].height = 45


# =====================================================================
# SHEET 2: 歸類規則參考
# =====================================================================
ws2 = wb.create_sheet("歸類規則參考")

headers2 = ["規則編號", "規則名稱", "規則說明", "應用示例"]
for col, h in enumerate(headers2, 1):
    ws2.cell(row=1, column=col, value=h)
style_header(ws2, len(headers2))

rules = [
    {
        "id": "規則一",
        "name": "品目条文及类注章注优先",
        "desc": (
            "類、章及分章的標題，僅為查找方便而設；具有法律效力的歸類，應按品目條文和有關類注或章注確定。"
            "如品目條文和類注、章注無其他規定，則按以下規則二至六確定。"
            "這是最基本、最重要的歸類規則，絕大多數商品可通過品目條文直接確定歸類。"
        ),
        "example": (
            "例1：不鏽鋼螺釘 → 品目7318明確列名「螺釘、螺栓」，直接歸入7318。\n"
            "例2：電動馬達 → 品目8501明確列名「電動機及發電機」，直接歸入8501。\n"
            "例3：活性碳 → 品目3801明確列名「活性碳」，直接歸入3801。"
        )
    },
    {
        "id": "規則二",
        "name": "不完整品、未組裝件及混合物",
        "desc": (
            "規則二(甲)：品目所列的貨品，應視為包括該貨品的不完整品或未製成品，"
            "只要在報驗時該不完整品或未完成品已具有完整品或製成品的基本特徵；"
            "還應視為包括該貨品的完整品或製成品（或按本規則可作為完整品或製成品歸類的貨品），"
            "在報驗時為未組裝或已拆開的。\n\n"
            "規則二(乙)：品目中對某種材料或物質的提及，應視為包括該種材料或物質的混合物或組合品。"
            "由一種以上材料或物質構成的貨品，應按規則三歸類。"
        ),
        "example": (
            "例1(甲)：未裝電機的泵(缺少電機) → 已具有泵的基本特徵，仍歸8413。\n"
            "例2(甲)：拆解運輸的整套設備(CBD散裝) → 按完整設備歸類。\n"
            "例3(乙)：含50%棉+50%聚酯的紗線 → 需按規則三進一步判定。"
        )
    },
    {
        "id": "規則三",
        "name": "兩種以上品目的歸類(從具體到一般)",
        "desc": (
            "當貨品按規則二(乙)或由於其他原因，看起來可歸入兩個或兩個以上品目時，應按以下規則歸類：\n\n"
            "規則三(甲) — 具體列名原則：列名比較具體的品目，優先於列名較為一般的品目。\n\n"
            "規則三(乙) — 基本特徵原則：混合物、不同材料構成或不同部件組成的組合物，"
            "以及零售的成套貨品，如不能按規則三(甲)歸類時，在本款可適用的條件下，"
            "應按構成貨品基本特徵的材料或部件歸類。\n\n"
            "規則三(丙) — 從後歸類原則：貨品不能按規則三(甲)或(乙)歸類時，"
            "應歸入該貨品可歸入的品目中，按品目編號順序歸入其最後一個品目。"
        ),
        "example": (
            "例1(甲)：電動剃鬚刀 → 品目8510(電動剃鬚刀)比品目8509(家用電動器具)更具體。\n"
            "例2(乙)：含望遠鏡的三腳架 → 望遠鏡賦予基本特徵，歸9005而非9620。\n"
            "例3(丙)：50%棉(品目5205)+50%聚酯(品目5509)紗線 → 從後歸5509。"
        )
    },
    {
        "id": "規則四",
        "name": "最類似品目歸類",
        "desc": (
            "根據上述規則無法歸類的貨品，應歸入與該貨品最相類似的貨品所歸入的品目。"
            "此規則極少使用，僅在品目條文和類注、章注、規則一至三均無法解決時才適用。"
            "判斷「最相類似」時，應考慮貨品的名稱、特徵、用途、功能等因素。"
        ),
        "example": (
            "例：新型智能穿戴設備(無明確品目) → 類比最相似功能的已有品目歸類。\n"
            "注意：此規則在實務中很少使用，因為絕大多數產品都能通過規則一到三確定歸類。"
        )
    },
    {
        "id": "規則五",
        "name": "包裝材料與容器的歸類",
        "desc": (
            "規則五(甲)：製成特殊形狀的容器(如照相機套、樂器盒、槍套等)，"
            "適合盛放某一物品或某一套物品，適於長期使用，與所裝物品一同報驗的，"
            "應與所裝物品一同歸類。但本身構成整個貨品基本特徵的容器除外。\n\n"
            "規則五(乙)：與所裝貨品同時報驗的包裝材料或包裝容器，"
            "如通常是用来包裝這類貨品的，應與所裝貨品一同歸類。"
            "但明顯可重複使用的包裝材料或容器不受本規則限制。"
        ),
        "example": (
            "例1(甲)：裝在小提琴盒中的小提琴 → 琴盒與小提琴一併歸入9202。\n"
            "例2(乙)：裝在紙箱中的機器設備 → 紙箱包裝隨設備歸類，不需單獨歸類。\n"
            "例外：可重複使用的鋼瓶(裝氣體) → 鋼瓶需單獨歸類。"
        )
    },
    {
        "id": "規則六",
        "name": "子目的歸類",
        "desc": (
            "貨品在某一品目項下各子目的歸類，應按子目條文和有關子目注釋，"
            "並在必要時參照上述各規則來確定，但子目的比較只能在同一數級上進行。"
            "除另有規定外，有關的類注和章注也適用於本規則。\n\n"
            "關鍵要點：\n"
            "1. 子目歸類必須先確定品目(規則一至五)\n"
            "2. 子目比較必須在同一層級(一槓子目與一槓子目比，二槓子目與二槓子目比)\n"
            "3. 類注和章注也適用於子目歸類"
        ),
        "example": (
            "例1：不鏽鋼法蘭 → 先確定品目7307(管配件)，再按子目層級比較：\n"
            "  7307.2 不鏽鋼製 → 7307.21 法蘭 → 歸入7307.21\n"
            "例2：離心泵 → 先確定品目8413(液體泵)，再按子目：\n"
            "  8413.7 其他離心泵 → 8413.70 → 歸入8413.70"
        )
    },
]

for row_idx, rule in enumerate(rules, 2):
    ws2.cell(row=row_idx, column=1, value=rule["id"])
    ws2.cell(row=row_idx, column=2, value=rule["name"])
    ws2.cell(row=row_idx, column=3, value=rule["desc"])
    ws2.cell(row=row_idx, column=4, value=rule["example"])
    for col in range(1, 5):
        cell = ws2.cell(row=row_idx, column=col)
        cell.font = data_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if col == 1:
            cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    if row_idx % 2 == 0:
        for col in range(1, 5):
            ws2.cell(row=row_idx, column=col).fill = PatternFill(
                start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid"
            )

ws2.column_dimensions["A"].width = 12
ws2.column_dimensions["B"].width = 32
ws2.column_dimensions["C"].width = 70
ws2.column_dimensions["D"].width = 60

ws2.freeze_panes = "A2"
ws2.auto_filter.ref = f"A1:D{len(rules)+1}"

ws2.row_dimensions[1].height = 32
for r in range(2, len(rules) + 2):
    ws2.row_dimensions[r].height = 140


# =====================================================================
# SHEET 3: 章節速查表
# =====================================================================
ws3 = wb.create_sheet("章節速查表")

headers3 = ["類號", "章號", "章節標題_中文", "章節標題_英文", "主要涵蓋產品", "常見歸類問題"]
for col, h in enumerate(headers3, 1):
    ws3.cell(row=1, column=col, value=h)
style_header(ws3, len(headers3))

chapters = [
    {
        "section": "VI", "ch": "27", "cn": "礦物燃料、礦物油及其蒸餾產品；瀝青物質；礦物蠟",
        "en": "Mineral fuels, mineral oils and products of their distillation; bituminous substances; mineral waxes",
        "products": "石油、天然氣、潤滑油、瀝青、石油焦",
        "issues": "潤滑油(2710)與合成潤滑劑(3403)的區分；生物柴油歸類"
    },
    {
        "section": "VI", "ch": "38", "cn": "化學工業及其相關工業的產品",
        "en": "Miscellaneous chemical products",
        "products": "活性炭、催化劑、黏合劑、化學試劑",
        "issues": "活性炭(3801)與活性碳纖維的區分；化學品混合物歸類"
    },
    {
        "section": "VII", "ch": "39", "cn": "塑膠及其製品",
        "en": "Plastics and articles thereof",
        "products": "塑膠原料(PE/PP/PVC/環氧樹脂等)、塑膠管材、塑膠薄膜、塑膠配件",
        "issues": "塑膠管(3917)與金屬增強管的區分；塑膠與橡膠的界定；共聚物歸類原則"
    },
    {
        "section": "VII", "ch": "40", "cn": "橡膠及其製品",
        "en": "Rubber and articles thereof",
        "products": "天然橡膠、合成橡膠、橡膠管、密封件、傳動帶、輪胎",
        "issues": "橡膠與塑膠混合材料的歸類；硫化與未硫化的區分"
    },
    {
        "section": "IX", "ch": "44", "cn": "木材及其製品；木炭",
        "en": "Wood and articles of wood; wood charcoal",
        "products": "原木、鋸材、膠合板、木質包裝箱、木地板",
        "issues": "木質包裝箱(4415)需ISPM-15處理；木家具歸94章"
    },
    {
        "section": "XIII", "ch": "68", "cn": "石料、石膏、水泥、石棉、雲母及類似材料的製品",
        "en": "Articles of stone, plaster, cement, asbestos, mica or similar materials",
        "products": "研磨材料、石棉製品、水泥製品、耐火材料",
        "issues": "岩棉保溫材料(6806)與玻璃纖維保溫(7019)的區分"
    },
    {
        "section": "XIII", "ch": "69", "cn": "陶瓷產品",
        "en": "Ceramic products",
        "products": "耐火磚、陶瓷磚、衛生陶瓷、陶瓷餐具",
        "issues": "吸水率對瓷磚子目的影響；工業陶瓷(如氧化鋁)歸類"
    },
    {
        "section": "XIII", "ch": "70", "cn": "玻璃及其製品",
        "en": "Glass and glassware",
        "products": "平板玻璃、玻璃纖維、玻璃容器、光學玻璃",
        "issues": "玻璃纖維保溫棉(7019)與礦物棉(6806)區分；鋼化玻璃的加工限度"
    },
    {
        "section": "XIV", "ch": "73", "cn": "鋼鐵製品",
        "en": "Articles of iron or steel",
        "products": "鋼鐵管及管件、法蘭、螺栓、結構件、彈簧、絲網",
        "issues": "無縫管(7304)與焊接管(7305/7306)區分；管配件中法蘭(7307.21)與對焊件(7307.23)區分；儲液罐300L分界(7309/7310)"
    },
    {
        "section": "XV", "ch": "76", "cn": "鋁及其製品",
        "en": "Aluminium and articles thereof",
        "products": "鋁錠、鋁板、鋁型材、鋁箔、鋁結構件",
        "issues": "鋁型材(7604)與加工後門窗(7610)的區分；鋁合金牌號對歸類的影響"
    },
    {
        "section": "XV", "ch": "83", "cn": "雜項金屬製品",
        "en": "Miscellaneous articles of base metal",
        "products": "鎖具、鉸鏈、焊條焊絲、金屬配件",
        "issues": "焊絲(8311)的歸類；五金配件按材質歸類 vs 按用途歸類"
    },
    {
        "section": "XVI", "ch": "84", "cn": "鍋爐、機器、機械器具及其零件",
        "en": "Nuclear reactors, boilers, machinery and mechanical appliances; parts thereof",
        "products": "泵、閥門、壓縮機、軸承、過濾設備、衡器、機械零件",
        "issues": "功能單元歸類(規則四)；通用零件(十六類注二)；泵(8413)與閥門(8481)的區分"
    },
    {
        "section": "XVI", "ch": "85", "cn": "電機、電氣設備及其零件；錄音機及放聲機、電視圖像、聲音的錄製和重放設備",
        "en": "Electrical machinery and equipment and parts thereof",
        "products": "電動機、發電機、變壓器、電纜、開關、照明、半導體、太陽能板",
        "issues": "PCB空板(8534)與PCBA(8537)的區分；LED燈(8539)與燈具(9405)的區分；光伏電池(8541.43)新版本"
    },
    {
        "section": "XVII", "ch": "87", "cn": "車輛(鐵道或電車道車輛除外)及其零件、附件",
        "en": "Vehicles other than railway or tramway rolling stock, and parts and accessories thereof",
        "products": "汽車、摩托車、自行車、車輛零部件",
        "issues": "汽車零部件歸類(8708)；新能源汽車歸類；通用零件排除條款"
    },
    {
        "section": "XVIII", "ch": "90", "cn": "光學、照相、電影、計量、檢驗、醫療或外科用儀器及設備",
        "en": "Optical, photographic, measuring, medical instruments and apparatus",
        "products": "壓力表、流量計、溫度計、分析儀器、醫療器械",
        "issues": "儀器儀表(90章)與電氣測量(9030)區分；零件歸類原則(90章注二)"
    },
    {
        "section": "XVIII", "ch": "91", "cn": "鐘錶及其零件",
        "en": "Clocks and watches and parts thereof",
        "products": "手錶、時鐘、計時器",
        "issues": "工業計時器歸91章或90章的區分"
    },
    {
        "section": "XX", "ch": "94", "cn": "家具；燈具、照明裝置；活動房屋",
        "en": "Furniture; bedding, lighting fittings; illuminated signs; prefabricated buildings",
        "products": "辦公家具、燈具照明、預製建築",
        "issues": "LED燈具(8539)與照明裝置(9405)的區分；預製建築(9406)的認定"
    },
    {
        "section": "XX", "ch": "96", "cn": "雜項製品",
        "en": "Miscellaneous manufactured articles",
        "products": "文具、鈕扣、拉鍊、打火機、掃把",
        "issues": "三腳架(9620)與光學儀器(90章)的區分"
    },
    {
        "section": "I", "ch": "01-05", "cn": "活動物；動物產品",
        "en": "Live animals; animal products",
        "products": "活動物、肉類、乳製品、蛋、蜂蜜",
        "issues": "加工限度：新鮮/冷藏(第1-2章) vs 進一步加工(第16章)"
    },
    {
        "section": "II", "ch": "06-14", "cn": "植物產品",
        "en": "Vegetable products",
        "products": "活植物、蔬菜、水果、穀物、茶、香料",
        "issues": "加工程度對歸類的影響：新鮮(第7-8章) vs 乾製/加工(第20章)"
    },
]

for row_idx, ch in enumerate(chapters, 2):
    ws3.cell(row=row_idx, column=1, value=ch["section"])
    ws3.cell(row=row_idx, column=2, value=ch["ch"])
    ws3.cell(row=row_idx, column=3, value=ch["cn"])
    ws3.cell(row=row_idx, column=4, value=ch["en"])
    ws3.cell(row=row_idx, column=5, value=ch["products"])
    ws3.cell(row=row_idx, column=6, value=ch["issues"])
    for col in range(1, 7):
        cell = ws3.cell(row=row_idx, column=col)
        cell.font = data_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        if col in (1, 2):
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if row_idx % 2 == 0:
        for col in range(1, 7):
            ws3.cell(row=row_idx, column=col).fill = PatternFill(
                start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid"
            )

ws3.column_dimensions["A"].width = 10
ws3.column_dimensions["B"].width = 10
ws3.column_dimensions["C"].width = 45
ws3.column_dimensions["D"].width = 55
ws3.column_dimensions["E"].width = 45
ws3.column_dimensions["F"].width = 55

ws3.freeze_panes = "A2"
ws3.auto_filter.ref = f"A1:F{len(chapters)+1}"

ws3.row_dimensions[1].height = 32
for r in range(2, len(chapters) + 2):
    ws3.row_dimensions[r].height = 65


# ── Save ─────────────────────────────────────────────────────────────
output_path = "/Users/chengshihhuai/.qoderwork/workspace/mqfdliaphm10z097/HS編碼歸類數據庫.xlsx"
wb.save(output_path)
print(f"Successfully created: {output_path}")
print(f"Sheet 1: HS歸類記錄 — {len(records)} records")
print(f"Sheet 2: 歸類規則參考 — {len(rules)} rules")
print(f"Sheet 3: 章節速查表 — {len(chapters)} chapters")
