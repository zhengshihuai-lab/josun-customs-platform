#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
==============================================================================
  HS 編碼智能歸類助手 v1.0
  上海御大集團 關務系統
==============================================================================

功能說明：
  命令行工具，關務人員輸入產品描述（中文或英文），系統自動：
  1. 分析產品的關鍵屬性（材質、用途、功能、加工方式）
  2. 匹配歷史歸類數據庫中的相似產品
  3. 提供 Top 3 推薦 HS 編碼及信心指數
  4. 給出申報要素的建議寫法
  5. 標註歸類風險點
  6. 顯示泰國/印尼進口關稅稅率

使用方式：
  python HS編碼歸類助手.py --interactive          # 互動模式
  python HS編碼歸類助手.py --batch input.xlsx     # 批量模式
  python HS編碼歸類助手.py --demo                 # 展示模式（自動執行3個查詢）

輸入檔案：
  HS編碼歸類數據庫.xlsx
  - Sheet "HS歸類記錄"
  - 欄位：記錄編號, 產品描述_中文, 產品描述_英文, 材質, 用途, 功能描述,
          加工方式, 已歸類HS編碼, HS編碼章節, 歸類依據, 中國進口關稅稅率%,
          泰國進口關稅稅率%, 印尼進口關稅稅率%, 增值稅率%, 申報要素_中文,
          申報要素_英文, 歸類日期, 確認狀態, 風險標記, 風險說明, 備註

批量模式輸入格式 (Excel)：
  | 序號 | 產品描述 | 材質 | 用途 | 功能 |

批量模式輸出格式 (Excel)：
  | 序號 | 產品描述 | 推薦HS編碼1 | 信心指數1 | 推薦HS編碼2 | 信心指數2 |
  | 推薦HS編碼3 | 信心指數3 | 建議申報要素 | 風險提示 | 中國關稅 |
  | 泰國關稅 | 印尼關稅 |

技術要求：
  - Python 3.7+
  - openpyxl (pip install openpyxl)
  - 純本地運行，不依賴外部 API 或網絡請求
==============================================================================
"""

import os
import sys
import re
import argparse
import math
from datetime import datetime
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[ERROR] 需要安裝 openpyxl 套件：pip install openpyxl")
    sys.exit(1)


# ============================================================================
# 關鍵詞字典庫
# ============================================================================

# 材質關鍵詞（中英對照）
MATERIAL_KEYWORDS = {
    "不鏽鋼": ["不鏽鋼", "stainless steel", "ss304", "ss316", "ss316l", "ss304l", "304不鏽鋼", "316l不鏽鋼", "316不鏽鋼", "ss201", "ss321"],
    "碳鋼": ["碳鋼", "carbon steel", "20#鋼", "q235", "q345", "a106", "a105", "低碳鋼", "中碳鋼", "高碳鋼"],
    "合金鋼": ["合金鋼", "alloy steel", "鉻鉬鋼", "cr-mo", "鉻鋼", "gcr15", "42crmo", "40cr"],
    "銅": ["銅", "copper", "黃銅", "brass", "青銅", "bronze", "紫銅", "red copper", "磷銅"],
    "鋁": ["鋁", "aluminum", "aluminium", "鋁合金", "6061", "6063", "7075", "al6063", "al6061"],
    "鑄鐵": ["鑄鐵", "cast iron", "球墨鑄鐵", "ductile iron", "灰鑄鐵", "grey iron"],
    "PVC": ["pvc", "聚氯乙烯", "polyvinyl chloride"],
    "橡膠": ["橡膠", "rubber", "丁腈橡膠", "nbr", "epdm", "三元乙丙", "矽橡膠", "silicone", "氟橡膠", "fkm", "viton"],
    "陶瓷": ["陶瓷", "ceramic", "氧化鋁", "alumina", "氧化鋯", "zirconia", "陶瓷纖維", "ceramic fiber"],
    "玻璃纖維": ["玻璃纖維", "fiberglass", "glass fiber", "e-glass", "e型玻璃纖維"],
    "聚丙烯": ["聚丙烯", "polypropylene", "pp", "聚丙烯酸", "polyacrylic", "pam", "pac"],
    "環氧樹脂": ["環氧樹脂", "epoxy", "epoxy resin"],
    "聚乙烯": ["聚乙烯", "polyethylene", "pe", "hdpe", "ldpe", "uhmwpe"],
    "礦物油": ["礦物油", "mineral oil", "潤滑油", "lubricating oil"],
    "鉑": ["鉑", "platinum", "pt100", "pt1000"],
    "塑料": ["塑料", "plastic", "聚合物", "polymer", "高分子"],
    "電子元件": ["電子元件", "electronic component", "半導體", "semiconductor", "ic", "芯片"],
    "合成橡膠": ["合成橡膠", "synthetic rubber", "丁苯橡膠", "sbr", "丁腈", "丁基橡膠"],
    "低碳鋼芯": ["低碳鋼芯", "焊條鋼芯", "welding wire"],
}

# 用途關鍵詞
USAGE_KEYWORDS = {
    "工業": ["工業", "industrial", "製造", "manufacturing", "工廠", "factory", "生產線", "production line"],
    "建築": ["建築", "construction", "building", "土木工程", "civil engineering"],
    "電力": ["電力", "electric power", "發電", "power generation", "輸配電", "power transmission"],
    "化工": ["化工", "chemical", "化學", "製藥", "pharmaceutical", "石化", "petrochemical"],
    "食品": ["食品", "food", "飲料", "beverage", "乳品", "dairy", "釀造", "brewing"],
    "水處理": ["水處理", "water treatment", "淨水", "water purification", "污水", "wastewater", "飲用水"],
    "石油天然氣": ["石油", "petroleum", "天然氣", "natural gas", "油氣", "oil and gas", "煉化", "refinery"],
    "暖通空調": ["暖通", "hvac", "空調", "air conditioning", "製冷", "refrigeration", "冷卻", "cooling"],
    "自動化": ["自動化", "automation", "控制", "control", "plc", "scada", "dcs", "工控"],
    "船舶": ["船舶", "ship", "marine", "船用", "海運", "maritime"],
    "礦業": ["礦業", "mining", "採礦", "選礦", "ore processing"],
    "農業": ["農業", "agriculture", "灌溉", "irrigation", "農機", "farm machinery"],
    "醫療": ["醫療", "medical", "醫院", "hospital", "診斷", "diagnostic"],
    "電子": ["電子", "electronic", "電路板", "pcb", "半導體", "semiconductor", "芯片"],
    "汽車": ["汽車", "automobile", "automotive", "車輛", "vehicle", "汽車零部件"],
    "氣動": ["氣動", "pneumatic", "氣缸", "cylinder", "氣動控制", "壓縮空氣"],
    "液壓": ["液壓", "hydraulic", "油壓", "液壓缸", "液壓系統"],
    "輸送": ["輸送", "conveyor", "輸送機", "輸送帶", "傳送", "物料搬運"],
    "管道": ["管道", "pipeline", "管路", "piping", "配管"],
    "焊接": ["焊接", "welding", "焊條", "electrode", "焊絲", "welding wire"],
    "潤滑": ["潤滑", "lubrication", "齒輪箱", "gearbox", "軸承", "bearing"],
    "鍋爐": ["鍋爐", "boiler", "蒸汽", "steam", "熱交換", "heat exchange"],
    "保溫": ["保溫", "insulation", "隔熱", "thermal insulation", "耐火", "refractory"],
    "清洗": ["清洗", "cleaning", "去污", "degreasing", "去油", "脫脂"],
    "散熱": ["散熱", "heatsink", "heat sink", "散熱器", "冷卻器", "cooler"],
    "傳動": ["傳動", "transmission", "drive", "驅動", "鏈條", "chain", "皮帶", "belt"],
    "密封": ["密封", "sealing", "seal", "墊片", "gasket", "o-ring", "油封"],
    "安全保護": ["安全", "safety", "保護", "protection", "安全閥", "relief valve"],
    "測量": ["測量", "measurement", "測量儀器", "計量", "metering", "檢測"],
    "儲存": ["儲存", "storage", "儲罐", "tank", "容器", "container", "水箱", "water tank"],
    "連接": ["連接", "connection", "connector", "接頭", "fitting", "法蘭", "flange"],
    "粘接": ["粘接", "bonding", "膠黏", "adhesive", "黏合", "glue"],
}

# 功能關鍵詞
FUNCTION_KEYWORDS = {
    "連接": ["連接", "connect", "connection", "耦合", "coupling", "對接", "mating"],
    "密封": ["密封", "seal", "sealing", "防漏", "leak-proof", "氣密"],
    "傳動": ["傳動", "drive", "transmission", "動力傳遞", "power transmission"],
    "測量": ["測量", "measure", "measurement", "檢測", "detect", "監測", "monitor"],
    "控制": ["控制", "control", "調節", "regulate", "邏輯控制", "logic control"],
    "輸送": ["輸送", "convey", "transport", "搬運", "handle", "送料"],
    "轉速調節": ["調速", "speed control", "變頻", "vfd", "frequency", "變速"],
    "壓力釋放": ["釋放壓力", "pressure relief", "超壓保護", "overpressure", "安全閥"],
    "截斷": ["截斷", "shutoff", "切斷", "on-off", "開關", "switch"],
    "溫度測量": ["測溫", "temperature measurement", "溫度傳感", "temperature sensor"],
    "壓力測量": ["測壓", "pressure measurement", "壓力傳感", "pressure sensor", "壓力變送"],
    "隔熱": ["隔熱", "thermal insulation", "保溫", "heat insulation", "節能"],
    "散熱": ["散熱", "heat dissipation", "冷卻", "cooling", "導熱"],
    "儲存": ["儲存", "store", "storage", "容納", "contain", "貯存"],
    "潤滑": ["潤滑", "lubricate", "減少摩擦", "reduce friction", "抗磨"],
    "粘接固定": ["粘接", "bond", "adhere", "黏合", "固定", "fix"],
    "清洗去污": ["清洗", "clean", "degrease", "去油", "脫脂", "去污"],
    "電能轉換": ["電能轉換", "power conversion", "變流", "整流", "inverter"],
    "信號傳輸": ["信號傳輸", "signal transmission", "通訊", "communication", "數據傳輸"],
    "防止沉積": ["分散", "disperse", "防止沉積", "anti-deposit", "阻垢", "scale inhibition"],
    "改變方向": ["改變方向", "change direction", "轉向", "彎曲", "redirect"],
    "支撐旋轉": ["支撐", "support", "軸承", "bearing", "旋轉支撐"],
    "動力驅動": ["驅動", "drive", "馬達", "motor", "電動機", "電機"],
}

# 加工方式關鍵詞
PROCESSING_KEYWORDS = {
    "鑄造": ["鑄造", "casting", "精密鑄造", "investment casting", "砂鑄", "sand casting"],
    "鍛造": ["鍛造", "forging", "forged", "熱鍛", "hot forging", "冷鍛", "cold forging"],
    "機加工": ["機加工", "machining", "車削", "turning", "銑削", "milling", "cnc"],
    "焊接": ["焊接", "welding", "welded", "氬弧焊", "tig", "mig", "電焊"],
    "熱處理": ["熱處理", "heat treatment", "淬火", "quenching", "回火", "tempering", "退火"],
    "表面處理": ["表面處理", "surface treatment", "電鍍", "plating", "陽極氧化", "anodizing", "拋光", "polishing"],
    "擠壓": ["擠壓", "extrusion", "擠型", "型材擠壓"],
    "注塑": ["注塑", "injection molding", "射出成型", "塑料成型"],
    "模壓": ["模壓", "compression molding", "壓製", "pressing", "模壓硫化"],
    "熱軋": ["熱軋", "hot rolling", "hot rolled", "熱軋無縫"],
    "冷軋": ["冷軋", "cold rolling", "cold rolled", "冷拔", "cold drawn"],
    "組裝": ["組裝", "assembly", "裝配", "組裝測試", "組裝校準"],
    "化學配製": ["化學配製", "chemical blending", "調和", "配製", "配製混合"],
    "聚合反應": ["聚合", "polymerization", "聚合反應", "合成"],
    "織造": ["織造", "weaving", "編織", "紡織", "textile"],
    "精密磨削": ["磨削", "grinding", "精密磨削", "precision grinding", "超精加工"],
    "SMT貼片": ["smt", "貼片", "表面貼裝", "smt assembly", "電子組裝"],
    "絞合": ["絞合", "stranding", "成纜", "cabling", "絞線"],
    "推制成型": ["推制", "push forming", "彎頭成型", "elbow forming"],
    "擠出": ["擠出", "extrusion", "擠出成型", "管材擠出", "tube extrusion"],
    "濕法成型": ["濕法成型", "wet process", "真空成型", "vacuum forming"],
    "衝壓": ["衝壓", "stamping", "pressing", "衝裁"],
}

# HS編碼章節描述映射（常用章節）
HS_CHAPTER_DESCRIPTIONS = {
    "7304": "無縫鋼鐵管",
    "7307": "鋼鐵管件附件（法蘭、彎頭、接頭等）",
    "7309": "鋼鐵製容器（容量 > 300L，有內襯/保溫）",
    "7310": "鋼鐵製容器（容量 < 300L 或無內襯）",
    "7315": "鋼鐵鏈及其零件",
    "7616": "其他鋁製品",
    "3402": "洗滌劑及清潔配製品",
    "3506": "膠黏劑",
    "3906": "丙烯酸聚合物",
    "4009": "橡膠管",
    "4016": "硫化橡膠製品",
    "2710": "石油及石油製品",
    "6903": "耐火陶瓷製品",
    "7011": "玻璃纖維（含玻璃棉）",
    "8311": "焊接材料（焊條、焊絲等）",
    "8481": "閥門及類似器具",
    "8482": "滾動軸承",
    "8501": "電動機及發電機",
    "8504": "靜態變流器（變頻器、整流器等）",
    "8537": "電氣控制裝置（配電盤、控制櫃等）",
    "8544": "電線及電纜",
    "9025": "溫度計、高溫計等",
    "9026": "流量、液位、壓力等測量儀器",
}


# ============================================================================
# HSCodeClassifier 核心引擎
# ============================================================================

class HSCodeClassifier:
    """HS編碼智能歸類引擎"""

    def __init__(self, db_path):
        """
        載入HS編碼數據庫
        
        Args:
            db_path: HS編碼歸類數據庫.xlsx 的路徑
        """
        self.db_path = db_path
        self.records = []
        self._load_database()

    def _load_database(self):
        """從 Excel 載入歸類記錄"""
        if not os.path.exists(self.db_path):
            print(f"[ERROR] 數據庫檔案不存在: {self.db_path}")
            print("[提示] 請先執行 create_hs_db.py 建立測試數據庫")
            sys.exit(1)

        try:
            wb = openpyxl.load_workbook(self.db_path, read_only=True, data_only=True)
        except Exception as e:
            print(f"[ERROR] 無法開啟數據庫檔案: {e}")
            sys.exit(1)

        # 尋找 HS歸類記錄 sheet
        sheet_name = None
        for name in wb.sheetnames:
            if "HS歸類記錄" in name or "hs" in name.lower():
                sheet_name = name
                break
        if sheet_name is None:
            sheet_name = wb.sheetnames[0]

        ws = wb[sheet_name]

        # 讀取標題行，建立欄位映射
        header_map = {}
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value:
                header_map[str(cell.value).strip()] = col_idx

        # 定義期望欄位
        field_names = [
            "記錄編號", "產品描述_中文", "產品描述_英文", "材質", "用途",
            "功能描述", "加工方式", "已歸類HS編碼", "HS編碼章節", "歸類依據",
            "中國進口關稅稅率%", "泰國進口關稅稅率%", "印尼進口關稅稅率%", "增值稅率%",
            "申報要素_中文", "申報要素_英文", "歸類日期", "確認狀態",
            "風險標記", "風險說明", "備註"
        ]

        # 讀取數據行
        for row in ws.iter_rows(min_row=2, values_only=False):
            record = {}
            for field in field_names:
                col_idx = header_map.get(field, None)
                if col_idx and col_idx <= len(row):
                    val = row[col_idx - 1].value
                    record[field] = str(val).strip() if val is not None else ""
                else:
                    record[field] = ""

            # 跳過空行
            if record.get("產品描述_中文") or record.get("已歸類HS編碼"):
                self.records.append(record)

        wb.close()

        if not self.records:
            print("[WARNING] 數據庫中無歸類記錄")

    def _normalize_text(self, text):
        """標準化文本（轉小寫、去空白）"""
        if not text:
            return ""
        text = text.lower().strip()
        text = re.sub(r'\s+', ' ', text)
        return text

    def _tokenize(self, text):
        """
        簡單分詞：按空白和常見分隔符切分，同時保留完整文本做子串匹配
        不依賴 jieba 等外部分詞庫
        """
        if not text:
            return set()
        text = self._normalize_text(text)
        # 按空白、逗號、斜線、括號等分隔
        tokens = set(re.split(r'[\s,;/|、，。()（）\[\]【】]+', text))
        tokens.discard('')
        return tokens

    def _extract_keywords(self, description):
        """
        從產品描述中提取關鍵屬性
        
        Returns:
            dict: {
                'materials': set of material keywords,
                'usages': set of usage keywords,
                'functions': set of function keywords,
                'processings': set of processing keywords,
                'raw_tokens': set of all tokens
            }
        """
        text = self._normalize_text(description)
        result = {
            'materials': set(),
            'usages': set(),
            'functions': set(),
            'processings': set(),
            'raw_tokens': self._tokenize(description)
        }

        # 匹配材質關鍵詞
        for category, variants in MATERIAL_KEYWORDS.items():
            for variant in variants:
                if variant.lower() in text:
                    result['materials'].add(category)
                    break

        # 匹配用途關鍵詞
        for category, variants in USAGE_KEYWORDS.items():
            for variant in variants:
                if variant.lower() in text:
                    result['usages'].add(category)
                    break

        # 匹配功能關鍵詞
        for category, variants in FUNCTION_KEYWORDS.items():
            for variant in variants:
                if variant.lower() in text:
                    result['functions'].add(category)
                    break

        # 匹配加工方式關鍵詞
        for category, variants in PROCESSING_KEYWORDS.items():
            for variant in variants:
                if variant.lower() in text:
                    result['processings'].add(category)
                    break

        return result

    def _jaccard_similarity(self, set_a, set_b):
        """計算 Jaccard 相似度"""
        if not set_a and not set_b:
            return 0.0
        intersection = len(set_a & set_b)
        union = len(set_a | set_b)
        if union == 0:
            return 0.0
        return intersection / union

    def _substring_similarity(self, text_a, text_b):
        """
        子串相似度：檢查較短文本的詞是否出現在較長文本中
        用於處理品名匹配。以較短集合為基準，衡量查詢詞在目標文本中的覆蓋率
        """
        if not text_a or not text_b:
            return 0.0
        a = self._normalize_text(text_a)
        b = self._normalize_text(text_b)
        tokens_a = set(re.split(r'[\s,;/|、，。()（）\[\]【】]+', a))
        tokens_b = set(re.split(r'[\s,;/|、，。()（）\[\]【】]+', b))
        tokens_a.discard('')
        tokens_b.discard('')
        if not tokens_a or not tokens_b:
            return 0.0

        # 以較小集合為基準（通常是查詢）
        smaller = tokens_a if len(tokens_a) <= len(tokens_b) else tokens_b
        larger = tokens_b if len(tokens_a) <= len(tokens_b) else tokens_a

        # 計算 smaller 中有多少 token 出現在 larger 中（含子串匹配）
        match_count = 0
        for ts in smaller:
            for tl in larger:
                if ts == tl or ts in tl or tl in ts:
                    match_count += 1
                    break

        return match_count / len(smaller) if smaller else 0.0

    def analyze_product(self, description):
        """
        分析產品描述，提取關鍵屬性
        
        Args:
            description: 產品描述（中文或英文）
            
        Returns:
            dict: 分析結果，包含品名、材質、用途、功能、加工方式等
        """
        keywords = self._extract_keywords(description)
        text = self._normalize_text(description)

        # 提取品名（取描述的前半部分或主要名詞）
        # 簡單策略：取第一個關鍵詞之前的文本作為品名
        product_name = description.strip()
        # 嘗試去除容量/規格等數字信息
        product_name_clean = re.sub(r'\d+\s*(mm|cm|m|l|ml|kw|w|v|a|kg|g|t|℃|°c|mpa|dn|pn|rpm|hz|mm2|°)', '', product_name, flags=re.IGNORECASE)
        product_name_clean = re.sub(r'\d+\s*(毫米|厘米|米|升|毫升|千瓦|瓦|伏|安|公斤|克|噸)', '', product_name_clean)
        product_name_clean = re.sub(r'\d+[xX×]\d+[xX×]?\d*', '', product_name_clean)
        product_name_clean = re.sub(r'\s+', ' ', product_name_clean).strip()

        analysis = {
            '品名': product_name_clean[:50] if product_name_clean else product_name[:50],
            '材質': ', '.join(sorted(keywords['materials'])) if keywords['materials'] else '未識別',
            '用途': ', '.join(sorted(keywords['usages'])) if keywords['usages'] else '未識別',
            '功能': ', '.join(sorted(keywords['functions'])) if keywords['functions'] else '未識別',
            '加工方式': ', '.join(sorted(keywords['processings'])) if keywords['processings'] else '未識別',
            'keywords': keywords,
            '原始描述': description,
        }

        # 嘗試提取容量/規格信息
        capacity_patterns = [
            r'(\d+\s*(?:L|ML|升|毫升|加侖))',
            r'(\d+\s*(?:KW|W|HP|馬力|千瓦|瓦))',
            r'(\d+\s*(?:V|伏|kV))',
            r'(DN\d+)',
            r'(PN\d+)',
            r'(\d+\s*(?:mm|cm|m|inch|英寸|毫米|厘米|米))',
            r'(\d+\s*(?:MPa|mpa|bar|psi))',
            r'(\d+\s*(?:kg|g|噸|公斤|克))',
        ]
        specs = []
        for pattern in capacity_patterns:
            matches = re.findall(pattern, description, re.IGNORECASE)
            specs.extend(matches)
        if specs:
            analysis['規格'] = ', '.join(specs)

        return analysis

    def search_similar(self, keywords, analysis=None):
        """
        在數據庫中搜索相似產品
        
        Args:
            keywords: 從 analyze_product 提取的關鍵詞字典
            analysis: 產品分析結果（可選，用於品名匹配）
            
        Returns:
            list: [(record, similarity_score)] 按分數降序排列
        """
        results = []
        query_desc = analysis.get('原始描述', '') if analysis else ''
        query_norm = self._normalize_text(query_desc)

        for record in self.records:
            # 提取數據庫記錄的關鍵詞
            desc_text = (record.get("產品描述_中文", "") + " " +
                         record.get("產品描述_英文", ""))
            db_keywords = self._extract_keywords(desc_text)

            # 1. 品名相似度（40% 權重）—— 使用子串相似度
            #    同時對中文描述和完整描述計算，取較高者
            product_sim = self._substring_similarity(
                query_desc, desc_text
            )
            # 也計算 raw tokens 的 Jaccard
            raw_jaccard = self._jaccard_similarity(
                keywords.get('raw_tokens', set()),
                db_keywords.get('raw_tokens', set())
            )
            product_score = max(product_sim, raw_jaccard)

            # 2. 材質匹配度（20% 權重）
            #    除類別匹配外，也做直接文本匹配
            material_cat = self._jaccard_similarity(
                keywords.get('materials', set()),
                db_keywords.get('materials', set())
            )
            # 直接文本匹配：查詢中的材質詞是否出現在記錄材質欄位
            db_material_text = self._normalize_text(record.get("材質", ""))
            material_direct = self._direct_keyword_match(query_norm, db_material_text, MATERIAL_KEYWORDS)
            material_score = max(material_cat, material_direct)

            # 3. 用途匹配度（20% 權重）
            usage_cat = self._jaccard_similarity(
                keywords.get('usages', set()),
                db_keywords.get('usages', set())
            )
            db_usage_text = self._normalize_text(record.get("用途", ""))
            usage_direct = self._direct_keyword_match(query_norm, db_usage_text, USAGE_KEYWORDS)
            usage_score = max(usage_cat, usage_direct)

            # 4. 功能匹配度（20% 權重）
            function_cat = self._jaccard_similarity(
                keywords.get('functions', set()),
                db_keywords.get('functions', set())
            )
            db_func_text = self._normalize_text(record.get("功能描述", ""))
            function_direct = self._direct_keyword_match(query_norm, db_func_text, FUNCTION_KEYWORDS)
            function_score = max(function_cat, function_direct)

            # 加權總分
            total_score = (
                product_score * 0.40 +
                material_score * 0.20 +
                usage_score * 0.20 +
                function_score * 0.20
            )

            if total_score > 0.05:  # 最低門檻
                results.append((record, total_score, {
                    'product': product_score,
                    'material': material_score,
                    'usage': usage_score,
                    'function': function_score
                }))

        # 按分數降序排列
        results.sort(key=lambda x: x[1], reverse=True)
        return results

    def _direct_keyword_match(self, query_text, db_field_text, keyword_dict):
        """
        直接關鍵詞匹配：檢查查詢文本和數據庫欄位文本是否命中相同的關鍵詞變體
        返回 0.0 ~ 1.0 的匹配分數
        """
        if not query_text or not db_field_text:
            return 0.0

        query_hits = set()
        db_hits = set()

        for category, variants in keyword_dict.items():
            for variant in variants:
                v = variant.lower()
                if v in query_text:
                    query_hits.add(category)
                if v in db_field_text:
                    db_hits.add(category)

        if not query_hits and not db_hits:
            return 0.0
        intersection = len(query_hits & db_hits)
        union = len(query_hits | db_hits)
        return intersection / union if union > 0 else 0.0

    def _calculate_confidence(self, record, raw_score, detail_scores):
        """
        計算信心指數
        
        基於匹配分數、歷史歸類確認狀態、風險等級綜合計算
        使用非線性縮放使分數分布更合理
        
        Returns:
            int: 0-100 的信心指數
        """
        # 非線性縮放：原始分數 0~1 映射到 0~100
        # 0.5 -> ~70, 0.8 -> ~92, 0.3 -> ~48
        base_score = raw_score * 100 * 1.35  # 基礎放大

        # 確認狀態加成
        status = record.get("確認狀態", "")
        if "已確認" in status:
            base_score += 8
        elif "待確認" in status:
            base_score -= 5
        elif "已駁回" in status or "退回" in status:
            base_score -= 20

        # 風險等級影響
        risk = record.get("風險標記", "")
        if "高風險" in risk:
            base_score -= 10
        elif "中風險" in risk:
            base_score -= 5

        # 各維度均勻性加成（如果所有維度都匹配較好，加分）
        dim_scores = [detail_scores.get('product', 0),
                      detail_scores.get('material', 0),
                      detail_scores.get('usage', 0),
                      detail_scores.get('function', 0)]
        non_zero_dims = sum(1 for s in dim_scores if s > 0.1)
        if non_zero_dims >= 3:
            base_score += 8
        elif non_zero_dims >= 2:
            base_score += 3
        elif non_zero_dims <= 1:
            base_score -= 5

        # 限制範圍
        confidence = max(0, min(100, int(base_score)))
        return confidence

    def classify(self, description, material=None, usage=None, function=None):
        """
        主歸類方法
        
        Args:
            description: 產品描述
            material: 材質（可選，補充描述中的不足）
            usage: 用途（可選）
            function: 功能（可選）
            
        Returns:
            dict: {
                'analysis': 產品分析結果,
                'recommendations': [
                    {
                        'hs_code': str,
                        'chapter': str,
                        'description': str,
                        'confidence': int,
                        'basis': str,
                        'tariff_cn': str,
                        'tariff_th': str,
                        'tariff_id': str,
                        'vat': str,
                        'declaration_elements': str,
                        'risk_level': str,
                        'risk_note': str,
                        'scores': dict
                    }
                ],
                'risks': list of risk strings,
                'suggested_declaration': str
            }
        """
        # 如果有補充信息，加入描述
        full_description = description
        if material:
            full_description += f" {material}"
        if usage:
            full_description += f" {usage}"
        if function:
            full_description += f" {function}"

        # 1. 分析產品描述
        analysis = self.analyze_product(full_description)
        keywords = analysis['keywords']

        # 2. 搜索相似產品
        similar = self.search_similar(keywords, analysis)

        # 3. 構建 Top 3 推薦（去重 HS 編碼）
        seen_codes = set()
        recommendations = []

        for record, score, detail_scores in similar:
            hs_code = record.get("已歸類HS編碼", "")
            if not hs_code or hs_code in seen_codes:
                continue
            seen_codes.add(hs_code)

            confidence = self._calculate_confidence(record, score, detail_scores)

            # 獲取 HS 編碼描述
            hs_prefix = hs_code[:4].replace(".", "")
            hs_desc = HS_CHAPTER_DESCRIPTIONS.get(hs_prefix, record.get("HS編碼章節", ""))

            rec = {
                'hs_code': hs_code,
                'chapter': record.get("HS編碼章節", ""),
                'description': hs_desc,
                'confidence': confidence,
                'basis': record.get("歸類依據", ""),
                'tariff_cn': record.get("中國進口關稅稅率%", ""),
                'tariff_th': record.get("泰國進口關稅稅率%", ""),
                'tariff_id': record.get("印尼進口關稅稅率%", ""),
                'vat': record.get("增值稅率%", ""),
                'declaration_elements': record.get("申報要素_中文", ""),
                'risk_level': record.get("風險標記", ""),
                'risk_note': record.get("風險說明", ""),
                'scores': detail_scores,
                'record': record,
            }
            recommendations.append(rec)

            if len(recommendations) >= 3:
                break

        # 4. 風險評估
        risks = self._assess_risks(recommendations, analysis, description)

        # 5. 生成申報要素建議
        suggested_decl = self._generate_declaration_suggestion(analysis, recommendations)

        return {
            'analysis': analysis,
            'recommendations': recommendations,
            'risks': risks,
            'suggested_declaration': suggested_decl,
        }

    def _assess_risks(self, recommendations, analysis, description):
        """評估歸類風險"""
        risks = []

        # 1. 檢查數據庫中的風險標記
        for rec in recommendations:
            risk_level = rec.get('risk_level', '')
            risk_note = rec.get('risk_note', '')
            if risk_note:
                risks.append(f"[{rec['hs_code']}] {risk_note}")
            if "高風險" in risk_level:
                risks.append(f"[{rec['hs_code']}] 此編碼在歷史歸類中被標記為高風險，需特別注意")

        # 2. 檢查是否有相似但不同編碼的記錄
        if len(recommendations) >= 2:
            code1 = recommendations[0]['hs_code'][:4]
            code2 = recommendations[1]['hs_code'][:4]
            if code1 != code2:
                conf1 = recommendations[0]['confidence']
                conf2 = recommendations[1]['confidence']
                if abs(conf1 - conf2) < 20:
                    risks.append(
                        f"存在不同章節的相似產品歸類（{recommendations[0]['hs_code']} vs "
                        f"{recommendations[1]['hs_code']}），信心指數接近，建議人工確認"
                    )

        # 3. 信心指數過低
        if recommendations and recommendations[0]['confidence'] < 70:
            risks.append(f"最高信心指數僅 {recommendations[0]['confidence']}%，建議由專業歸類人員判斷")

        # 4. 材質不明
        if analysis.get('材質') == '未識別':
            risks.append("產品描述中未識別到明確材質信息，材質是 HS 歸類的重要依據，建議補充")

        # 5. 通用風險提示
        text = self._normalize_text(description)
        if any(kw in text for kw in ['反傾銷', 'anti-dumping', '反補貼']):
            risks.append("此產品可能涉及反傾銷/反補貼調查，進口前請確認相關稅率")
        if any(kw in text for kw in ['許可證', 'license', 'permit', '認證']):
            risks.append("此產品可能涉及進口許可證或認證要求")

        return risks

    def _generate_declaration_suggestion(self, analysis, recommendations):
        """生成申報要素建議"""
        if recommendations:
            # 基於最高信心指數的記錄
            best_rec = recommendations[0]
            template = best_rec.get('declaration_elements', '')
            if template:
                return template

        # 如果沒有模板，基於分析結果生成
        parts = []
        if analysis.get('品名'):
            parts.append(f"品名:{analysis['品名'][:20]}")
        if analysis.get('材質') and analysis['材質'] != '未識別':
            parts.append(f"材質:{analysis['材質']}")
        if analysis.get('用途') and analysis['用途'] != '未識別':
            parts.append(f"用途:{analysis['用途']}")
        if analysis.get('規格'):
            parts.append(f"規格:{analysis['規格']}")

        return '|'.join(parts) if parts else "請根據實際產品信息填寫申報要素"

    def get_declaration_elements(self, hs_code):
        """根據HS編碼獲取申報要素模板"""
        for record in self.records:
            if record.get("已歸類HS編碼", "") == hs_code:
                return record.get("申報要素_中文", "")
        return ""

    def get_tariff_info(self, hs_code):
        """獲取關稅稅率資訊"""
        for record in self.records:
            if record.get("已歸類HS編碼", "") == hs_code:
                return {
                    '中國關稅': record.get("中國進口關稅稅率%", ""),
                    '泰國關稅': record.get("泰國進口關稅稅率%", ""),
                    '印尼關稅': record.get("印尼進口關稅稅率%", ""),
                    '增值稅率': record.get("增值稅率%", ""),
                }
        return {}

    def check_risks(self, hs_code, description):
        """檢查歸類風險"""
        risks = []
        for record in self.records:
            if record.get("已歸類HS編碼", "") == hs_code:
                risk_note = record.get("風險說明", "")
                risk_level = record.get("風險標記", "")
                if risk_note:
                    risks.append(risk_note)
                if "高風險" in risk_level:
                    risks.append("此編碼在歷史歸類中被標記為高風險")
        return risks


# ============================================================================
# InteractiveClassifier 互動式歸類介面
# ============================================================================

class InteractiveClassifier:
    """互動式歸類介面"""

    # ANSI 顏色碼
    BOLD = '\033[1m'
    GREEN = '\033[92m'
    YELLOW = '\033[93m'
    RED = '\033[91m'
    CYAN = '\033[96m'
    DIM = '\033[2m'
    RESET = '\033[0m'
    BLUE = '\033[94m'

    def __init__(self, db_path):
        self.classifier = HSCodeClassifier(db_path)

    def _print_banner(self):
        """顯示橫幅"""
        print()
        print(f"{self.BOLD}{self.GREEN}╔═══════════════════════════════════════════════════╗{self.RESET}")
        print(f"{self.BOLD}{self.GREEN}║        HS 編碼智能歸類助手 v1.0                  ║{self.RESET}")
        print(f"{self.BOLD}{self.GREEN}║        上海御大集團 關務系統                      ║{self.RESET}")
        print(f"{self.BOLD}{self.GREEN}╚═══════════════════════════════════════════════════╝{self.RESET}")
        print()
        print(f"  數據庫記錄數: {len(self.classifier.records)} 筆")
        print()

    def _print_separator(self):
        print(f"{self.DIM}━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━{self.RESET}")

    def _confidence_bar(self, confidence):
        """生成信心指數進度條"""
        filled = confidence // 10
        empty = 10 - filled
        bar = "█" * filled + "░" * empty

        if confidence >= 90:
            color = self.GREEN
        elif confidence >= 70:
            color = self.YELLOW
        else:
            color = self.RED

        return f"{color}{bar}{self.RESET}"

    def _confidence_label(self, confidence):
        """信心指數文字標籤"""
        if confidence >= 90:
            return f"{self.GREEN}高信心（可直接使用）{self.RESET}"
        elif confidence >= 70:
            return f"{self.YELLOW}中信心（建議與報關行確認）{self.RESET}"
        else:
            return f"{self.RED}低信心（需要專業歸類人員判斷）{self.RESET}"

    def format_report(self, result):
        """格式化歸類報告（單筆）"""
        analysis = result['analysis']
        recommendations = result['recommendations']
        risks = result['risks']
        suggested = result['suggested_declaration']

        lines = []

        # 產品分析
        lines.append(f"{self.BOLD}{self.CYAN}📋 產品分析結果：{self.RESET}")
        lines.append(f"  品名：{analysis.get('品名', 'N/A')}")
        lines.append(f"  材質：{analysis.get('材質', 'N/A')}")
        lines.append(f"  用途：{analysis.get('用途', 'N/A')}")
        lines.append(f"  功能：{analysis.get('功能', 'N/A')}")
        if analysis.get('規格'):
            lines.append(f"  規格：{analysis['規格']}")
        if analysis.get('加工方式') and analysis['加工方式'] != '未識別':
            lines.append(f"  加工方式：{analysis['加工方式']}")
        lines.append("")
        lines.append(self._print_separator_inline())
        lines.append("")

        # 推薦編碼
        lines.append(f"{self.BOLD}🏆 推薦 HS 編碼：{self.RESET}")
        lines.append("")

        for i, rec in enumerate(recommendations, 1):
            star = "⭐ " if i == 1 else ""
            conf = rec['confidence']
            bar = self._confidence_bar(conf)
            label = self._confidence_label(conf)

            lines.append(f"  {self.BOLD}[{i}] {star}{rec['hs_code']} — {rec['description']}{self.RESET}")
            lines.append(f"      信心指數：{conf}% {bar}  {label}")
            if rec.get('basis'):
                lines.append(f"      歸類依據：{rec['basis']}")
            tariff_cn = rec.get('tariff_cn', 'N/A')
            tariff_th = rec.get('tariff_th', 'N/A')
            tariff_id = rec.get('tariff_id', 'N/A')
            vat = rec.get('vat', 'N/A')
            lines.append(f"      {self.CYAN}中國關稅：{tariff_cn}%  |  泰國關稅：{tariff_th}%  |  印尼關稅：{tariff_id}%{self.RESET}")
            lines.append(f"      {self.CYAN}增值稅：{vat}{self.RESET}")
            if rec.get('declaration_elements'):
                lines.append(f"      申報要素：{rec['declaration_elements']}")
            if rec.get('risk_note'):
                lines.append(f"      {self.YELLOW}⚠️  風險：{rec['risk_note']}{self.RESET}")
            lines.append("")

        lines.append(self._print_separator_inline())
        lines.append("")

        # 風險提示
        if risks:
            lines.append(f"{self.YELLOW}{self.BOLD}⚠️  風險提示：{self.RESET}")
            for risk in risks:
                lines.append(f"  {self.YELLOW}• {risk}{self.RESET}")
            lines.append("")
            lines.append(self._print_separator_inline())
            lines.append("")

        # 申報要素建議
        lines.append(f"{self.BOLD}📝 建議申報要素寫法：{self.RESET}")
        lines.append(f"  {suggested}")
        lines.append("")

        return '\n'.join(lines)

    def _print_separator_inline(self):
        return f"{self.DIM}━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━{self.RESET}"

    def _query_and_display(self, description, material=None, usage=None, function=None):
        """執行查詢並顯示結果"""
        print(f"\n{self.DIM}🔍 分析中...{self.RESET}\n")
        print(self._print_separator_inline())
        print()

        result = self.classifier.classify(description, material, usage, function)
        report = self.format_report(result)
        print(report)
        print(self._print_separator_inline())

        return result

    def run_interactive(self):
        """互動模式：引導用戶輸入並顯示結果"""
        self._print_banner()

        while True:
            try:
                print(f"{self.BOLD}請輸入產品描述（中文或英文）：{self.RESET}", end="")
                description = input().strip()

                if not description:
                    print(f"{self.YELLOW}[提示] 請輸入產品描述{self.RESET}")
                    continue

                if description.lower() in ('q', 'quit', 'exit', '結束', '退出'):
                    print(f"\n{self.GREEN}感謝使用 HS 編碼智能歸類助手，再見！{self.RESET}\n")
                    break

                # 可選：補充材質/用途/功能
                print(f"{self.DIM}（可選）補充材質信息，直接按 Enter 跳過：{self.RESET}", end="")
                material = input().strip() or None

                print(f"{self.DIM}（可選）補充用途信息，直接按 Enter 跳過：{self.RESET}", end="")
                usage = input().strip() or None

                print(f"{self.DIM}（可選）補充功能信息，直接按 Enter 跳過：{self.RESET}", end="")
                function = input().strip() or None

                self._query_and_display(description, material, usage, function)

                print(f"\n是否繼續查詢？(y/n): ", end="")
                cont = input().strip().lower()
                if cont in ('n', 'no', '否', '不'):
                    print(f"\n{self.GREEN}感謝使用 HS 編碼智能歸類助手，再見！{self.RESET}\n")
                    break

            except KeyboardInterrupt:
                print(f"\n\n{self.GREEN}感謝使用 HS 編碼智能歸類助手，再見！{self.RESET}\n")
                break
            except EOFError:
                print(f"\n\n{self.GREEN}感謝使用 HS 編碼智能歸類助手，再見！{self.RESET}\n")
                break

    def run_batch(self, input_file):
        """
        批量模式：從 Excel 讀入多筆產品批量歸類
        
        輸入格式：| 序號 | 產品描述 | 材質 | 用途 | 功能 |
        輸出格式：歸類結果報告.xlsx
        """
        if not os.path.exists(input_file):
            print(f"[ERROR] 輸入檔案不存在: {input_file}")
            return

        print(f"\n{self.BOLD}[批量模式] 載入檔案: {input_file}{self.RESET}")

        try:
            wb_in = openpyxl.load_workbook(input_file, read_only=True, data_only=True)
        except Exception as e:
            print(f"[ERROR] 無法開啟輸入檔案: {e}")
            return

        ws_in = wb_in.active

        # 讀取輸入數據
        products = []
        for row in ws_in.iter_rows(min_row=2, values_only=True):
            if row and row[1]:  # 產品描述欄位不為空
                products.append({
                    '序號': row[0] if row[0] else len(products) + 1,
                    '產品描述': str(row[1]).strip() if row[1] else "",
                    '材質': str(row[2]).strip() if len(row) > 2 and row[2] else None,
                    '用途': str(row[3]).strip() if len(row) > 3 and row[3] else None,
                    '功能': str(row[4]).strip() if len(row) > 4 and row[4] else None,
                })
        wb_in.close()

        if not products:
            print("[ERROR] 輸入檔案中無有效產品記錄")
            return

        print(f"共 {len(products)} 筆產品待歸類")
        print()

        # 建立輸出工作簿
        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active
        ws_out.title = "歸類結果"

        # 標題
        out_headers = [
            "序號", "產品描述", "推薦HS編碼1", "信心指數1",
            "推薦HS編碼2", "信心指數2", "推薦HS編碼3", "信心指數3",
            "建議申報要素", "風險提示", "中國關稅", "泰國關稅", "印尼關稅"
        ]

        header_fill = PatternFill(start_color="014725", end_color="014725", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col, header in enumerate(out_headers, 1):
            cell = ws_out.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

        # 逐筆歸類
        for idx, product in enumerate(products):
            desc = product['產品描述']
            mat = product['材質']
            use = product['用途']
            func = product['功能']

            print(f"  [{idx+1}/{len(products)}] 歸類中: {desc[:40]}...", end="")

            result = self.classifier.classify(desc, mat, use, func)
            recs = result['recommendations']
            risks = result['risks']

            row_data = [
                product['序號'],
                desc,
                recs[0]['hs_code'] if len(recs) > 0 else "",
                recs[0]['confidence'] if len(recs) > 0 else "",
                recs[1]['hs_code'] if len(recs) > 1 else "",
                recs[1]['confidence'] if len(recs) > 1 else "",
                recs[2]['hs_code'] if len(recs) > 2 else "",
                recs[2]['confidence'] if len(recs) > 2 else "",
                result['suggested_declaration'],
                '; '.join(risks) if risks else "",
                recs[0]['tariff_cn'] if len(recs) > 0 else "",
                recs[0]['tariff_th'] if len(recs) > 0 else "",
                recs[0]['tariff_id'] if len(recs) > 0 else "",
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws_out.cell(row=idx+2, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)

            # 信心指數著色
            for col_idx in [4, 6, 8]:
                cell = ws_out.cell(row=idx+2, column=col_idx)
                if cell.value and isinstance(cell.value, int):
                    if cell.value >= 90:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif cell.value >= 70:
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            best_conf = recs[0]['confidence'] if recs else 0
            if best_conf >= 90:
                print(f" {self.GREEN}✓ {recs[0]['hs_code']} ({best_conf}%){self.RESET}")
            elif best_conf >= 70:
                print(f" {self.YELLOW}~ {recs[0]['hs_code']} ({best_conf}%){self.RESET}")
            else:
                print(f" {self.RED}? {recs[0]['hs_code'] if recs else 'N/A'} ({best_conf}%){self.RESET}")

        # 調整欄寬
        out_widths = [6, 40, 12, 10, 12, 10, 12, 10, 50, 50, 10, 10, 10]
        for i, width in enumerate(out_widths, 1):
            ws_out.column_dimensions[get_column_letter(i)].width = width

        ws_out.freeze_panes = 'A2'

        # 儲存
        output_dir = os.path.dirname(input_file)
        output_name = f"歸類結果報告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_path = os.path.join(output_dir, output_name)
        wb_out.save(output_path)

        print()
        print(f"{self.GREEN}{self.BOLD}[完成] 歸類結果已儲存至: {output_path}{self.RESET}")
        print(f"  共處理 {len(products)} 筆產品")
        high_conf = sum(1 for p in products if True)  # placeholder
        print()

    def run_demo(self):
        """展示模式：自動執行 3 個查詢"""
        self._print_banner()

        demo_queries = [
            "電動馬達 550W 用於驅動工業泵",
            "不鏽鋼法蘭 DN80 管道連接用",
            "化工用聚丙烯酸添加劑 25kg裝",
        ]

        for i, query in enumerate(demo_queries, 1):
            print(f"\n{self.BOLD}{self.BLUE}{'='*55}{self.RESET}")
            print(f"{self.BOLD}{self.BLUE}  查詢 {i}/3: {query}{self.RESET}")
            print(f"{self.BOLD}{self.BLUE}{'='*55}{self.RESET}")

            self._query_and_display(query)

            if i < len(demo_queries):
                print(f"\n{self.DIM}--- 按 Enter 繼續下一個查詢 ---{self.RESET}", end="")
                try:
                    input()
                except (EOFError, KeyboardInterrupt):
                    pass

        print(f"\n{self.GREEN}{self.BOLD}展示完成！共執行 3 個查詢。{self.RESET}\n")


# ============================================================================
# 主程式
# ============================================================================

def get_default_db_path():
    """獲取默認數據庫路徑（與腳本同目錄）"""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(script_dir, "HS編碼歸類數據庫.xlsx")


def main():
    parser = argparse.ArgumentParser(
        description="HS 編碼智能歸類助手 v1.0 - 上海御大集團關務系統",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用範例：
  python HS編碼歸類助手.py --interactive          互動模式
  python HS編碼歸類助手.py --batch input.xlsx     批量模式
  python HS編碼歸類助手.py --demo                 展示模式
        """
    )
    parser.add_argument('--interactive', '-i', action='store_true',
                        help='啟動互動模式')
    parser.add_argument('--batch', '-b', type=str, metavar='INPUT.xlsx',
                        help='批量模式：指定輸入 Excel 檔案')
    parser.add_argument('--demo', '-d', action='store_true',
                        help='展示模式：自動執行 3 個查詢')
    parser.add_argument('--db', type=str, default=None,
                        help='指定數據庫路徑（預設同目錄下的 HS編碼歸類數據庫.xlsx）')

    args = parser.parse_args()

    db_path = args.db or get_default_db_path()
    interface = InteractiveClassifier(db_path)

    if args.batch:
        interface.run_batch(args.batch)
    elif args.demo:
        interface.run_demo()
    elif args.interactive:
        interface.run_interactive()
    else:
        # 預設為互動模式
        interface.run_interactive()


if __name__ == '__main__':
    main()
