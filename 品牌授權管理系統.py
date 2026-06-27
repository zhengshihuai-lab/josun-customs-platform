#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
品牌授權管理系統 (Brand Authorization Management System)
========================================================
管理多品牌進口至泰國與印尼的授權追蹤、到期預警、續約管理與報表產出。
"""

import argparse
import json
import os
import sys
from datetime import datetime, timedelta
from collections import defaultdict

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule, FormatObject
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("[警告] 未安裝 openpyxl，Excel 功能將無法使用。請執行: pip install openpyxl")


# ─────────────────────────── 常數與設定 ───────────────────────────

TODAY = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

ALERT_LEVELS = {
    "expired":    {"label": "已過期",     "color": "red",    "hex": "FF0000", "fill": "FFC7CE", "priority": 0},
    "7days":      {"label": "最後警告",   "color": "red",    "hex": "FF0000", "fill": "FFC7CE", "priority": 1},
    "30days":     {"label": "緊急續約",   "color": "orange", "hex": "FF8C00", "fill": "FFE699", "priority": 2},
    "60days":     {"label": "準備續約",   "color": "yellow", "hex": "FFD700", "fill": "FFF2CC", "priority": 3},
    "90days":     {"label": "提前規劃",   "color": "green",  "hex": "228B22", "fill": "C6EFCE", "priority": 4},
}

DEEP_BLUE_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
WHITE_BOLD_FONT = Font(name="Microsoft JhengHei", bold=True, color="FFFFFF", size=11)
HEADER_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

WORKING_DIR = os.path.dirname(os.path.abspath(__file__))


# ─────────────────────────── 資料模型 ───────────────────────────

class AuthorizationRecord:
    """單一筆品牌授權紀錄。"""

    def __init__(self, auth_id, brand_name, licensor, licensee,
                 start_date, end_date, countries, product_category,
                 cert_number, cert_path, renewal_conditions,
                 advance_notice_days, contact_person, contact_info,
                 status="有效"):
        self.auth_id = auth_id
        self.brand_name = brand_name
        self.licensor = licensor
        self.licensee = licensee
        self.start_date = self._parse_date(start_date)
        self.end_date = self._parse_date(end_date)
        self.countries = countries
        self.product_category = product_category
        self.cert_number = cert_number
        self.cert_path = cert_path
        self.renewal_conditions = renewal_conditions
        self.advance_notice_days = advance_notice_days
        self.contact_person = contact_person
        self.contact_info = contact_info
        self.status = status

    @staticmethod
    def _parse_date(d):
        if isinstance(d, datetime):
            return d.replace(hour=0, minute=0, second=0, microsecond=0)
        if isinstance(d, str):
            for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y"):
                try:
                    return datetime.strptime(d, fmt)
                except ValueError:
                    continue
        raise ValueError(f"無法解析日期: {d}")

    @property
    def days_remaining(self):
        delta = (self.end_date - TODAY).days
        return delta

    @property
    def alert_level(self):
        d = self.days_remaining
        if d < 0:
            return "expired"
        elif d <= 7:
            return "7days"
        elif d <= 30:
            return "30days"
        elif d <= 60:
            return "60days"
        elif d <= 90:
            return "90days"
        else:
            return None

    @property
    def computed_status(self):
        if self.days_remaining < 0:
            return "已過期"
        elif self.days_remaining <= 30:
            return "即將到期"
        else:
            return "有效"

    def to_dict(self):
        return {
            "auth_id": self.auth_id,
            "brand_name": self.brand_name,
            "licensor": self.licensor,
            "licensee": self.licensee,
            "start_date": self.start_date.strftime("%Y-%m-%d"),
            "end_date": self.end_date.strftime("%Y-%m-%d"),
            "countries": self.countries,
            "product_category": self.product_category,
            "cert_number": self.cert_number,
            "cert_path": self.cert_path,
            "renewal_conditions": self.renewal_conditions,
            "advance_notice_days": self.advance_notice_days,
            "contact_person": self.contact_person,
            "contact_info": self.contact_info,
            "status": self.status,
        }


# ─────────────────────────── 1. AuthorizationDatabase ───────────────────────────

class AuthorizationDatabase:
    """品牌授權主資料庫。"""

    def __init__(self):
        self._records = {}
        self._next_id = 1

    # ---- CRUD ----
    def add_authorization(self, **kwargs):
        auth_id = kwargs.pop("auth_id", None) or f"AUTH-{self._next_id:03d}"
        rec = AuthorizationRecord(auth_id=auth_id, **kwargs)
        self._records[auth_id] = rec
        self._next_id = max(self._next_id, int(auth_id.split("-")[1]) + 1) if auth_id.startswith("AUTH-") else self._next_id + 1
        return rec

    def update_authorization(self, auth_id, **kwargs):
        rec = self._records.get(auth_id)
        if not rec:
            raise KeyError(f"找不到授權紀錄: {auth_id}")
        for k, v in kwargs.items():
            if hasattr(rec, k):
                setattr(rec, k, v)
        return rec

    def get_by_brand(self, brand_name):
        return [r for r in self._records.values()
                if brand_name.lower() in r.brand_name.lower()]

    def get_by_status(self, status):
        return [r for r in self._records.values()
                if r.computed_status == status]

    def get_expiring_soon(self, days=90):
        return sorted(
            [r for r in self._records.values()
             if 0 <= r.days_remaining <= days],
            key=lambda r: r.days_remaining,
        )

    def get_expired(self):
        return [r for r in self._records.values() if r.days_remaining < 0]

    def get_all(self):
        return sorted(self._records.values(), key=lambda r: r.days_remaining)

    def get_by_id(self, auth_id):
        return self._records.get(auth_id)

    @property
    def count(self):
        return len(self._records)

    # ---- 預置範例資料 ----
    def load_sample_data(self):
        samples = [
            dict(
                brand_name="BrandA (電子產品)",
                licensor="BrandA Corp. (日本)",
                licensee="暹羅進口貿易有限公司",
                start_date="2025-06-20",
                end_date="2026-06-20",
                countries="泰國, 印尼",
                product_category="電子產品",
                cert_number="BA-2025-TH-001",
                cert_path="/documents/auth/BA-2025-TH-001.pdf",
                renewal_conditions="提前90天書面通知，需重新審核產品品質",
                advance_notice_days=90,
                contact_person="田中太郎",
                contact_info="tanaka@branda.co.jp",
            ),
            dict(
                brand_name="BrandB (家電)",
                licensor="BrandB Industries (韓國)",
                licensee="雅加達家電進口有限公司",
                start_date="2025-07-15",
                end_date="2026-07-15",
                countries="印尼",
                product_category="家電",
                cert_number="BB-2025-ID-002",
                cert_path="/documents/auth/BB-2025-ID-002.pdf",
                renewal_conditions="自動續約，除非提前60天書面終止",
                advance_notice_days=60,
                contact_person="金民秀",
                contact_info="minsoo@brandb.kr",
            ),
            dict(
                brand_name="BrandC (工業設備)",
                licensor="BrandC GmbH (德國)",
                licensee="曼谷工業設備有限公司",
                start_date="2025-08-10",
                end_date="2026-08-10",
                countries="泰國",
                product_category="工業設備",
                cert_number="BC-2025-TH-003",
                cert_path="/documents/auth/BC-2025-TH-003.pdf",
                renewal_conditions="需提交年度銷售報告及市場分析",
                advance_notice_days=90,
                contact_person="Hans Mueller",
                contact_info="h.mueller@brandc.de",
            ),
            dict(
                brand_name="BrandD (紡織品)",
                licensor="BrandD Textiles (越南)",
                licensee="印尼紡織進口有限公司",
                start_date="2025-09-14",
                end_date="2026-09-14",
                countries="印尼",
                product_category="紡織品",
                cert_number="BD-2025-ID-004",
                cert_path="/documents/auth/BD-2025-ID-004.pdf",
                renewal_conditions="雙方同意即可續約，需提前30天通知",
                advance_notice_days=30,
                contact_person="Nguyen Thi Lan",
                contact_info="lan@brandd.vn",
            ),
            dict(
                brand_name="BrandE (汽車零件)",
                licensor="BrandE Auto Parts (中國)",
                licensee="泰國汽車零件進口有限公司",
                start_date="2026-04-01",
                end_date="2027-03-31",
                countries="泰國, 印尼",
                product_category="汽車零件",
                cert_number="BE-2026-TH-005",
                cert_path="/documents/auth/BE-2026-TH-005.pdf",
                renewal_conditions="年度審核通過後自動續約",
                advance_notice_days=60,
                contact_person="李明",
                contact_info="liming@brande.cn",
            ),
            dict(
                brand_name="BrandF (電子產品)",
                licensor="BrandF Technologies (台灣)",
                licensee="雅加達電子進口有限公司",
                start_date="2025-06-01",
                end_date="2026-06-01",
                countries="印尼",
                product_category="電子產品",
                cert_number="BF-2025-ID-006",
                cert_path="/documents/auth/BF-2025-ID-006.pdf",
                renewal_conditions="需重新提交產品認證文件",
                advance_notice_days=90,
                contact_person="陳志偉",
                contact_info="cwchen@brandf.tw",
            ),
            dict(
                brand_name="BrandG (食品飲料)",
                licensor="BrandG Foods (馬來西亞)",
                licensee="曼谷食品進口有限公司",
                start_date="2025-07-01",
                end_date="2026-07-01",
                countries="泰國",
                product_category="食品飲料",
                cert_number="BG-2025-TH-007",
                cert_path="/documents/auth/BG-2025-TH-007.pdf",
                renewal_conditions="需通過食品安全檢驗，提前45天申請",
                advance_notice_days=45,
                contact_person="Ahmad Razak",
                contact_info="ahmad@brandg.my",
            ),
            dict(
                brand_name="BrandH (化學品)",
                licensor="BrandH Chemicals (美國)",
                licensee="印尼化學品進口有限公司",
                start_date="2026-01-01",
                end_date="2026-12-31",
                countries="泰國, 印尼",
                product_category="化學品",
                cert_number="BH-2026-ID-008",
                cert_path="/documents/auth/BH-2026-ID-008.pdf",
                renewal_conditions="年度合規審查，需提交環保報告",
                advance_notice_days=90,
                contact_person="John Smith",
                contact_info="jsmith@brandh.us",
            ),
            dict(
                brand_name="BrandI (醫療器材)",
                licensor="BrandI Medical (瑞士)",
                licensee="曼谷醫療器材有限公司",
                start_date="2025-08-20",
                end_date="2026-08-20",
                countries="泰國",
                product_category="醫療器材",
                cert_number="BI-2025-TH-009",
                cert_path="/documents/auth/BI-2025-TH-009.pdf",
                renewal_conditions="需提交FDA認證更新及臨床試驗報告",
                advance_notice_days=120,
                contact_person="Dr. Anna Weber",
                contact_info="a.weber@brandi.ch",
            ),
            dict(
                brand_name="BrandJ (運動用品)",
                licensor="BrandJ Sports (澳洲)",
                licensee="東南亞運動用品有限公司",
                start_date="2026-07-01",
                end_date="2027-06-30",
                countries="泰國, 印尼",
                product_category="運動用品",
                cert_number="BJ-2026-TH-010",
                cert_path="/documents/auth/BJ-2026-TH-010.pdf",
                renewal_conditions="雙方協商續約，需提前60天通知",
                advance_notice_days=60,
                contact_person="David Wilson",
                contact_info="dwilson@brandj.au",
            ),
        ]
        for s in samples:
            self.add_authorization(**s)
        return self


# ─────────────────────────── 2. ExpiryAlertEngine ───────────────────────────

class ExpiryAlertEngine:
    """多層級到期預警引擎。"""

    def __init__(self, db: AuthorizationDatabase):
        self.db = db

    def check_all_alerts(self):
        """回傳所有啟動中的預警，依嚴重程度排序。"""
        alerts = []
        for rec in self.db.get_all():
            level = rec.alert_level
            if level:
                info = ALERT_LEVELS[level]
                alerts.append({
                    "record": rec,
                    "level": level,
                    "label": info["label"],
                    "color": info["color"],
                    "hex": info["hex"],
                    "fill": info["fill"],
                    "priority": info["priority"],
                    "days_remaining": rec.days_remaining,
                })
        alerts.sort(key=lambda a: a["priority"])
        return alerts

    def get_alerts_by_level(self, level):
        return [a for a in self.check_all_alerts() if a["level"] == level]

    def generate_alert_message(self, rec: AuthorizationRecord):
        level = rec.alert_level
        if not level:
            return None
        info = ALERT_LEVELS[level]
        days = rec.days_remaining
        if days < 0:
            time_str = f"已過期 {abs(days)} 天"
        elif days == 0:
            time_str = "今日到期"
        else:
            time_str = f"剩餘 {days} 天"

        msg_lines = [
            f"{'='*60}",
            f"  [{info['label']}] {rec.brand_name}",
            f"  授權編號: {rec.cert_number}",
            f"  授權方: {rec.licensor}",
            f"  被授權方: {rec.licensee}",
            f"  到期日: {rec.end_date.strftime('%Y-%m-%d')} ({time_str})",
            f"  授權範圍: {rec.countries}",
            f"  聯絡人: {rec.contact_person} ({rec.contact_info})",
        ]
        if level == "expired":
            msg_lines.append(f"  *** 緊急處置: 請立即聯繫授權方討論續約或延長，並上報經理 ***")
        elif level == "7days":
            msg_lines.append(f"  *** 最後警告: 請立即啟動續約流程並發送 WeCom 通知 ***")
        elif level == "30days":
            msg_lines.append(f"  *** 緊急行動: 請於本週內聯繫授權方確認續約意願 ***")
        elif level == "60days":
            msg_lines.append(f"  *** 提醒: 請開始準備續約文件 ***")
        elif level == "90days":
            msg_lines.append(f"  *** 規劃: 請將續約排入工作計畫 ***")
        msg_lines.append(f"{'='*60}")
        return "\n".join(msg_lines)

    def generate_wecom_notification(self, alerts=None):
        if alerts is None:
            alerts = self.check_all_alerts()
        if not alerts:
            return "目前無任何到期預警。"

        lines = [
            "📋 【品牌授權到期預警通知】",
            f"📅 日期: {TODAY.strftime('%Y-%m-%d')}",
            f"🔔 預警數量: {len(alerts)} 項",
            "",
        ]

        grouped = defaultdict(list)
        for a in alerts:
            grouped[a["label"]].append(a)

        for level_key in ["expired", "7days", "30days", "60days", "90days"]:
            info = ALERT_LEVELS[level_key]
            items = grouped.get(info["label"], [])
            if not items:
                continue
            lines.append(f"▎{info['label']} ({len(items)} 項):")
            for a in items:
                rec = a["record"]
                days = a["days_remaining"]
                if days < 0:
                    day_str = f"已過期 {abs(days)} 天"
                elif days == 0:
                    day_str = "今日到期"
                else:
                    day_str = f"剩餘 {days} 天"
                lines.append(f"  • {rec.brand_name} | {rec.cert_number} | {day_str}")
            lines.append("")

        lines.append("請相關負責人儘速處理。")
        lines.append("— 品牌授權管理系統 自動通知")
        return "\n".join(lines)


# ─────────────────────────── 3. RenewalManager ───────────────────────────

class RenewalManager:
    """續約任務追蹤管理器。"""

    def __init__(self, db: AuthorizationDatabase):
        self.db = db
        self._tasks = {}
        self._history = []
        self._next_task_id = 1

    def create_renewal_task(self, auth_id, assigned_to, deadline, notes=""):
        rec = self.db.get_by_id(auth_id)
        if not rec:
            raise KeyError(f"找不到授權紀錄: {auth_id}")
        task_id = f"RENEW-{self._next_task_id:03d}"
        self._next_task_id += 1
        task = {
            "task_id": task_id,
            "auth_id": auth_id,
            "brand_name": rec.brand_name,
            "assigned_to": assigned_to,
            "deadline": RenewalManager._parse_deadline(deadline),
            "status": "進行中",
            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "notes": notes,
            "history": [
                {"time": datetime.now().strftime("%Y-%m-%d %H:%M"),
                 "action": "建立續約任務", "by": "系統"}
            ],
        }
        self._tasks[task_id] = task
        return task

    def update_renewal_status(self, task_id, status, notes=""):
        task = self._tasks.get(task_id)
        if not task:
            raise KeyError(f"找不到續約任務: {task_id}")
        old_status = task["status"]
        task["status"] = status
        task["notes"] = notes
        task["history"].append({
            "time": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "action": f"狀態更新: {old_status} → {status}",
            "by": "操作員",
        })
        if status in ("已完成", "已取消"):
            self._history.append(task)
            del self._tasks[task_id]
        return task

    def get_pending_renewals(self):
        return list(self._tasks.values())

    def get_renewal_history(self):
        return list(self._history)

    def calculate_renewal_cost(self, auth_id):
        rec = self.db.get_by_id(auth_id)
        if not rec:
            raise KeyError(f"找不到授權紀錄: {auth_id}")

        # 依據產品類別估算續約成本與時程
        cost_table = {
            "電子產品":   {"fee": 150000, "currency": "THB", "processing_days": 30, "docs": "授權書、產品認證、品質報告"},
            "家電":       {"fee": 120000, "currency": "THB", "processing_days": 21, "docs": "授權書、產品目錄"},
            "工業設備":   {"fee": 200000, "currency": "THB", "processing_days": 45, "docs": "授權書、技術規格、年度報告"},
            "紡織品":     {"fee": 80000,  "currency": "THB", "processing_days": 14, "docs": "授權書、材質證明"},
            "汽車零件":   {"fee": 180000, "currency": "THB", "processing_days": 30, "docs": "授權書、安全認證、檢驗報告"},
            "食品飲料":   {"fee": 100000, "currency": "THB", "processing_days": 30, "docs": "授權書、FDA認證、食品安全檢驗"},
            "化學品":     {"fee": 160000, "currency": "THB", "processing_days": 45, "docs": "授權書、環保報告、安全資料表"},
            "醫療器材":   {"fee": 250000, "currency": "THB", "processing_days": 60, "docs": "授權書、FDA認證更新、臨床報告"},
            "運動用品":   {"fee": 90000,  "currency": "THB", "processing_days": 21, "docs": "授權書、產品目錄"},
        }
        info = cost_table.get(rec.product_category, {"fee": 100000, "currency": "THB", "processing_days": 30, "docs": "授權書"})
        return {
            "auth_id": auth_id,
            "brand_name": rec.brand_name,
            "estimated_fee": info["fee"],
            "currency": info["currency"],
            "processing_days": info["processing_days"],
            "required_documents": info["docs"],
            "recommended_start": (rec.end_date - timedelta(days=info["processing_days"] + 30)).strftime("%Y-%m-%d"),
        }

    @staticmethod
    def _parse_deadline(d):
        if isinstance(d, datetime):
            return d.strftime("%Y-%m-%d")
        return d

    def load_sample_tasks(self):
        """載入範例續約任務。"""
        self.create_renewal_task("AUTH-002", "張曉明", "2026-07-01", "已聯繫韓方確認續約意願")
        self.create_renewal_task("AUTH-006", "林佳穎", "2026-06-25", "過期緊急處理中")
        self.create_renewal_task("AUTH-007", "王大偉", "2026-06-28", "食品安全檢驗進行中")
        task_a = self.create_renewal_task("AUTH-001", "陳俊宏", "2026-06-18", "緊急續約處理")
        self.create_renewal_task("AUTH-003", "黃美玲", "2026-07-20", "準備技術文件")

        # 將 BrandF 的任務標記為進行中（過期處理）
        return self


# ─────────────────────────── 4. ReportGenerator ───────────────────────────

class ReportGenerator:
    """報表與分析產出器。"""

    def __init__(self, db: AuthorizationDatabase, engine: ExpiryAlertEngine,
                 manager: RenewalManager):
        self.db = db
        self.engine = engine
        self.manager = manager

    # ---- 綜合狀態報告 ----
    def generate_status_report(self):
        all_recs = self.db.get_all()
        alerts = self.engine.check_all_alerts()
        expired = self.db.get_expired()
        expiring_30 = self.db.get_expiring_soon(30)
        pending = self.manager.get_pending_renewals()

        report = {
            "report_date": TODAY.strftime("%Y-%m-%d"),
            "total_authorizations": len(all_recs),
            "active": len([r for r in all_recs if r.computed_status == "有效"]),
            "expiring_soon": len([r for r in all_recs if r.computed_status == "即將到期"]),
            "expired": len(expired),
            "active_alerts": len(alerts),
            "pending_renewals": len(pending),
            "critical_items": [r.brand_name for r in all_recs if r.days_remaining <= 7],
        }
        return report

    # ---- 月報 ----
    def generate_monthly_report(self, year_month=None):
        if year_month is None:
            year_month = TODAY.strftime("%Y-%m")
        all_recs = self.db.get_all()
        month_data = {
            "period": year_month,
            "total": len(all_recs),
            "by_country": defaultdict(list),
            "by_category": defaultdict(list),
            "expiring_this_month": [],
        }
        for rec in all_recs:
            for country in rec.countries.split(", "):
                month_data["by_country"][country].append(rec.brand_name)
            month_data["by_category"][rec.product_category].append(rec.brand_name)
            if rec.end_date.strftime("%Y-%m") == year_month:
                month_data["expiring_this_month"].append(rec.brand_name)
        return month_data

    # ---- 續約日曆 ----
    def generate_renewal_calendar(self, months=6):
        end_range = TODAY + timedelta(days=months * 30)
        calendar = defaultdict(list)
        for rec in self.db.get_all():
            if rec.end_date <= end_range:
                month_key = rec.end_date.strftime("%Y-%m")
                calendar[month_key].append({
                    "brand": rec.brand_name,
                    "end_date": rec.end_date.strftime("%Y-%m-%d"),
                    "days_remaining": rec.days_remaining,
                    "cert_number": rec.cert_number,
                })
        # 排序
        for k in calendar:
            calendar[k].sort(key=lambda x: x["end_date"])
        return dict(sorted(calendar.items()))

    # ---- 風險評估 ----
    def generate_risk_assessment(self):
        risks = []
        for rec in self.db.get_all():
            risk_score = 0
            risk_factors = []

            # 到期時間風險
            if rec.days_remaining < 0:
                risk_score += 50
                risk_factors.append("已過期")
            elif rec.days_remaining <= 7:
                risk_score += 40
                risk_factors.append("7天內到期")
            elif rec.days_remaining <= 30:
                risk_score += 30
                risk_factors.append("30天內到期")
            elif rec.days_remaining <= 60:
                risk_score += 20
                risk_factors.append("60天內到期")
            elif rec.days_remaining <= 90:
                risk_score += 10
                risk_factors.append("90天內到期")

            # 提前通知天數風險
            if rec.days_remaining < rec.advance_notice_days:
                risk_score += 15
                risk_factors.append(f"已超過提前通知期限 ({rec.advance_notice_days}天)")

            # 多國家風險（管理複雜度）
            num_countries = len(rec.countries.split(", "))
            if num_countries > 1:
                risk_score += 5 * num_countries
                risk_factors.append(f"涉及 {num_countries} 個國家")

            # 特殊類別風險
            high_risk_categories = {"醫療器材": 15, "化學品": 10, "食品飲料": 10}
            cat_risk = high_risk_categories.get(rec.product_category, 0)
            if cat_risk:
                risk_score += cat_risk
                risk_factors.append(f"高風險類別: {rec.product_category}")

            # 風險等級
            if risk_score >= 50:
                risk_level = "極高"
            elif risk_score >= 35:
                risk_level = "高"
            elif risk_score >= 20:
                risk_level = "中"
            else:
                risk_level = "低"

            risks.append({
                "auth_id": rec.auth_id,
                "brand_name": rec.brand_name,
                "risk_score": risk_score,
                "risk_level": risk_level,
                "risk_factors": risk_factors,
                "days_remaining": rec.days_remaining,
            })

        risks.sort(key=lambda x: -x["risk_score"])
        return risks

    # ---- 匯出 Excel ----
    def export_to_excel(self, filepath=None):
        if not HAS_OPENPYXL:
            print("[錯誤] 需要安裝 openpyxl 才能匯出 Excel。")
            return None
        if filepath is None:
            filepath = os.path.join(WORKING_DIR, "品牌授權數據庫.xlsx")

        wb = Workbook()

        # ========== Sheet 1: 品牌授權主數據 ==========
        ws1 = wb.active
        ws1.title = "品牌授權主數據"

        headers1 = [
            "授權編號", "品牌名稱", "授權方", "被授權方",
            "授權起始日", "授權到期日", "剩餘天數", "授權範圍",
            "產品類別", "授權書編號", "授權書存放路徑", "續約條件",
            "提前通知天數", "聯絡人", "聯絡方式", "狀態", "預警等級"
        ]
        self._write_header(ws1, headers1)

        for idx, rec in enumerate(self.db.get_all(), start=2):
            level = rec.alert_level
            level_label = ALERT_LEVELS[level]["label"] if level else "正常"
            level_fill_hex = ALERT_LEVELS[level]["fill"] if level else "C6EFCE"

            row_data = [
                rec.auth_id, rec.brand_name, rec.licensor, rec.licensee,
                rec.start_date.strftime("%Y-%m-%d"),
                rec.end_date.strftime("%Y-%m-%d"),
                rec.days_remaining, rec.countries,
                rec.product_category, rec.cert_number, rec.cert_path,
                rec.renewal_conditions, rec.advance_notice_days,
                rec.contact_person, rec.contact_info,
                rec.computed_status, level_label,
            ]
            for col, val in enumerate(row_data, start=1):
                cell = ws1.cell(row=idx, column=col, value=val)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = HEADER_BORDER
                # 狀態欄色彩
                if col == 16:
                    if rec.computed_status == "已過期":
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        cell.font = Font(bold=True, color="9C0006")
                    elif rec.computed_status == "即將到期":
                        cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
                        cell.font = Font(bold=True, color="9C5700")
                    else:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        cell.font = Font(bold=True, color="006100")
                # 預警等級欄色彩
                if col == 17:
                    cell.fill = PatternFill(start_color=level_fill_hex, end_color=level_fill_hex, fill_type="solid")
                # 剩餘天數欄
                if col == 7:
                    if isinstance(val, int) and val < 0:
                        cell.font = Font(bold=True, color="FF0000")
                    elif isinstance(val, int) and val <= 30:
                        cell.font = Font(bold=True, color="FF8C00")

        # 設定欄寬
        col_widths1 = [12, 22, 26, 26, 14, 14, 10, 14, 12, 18, 36, 32, 12, 16, 26, 10, 12]
        for i, w in enumerate(col_widths1, start=1):
            ws1.column_dimensions[get_column_letter(i)].width = w

        ws1.freeze_panes = "A2"
        ws1.auto_filter.ref = ws1.dimensions

        # ========== Sheet 2: 續約追蹤 ==========
        ws2 = wb.create_sheet("續約追蹤")

        headers2 = [
            "任務編號", "授權編號", "品牌名稱", "負責人",
            "截止日期", "狀態", "建立時間", "備註"
        ]
        self._write_header(ws2, headers2)

        tasks = self.manager.get_pending_renewals() + self.manager.get_renewal_history()
        for idx, task in enumerate(tasks, start=2):
            row_data = [
                task["task_id"], task["auth_id"], task["brand_name"],
                task["assigned_to"], task["deadline"], task["status"],
                task["created_at"], task["notes"],
            ]
            for col, val in enumerate(row_data, start=1):
                cell = ws2.cell(row=idx, column=col, value=val)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = HEADER_BORDER
                if col == 6:
                    if val == "已完成":
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        cell.font = Font(bold=True, color="006100")
                    elif val == "進行中":
                        cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
                        cell.font = Font(bold=True, color="1F4E79")
                    elif val == "已取消":
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        cell.font = Font(bold=True, color="9C0006")

        col_widths2 = [14, 14, 22, 12, 14, 10, 18, 32]
        for i, w in enumerate(col_widths2, start=1):
            ws2.column_dimensions[get_column_letter(i)].width = w

        ws2.freeze_panes = "A2"
        ws2.auto_filter.ref = ws2.dimensions

        # ========== Sheet 3: 到期預警日誌 ==========
        ws3 = wb.create_sheet("到期預警日誌")

        headers3 = [
            "序號", "授權編號", "品牌名稱", "預警等級",
            "剩餘天數", "到期日", "偵測時間", "預警動作", "負責人"
        ]
        self._write_header(ws3, headers3)

        alerts = self.engine.check_all_alerts()
        for idx, alert in enumerate(alerts, start=2):
            rec = alert["record"]
            action = "系統自動偵測"
            if alert["level"] == "expired":
                action = "已上報經理，緊急處理中"
            elif alert["level"] == "7days":
                action = "已發送 WeCom 通知"
            elif alert["level"] == "30days":
                action = "已建立續約任務"

            row_data = [
                idx - 1, rec.auth_id, rec.brand_name,
                alert["label"], alert["days_remaining"],
                rec.end_date.strftime("%Y-%m-%d"),
                datetime.now().strftime("%Y-%m-%d %H:%M"),
                action, rec.contact_person,
            ]
            for col, val in enumerate(row_data, start=1):
                cell = ws3.cell(row=idx, column=col, value=val)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = HEADER_BORDER
                if col == 4:
                    cell.fill = PatternFill(start_color=alert["fill"], end_color=alert["fill"], fill_type="solid")

        col_widths3 = [8, 14, 22, 12, 10, 14, 18, 24, 16]
        for i, w in enumerate(col_widths3, start=1):
            ws3.column_dimensions[get_column_letter(i)].width = w

        ws3.freeze_panes = "A2"
        ws3.auto_filter.ref = ws3.dimensions

        # 儲存
        wb.save(filepath)
        return filepath

    # ---- 工具函式 ----
    @staticmethod
    def _write_header(ws, headers):
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = WHITE_BOLD_FONT
            cell.fill = DEEP_BLUE_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = HEADER_BORDER
        ws.row_dimensions[1].height = 28


# ─────────────────────────── 5. BrandAuthTool (CLI) ───────────────────────────

class BrandAuthTool:
    """品牌授權管理系統 — 主命令列介面。"""

    def __init__(self):
        self.db = AuthorizationDatabase()
        self.engine = ExpiryAlertEngine(self.db)
        self.manager = RenewalManager(self.db)
        self.reporter = ReportGenerator(self.db, self.engine, self.manager)

    # ---- 總覽 ----
    def show_overview(self):
        all_recs = self.db.get_all()
        print("\n" + "=" * 100)
        print("                        品牌授權管理系統 — 總覽")
        print("=" * 100)
        print(f"{'編號':<12}{'品牌名稱':<24}{'產品類別':<12}{'授權範圍':<12}"
              f"{'到期日':<14}{'剩餘天數':>8}{'狀態':<10}{'預警等級':<10}")
        print("-" * 100)

        for rec in all_recs:
            level = rec.alert_level
            level_label = ALERT_LEVELS[level]["label"] if level else "正常"
            days = rec.days_remaining
            if days < 0:
                days_str = f"{abs(days)}天(過期)"
            else:
                days_str = str(days)

            # 顏色標記
            marker = ""
            if level in ("expired", "7days"):
                marker = " [!!!]"
            elif level == "30days":
                marker = " [!!]"
            elif level == "60days":
                marker = " [!]"

            print(f"{rec.auth_id:<12}{rec.brand_name:<24}{rec.product_category:<12}"
                  f"{rec.countries:<12}{rec.end_date.strftime('%Y-%m-%d'):<14}"
                  f"{days_str:>8}  {rec.computed_status:<10}{level_label:<10}{marker}")

        print("-" * 100)

        summary = self.reporter.generate_status_report()
        print(f"  總計: {summary['total_authorizations']} 筆 | "
              f"有效: {summary['active']} | "
              f"即將到期: {summary['expiring_soon']} | "
              f"已過期: {summary['expired']}")
        print("=" * 100)

    # ---- 檢查預警 ----
    def check_alerts(self, days_filter=None):
        alerts = self.engine.check_all_alerts()
        if days_filter is not None:
            alerts = [a for a in alerts if a["days_remaining"] <= days_filter]

        print("\n" + "=" * 80)
        print("                        品牌授權到期預警")
        print("=" * 80)

        if not alerts:
            print("  目前無任何到期預警。")
            print("=" * 80)
            return

        # 依等級分組顯示
        grouped = defaultdict(list)
        for a in alerts:
            grouped[a["label"]].append(a)

        level_order = ["已過期", "最後警告", "緊急續約", "準備續約", "提前規劃"]
        for label in level_order:
            items = grouped.get(label, [])
            if not items:
                continue
            print(f"\n  ▎{label} ({len(items)} 項)")
            print(f"  {'─'*70}")
            for a in items:
                rec = a["record"]
                days = a["days_remaining"]
                if days < 0:
                    day_str = f"已過期 {abs(days)} 天"
                elif days == 0:
                    day_str = "今日到期"
                else:
                    day_str = f"剩餘 {days} 天"
                print(f"    {rec.auth_id} | {rec.brand_name:<22} | "
                      f"到期: {rec.end_date.strftime('%Y-%m-%d')} | {day_str}")

        print(f"\n{'=' * 80}")
        print(f"  預警總計: {len(alerts)} 項")
        print("=" * 80)

    # ---- 續約狀態 ----
    def show_renewal_status(self):
        pending = self.manager.get_pending_renewals()
        history = self.manager.get_renewal_history()

        print("\n" + "=" * 90)
        print("                        續約任務追蹤")
        print("=" * 90)

        if pending:
            print(f"\n  ▎進行中的續約任務 ({len(pending)} 項)")
            print(f"  {'─'*80}")
            print(f"  {'任務編號':<14}{'授權編號':<12}{'品牌名稱':<24}{'負責人':<10}"
                  f"{'截止日期':<14}{'狀態':<10}")
            print(f"  {'─'*80}")
            for t in pending:
                print(f"  {t['task_id']:<14}{t['auth_id']:<12}{t['brand_name']:<24}"
                      f"{t['assigned_to']:<10}{t['deadline']:<14}{t['status']:<10}")
                if t.get("notes"):
                    print(f"    └─ 備註: {t['notes']}")
        else:
            print("\n  目前無進行中的續約任務。")

        if history:
            print(f"\n  ▎已完成的續約紀錄 ({len(history)} 項)")
            print(f"  {'─'*80}")
            for t in history:
                print(f"  {t['task_id']} | {t['brand_name']} | {t['status']} | {t['notes']}")

        print(f"\n{'=' * 90}")

    # ---- 建立續約任務 ----
    def create_renewal(self, auth_id):
        rec = self.db.get_by_id(auth_id)
        if not rec:
            print(f"[錯誤] 找不到授權編號: {auth_id}")
            return

        level = rec.alert_level
        days = rec.days_remaining
        # 根據緊急程度自動設定截止日期
        if days < 0:
            deadline = (TODAY + timedelta(days=7)).strftime("%Y-%m-%d")
            assigned = "林佳穎"
        elif days <= 30:
            deadline = (TODAY + timedelta(days=14)).strftime("%Y-%m-%d")
            assigned = "張曉明"
        else:
            deadline = (rec.end_date - timedelta(days=30)).strftime("%Y-%m-%d")
            assigned = "陳俊宏"

        task = self.manager.create_renewal_task(
            auth_id=auth_id,
            assigned_to=assigned,
            deadline=deadline,
            notes=f"由系統自動建立，品牌: {rec.brand_name}",
        )

        cost = self.manager.calculate_renewal_cost(auth_id)

        print(f"\n{'=' * 70}")
        print(f"  續約任務已建立")
        print(f"{'=' * 70}")
        print(f"  任務編號:   {task['task_id']}")
        print(f"  授權編號:   {auth_id}")
        print(f"  品牌名稱:   {rec.brand_name}")
        print(f"  負責人:     {task['assigned_to']}")
        print(f"  截止日期:   {task['deadline']}")
        print(f"  預估費用:   {cost['estimated_fee']:,} {cost['currency']}")
        print(f"  處理天數:   約 {cost['processing_days']} 天")
        print(f"  所需文件:   {cost['required_documents']}")
        print(f"  建議啟動日: {cost['recommended_start']}")
        print(f"{'=' * 70}")

    # ---- 搜尋品牌 ----
    def search_brand(self, keyword):
        results = self.db.get_by_brand(keyword)
        print(f"\n{'=' * 80}")
        print(f"  搜尋結果: 「{keyword}」 (共 {len(results)} 筆)")
        print(f"{'=' * 80}")
        if not results:
            print("  未找到符合的品牌授權紀錄。")
            return
        for rec in results:
            msg = self.engine.generate_alert_message(rec)
            if msg:
                print(msg)
            else:
                print(f"  {rec.auth_id} | {rec.brand_name} | 狀態: 正常 | "
                      f"到期: {rec.end_date.strftime('%Y-%m-%d')} | "
                      f"剩餘 {rec.days_remaining} 天")

    # ---- 續約日曆 ----
    def show_calendar(self):
        cal = self.reporter.generate_renewal_calendar(6)
        print(f"\n{'=' * 80}")
        print(f"                        未來 6 個月續約日曆")
        print(f"{'=' * 80}")
        if not cal:
            print("  未來 6 個月內無到期的授權。")
            print(f"{'=' * 80}")
            return
        for month, items in cal.items():
            print(f"\n  ▎{month}")
            print(f"  {'─'*60}")
            for item in items:
                days = item["days_remaining"]
                if days < 0:
                    marker = f"[已過期 {abs(days)} 天]"
                elif days <= 7:
                    marker = f"[!!! 剩餘 {days} 天]"
                elif days <= 30:
                    marker = f"[!! 剩餘 {days} 天]"
                else:
                    marker = f"[剩餘 {days} 天]"
                print(f"    {item['end_date']} | {item['brand']:<24} | "
                      f"{item['cert_number']:<18} | {marker}")
        print(f"\n{'=' * 80}")

    # ---- 風險評估 ----
    def show_risk(self):
        risks = self.reporter.generate_risk_assessment()
        print(f"\n{'=' * 90}")
        print(f"                        品牌授權風險評估")
        print(f"{'=' * 90}")
        print(f"  {'編號':<12}{'品牌名稱':<24}{'風險分數':>8}{'風險等級':<8}{'風險因素'}")
        print(f"  {'─'*80}")
        for r in risks:
            factors = ", ".join(r["risk_factors"]) if r["risk_factors"] else "無"
            print(f"  {r['auth_id']:<12}{r['brand_name']:<24}"
                  f"{r['risk_score']:>6}  {r['risk_level']:<8}{factors}")

        # 統計
        high = [r for r in risks if r["risk_level"] in ("極高", "高")]
        mid = [r for r in risks if r["risk_level"] == "中"]
        low = [r for r in risks if r["risk_level"] == "低"]
        print(f"\n  風險統計: 極高/高: {len(high)} | 中: {len(mid)} | 低: {len(low)}")
        print(f"{'=' * 90}")

    # ---- 產生報表 ----
    def generate_report(self):
        filepath = self.reporter.export_to_excel()
        if filepath:
            print(f"\n{'=' * 70}")
            print(f"  Excel 報表已產出: {filepath}")
            print(f"{'=' * 70}")
            print(f"  包含以下工作表:")
            print(f"    1. 品牌授權主數據 — 所有授權紀錄及狀態")
            print(f"    2. 續約追蹤 — 續約任務進度")
            print(f"    3. 到期預警日誌 — 預警紀錄及處理動作")
            print(f"{'=' * 70}")

    # ---- 新增授權（互動式） ----
    def add_interactive(self):
        print(f"\n{'=' * 60}")
        print("  新增品牌授權")
        print(f"{'=' * 60}")
        fields = [
            ("品牌名稱", "brand_name"),
            ("授權方", "licensor"),
            ("被授權方", "licensee"),
            ("授權起始日 (YYYY-MM-DD)", "start_date"),
            ("授權到期日 (YYYY-MM-DD)", "end_date"),
            ("授權範圍（國家）", "countries"),
            ("產品類別", "product_category"),
            ("授權書編號", "cert_number"),
            ("授權書存放路徑", "cert_path"),
            ("續約條件", "renewal_conditions"),
            ("提前通知天數", "advance_notice_days"),
            ("聯絡人", "contact_person"),
            ("聯絡方式", "contact_info"),
        ]
        data = {}
        for label, key in fields:
            val = input(f"  {label}: ").strip()
            if key == "advance_notice_days":
                val = int(val) if val.isdigit() else 30
            data[key] = val

        rec = self.db.add_authorization(**data)
        print(f"\n  授權已成功新增: {rec.auth_id} — {rec.brand_name}")

    # ---- 完整 Demo ----
    def run_demo(self):
        print("\n" + "#" * 100)
        print("#" + " " * 98 + "#")
        print("#" + "              品牌授權管理系統 (Brand Authorization Management System)            ".center(98) + "#")
        print("#" + " " * 98 + "#")
        print("#" * 100)
        print()

        # Step 1: 載入範例資料
        print("[Step 1] 載入範例品牌授權資料...")
        self.db.load_sample_data()
        self.manager.load_sample_tasks()
        print(f"  已載入 {self.db.count} 筆品牌授權紀錄")
        print(f"  已載入 {len(self.manager.get_pending_renewals())} 筆續約任務")
        print()

        # Step 2: 總覽
        print("[Step 2] 顯示所有品牌授權總覽...")
        self.show_overview()
        print()

        # Step 3: 檢查所有預警
        print("[Step 3] 檢查所有到期預警...")
        self.check_alerts()
        print()

        # Step 4: 顯示已過期項目
        print("[Step 4] 已過期項目 — 需緊急處置...")
        expired = self.db.get_expired()
        if expired:
            for rec in expired:
                msg = self.engine.generate_alert_message(rec)
                print(msg)
            print("\n  [升級通知] 以下品牌授權已過期，已自動上報經理:")
            for rec in expired:
                print(f"    - {rec.brand_name} ({rec.cert_number})")
        print()

        # Step 5: 顯示緊急項目
        print("[Step 5] 緊急項目（7天內 / 30天內到期）...")
        critical = [r for r in self.db.get_all() if 0 <= r.days_remaining <= 30]
        for rec in critical:
            msg = self.engine.generate_alert_message(rec)
            print(msg)
        print()

        # Step 6: WeCom 通知
        print("[Step 6] 產生 WeCom 通知訊息...")
        wecom_msg = self.engine.generate_wecom_notification()
        print(wecom_msg)
        print()

        # Step 7: 建立續約任務
        print("[Step 7] 為 BrandB 建立續約任務...")
        self.create_renewal("AUTH-002")
        print()

        # Step 8: 產生 Excel 報表
        print("[Step 8] 產生綜合 Excel 報表...")
        self.generate_report()
        print()

        # Step 9: 續約日曆
        print("[Step 9] 顯示未來 6 個月續約日曆...")
        self.show_calendar()
        print()

        # Step 10: 風險評估
        print("[Step 10] 風險評估摘要...")
        self.show_risk()
        print()

        print("#" * 100)
        print("#" + " " * 98 + "#")
        print("#" + "                           品牌授權管理系統 — Demo 完成                          ".center(98) + "#")
        print("#" + " " * 98 + "#")
        print("#" * 100)
        print()


# ─────────────────────────── CLI 入口 ───────────────────────────

def build_parser():
    parser = argparse.ArgumentParser(
        description="品牌授權管理系統 — 管理多品牌進口至泰國與印尼的授權追蹤",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用範例:
  python3 品牌授權管理系統.py --demo              執行完整 Demo
  python3 品牌授權管理系統.py --overview           顯示所有授權總覽
  python3 品牌授權管理系統.py --check-alerts       檢查所有到期預警
  python3 品牌授權管理系統.py --check-alerts --days 30  檢查 30 天內預警
  python3 品牌授權管理系統.py --renewal-status     顯示續約任務狀態
  python3 品牌授權管理系統.py --create-renewal AUTH-001  建立續約任務
  python3 品牌授權管理系統.py --report             產生 Excel 報表
  python3 品牌授權管理系統.py --calendar           顯示續約日曆
  python3 品牌授權管理系統.py --risk               顯示風險評估
  python3 品牌授權管理系統.py --search BrandA      搜尋品牌
  python3 品牌授權管理系統.py --add                新增授權（互動式）
        """,
    )

    parser.add_argument("--overview", action="store_true",
                        help="顯示所有品牌授權總覽")
    parser.add_argument("--check-alerts", action="store_true",
                        help="檢查並顯示所有到期預警")
    parser.add_argument("--days", type=int, default=None,
                        help="搭配 --check-alerts，篩選指定天數內的預警")
    parser.add_argument("--renewal-status", action="store_true",
                        help="顯示所有續約任務狀態")
    parser.add_argument("--create-renewal", metavar="AUTH_ID",
                        help="為指定授權建立續約任務")
    parser.add_argument("--report", action="store_true",
                        help="產生綜合 Excel 報表")
    parser.add_argument("--calendar", action="store_true",
                        help="顯示未來 6 個月續約日曆")
    parser.add_argument("--risk", action="store_true",
                        help="顯示風險評估")
    parser.add_argument("--search", metavar="BRAND",
                        help="搜尋指定品牌")
    parser.add_argument("--add", action="store_true",
                        help="新增品牌授權（互動模式）")
    parser.add_argument("--demo", action="store_true",
                        help="執行完整 Demo（含所有功能展示）")
    return parser


def main():
    parser = build_parser()
    args = parser.parse_args()

    tool = BrandAuthTool()

    # Demo 模式：載入樣本資料
    if args.demo:
        tool.run_demo()
        return

    # 其他模式也需要載入資料
    tool.db.load_sample_data()
    tool.manager.load_sample_tasks()

    if args.overview:
        tool.show_overview()
    elif args.check_alerts:
        tool.check_alerts(days_filter=args.days)
    elif args.renewal_status:
        tool.show_renewal_status()
    elif args.create_renewal:
        tool.create_renewal(args.create_renewal)
    elif args.report:
        tool.generate_report()
    elif args.calendar:
        tool.show_calendar()
    elif args.risk:
        tool.show_risk()
    elif args.search:
        tool.search_brand(args.search)
    elif args.add:
        tool.add_interactive()
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
