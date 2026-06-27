#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
  海關政策變動自動追蹤工具 (Customs Policy Change Auto-Tracker)
================================================================================

使用說明 (Usage):
-----------------

  1. 查看政策總覽:
     python3 政策追蹤工具.py --overview

  2. 搜尋政策:
     python3 政策追蹤工具.py --search "電子產品" --country 泰國
     python3 政策追蹤工具.py --search "SNI" --country 印尼
     python3 政策追蹤工具.py --search "關稅" --category 關稅調整

  3. 檢查到期預警 (30天內到期的授權/許可):
     python3 政策追蹤工具.py --expiry-check --days 30

  4. 生成月度報告:
     python3 政策追蹤工具.py --monthly-report 2026-06

  5. 新增政策記錄:
     python3 政策追蹤工具.py --add --title "泰國新關稅政策" --country 泰國 --category 關稅調整

  6. 指定資料庫路徑 (預設為同目錄下的 政策資料庫.xlsx):
     python3 政策追蹤工具.py --overview --db /path/to/政策資料庫.xlsx

功能概述:
---------
  - 政策資料庫管理: 新增、搜尋、統計政策記錄
  - 政策影響分析: 分析政策變動對業務的影響，生成月度報告
  - 到期預警管理: 追蹤品牌授權、進口許可、海關資質到期

技術要求:
---------
  - Python 3.6+
  - openpyxl (pip3 install openpyxl)

================================================================================
"""

import argparse
import os
import sys
from datetime import datetime, timedelta
from copy import copy

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("\033[91m[錯誤] 缺少 openpyxl 套件，請執行: pip3 install openpyxl\033[0m")
    sys.exit(1)


# =============================================================================
#  常數與設定
# =============================================================================

# 路徑設定
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_DB_PATH = os.path.join(SCRIPT_DIR, "政策資料庫.xlsx")
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "輸出")

# 國家選項
COUNTRIES = ["泰國", "印尼", "中國"]

# 政策類別
POLICY_CATEGORIES = [
    "關稅調整", "法規變更", "貿易協定",
    "檢驗要求", "許可證要求", "其他"
]

# 影響程度
IMPACT_LEVELS = ["高", "中", "低"]

# 政策狀態
POLICY_STATUSES = ["已生效", "待生效", "過渡中", "已過期"]

# 貿易方向
TRADE_DIRECTIONS = ["出口", "進口", "兩者"]

# HS 編碼章節對應描述
HS_CHAPTER_DESC = {
    "第01章": "活動物",
    "第02章": "肉類",
    "第03章": "魚類及甲殼類",
    "第07章": "蔬菜",
    "第08章": "水果",
    "第10章": "穀物",
    "第15章": "動植物油脂",
    "第20章": "蔬菜/水果製品",
    "第22章": "飲料",
    "第25章": "鹽/硫磺/石灰",
    "第28章": "无机化學品",
    "第29章": "有機化學品",
    "第30章": "藥品",
    "第33章": "精油及化妝品",
    "第34章": "肥皂及洗滌劑",
    "第38章": "雜項化學品",
    "第39章": "塑膠及其製品",
    "第40章": "橡膠及其製品",
    "第42章": "皮革製品",
    "第44章": "木材及木製品",
    "第48章": "紙及紙板",
    "第54章": "化學纖維",
    "第61章": "針織服裝",
    "第62章": "梭織服裝",
    "第64章": "鞋類",
    "第68章": "石材製品",
    "第69章": "陶瓷產品",
    "第70章": "玻璃及其製品",
    "第73章": "鋼鐵製品",
    "第74章": "銅及其製品",
    "第76章": "鋁及其製品",
    "第82章": "工具/刀具",
    "第84章": "機械及機械器具",
    "第85章": "電機及電氣設備",
    "第87章": "車輛",
    "第90章": "光學/精密儀器",
    "第94章": "家具/燈具",
    "第95章": "玩具/運動用品",
    "第96章": "雜項製品",
}

# =============================================================================
#  終端彩色輸出工具
# =============================================================================

class ColorPrinter:
    """終端彩色輸出"""

    RESET   = "\033[0m"
    BOLD    = "\033[1m"
    DIM     = "\033[2m"
    RED     = "\033[91m"
    GREEN   = "\033[92m"
    YELLOW  = "\033[93m"
    BLUE    = "\033[94m"
    MAGENTA = "\033[95m"
    CYAN    = "\033[96m"
    WHITE   = "\033[97m"
    BG_RED    = "\033[41m"
    BG_GREEN  = "\033[42m"
    BG_YELLOW = "\033[43m"
    BG_BLUE   = "\033[44m"

    @staticmethod
    def header(title):
        """印出大標題"""
        width = 72
        print()
        print(f"{ColorPrinter.CYAN}{ColorPrinter.BOLD}{'=' * width}{ColorPrinter.RESET}")
        print(f"{ColorPrinter.CYAN}{ColorPrinter.BOLD}{title:^{width}}{ColorPrinter.RESET}")
        print(f"{ColorPrinter.CYAN}{ColorPrinter.BOLD}{'=' * width}{ColorPrinter.RESET}")
        print()

    @staticmethod
    def section(title):
        """印出區段標題"""
        width = 60
        print()
        print(f"  {ColorPrinter.BLUE}{ColorPrinter.BOLD}{'─' * width}{ColorPrinter.RESET}")
        print(f"  {ColorPrinter.BLUE}{ColorPrinter.BOLD}  {title}{ColorPrinter.RESET}")
        print(f"  {ColorPrinter.BLUE}{ColorPrinter.BOLD}{'─' * width}{ColorPrinter.RESET}")

    @staticmethod
    def success(msg):
        print(f"  {ColorPrinter.GREEN}[OK] {msg}{ColorPrinter.RESET}")

    @staticmethod
    def warning(msg):
        print(f"  {ColorPrinter.YELLOW}[!] {msg}{ColorPrinter.RESET}")

    @staticmethod
    def error(msg):
        print(f"  {ColorPrinter.RED}[X] {msg}{ColorPrinter.RESET}")

    @staticmethod
    def info(msg):
        print(f"  {ColorPrinter.CYAN}[i] {msg}{ColorPrinter.RESET}")

    @staticmethod
    def stat(label, value, color=None):
        c = color or ColorPrinter.WHITE
        print(f"    {ColorPrinter.DIM}{label}:{ColorPrinter.RESET} {c}{ColorPrinter.BOLD}{value}{ColorPrinter.RESET}")

    @staticmethod
    def table_row(cells, widths, bold=False, color=None):
        c = color or ""
        b = ColorPrinter.BOLD if bold else ""
        parts = []
        for cell, w in zip(cells, widths):
            text = str(cell)
            # 粗略估算：中文字元佔 2 個字寬
            display_len = sum(2 if ord(ch) > 127 else 1 for ch in text)
            padding = max(0, w - display_len)
            parts.append(f"{c}{b}{text}{ColorPrinter.RESET}{' ' * padding}")
        print(f"    {' | '.join(parts)}")

    @staticmethod
    def table_sep(widths):
        print(f"    {'-+-'.join('-' * w for w in widths)}")


# =============================================================================
#  日期工具
# =============================================================================

def parse_date(date_str):
    """解析日期字串，支援多種格式"""
    if date_str is None:
        return None
    if isinstance(date_str, datetime):
        return date_str.date() if hasattr(date_str, 'date') else date_str
    date_str = str(date_str).strip()
    if not date_str or date_str.lower() in ("none", "", "n/a"):
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d"):
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    return None


def format_date(d):
    """格式化日期"""
    if d is None:
        return "N/A"
    if isinstance(d, datetime):
        return d.strftime("%Y-%m-%d")
    return str(d)


def days_until(d):
    """計算距離今天的天數"""
    if d is None:
        return None
    today = datetime.now().date()
    if isinstance(d, datetime):
        d = d.date()
    delta = d - today
    return delta.days


# =============================================================================
#  Excel 樣式工具
# =============================================================================

# 預設樣式
STYLE_HEADER_FONT = Font(name="Microsoft JhengHei", bold=True, size=11, color="FFFFFF")
STYLE_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
STYLE_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
STYLE_CELL_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
STYLE_THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

# 影響程度顏色
IMPACT_FILLS = {
    "高": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "中": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "低": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
}

STATUS_FILLS = {
    "已生效": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "待生效": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "過渡中": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
    "已過期": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "有效":   PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "即將到期": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
}


def apply_header_style(ws, headers, row=1):
    """為標題列套用樣式"""
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = STYLE_HEADER_FONT
        cell.fill = STYLE_HEADER_FILL
        cell.alignment = STYLE_HEADER_ALIGN
        cell.border = STYLE_THIN_BORDER


def apply_cell_style(cell, wrap=True):
    """為儲存格套用基本樣式"""
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=wrap)
    cell.border = STYLE_THIN_BORDER


def auto_column_width(ws, min_width=10, max_width=45):
    """自動調整欄寬"""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                val = str(cell.value)
                # 中文字元算 2 個字寬
                length = sum(2 if ord(ch) > 127 else 1 for ch in val)
                max_len = max(max_len, length)
        adjusted = max(min_width, min(max_len + 4, max_width))
        ws.column_dimensions[col_letter].width = adjusted


# =============================================================================
#  類別一：PolicyDatabase - 政策資料庫管理
# =============================================================================

class PolicyDatabase:
    """政策資料庫管理"""

    # Sheet 名稱
    SHEET_POLICIES = "政策記錄"
    SHEET_BRANDS = "品牌授權追蹤"
    SHEET_LICENSES = "進口許可追蹤"

    # 各 Sheet 欄位定義
    POLICY_HEADERS = [
        "政策編號", "發布日期", "國家", "政策類別", "政策標題",
        "政策摘要", "影響HS章節", "影響貿易方向", "影響程度",
        "生效日期", "過渡期截止日", "應對措施", "狀態",
        "記錄人", "備註"
    ]

    BRAND_HEADERS = [
        "品牌名稱", "授權編號", "授權方", "授權開始日", "授權到期日",
        "涵蓋產品", "涵蓋國家", "狀態", "續辦進度", "負責人", "備註"
    ]

    LICENSE_HEADERS = [
        "許可證名稱", "許可證編號", "發證機關", "國家", "涵蓋產品",
        "許可開始日", "許可到期日", "狀態", "續辦進度", "負責人", "備註"
    ]

    def __init__(self, db_path=None):
        """載入/建立政策資料庫 Excel"""
        self.db_path = db_path or DEFAULT_DB_PATH
        self.wb = None
        self._ensure_db()

    def _ensure_db(self):
        """確保資料庫檔案存在，不存在則建立"""
        if os.path.exists(self.db_path):
            try:
                self.wb = load_workbook(self.db_path)
                ColorPrinter.success(f"已載入資料庫: {self.db_path}")
            except Exception as e:
                ColorPrinter.error(f"載入資料庫失敗: {e}")
                self._create_new_db()
        else:
            self._create_new_db()

    def _create_new_db(self):
        """建立全新的資料庫"""
        self.wb = Workbook()

        # --- 政策記錄 Sheet ---
        ws_pol = self.wb.active
        ws_pol.title = self.SHEET_POLICIES
        apply_header_style(ws_pol, self.POLICY_HEADERS)
        # 資料驗證
        dv_country = DataValidation(type="list", formula1='"泰國,印尼,中國"', allow_blank=True)
        dv_category = DataValidation(type="list", formula1='"關稅調整,法規變更,貿易協定,檢驗要求,許可證要求,其他"', allow_blank=True)
        dv_impact = DataValidation(type="list", formula1='"高,中,低"', allow_blank=True)
        dv_status = DataValidation(type="list", formula1='"已生效,待生效,過渡中,已過期"', allow_blank=True)
        dv_direction = DataValidation(type="list", formula1='"出口,進口,兩者"', allow_blank=True)
        for dv in [dv_country, dv_category, dv_impact, dv_status, dv_direction]:
            dv.showErrorMessage = True
            ws_pol.add_data_validation(dv)
        dv_country.add(f"C2:C1000")
        dv_category.add(f"D2:D1000")
        dv_impact.add(f"I2:I1000")
        dv_status.add(f"M2:M1000")
        dv_direction.add(f"H2:H1000")
        # 預填範例數據
        self._fill_sample_policies(ws_pol)

        # --- 品牌授權追蹤 Sheet ---
        ws_brand = self.wb.create_sheet(self.SHEET_BRANDS)
        apply_header_style(ws_brand, self.BRAND_HEADERS)
        dv_brand_status = DataValidation(type="list", formula1='"有效,即將到期,已過期"', allow_blank=True)
        ws_brand.add_data_validation(dv_brand_status)
        dv_brand_status.add("H2:H1000")
        self._fill_sample_brands(ws_brand)

        # --- 進口許可追蹤 Sheet ---
        ws_lic = self.wb.create_sheet(self.SHEET_LICENSES)
        apply_header_style(ws_lic, self.LICENSE_HEADERS)
        dv_lic_status = DataValidation(type="list", formula1='"有效,即將到期,已過期"', allow_blank=True)
        ws_lic.add_data_validation(dv_lic_status)
        dv_lic_status.add("H2:H1000")
        self._fill_sample_licenses(ws_lic)

        # 調整欄寬
        for ws in [ws_pol, ws_brand, ws_lic]:
            auto_column_width(ws)

        self.wb.save(self.db_path)
        ColorPrinter.success(f"已建立新資料庫: {self.db_path}")

    # -------------------------------------------------------------------------
    #  範例數據
    # -------------------------------------------------------------------------

    def _fill_sample_policies(self, ws):
        """預填 15 筆政策範例"""
        samples = [
            ["POL-2026-001", "2026-05-10", "泰國", "檢驗要求",
             "泰國電子產品附加安全認證要求",
             "泰國工業標準研究院(TISI)发布公告，要求所有進口電子產品須通過新版安全認證（TIS 2xxx-2569），包括EMC測試及能效標示。",
             "第85章", "進口", "高", "2026-07-01", "2026-12-31",
             "1.確認產品是否符合新認證標準\n2.聯繫TISI認可實驗室安排測試\n3.準備技術文件及樣品送檢", "待生效", "王小明", "影響所有電子零組件及成品出口泰國"],

            ["POL-2026-002", "2026-04-15", "印尼", "法規變更",
             "印尼SNI標籤新規",
             "印尼國家標準化機構(BSN)修訂SNI標籤規範，要求所有消費品外包裝須標示印尼語SNI標誌及QR追溯碼。",
             "第01章,第02章,第03章,第07章,第08章,第20章,第22章,第33章,第34章,第39章,第40章,第69章,第84章,第85章,第94章,第95章", "出口", "中", "2026-06-01", "2026-09-01",
             "1.更新產品包裝設計\n2.申請SNI標誌使用授權\n3.建立QR追溯碼系統", "過渡中", "李大華", "需於過渡期內完成包裝更新"],

            ["POL-2026-003", "2026-05-20", "泰國", "貿易協定",
             "RCEP原產地規則修訂",
             "RCEP聯合委員會通過原產地規則修訂案，放寬部分產品的區域價值成分(RVC)計算標準，並新增化學品章節的稅則分類改變(CTC)規則。",
             "第28章,第29章,第30章,第38章,第84章,第85章", "兩者", "高", "2026-08-01", None,
             "1.重新評估產品的RCEP原產地資格\n2.更新原產地證明文件\n3.利用放寬規則擴大關稅優惠適用範圍", "待生效", "張美玲", "可能影響現有關稅優惠安排"],

            ["POL-2026-004", "2026-03-25", "泰國", "關稅調整",
             "泰國BOI投資優惠政策更新",
             "泰國投資促進委員會(BOI)更新投資優惠措施，調整進口生產設備及原材料免稅額度，並新增智慧製造設備加速折舊優惠。",
             "第84章,第85章,第90章", "進口", "中", "2026-05-15", None,
             "1.檢視現有BOI優惠是否受影響\n2.評估是否申請新增優惠項目\n3.調整進口設備採購計畫", "已生效", "陳志偉", "有利於在泰國設廠的企業"],

            ["POL-2026-005", "2026-05-30", "印尼", "法規變更",
             "印尼禁止特定塑膠原料進口",
             "印尼貿易部發布禁令，禁止進口特定類型塑膠原料（HS 3901-3914項下部分稅號），以推動國內石化產業發展及環保目標。",
             "第39章", "出口", "高", "2026-09-01", "2027-03-01",
             "1.確認出口產品是否涉及禁止進口清單\n2.尋找印尼當地替代供應商\n3.評估轉口或加工區方案", "待生效", "王小明", "嚴重影響塑膠相關產品出口印尼"],

            ["POL-2026-006", "2026-04-01", "印尼", "關稅調整",
             "印尼調整鋼鐵產品進口關稅",
             "印尼財政部調高部分鋼鐵產品(HS 72-73章)進口關稅5-10%，以保護國內鋼鐵產業，同時對東盟自貿區夥伴實施差別稅率。",
             "第73章", "出口", "中", "2026-06-15", None,
             "1.評估關稅調高對產品定價的影響\n2.確認是否適用ACFTA優惠稅率\n3.考慮調整產品結構", "待生效", "李大華", "需重新計算出口成本"],

            ["POL-2026-007", "2026-02-20", "中國", "法規變更",
             "中國海關AEO認證標準升級",
             "中國海關總署發布新版AEO(經認證的經營者)認證標準，強化供應鏈安全要求及數位化合規能力評估。",
             "", "兩者", "高", "2026-06-01", "2027-06-01",
             "1.對照新標準進行自評\n2.補強供應鏈安全管理制度\n3.升級關務資訊系統", "過渡中", "張美玲", "需在過渡期內完成重新認證"],

            ["POL-2026-008", "2026-05-05", "泰國", "檢驗要求",
             "泰國FDA加強食品接觸材料檢驗",
             "泰國食品藥品監督管理局(FDA)加強對進口食品接觸材料(塑膠、陶瓷、玻璃等)的檢驗標準，新增重金屬遷移量及塑化劑檢測項目。",
             "第39章,第69章,第70章", "出口", "中", "2026-08-15", "2027-02-15",
             "1.安排產品送檢\n2.確認原材料合規性\n3.取得泰國FDA認可實驗室報告", "待生效", "陳志偉", "影響食品包裝及餐具類產品"],

            ["POL-2026-009", "2026-03-10", "印尼", "許可證要求",
             "印尼進口配額管理制度變更",
             "印尼貿易部修訂進口配額管理辦法，部分產品改為線上申請系統，並調整配額分配週期為季度制。",
             "第84章,第85章,第87章", "出口", "高", "2026-07-01", None,
             "1.註冊新的線上申請系統\n2.提前申請下季配額\n3.與印尼進口商確認配額安排", "待生效", "王小明", "影響電子產品及汽車零組件出口"],

            ["POL-2026-010", "2026-01-15", "中國", "貿易協定",
             "中國-東盟自貿區升級版實施細則",
             "中國-東盟自貿區升級版協議實施細則公布，進一步降低部分產品關稅並簡化原產地證明程序。",
             "第84章,第85章,第90章", "兩者", "中", "2026-03-01", None,
             "1.確認產品是否享有新增關稅優惠\n2.更新原產地證明格式\n3.優化供應鏈佈局以最大化優惠", "已生效", "張美玲", "有利於降低貿易成本"],

            ["POL-2026-011", "2026-04-20", "泰國", "法規變更",
             "泰國有害物質管理法規修訂",
             "泰國有害物質管控委員會修訂有害物質清單，新增多種化學物質為管控物質，影響含相關化學品的產品進口。",
             "第28章,第29章,第38章", "出口", "中", "2026-10-01", "2027-04-01",
             "1.排查產品是否含有新增管控物質\n2.準備替代材料方案\n3.申請有害物質進口許可(如適用)", "待生效", "李大華", "影響化工產品及部分電子零組件"],

            ["POL-2026-012", "2026-05-25", "印尼", "檢驗要求",
             "印尼SNI強制認證擴增產品範圍",
             "印尼工業部公告新增50項產品列入SNI強制認證範圍，包括部分家電、建材及兒童用品。",
             "第84章,第85章,第68章,第69章,第94章,第95章", "出口", "高", "2026-11-01", "2027-05-01",
             "1.確認產品是否在新認證範圍內\n2.聯繫印尼認證機構安排認證\n3.準備工廠審查文件", "待生效", "陳志偉", "需盡早啟動認證程序"],

            ["POL-2026-013", "2026-06-01", "中國", "法規變更",
             "中國出口退稅率調整",
             "財政部及稅務總局聯合公告調整部分產品出口退稅率，機電產品維持13%退稅率，部分資源性產品退稅率下調。",
             "第84章,第85章,第90章", "出口", "中", "2026-07-01", None,
             "1.確認產品出口退稅率變動\n2.調整出口定價策略\n3.優化出口產品結構", "待生效", "張美玲", "需關注後續細則公告"],

            ["POL-2026-014", "2026-03-15", "泰國", "關稅調整",
             "泰國暫定降低能源設備進口關稅",
             "泰國財政部公告暫定降低太陽能板、風力發電設備等綠色能源產品進口關稅至0-2%，有效期2年。",
             "第85章", "出口", "低", "2026-04-01", None,
             "1.評估綠色能源產品出口泰國的商機\n2.確認產品是否符合零關稅條件\n3.準備相關進口文件", "已生效", "王小明", "綠色能源產品出口良機"],

            ["POL-2026-015", "2026-06-05", "印尼", "貿易協定",
             "印尼-中國雙邊貿易便利化協議",
             "印尼與中國簽署貿易便利化協議，簡化海關清關程序、縮短檢驗時間，並建立電子數據交換系統。",
             "", "兩者", "中", "2026-09-01", None,
             "1.了解新電子通關系統操作流程\n2.註冊電子數據交換平台\n3.培訓相關人員使用新系統", "待生效", "李大華", "有望加速通關效率"],
        ]

        for row_idx, row_data in enumerate(samples, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                apply_cell_style(cell)
                # 影響程度着色
                if col_idx == 9 and value in IMPACT_FILLS:
                    cell.fill = IMPACT_FILLS[value]
                # 狀態着色
                if col_idx == 13 and value in STATUS_FILLS:
                    cell.fill = STATUS_FILLS[value]

    def _fill_sample_brands(self, ws):
        """預填 8 筆品牌授權範例"""
        samples = [
            ["BrandA", "AUTH-2024-001", "BrandA Inc. (USA)", "2024-01-01", "2026-12-31",
             "家電全系列", "泰國,印尼", "有效", "N/A", "林經理", ""],
            ["BrandB", "AUTH-2025-002", "BrandB Co. (Japan)", "2025-04-01", "2026-07-15",
             "廚房電器", "泰國", "有效", "N/A", "林經理", "即將到期需續約"],
            ["BrandC", "AUTH-2024-003", "BrandC Ltd. (UK)", "2024-06-01", "2026-06-20",
             "照明設備", "印尼", "有效", "N/A", "陳經理", "需盡快確認續約意向"],
            ["BrandD", "AUTH-2025-004", "BrandD GmbH (Germany)", "2025-01-01", "2027-12-31",
             "工業設備", "泰國,印尼", "有效", "N/A", "張經理", ""],
            ["BrandE", "AUTH-2023-005", "BrandE S.p.A. (Italy)", "2023-08-01", "2026-07-31",
             "家具系列", "泰國", "有效", "N/A", "林經理", ""],
            ["BrandF", "AUTH-2024-006", "BrandF Pty (Australia)", "2024-03-01", "2026-08-15",
             "運動用品", "印尼", "有效", "N/A", "陳經理", ""],
            ["BrandG", "AUTH-2023-007", "BrandG SA (France)", "2023-01-01", "2026-06-30",
             "美妝產品", "泰國,印尼", "有效", "已啟動續約洽談", "張經理", "續約中"],
            ["BrandH", "AUTH-2024-008", "BrandH Corp. (Korea)", "2024-09-01", "2027-08-31",
             "電子配件", "泰國,印尼", "有效", "N/A", "林經理", ""],
        ]

        for row_idx, row_data in enumerate(samples, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                apply_cell_style(cell)

    def _fill_sample_licenses(self, ws):
        """預填 5 筆進口許可範例"""
        samples = [
            ["泰國電子產品進口許可", "LIC-TH-2025-001", "泰國商業部", "泰國",
             "電子產品(HS 85章)", "2025-01-01", "2026-07-31", "有效", "N/A", "王專員", ""],
            ["印尼SNI認證許可", "LIC-ID-2024-002", "印尼國家標準局(BSN)", "印尼",
             "家電及建材", "2024-06-01", "2026-06-25", "有效", "已提交續辦申請", "李專員", "需加緊處理"],
            ["泰國FDA進口許可", "LIC-TH-2025-003", "泰國FDA", "泰國",
             "食品接觸材料", "2025-04-01", "2027-03-31", "有效", "N/A", "王專員", ""],
            ["印尼API進口執照", "LIC-ID-2024-004", "印尼貿易部", "印尼",
             "一般商品", "2024-01-01", "2026-12-31", "有效", "N/A", "陳專員", ""],
            ["印尼特殊化學品進口許可", "LIC-ID-2025-005", "印尼工業部", "印尼",
             "化學品(HS 28-29章)", "2025-07-01", "2026-07-10", "有效", "N/A", "李專員", "即將到期"],
        ]

        for row_idx, row_data in enumerate(samples, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                apply_cell_style(cell)

    # -------------------------------------------------------------------------
    #  資料操作
    # -------------------------------------------------------------------------

    def _get_sheet(self, sheet_name):
        """取得指定 Sheet，不存在則回傳 None"""
        if self.wb and sheet_name in self.wb.sheetnames:
            return self.wb[sheet_name]
        return None

    def _read_rows(self, ws):
        """讀取 Sheet 所有資料列 (不含標題列)，回傳 list of dict"""
        if ws is None:
            return []
        headers = [cell.value for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=False):
            vals = [cell.value for cell in row]
            if all(v is None for v in vals):
                continue
            rows.append(dict(zip(headers, vals)))
        return rows

    def save(self):
        """儲存資料庫"""
        try:
            self.wb.save(self.db_path)
        except Exception as e:
            ColorPrinter.error(f"儲存失敗: {e}")

    def add_policy(self, policy_data):
        """
        新增一筆政策記錄。
        policy_data: dict，鍵對應 POLICY_HEADERS
        """
        ws = self._get_sheet(self.SHEET_POLICIES)
        if ws is None:
            ColorPrinter.error(f"找不到 Sheet: {self.SHEET_POLICIES}")
            return False

        # 自動產生編號
        existing = self._read_rows(ws)
        next_num = len(existing) + 1
        pol_id = f"POL-{datetime.now().strftime('%Y')}-{next_num:03d}"

        # 確保所有欄位都有值
        row_data = []
        for header in self.POLICY_HEADERS:
            if header == "政策編號":
                row_data.append(pol_id)
            elif header == "發布日期":
                row_data.append(policy_data.get(header, datetime.now().strftime("%Y-%m-%d")))
            elif header == "狀態":
                row_data.append(policy_data.get(header, "待生效"))
            else:
                row_data.append(policy_data.get(header, ""))

        new_row = ws.max_row + 1
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=new_row, column=col_idx, value=value)
            apply_cell_style(cell)
            if col_idx == 9 and value in IMPACT_FILLS:
                cell.fill = IMPACT_FILLS[value]
            if col_idx == 13 and value in STATUS_FILLS:
                cell.fill = STATUS_FILLS[value]

        self.save()
        ColorPrinter.success(f"已新增政策: {pol_id} - {policy_data.get('政策標題', '')}")
        return pol_id

    def search(self, keywords=None, country=None, category=None, date_range=None):
        """
        搜尋政策。
        keywords: 關鍵字 (搜尋標題、摘要、應對措施)
        country: 國家篩選
        category: 類別篩選
        date_range: (start_date, end_date) 日期範圍
        """
        ws = self._get_sheet(self.SHEET_POLICIES)
        rows = self._read_rows(ws)

        results = []
        for row in rows:
            # 國家篩選
            if country and str(row.get("國家", "")).strip() != country:
                continue
            # 類別篩選
            if category and str(row.get("政策類別", "")).strip() != category:
                continue
            # 日期範圍篩選
            if date_range:
                start_d, end_d = date_range
                pub_date = parse_date(row.get("發布日期"))
                if pub_date:
                    if start_d and pub_date < start_d:
                        continue
                    if end_d and pub_date > end_d:
                        continue
            # 關鍵字搜尋
            if keywords:
                kw = keywords.lower()
                searchable = " ".join([
                    str(row.get("政策標題", "")),
                    str(row.get("政策摘要", "")),
                    str(row.get("應對措施", "")),
                    str(row.get("影響HS章節", "")),
                    str(row.get("備註", "")),
                ]).lower()
                if kw not in searchable:
                    continue
            results.append(row)

        return results

    def get_active_policies(self, country=None):
        """獲取當前有效的政策列表 (狀態為已生效或過渡中)"""
        ws = self._get_sheet(self.SHEET_POLICIES)
        rows = self._read_rows(ws)

        active = []
        for row in rows:
            status = str(row.get("狀態", "")).strip()
            if status not in ("已生效", "過渡中"):
                continue
            if country and str(row.get("國家", "")).strip() != country:
                continue
            active.append(row)

        return active

    def get_impact_alerts(self):
        """獲取需要注意的政策變動 (高影響程度且待生效/過渡中)"""
        ws = self._get_sheet(self.SHEET_POLICIES)
        rows = self._read_rows(ws)

        alerts = []
        for row in rows:
            impact = str(row.get("影響程度", "")).strip()
            status = str(row.get("狀態", "")).strip()
            if impact == "高" and status in ("待生效", "過渡中"):
                effective = parse_date(row.get("生效日期"))
                d = days_until(effective)
                row["_days_until"] = d
                alerts.append(row)

        # 按生效日期排序 (越急迫越前面)
        alerts.sort(key=lambda x: x.get("_days_until") if x.get("_days_until") is not None else 9999)
        return alerts

    def get_all_policies(self):
        """獲取所有政策記錄"""
        ws = self._get_sheet(self.SHEET_POLICIES)
        return self._read_rows(ws)

    def get_all_brands(self):
        """獲取所有品牌授權"""
        ws = self._get_sheet(self.SHEET_BRANDS)
        return self._read_rows(ws)

    def get_all_licenses(self):
        """獲取所有進口許可"""
        ws = self._get_sheet(self.SHEET_LICENSES)
        return self._read_rows(ws)


# =============================================================================
#  類別二：ImpactAnalyzer - 政策影響分析
# =============================================================================

class ImpactAnalyzer:
    """政策影響分析"""

    def __init__(self, db):
        """
        db: PolicyDatabase 實例
        """
        self.db = db

    def analyze_impact(self, policy, product_db=None):
        """
        分析一筆政策變動對公司業務的影響。
        回傳 dict 包含:
          - affected_products: 受影響的產品類別
          - trade_direction: 影響的貿易方向
          - impact_level: 影響程度
          - recommendations: 建議的應對措施
          - effective_date: 生效日期
          - transition_deadline: 過渡期截止日
          - urgency_days: 距生效/過渡截止的天數
        """
        result = {
            "policy_id": policy.get("政策編號", ""),
            "title": policy.get("政策標題", ""),
            "country": policy.get("國家", ""),
            "category": policy.get("政策類別", ""),
            "affected_products": [],
            "trade_direction": policy.get("影響貿易方向", ""),
            "impact_level": policy.get("影響程度", "中"),
            "recommendations": policy.get("應對措施", ""),
            "effective_date": parse_date(policy.get("生效日期")),
            "transition_deadline": parse_date(policy.get("過渡期截止日")),
            "urgency_days": None,
            "hs_chapters": [],
        }

        # 解析 HS 章節
        hs_str = str(policy.get("影響HS章節", ""))
        if hs_str:
            chapters = [ch.strip() for ch in hs_str.split(",") if ch.strip()]
            result["hs_chapters"] = chapters
            for ch in chapters:
                desc = HS_CHAPTER_DESC.get(ch, "")
                if desc:
                    result["affected_products"].append(f"{ch} - {desc}")
                else:
                    result["affected_products"].append(ch)

        # 計算緊迫程度
        deadline = result["transition_deadline"] or result["effective_date"]
        if deadline:
            result["urgency_days"] = days_until(deadline)

        return result

    def generate_monthly_report(self, month_str, output_path=None):
        """
        生成月度政策變動報告。
        month_str: 'YYYY-MM' 格式
        output_path: 輸出 Excel 路徑
        """
        # 解析月份
        try:
            year, month = month_str.split("-")
            year, month = int(year), int(month)
            month_start = datetime(year, month, 1).date()
            if month == 12:
                month_end = datetime(year + 1, 1, 1).date() - timedelta(days=1)
            else:
                month_end = datetime(year, month + 1, 1).date() - timedelta(days=1)
        except Exception:
            ColorPrinter.error(f"月份格式錯誤: {month_str}，應為 YYYY-MM")
            return None

        ColorPrinter.section(f"生成 {month_str} 月度報告")

        # 搜尋本月發布的政策
        policies = self.db.search(date_range=(month_start, month_end))

        # 統計
        total = len(policies)
        by_country = {}
        by_impact = {"高": 0, "中": 0, "低": 0}
        action_items = []

        for p in policies:
            country = str(p.get("國家", "未知"))
            by_country[country] = by_country.get(country, 0) + 1
            impact = str(p.get("影響程度", "中"))
            by_impact[impact] = by_impact.get(impact, 0) + 1
            if impact == "高":
                action_items.append(p)

        # 所有待生效的高影響政策
        alerts = self.db.get_impact_alerts()

        # 建立輸出 Excel
        if output_path is None:
            os.makedirs(OUTPUT_DIR, exist_ok=True)
            output_path = os.path.join(OUTPUT_DIR, f"月度報告_{month_str}.xlsx")

        wb = Workbook()

        # --- Sheet 1: 報告摘要 ---
        ws_summary = wb.active
        ws_summary.title = "報告摘要"

        # 標題
        ws_summary.merge_cells("A1:F1")
        title_cell = ws_summary["A1"]
        title_cell.value = f"海關政策變動月度報告 - {month_str}"
        title_cell.font = Font(name="Microsoft JhengHei", bold=True, size=16, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_summary.row_dimensions[1].height = 40

        # 統計資訊
        ws_summary["A3"] = "統計項目"
        ws_summary["B3"] = "數值"
        ws_summary["A3"].font = Font(bold=True)
        ws_summary["B3"].font = Font(bold=True)

        stats = [
            ("本月新增/修改政策總數", total),
            ("高影響政策數", by_impact.get("高", 0)),
            ("中影響政策數", by_impact.get("中", 0)),
            ("低影響政策數", by_impact.get("低", 0)),
        ]
        for country in COUNTRIES:
            stats.append((f"{country}政策數", by_country.get(country, 0)))

        for i, (label, val) in enumerate(stats, start=4):
            ws_summary.cell(row=i, column=1, value=label).border = STYLE_THIN_BORDER
            c = ws_summary.cell(row=i, column=2, value=val)
            c.border = STYLE_THIN_BORDER
            c.alignment = Alignment(horizontal="center")

        # 報告日期
        row_n = 4 + len(stats) + 2
        ws_summary.cell(row=row_n, column=1, value="報告生成日期").font = Font(bold=True)
        ws_summary.cell(row=row_n, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M"))

        # --- Sheet 2: 本月政策明細 ---
        ws_detail = wb.create_sheet("本月政策明細")
        detail_headers = [
            "政策編號", "發布日期", "國家", "政策類別", "政策標題",
            "影響程度", "生效日期", "狀態"
        ]
        apply_header_style(ws_detail, detail_headers)

        for row_idx, p in enumerate(policies, start=2):
            for col_idx, h in enumerate(detail_headers, start=1):
                val = p.get(h, "")
                if h in ("發布日期", "生效日期") and val:
                    val = format_date(parse_date(val))
                cell = ws_detail.cell(row=row_idx, column=col_idx, value=val)
                apply_cell_style(cell)
                if h == "影響程度" and val in IMPACT_FILLS:
                    cell.fill = IMPACT_FILLS[val]
                if h == "狀態" and val in STATUS_FILLS:
                    cell.fill = STATUS_FILLS[val]

        # --- Sheet 3: 需立即行動事項 ---
        ws_action = wb.create_sheet("需立即行動事項")
        action_headers = [
            "政策編號", "國家", "政策標題", "影響程度",
            "生效日期", "剩餘天數", "應對措施"
        ]
        apply_header_style(ws_action, action_headers)

        for row_idx, p in enumerate(alerts, start=2):
            for col_idx, h in enumerate(action_headers, start=1):
                if h == "剩餘天數":
                    effective = parse_date(p.get("生效日期"))
                    val = days_until(effective) if effective else "N/A"
                    if isinstance(val, int):
                        val = f"{val} 天"
                else:
                    val = p.get(h, "")
                    if h in ("生效日期",) and val:
                        val = format_date(parse_date(val))
                cell = ws_action.cell(row=row_idx, column=col_idx, value=val)
                apply_cell_style(cell)
                if h == "影響程度" and val in IMPACT_FILLS:
                    cell.fill = IMPACT_FILLS[val]

        # 調整欄寬
        for ws in [ws_summary, ws_detail, ws_action]:
            auto_column_width(ws)

        wb.save(output_path)
        ColorPrinter.success(f"月度報告已生成: {output_path}")

        # 終端顯示摘要
        ColorPrinter.stat("本月政策總數", total, ColorPrinter.CYAN)
        ColorPrinter.stat("高影響", by_impact.get("高", 0), ColorPrinter.RED)
        ColorPrinter.stat("中影響", by_impact.get("中", 0), ColorPrinter.YELLOW)
        ColorPrinter.stat("低影響", by_impact.get("低", 0), ColorPrinter.GREEN)
        for country in COUNTRIES:
            ColorPrinter.stat(f"{country}", by_country.get(country, 0))
        ColorPrinter.stat("待處理高影響項目", len(alerts), ColorPrinter.RED)

        return output_path


# =============================================================================
#  類別三：ExpiryManager - 到期預警管理
# =============================================================================

class ExpiryManager:
    """到期預警管理"""

    def __init__(self, db):
        """
        db: PolicyDatabase 實例
        """
        self.db = db

    def check_brand_authorizations(self, warning_days=30):
        """
        檢查品牌授權到期。
        回傳 list of dict，包含到期資訊。
        """
        brands = self.db.get_all_brands()
        results = []

        for b in brands:
            expiry = parse_date(b.get("授權到期日"))
            if expiry is None:
                continue
            d = days_until(expiry)
            status = "正常"
            if d is not None:
                if d < 0:
                    status = "已過期"
                elif d <= 7:
                    status = "緊急 (7天內)"
                elif d <= 15:
                    status = "警告 (15天內)"
                elif d <= warning_days:
                    status = "注意 (30天內)"
                else:
                    status = "正常"

            results.append({
                "brand": b.get("品牌名稱", ""),
                "auth_id": b.get("授權編號", ""),
                "licensor": b.get("授權方", ""),
                "products": b.get("涵蓋產品", ""),
                "countries": b.get("涵蓋國家", ""),
                "expiry_date": expiry,
                "days_remaining": d,
                "alert_status": status,
                "renewal_progress": b.get("續辦進度", ""),
                "responsible": b.get("負責人", ""),
            })

        results.sort(key=lambda x: x["days_remaining"] if x["days_remaining"] is not None else 9999)
        return results

    def check_import_licenses(self, warning_days=30):
        """
        檢查進口許可證到期。
        回傳 list of dict。
        """
        licenses = self.db.get_all_licenses()
        results = []

        for lic in licenses:
            expiry = parse_date(lic.get("許可到期日"))
            if expiry is None:
                continue
            d = days_until(expiry)
            status = "正常"
            if d is not None:
                if d < 0:
                    status = "已過期"
                elif d <= 7:
                    status = "緊急 (7天內)"
                elif d <= 15:
                    status = "警告 (15天內)"
                elif d <= warning_days:
                    status = "注意 (30天內)"
                else:
                    status = "正常"

            results.append({
                "license_name": lic.get("許可證名稱", ""),
                "license_id": lic.get("許可證編號", ""),
                "authority": lic.get("發證機關", ""),
                "country": lic.get("國家", ""),
                "products": lic.get("涵蓋產品", ""),
                "expiry_date": expiry,
                "days_remaining": d,
                "alert_status": status,
                "renewal_progress": lic.get("續辦進度", ""),
                "responsible": lic.get("負責人", ""),
            })

        results.sort(key=lambda x: x["days_remaining"] if x["days_remaining"] is not None else 9999)
        return results

    def check_customs_qualifications(self, warning_days=30):
        """
        檢查海關資質到期（AEO等）。
        從政策資料庫中篩選許可證要求類別的政策。
        """
        policies = self.db.search(category="許可證要求")
        results = []

        for p in policies:
            deadline = parse_date(p.get("過渡期截止日")) or parse_date(p.get("生效日期"))
            if deadline is None:
                continue
            d = days_until(deadline)
            status = "正常"
            if d is not None:
                if d < 0:
                    status = "已過期"
                elif d <= 7:
                    status = "緊急 (7天內)"
                elif d <= 15:
                    status = "警告 (15天內)"
                elif d <= warning_days:
                    status = "注意 (30天內)"
                else:
                    status = "正常"

            results.append({
                "policy_id": p.get("政策編號", ""),
                "title": p.get("政策標題", ""),
                "country": p.get("國家", ""),
                "expiry_date": deadline,
                "days_remaining": d,
                "alert_status": status,
                "responsible": p.get("記錄人", ""),
            })

        results.sort(key=lambda x: x["days_remaining"] if x["days_remaining"] is not None else 9999)
        return results

    def generate_expiry_calendar(self, output_path=None, warning_days=30):
        """
        生成到期日曆 Excel，標注 30/15/7 天預警。
        """
        if output_path is None:
            os.makedirs(OUTPUT_DIR, exist_ok=True)
            output_path = os.path.join(
                OUTPUT_DIR,
                f"到期預警_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            )

        ColorPrinter.section("到期預警檢查")

        brands = self.check_brand_authorizations(warning_days)
        licenses = self.check_import_licenses(warning_days)
        customs = self.check_customs_qualifications(warning_days)

        wb = Workbook()

        # --- 顏色定義 ---
        fill_urgent = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
        fill_warn = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        fill_caution = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        fill_ok = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fill_expired = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        font_white = Font(color="FFFFFF", bold=True)

        def get_alert_fill(item):
            status = item.get("alert_status", "")
            if "已過期" in status:
                return fill_expired, font_white
            elif "緊急" in status:
                return fill_urgent, font_white
            elif "警告" in status:
                return fill_warn, None
            elif "注意" in status:
                return fill_caution, None
            return fill_ok, None

        # --- Sheet 1: 總覽 ---
        ws_overview = wb.active
        ws_overview.title = "到期預警總覽"

        ws_overview.merge_cells("A1:G1")
        title_cell = ws_overview["A1"]
        title_cell.value = f"到期預警總覽 (截至 {datetime.now().strftime('%Y-%m-%d')})"
        title_cell.font = Font(name="Microsoft JhengHei", bold=True, size=16, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_overview.row_dimensions[1].height = 40

        # 圖例
        ws_overview["A3"] = "預警等級說明："
        ws_overview["A3"].font = Font(bold=True)
        legends = [
            ("已過期", fill_expired, font_white),
            ("緊急 (7天內)", fill_urgent, font_white),
            ("警告 (15天內)", fill_warn, None),
            ("注意 (30天內)", fill_caution, None),
            ("正常", fill_ok, None),
        ]
        for i, (label, fill, font) in enumerate(legends, start=4):
            cell = ws_overview.cell(row=i, column=1, value=label)
            cell.fill = fill
            if font:
                cell.font = font
            cell.border = STYLE_THIN_BORDER

        # 統計
        all_items = brands + licenses + customs
        expired_count = sum(1 for x in all_items if "已過期" in x.get("alert_status", ""))
        urgent_count = sum(1 for x in all_items if "緊急" in x.get("alert_status", ""))
        warn_count = sum(1 for x in all_items if "警告" in x.get("alert_status", ""))
        caution_count = sum(1 for x in all_items if "注意" in x.get("alert_status", ""))

        stat_row = 4 + len(legends) + 1
        stats = [
            ("總計項目", len(all_items)),
            ("已過期", expired_count),
            ("緊急", urgent_count),
            ("警告", warn_count),
            ("注意", caution_count),
        ]
        for i, (label, val) in enumerate(stats, start=stat_row):
            ws_overview.cell(row=i, column=1, value=label).font = Font(bold=True)
            ws_overview.cell(row=i, column=2, value=val).alignment = Alignment(horizontal="center")

        # --- Sheet 2: 品牌授權到期 ---
        ws_brand = wb.create_sheet("品牌授權到期")
        brand_headers = [
            "品牌名稱", "授權編號", "授權方", "涵蓋產品", "涵蓋國家",
            "到期日", "剩餘天數", "預警狀態", "續辦進度", "負責人"
        ]
        apply_header_style(ws_brand, brand_headers)
        for row_idx, item in enumerate(brands, start=2):
            values = [
                item["brand"], item["auth_id"], item["licensor"],
                item["products"], item["countries"],
                format_date(item["expiry_date"]),
                f'{item["days_remaining"]} 天' if item["days_remaining"] is not None else "N/A",
                item["alert_status"], item["renewal_progress"], item["responsible"]
            ]
            fill, font = get_alert_fill(item)
            for col_idx, val in enumerate(values, start=1):
                cell = ws_brand.cell(row=row_idx, column=col_idx, value=val)
                apply_cell_style(cell)
                # 預警狀態欄著色
                if col_idx == 8:
                    cell.fill = fill
                    if font:
                        cell.font = font

        # --- Sheet 3: 進口許可到期 ---
        ws_lic = wb.create_sheet("進口許可到期")
        lic_headers = [
            "許可證名稱", "許可證編號", "發證機關", "國家", "涵蓋產品",
            "到期日", "剩餘天數", "預警狀態", "續辦進度", "負責人"
        ]
        apply_header_style(ws_lic, lic_headers)
        for row_idx, item in enumerate(licenses, start=2):
            values = [
                item["license_name"], item["license_id"], item["authority"],
                item["country"], item["products"],
                format_date(item["expiry_date"]),
                f'{item["days_remaining"]} 天' if item["days_remaining"] is not None else "N/A",
                item["alert_status"], item["renewal_progress"], item["responsible"]
            ]
            fill, font = get_alert_fill(item)
            for col_idx, val in enumerate(values, start=1):
                cell = ws_lic.cell(row=row_idx, column=col_idx, value=val)
                apply_cell_style(cell)
                if col_idx == 8:
                    cell.fill = fill
                    if font:
                        cell.font = font

        # --- Sheet 4: 海關資質到期 ---
        ws_customs = wb.create_sheet("海關資質到期")
        customs_headers = [
            "政策編號", "政策標題", "國家",
            "到期日", "剩餘天數", "預警狀態", "負責人"
        ]
        apply_header_style(ws_customs, customs_headers)
        for row_idx, item in enumerate(customs, start=2):
            values = [
                item["policy_id"], item["title"], item["country"],
                format_date(item["expiry_date"]),
                f'{item["days_remaining"]} 天' if item["days_remaining"] is not None else "N/A",
                item["alert_status"], item["responsible"]
            ]
            fill, font = get_alert_fill(item)
            for col_idx, val in enumerate(values, start=1):
                cell = ws_customs.cell(row=row_idx, column=col_idx, value=val)
                apply_cell_style(cell)
                if col_idx == 6:
                    cell.fill = fill
                    if font:
                        cell.font = font

        # 調整欄寬
        for ws in [ws_overview, ws_brand, ws_lic, ws_customs]:
            auto_column_width(ws)

        wb.save(output_path)
        ColorPrinter.success(f"到期預警日曆已生成: {output_path}")

        # 終端輸出摘要
        self._print_expiry_summary(brands, licenses, customs, warning_days)

        return output_path

    def _print_expiry_summary(self, brands, licenses, customs, warning_days):
        """終端顯示到期預警摘要"""
        all_items = brands + licenses + customs
        within_warning = [x for x in all_items if x.get("days_remaining") is not None and x["days_remaining"] <= warning_days]

        if not within_warning:
            ColorPrinter.success(f"未來 {warning_days} 天內無到期項目")
            return

        ColorPrinter.warning(f"未來 {warning_days} 天內有 {len(within_warning)} 項即將到期：")
        print()

        widths = [22, 12, 10, 10, 14]
        ColorPrinter.table_row(["項目", "到期日", "剩餘天數", "預警", "負責人"], widths, bold=True)
        ColorPrinter.table_sep(widths)

        for item in within_warning:
            name = item.get("brand") or item.get("license_name") or item.get("title", "")
            if len(name) > 20:
                name = name[:18] + ".."
            expiry_str = format_date(item.get("expiry_date"))
            days_str = f'{item["days_remaining"]}天' if item["days_remaining"] is not None else "N/A"
            status = item.get("alert_status", "")
            responsible = item.get("responsible", "")

            color = None
            if "已過期" in status:
                color = ColorPrinter.RED
            elif "緊急" in status:
                color = ColorPrinter.RED
            elif "警告" in status:
                color = ColorPrinter.YELLOW
            elif "注意" in status:
                color = ColorPrinter.CYAN

            ColorPrinter.table_row([name, expiry_str, days_str, status, responsible], widths, color=color)

        print()


# =============================================================================
#  命令列介面 (CLI)
# =============================================================================

def cmd_overview(db):
    """顯示政策總覽"""
    ColorPrinter.header("海關政策變動追蹤系統 - 政策總覽")

    policies = db.get_all_policies()
    brands = db.get_all_brands()
    licenses = db.get_all_licenses()

    if not policies:
        ColorPrinter.warning("資料庫中尚無政策記錄")
        return

    # --- 基本統計 ---
    ColorPrinter.section("基本統計")
    total = len(policies)
    ColorPrinter.stat("政策總數", total, ColorPrinter.CYAN)

    # 按國家統計
    by_country = {}
    by_impact = {"高": 0, "中": 0, "低": 0}
    by_status = {}
    by_category = {}

    for p in policies:
        country = str(p.get("國家", "未知"))
        by_country[country] = by_country.get(country, 0) + 1

        impact = str(p.get("影響程度", "中"))
        by_impact[impact] = by_impact.get(impact, 0) + 1

        status = str(p.get("狀態", "未知"))
        by_status[status] = by_status.get(status, 0) + 1

        cat = str(p.get("政策類別", "其他"))
        by_category[cat] = by_category.get(cat, 0) + 1

    ColorPrinter.section("按國家分類")
    for country in COUNTRIES:
        count = by_country.get(country, 0)
        bar = "#" * count
        ColorPrinter.stat(country, f"{count:>2}  {ColorPrinter.BLUE}{bar}{ColorPrinter.RESET}")

    ColorPrinter.section("按影響程度分類")
    colors = {"高": ColorPrinter.RED, "中": ColorPrinter.YELLOW, "低": ColorPrinter.GREEN}
    for level in IMPACT_LEVELS:
        count = by_impact.get(level, 0)
        bar = "#" * count
        ColorPrinter.stat(level, f"{count:>2}  {colors.get(level, '')}{bar}{ColorPrinter.RESET}")

    ColorPrinter.section("按狀態分類")
    for status, count in sorted(by_status.items()):
        ColorPrinter.stat(status, count)

    ColorPrinter.section("按類別分類")
    for cat, count in sorted(by_category.items(), key=lambda x: -x[1]):
        ColorPrinter.stat(cat, count)

    # --- 高影響政策預警 ---
    analyzer = ImpactAnalyzer(db)
    alerts = db.get_impact_alerts()
    if alerts:
        ColorPrinter.section("高影響待處理政策 (需優先關注)")
        widths = [14, 8, 30, 10, 10, 10]
        ColorPrinter.table_row(
            ["編號", "國家", "標題", "影響", "生效日", "剩餘天數"],
            widths, bold=True
        )
        ColorPrinter.table_sep(widths)
        for a in alerts[:8]:
            title = str(a.get("政策標題", ""))
            if len(title) > 28:
                title = title[:26] + ".."
            eff = format_date(parse_date(a.get("生效日期")))
            d = a.get("_days_until")
            d_str = f"{d}天" if d is not None else "N/A"
            color = ColorPrinter.RED if (d is not None and d <= 30) else ColorPrinter.YELLOW
            ColorPrinter.table_row(
                [a.get("政策編號", ""), a.get("國家", ""), title,
                 a.get("影響程度", ""), eff, d_str],
                widths, color=color
            )

    # --- 最近 5 筆政策 ---
    ColorPrinter.section("最近發布的政策")
    sorted_policies = sorted(
        policies,
        key=lambda x: parse_date(x.get("發布日期")) or datetime(2000, 1, 1).date(),
        reverse=True
    )
    widths2 = [14, 12, 8, 10, 30, 8]
    ColorPrinter.table_row(
        ["編號", "發布日期", "國家", "類別", "標題", "影響"],
        widths2, bold=True
    )
    ColorPrinter.table_sep(widths2)
    for p in sorted_policies[:5]:
        title = str(p.get("政策標題", ""))
        if len(title) > 28:
            title = title[:26] + ".."
        pub = format_date(parse_date(p.get("發布日期")))
        ColorPrinter.table_row(
            [p.get("政策編號", ""), pub, p.get("國家", ""),
             p.get("政策類別", ""), title, p.get("影響程度", "")],
            widths2
        )

    # --- 品牌授權 & 進口許可摘要 ---
    ColorPrinter.section("授權/許可概覽")
    ColorPrinter.stat("品牌授權數量", len(brands))
    ColorPrinter.stat("進口許可數量", len(licenses))

    exp_mgr = ExpiryManager(db)
    brand_warnings = [b for b in exp_mgr.check_brand_authorizations()
                      if b["days_remaining"] is not None and b["days_remaining"] <= 30]
    lic_warnings = [l for l in exp_mgr.check_import_licenses()
                    if l["days_remaining"] is not None and l["days_remaining"] <= 30]

    if brand_warnings:
        ColorPrinter.warning(f"30天內到期品牌授權: {len(brand_warnings)} 項")
    else:
        ColorPrinter.success("30天內無品牌授權到期")

    if lic_warnings:
        ColorPrinter.warning(f"30天內到期進口許可: {len(lic_warnings)} 項")
    else:
        ColorPrinter.success("30天內無進口許可到期")

    print()
    ColorPrinter.info(f"資料庫路徑: {db.db_path}")
    print()


def cmd_search(db, args):
    """搜尋政策"""
    ColorPrinter.header("政策搜尋")

    results = db.search(
        keywords=args.search,
        country=args.country,
        category=args.category,
    )

    if not results:
        ColorPrinter.warning("找不到符合條件的政策")
        return

    ColorPrinter.info(f"找到 {len(results)} 筆符合結果")
    print()

    widths = [14, 12, 8, 10, 32, 8, 8]
    ColorPrinter.table_row(
        ["編號", "發布日期", "國家", "類別", "標題", "影響", "狀態"],
        widths, bold=True
    )
    ColorPrinter.table_sep(widths)

    for p in results:
        title = str(p.get("政策標題", ""))
        if len(title) > 30:
            title = title[:28] + ".."
        pub = format_date(parse_date(p.get("發布日期")))

        impact = str(p.get("影響程度", ""))
        color = None
        if impact == "高":
            color = ColorPrinter.RED
        elif impact == "中":
            color = ColorPrinter.YELLOW

        ColorPrinter.table_row(
            [p.get("政策編號", ""), pub, p.get("國家", ""),
             p.get("政策類別", ""), title,
             impact, p.get("狀態", "")],
            widths, color=color
        )

    # 詳細資訊
    print()
    for p in results:
        ColorPrinter.section(f'{p.get("政策編號", "")} - {p.get("政策標題", "")}')
        print(f"    {ColorPrinter.DIM}國家:{ColorPrinter.RESET} {p.get('國家', '')}")
        print(f"    {ColorPrinter.DIM}類別:{ColorPrinter.RESET} {p.get('政策類別', '')}")
        print(f"    {ColorPrinter.DIM}發布日期:{ColorPrinter.RESET} {format_date(parse_date(p.get('發布日期')))}")
        print(f"    {ColorPrinter.DIM}生效日期:{ColorPrinter.RESET} {format_date(parse_date(p.get('生效日期')))}")
        print(f"    {ColorPrinter.DIM}影響程度:{ColorPrinter.RESET} {p.get('影響程度', '')}")
        print(f"    {ColorPrinter.DIM}貿易方向:{ColorPrinter.RESET} {p.get('影響貿易方向', '')}")
        print(f"    {ColorPrinter.DIM}影響HS章節:{ColorPrinter.RESET} {p.get('影響HS章節', '')}")
        print(f"    {ColorPrinter.DIM}狀態:{ColorPrinter.RESET} {p.get('狀態', '')}")
        print(f"    {ColorPrinter.DIM}摘要:{ColorPrinter.RESET}")
        summary = str(p.get("政策摘要", ""))
        # 自動換行
        line_width = 60
        for i in range(0, len(summary), line_width):
            print(f"      {summary[i:i+line_width]}")
        print(f"    {ColorPrinter.DIM}應對措施:{ColorPrinter.RESET}")
        measures = str(p.get("應對措施", ""))
        for line in measures.split("\\n"):
            line = line.strip()
            if line:
                print(f"      {line}")
        if p.get("備註"):
            print(f"    {ColorPrinter.DIM}備註:{ColorPrinter.RESET} {p.get('備註', '')}")
        print()


def cmd_add(db, args):
    """新增政策"""
    ColorPrinter.header("新增政策記錄")

    policy_data = {
        "政策標題": args.title,
        "國家": args.country or "",
        "政策類別": args.category or "",
        "發布日期": args.date or datetime.now().strftime("%Y-%m-%d"),
        "政策摘要": args.summary or "",
        "影響HS章節": args.hs or "",
        "影響貿易方向": args.direction or "",
        "影響程度": args.impact or "中",
        "生效日期": args.effective or "",
        "過渡期截止日": args.transition or "",
        "應對措施": args.measures or "",
        "狀態": args.status or "待生效",
        "記錄人": args.recorder or "",
        "備註": args.note or "",
    }

    # 驗證
    if not args.title:
        ColorPrinter.error("請提供政策標題 (--title)")
        return

    if args.country and args.country not in COUNTRIES:
        ColorPrinter.warning(f"國家 '{args.country}' 不在預設清單中: {COUNTRIES}")

    pol_id = db.add_policy(policy_data)
    if pol_id:
        ColorPrinter.success(f"成功新增政策 {pol_id}")
    else:
        ColorPrinter.error("新增失敗")


def cmd_expiry_check(db, args):
    """執行到期檢查"""
    warning_days = args.days if hasattr(args, 'days') and args.days else 30
    mgr = ExpiryManager(db)
    output_path = mgr.generate_expiry_calendar(warning_days=warning_days)
    return output_path


def cmd_monthly_report(db, args):
    """生成月度報告"""
    month = args.monthly_report
    analyzer = ImpactAnalyzer(db)
    output_path = analyzer.generate_monthly_report(month)
    return output_path


# =============================================================================
#  主程式
# =============================================================================

def build_parser():
    """建立命令列參數解析器"""
    parser = argparse.ArgumentParser(
        description="海關政策變動自動追蹤工具",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
範例:
  python3 政策追蹤工具.py --overview
  python3 政策追蹤工具.py --search "電子產品" --country 泰國
  python3 政策追蹤工具.py --expiry-check --days 30
  python3 政策追蹤工具.py --monthly-report 2026-06
  python3 政策追蹤工具.py --add --title "新政策" --country 泰國 --category 關稅調整
        """
    )

    parser.add_argument("--db", type=str, default=None,
                        help="資料庫路徑 (預設: 政策資料庫.xlsx)")

    # 模式選擇
    group = parser.add_mutually_exclusive_group()
    group.add_argument("--overview", action="store_true",
                       help="顯示政策總覽")
    group.add_argument("--search", type=str, default=None,
                       help="搜尋關鍵字")
    group.add_argument("--expiry-check", action="store_true",
                       help="執行到期預警檢查")
    group.add_argument("--monthly-report", type=str, default=None,
                       help="生成月度報告 (格式: YYYY-MM)")
    group.add_argument("--add", action="store_true",
                       help="新增政策記錄")

    # 搜尋/新增共用參數
    parser.add_argument("--country", type=str, default=None,
                        help="國家 (泰國/印尼/中國)")
    parser.add_argument("--category", type=str, default=None,
                        help="政策類別")

    # 新增專用參數
    parser.add_argument("--title", type=str, default=None,
                        help="政策標題")
    parser.add_argument("--date", type=str, default=None,
                        help="發布日期 (YYYY-MM-DD)")
    parser.add_argument("--summary", type=str, default=None,
                        help="政策摘要")
    parser.add_argument("--hs", type=str, default=None,
                        help="影響HS章節")
    parser.add_argument("--direction", type=str, default=None,
                        help="影響貿易方向 (出口/進口/兩者)")
    parser.add_argument("--impact", type=str, default=None,
                        help="影響程度 (高/中/低)")
    parser.add_argument("--effective", type=str, default=None,
                        help="生效日期 (YYYY-MM-DD)")
    parser.add_argument("--transition", type=str, default=None,
                        help="過渡期截止日 (YYYY-MM-DD)")
    parser.add_argument("--measures", type=str, default=None,
                        help="應對措施")
    parser.add_argument("--status", type=str, default=None,
                        help="狀態")
    parser.add_argument("--recorder", type=str, default=None,
                        help="記錄人")
    parser.add_argument("--note", type=str, default=None,
                        help="備註")

    # 到期檢查參數
    parser.add_argument("--days", type=int, default=30,
                        help="預警天數 (預設: 30)")

    return parser


def main():
    """主程式入口"""
    parser = build_parser()
    args = parser.parse_args()

    # 若未提供任何參數，顯示說明
    if len(sys.argv) == 1:
        parser.print_help()
        print()
        ColorPrinter.info("提示: 使用 --overview 查看政策總覽")
        return

    # 確保輸出目錄存在
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # 初始化資料庫
    db_path = args.db or DEFAULT_DB_PATH
    db = PolicyDatabase(db_path)

    # 執行對應命令
    if args.overview:
        cmd_overview(db)

    elif args.search:
        cmd_search(db, args)

    elif args.add:
        cmd_add(db, args)

    elif args.expiry_check:
        cmd_expiry_check(db, args)

    elif args.monthly_report:
        cmd_monthly_report(db, args)

    else:
        parser.print_help()


if __name__ == "__main__":
    main()
