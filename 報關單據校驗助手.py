#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
==========================================================================
  報關模板智能填充與一致性校驗工具
  Customs Declaration Smart Fill & Cross-Document Validation Tool
==========================================================================

  功能說明 / Features:
    1. 智能填充 (Smart Fill)   -- 自動從產品主數據庫補全訂單缺漏欄位
    2. 跨單據一致性校驗         -- 對全套6 Sheet報關單據進行20項交叉比對
    3. 自動修正建議             -- 發現不一致時給出具體修正方向
    4. 校驗報告產出             -- 輸出 Excel 格式校驗報告

  命令行用法 / CLI Usage:
    # 校驗模式
    python3 報關單據校驗助手.py --validate 報關單據.xlsx

    # 智能填充模式
    python3 報關單據校驗助手.py --fill 不完整訂單.xlsx --db 產品主數據庫.xlsx

    # 完整模式：填充 + 生成 + 校驗一條龍
    python3 報關單據校驗助手.py --full 訂單.xlsx --db 產品主數據庫.xlsx

    # 測試模式：注入不一致數據
    python3 報關單據校驗助手.py --validate 報關單據.xlsx --inject-errors

  依賴 / Dependencies:  pip install openpyxl
  版本 / Version: 1.0.0   日期 / Date: 2026-06-16
==========================================================================
"""

import argparse, os, sys, re
from datetime import datetime
from collections import OrderedDict

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[ERROR] 缺少 openpyxl。請執行: pip install openpyxl")
    sys.exit(1)


# ====================================================================
#  輔助函數
# ====================================================================

def _to_str(v):
    return str(v).strip() if v is not None else ''

def _parse_num(v):
    if v is None: return 0.0
    if isinstance(v, (int, float)): return float(v)
    s = str(v).replace(',', '').replace(' ', '').strip()
    if not s or s == '-': return 0.0
    try: return float(s)
    except: return 0.0

def _close(a, b, tol=0.001):
    if abs(b) < 1e-9: return abs(a) < tol
    return abs(a - b) / max(abs(b), 1e-9) < tol

def _match(a, b):
    return _to_str(a).lower() == _to_str(b).lower()

def _contains(a, b):
    s, t = _to_str(a).lower(), _to_str(b).lower()
    return (s in t or t in s) if s and t else False

def _after_colon(text):
    if not text: return ''
    for sep in [':', '\uff1a']:
        if sep in str(text):
            return str(text).split(sep, 1)[1].strip()
    return str(text).strip()

def _between_colon_pipe(text):
    if not text: return ''
    t = str(text)
    for sep in [':', '\uff1a']:
        if sep in t:
            rest = t.split(sep, 1)[1].strip()
            return rest.split('|')[0].strip() if '|' in rest else rest
    return t.strip()

def _after_last_colon(text):
    if not text: return ''
    parts = re.split(r'[:\\uff1a]', str(text))
    return parts[-1].strip() if len(parts) > 1 else str(text).strip()


# ====================================================================
#  Sheet 名稱與欄位定義
# ====================================================================

SHEET_NAMES = OrderedDict([
    ('commercial_invoice', '商業發票'),
    ('packing_list', '裝箱單'),
    ('bl_info', '提單資訊'),
    ('co_application', '原產地證'),
    ('declaration_elements', '申報要素'),
    ('checklist', '核對清單'),
])

SHEET_EXCEL = {
    'commercial_invoice': '商業發票 Commercial Invoice',
    'packing_list': '裝箱單 Packing List',
    'bl_info': '提單資訊 BL Info',
    'co_application': '原產地證 CO Application',
    'declaration_elements': '申報要素 Declaration Elements',
    'checklist': '核對清單 Checklist',
}

# --- 商業發票 ---
CI = {
    'invoice_no':   {'r':5,'c':3,'t':'m'},
    'date':         {'r':5,'c':8,'t':'m'},
    'seller':       {'r':6,'c':3,'t':'m'},
    'buyer':        {'r':6,'c':8,'t':'m'},
    'contract_no':  {'r':8,'c':3,'t':'m'},
    'trade_terms':  {'r':8,'c':8,'t':'m'},
    'from_port':    {'r':10,'c':3,'t':'m'},
    'to_port':      {'r':10,'c':8,'t':'m'},
    'payment':      {'r':11,'c':8,'t':'m'},
    'items':        {'sr':14,'er':18,
                     'cols':{0:'item_no',1:'desc',2:'model',3:'hs_code',
                             4:'qty',5:'unit',6:'unit_price',7:'amount',
                             8:'origin',9:'remarks'}},
    'total_qty':    {'r':19,'c':5,'t':'m'},
    'total_amount': {'r':19,'c':8,'t':'m'},
    'order_no':     {'r':5,'c':3,'t':'m','x':'strip_ci'},
}
# --- 裝箱單 ---
PL = {
    'invoice_no':   {'r':4,'c':3,'t':'m'},
    'date':         {'r':4,'c':8,'t':'m'},
    'buyer':        {'r':5,'c':8,'t':'m'},
    'contract_no':  {'r':7,'c':8,'t':'m'},
    'items':        {'sr':10,'er':14,
                     'cols':{0:'item_no',1:'desc',2:'model',3:'qty',
                             4:'packages',5:'nw',6:'gw',7:'meas',
                             8:'cbm',9:'remarks'}},
    'total_qty':    {'r':15,'c':4,'t':'m'},
    'total_nw':     {'r':15,'c':6,'t':'m'},
    'total_gw':     {'r':15,'c':7,'t':'m'},
    'total_cbm':    {'r':15,'c':9,'t':'m'},
}
# --- 提單資訊 ---
BL = {
    'consignee':    {'r':4,'c':7,'t':'m'},
    'port_disch':   {'r':9,'c':7,'t':'m'},
    'port_load':    {'r':9,'c':3,'t':'m'},
    'items':        {'sr':14,'er':18,
                     'cols':{0:'marks',1:'desc',2:'qty',3:'unit',
                             4:'nw',5:'gw',6:'meas',7:'cbm'}},
    'total_qty':    {'r':19,'c':3,'t':'m'},
    'total_nw':     {'r':19,'c':5,'t':'m'},
    'total_gw':     {'r':19,'c':6,'t':'m'},
    'total_cbm':    {'r':19,'c':8,'t':'m'},
}
# --- 原產地證 ---
CO = {
    'consignee':    {'r':5,'c':7,'t':'m'},
    'invoice_no':   {'r':7,'c':7,'t':'m'},
    'date':         {'r':8,'c':3,'t':'m'},
    'items':        {'sr':11,'er':15,
                     'cols':{0:'item_no',1:'desc',2:'hs_code',
                             3:'origin_crit',4:'qty',5:'unit',
                             6:'fob',7:'origin'}},
    'total_fob':    {'r':16,'c':7,'t':'m'},
}
# --- 申報要素 ---
DE = {
    'order_no':     {'r':5,'c':3,'t':'m'},
    'contract_no':  {'r':5,'c':7,'t':'m'},
    'customer':     {'r':6,'c':3,'t':'m'},
    'dest_port':    {'r':6,'c':7,'t':'m'},
    'invoice_no':   {'r':3,'c':1,'t':'m','x':'bcp'},
    'date':         {'r':3,'c':1,'t':'m','x':'alc'},
}
# --- 核對清單 ---
CL = {
    'order_no':     {'r':5,'c':3,'t':'m'},
    'order_date':   {'r':5,'c':6,'t':'m'},
    'contract_no':  {'r':6,'c':3,'t':'m'},
    'customer':     {'r':6,'c':6,'t':'m'},
    'dest_port':    {'r':7,'c':3,'t':'m'},
    'trade_terms':  {'r':7,'c':6,'t':'m'},
}

ALL_FIELDS = {'commercial_invoice':CI,'packing_list':PL,
              'bl_info':BL,'co_application':CO,
              'declaration_elements':DE,'checklist':CL}


# ====================================================================
#  20 條校驗規則
# ====================================================================

RULES = [
  {'id':'V-001','d':'裝箱單發票號碼應與商業發票一致',
   'ss':'commercial_invoice','sf':'invoice_no',
   'ts':'packing_list','tf':'invoice_no','rt':'exact'},
  {'id':'V-002','d':'裝箱單產品項數應與商業發票一致',
   'ss':'commercial_invoice','sf':'_count',
   'ts':'packing_list','tf':'_count','rt':'exact'},
  {'id':'V-003','d':'裝箱單數量合計應與商業發票一致',
   'ss':'commercial_invoice','sf':'total_qty',
   'ts':'packing_list','tf':'total_qty','rt':'num'},
  {'id':'V-004','d':'商業發票各品金額加總應與總金額一致',
   'ss':'commercial_invoice','sf':'_sum_amount',
   'ts':'commercial_invoice','tf':'total_amount','rt':'num'},
  {'id':'V-005','d':'原產地證發票號碼應與商業發票一致',
   'ss':'commercial_invoice','sf':'invoice_no',
   'ts':'co_application','tf':'invoice_no','rt':'exact'},
  {'id':'V-006','d':'原產地證收貨人應與商業發票買方一致',
   'ss':'commercial_invoice','sf':'buyer',
   'ts':'co_application','tf':'consignee','rt':'exact'},
  {'id':'V-007','d':'原產地證FOB總金額應與商業發票總金額一致',
   'ss':'commercial_invoice','sf':'total_amount',
   'ts':'co_application','tf':'total_fob','rt':'num'},
  {'id':'V-008','d':'原產地證產品項數應與商業發票一致',
   'ss':'commercial_invoice','sf':'_count',
   'ts':'co_application','tf':'_count','rt':'exact'},
  {'id':'V-009','d':'提單資訊收貨人應與商業發票買方一致',
   'ss':'commercial_invoice','sf':'buyer',
   'ts':'bl_info','tf':'consignee','rt':'exact'},
  {'id':'V-010','d':'提單資訊卸貨港應與商業發票目的港一致',
   'ss':'commercial_invoice','sf':'to_port',
   'ts':'bl_info','tf':'port_disch','rt':'contains'},
  {'id':'V-011','d':'提單資訊數量合計應與商業發票一致',
   'ss':'commercial_invoice','sf':'total_qty',
   'ts':'bl_info','tf':'total_qty','rt':'num'},
  {'id':'V-012','d':'提單資訊淨重合計應與裝箱單一致',
   'ss':'packing_list','sf':'total_nw',
   'ts':'bl_info','tf':'total_nw','rt':'num'},
  {'id':'V-013','d':'提單資訊毛重合計應與裝箱單一致',
   'ss':'packing_list','sf':'total_gw',
   'ts':'bl_info','tf':'total_gw','rt':'num'},
  {'id':'V-014','d':'提單資訊CBM合計應與裝箱單一致',
   'ss':'packing_list','sf':'total_cbm',
   'ts':'bl_info','tf':'total_cbm','rt':'num'},
  {'id':'V-015','d':'提單資訊產品項數應與裝箱單一致',
   'ss':'packing_list','sf':'_count',
   'ts':'bl_info','tf':'_count','rt':'exact'},
  {'id':'V-016','d':'原產地證HS編碼應與商業發票逐項一致',
   'ss':'commercial_invoice','sf':'_hs_list',
   'ts':'co_application','tf':'_hs_list','rt':'list'},
  {'id':'V-017','d':'申報要素訂單編號應與商業發票一致',
   'ss':'commercial_invoice','sf':'order_no',
   'ts':'declaration_elements','tf':'order_no','rt':'contains'},
  {'id':'V-018','d':'核對清單訂單編號應與商業發票一致',
   'ss':'commercial_invoice','sf':'order_no',
   'ts':'checklist','tf':'order_no','rt':'contains'},
  {'id':'V-019','d':'申報要素日期應與商業發票日期一致',
   'ss':'commercial_invoice','sf':'date',
   'ts':'declaration_elements','tf':'date','rt':'contains'},
  {'id':'V-020','d':'核對清單日期應與商業發票日期一致',
   'ss':'commercial_invoice','sf':'date',
   'ts':'checklist','tf':'order_date','rt':'contains'},
]


# ====================================================================
#  SmartFiller -- 智能填充引擎
# ====================================================================

class SmartFiller:
    """智能填充引擎：自動偵測欄位 + 從產品主數據庫補全"""

    FIELD_PAT = OrderedDict([
        ('product_code', ['product code','part no','part number','item code',
                          'item no','code','sku','產品編碼','編碼','產品內部編碼']),
        ('product_name_cn', ['中文品名','品名','中文名稱','產品名稱']),
        ('product_name_en', ['英文品名','english name','description','product name',
                             'goods description']),
        ('model', ['型號規格','型號','model','spec','specification','規格']),
        ('hs_code', ['hs code','hs編碼','hs 編碼','harmonized','海關編碼','稅則號']),
        ('quantity', ['數量','qty','quantity']),
        ('unit', ['單位','unit','uom']),
        ('unit_price', ['單價','unit price','price','單價usd']),
        ('amount', ['金額','amount','total price','total amount','金額usd']),
        ('remarks', ['備註','remarks','notes','comment']),
    ])
    META_PAT = OrderedDict([
        ('order_no', ['order no','訂單編號','po no','purchase order']),
        ('order_date', ['order date','訂單日期','po date']),
        ('customer', ['customer','客戶名稱','buyer','consignee','客戶']),
        ('contract_no', ['contract no','合同編號','contract']),
        ('address', ['address','客戶地址','地址']),
        ('dest_port', ['dest','destination','目的港','to port']),
        ('trade_terms', ['terms','貿易條款','incoterms']),
        ('payment', ['payment','付款方式','付款條件']),
    ])

    def auto_detect_fields(self, path):
        """自動偵測欄位映射，回傳 (wb, hdr_row, fmap, items, meta)"""
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        hdr_row = None; fmap = {}; meta = {}
        for ri in range(1, min(ws.max_row+1, 30)):
            cells = [(c, str(ws.cell(ri, c).value).strip())
                     for c in range(1, ws.max_column+1)
                     if ws.cell(ri, c).value is not None]
            for ci, cs in cells:
                cl = cs.lower()
                for fn, pats in self.FIELD_PAT.items():
                    for p in pats:
                        if p.lower() in cl and fn not in fmap:
                            fmap[fn] = ci; hdr_row = ri; break
                for mn, pats in self.META_PAT.items():
                    for p in pats:
                        if p.lower() in cl and mn not in meta:
                            v = self._find_val(ws, ri, ci, ws.max_column)
                            if v: meta[mn] = v
                            break
            if len(fmap) >= 3: break
        items = []
        if hdr_row:
            for ri in range(hdr_row+1, ws.max_row+1):
                row = {fn: ws.cell(ri, ci).value for fn, ci in fmap.items()}
                if all(v is None for v in row.values()): continue
                if not row.get('product_code') and not row.get('product_name_en'):
                    continue
                for mk, mv in meta.items():
                    row.setdefault(mk, mv)
                items.append(row)
        return wb, hdr_row, fmap, items, meta

    def _find_val(self, ws, r, lc, mc):
        for c in range(lc+1, min(lc+6, mc+1)):
            v = ws.cell(r, c).value
            if v is not None and str(v).strip(): return str(v).strip()
        return None

    def fill_from_product_master(self, order_path, db_path):
        """從產品主數據庫補全訂單"""
        wb, hr, fm, items, meta = self.auto_detect_fields(order_path)
        pdb = self._load_db(db_path)
        filled = 0
        for it in items:
            code = _to_str(it.get('product_code'))
            if not code: continue
            prod = pdb.get(code) or pdb.get(code.upper()) or pdb.get(code.lower())
            if not prod: continue
            filled += 1
            fills = [
                ('product_name_cn','中文品名'),('product_name_en','英文品名'),
                ('model','型號規格'),('hs_code','HS編碼'),('unit','單位'),
                ('unit_price','單價USD'),('net_weight','淨重kg'),
                ('gross_weight','毛重kg'),('length','長cm'),('width','寬cm'),
                ('height','高cm'),('brand','品牌'),('material','材質'),
                ('purpose','用途'),('declaration_cn','申報要素描述_中文'),
                ('declaration_en','申報要素描述_英文'),
                ('func_cn','功能描述_中文'),('func_en','功能描述_英文'),
                ('origin','原產地'),
            ]
            for ikey, dkey in fills:
                if not _to_str(it.get(ikey)) and prod.get(dkey) is not None:
                    it[ikey] = prod[dkey]
            if not it.get('amount') and it.get('quantity') and it.get('unit_price'):
                it['amount'] = round(_parse_num(it['quantity'])*_parse_num(it['unit_price']),2)
        print(f"[INFO] 共 {len(items)} 項，匹配數據庫 {filled} 項")
        return wb, hr, fm, items, meta

    def _load_db(self, path):
        try: wb = openpyxl.load_workbook(path)
        except Exception as e: print(f"[ERROR] {e}"); return {}
        ws = wb.active
        hdrs = [str(ws.cell(1,c).value).strip() for c in range(1, ws.max_column+1)]
        db = {}
        for ri in range(2, ws.max_row+1):
            rd = {hdrs[ci]: ws.cell(ri, ci+1).value for ci in range(len(hdrs))}
            code = rd.get('產品內部編碼')
            if code: db[str(code).strip()] = rd
        return db

    def save(self, wb, hr, fm, items, meta, path):
        ws = wb.active
        if hr:
            for i, it in enumerate(items):
                for fn, ci in fm.items():
                    v = it.get(fn)
                    if v is not None: ws.cell(hr+1+i, ci).value = v
        os.makedirs(os.path.dirname(path) or '.', exist_ok=True)
        wb.save(path)
        print(f"[INFO] 填充結果: {path}")


# ====================================================================
#  DeclarationParser -- 報關單據解析器
# ====================================================================

class DeclarationParser:
    """解析含6個Sheet的報關單據Excel為結構化字典"""

    def parse(self, filepath):
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"找不到檔案: {filepath}")
        wb = openpyxl.load_workbook(filepath, data_only=True)
        parsed = {}
        for key in SHEET_NAMES:
            name = SHEET_EXCEL[key]
            if name in wb.sheetnames:
                ws = wb[name]
                parsed[key] = self._parse_sheet(ws, key)
            else:
                parsed[key] = {'_missing': True}
        wb.close()
        return parsed

    def _parse_sheet(self, ws, key):
        fdefs = ALL_FIELDS.get(key, {})
        result = {}; items_list = []
        for fname, fd in fdefs.items():
            if fname == 'items':
                items_list = self._read_items(ws, fd)
                result['items'] = items_list
                result['_count'] = len(items_list)
            elif fd.get('t') == 'm':
                v = ws.cell(fd['r'], fd['c']).value
                x = fd.get('x')
                if x == 'ac': v = _after_colon(v)
                elif x == 'bcp': v = _between_colon_pipe(v)
                elif x == 'alc': v = _after_last_colon(v)
                elif x == 'strip_ci':
                    v = _to_str(v).replace('CI-', '').replace('ci-', '')
                result[fname] = v
        # 申報要素特殊計算 product count
        if key == 'declaration_elements':
            cnt = 0
            for ri in range(1, ws.max_row+1):
                v = ws.cell(ri, 1).value
                if v and str(v).strip().startswith('Product #'): cnt += 1
            result['_count'] = cnt
        return result

    def _read_items(self, ws, fd):
        sr = fd['sr']; er = fd['er']; cmap = fd['cols']
        items = []
        for ri in range(sr, er+1):
            row = {}
            all_none = True
            for ci, cn in cmap.items():
                v = ws.cell(ri, ci+1).value
                row[cn] = v
                if v is not None: all_none = False
            if all_none: continue
            desc = row.get('desc','')
            if desc and ('total' in str(desc).lower() or '合計' in str(desc)):
                continue
            items.append(row)
        return items


# ====================================================================
#  CrossValidator -- 跨單據一致性校驗引擎
# ====================================================================

class CrossValidator:
    """20條規則的跨單據校驗引擎"""

    def validate(self, data):
        results = []
        for rule in RULES:
            r = self._check(rule, data)
            results.append(r)
        return results

    def _check(self, rule, data):
        ss, sf = rule['ss'], rule['sf']
        ts, tf = rule['ts'], rule['tf']
        rt = rule['rt']
        sd = data.get(ss, {}); td = data.get(ts, {})
        if sd.get('_missing'):
            return self._res(rule, '跳過', '', '', f'缺少來源: {SHEET_NAMES.get(ss,ss)}')
        if td.get('_missing'):
            return self._res(rule, '跳過', '', '', f'缺少目標: {SHEET_NAMES.get(ts,ts)}')
        sv = self._val(sd, sf, ss)
        tv = self._val(td, tf, ts)
        sd_ = self._fmt(sv); td_ = self._fmt(tv)
        if sv is None or (isinstance(sv,str) and not sv.strip()):
            return self._res(rule, '警告', sd_, td_, '來源值為空')
        if tv is None or (isinstance(tv,str) and not tv.strip()):
            return self._res(rule, '警告', sd_, td_, '目標值為空')
        ok = False
        if rt == 'exact': ok = _match(sv, tv)
        elif rt == 'num': ok = _close(_parse_num(sv), _parse_num(tv))
        elif rt == 'contains': ok = _contains(sv, tv)
        elif rt == 'list':
            a = sorted([str(x).lower() for x in sv]) if isinstance(sv,list) else []
            b = sorted([str(x).lower() for x in tv]) if isinstance(tv,list) else []
            ok = a == b
        if ok:
            return self._res(rule, '通過', sd_, td_, '校驗通過')
        return self._res(rule, '錯誤', sd_, td_, '數值不一致，需要修正')

    def _val(self, d, f, sheet):
        if f == '_sum_amount':
            return sum(_parse_num(it.get('amount',0)) for it in d.get('items',[]))
        if f == '_hs_list':
            return [str(it.get('hs_code','')).strip() for it in d.get('items',[]) if it.get('hs_code') is not None]
        if f == '_count':
            return d.get('_count', 0)
        return d.get(f)

    def _fmt(self, v):
        if v is None: return '(空)'
        if isinstance(v, list): return ', '.join(str(x) for x in v)
        if isinstance(v, float):
            if v == int(v) and abs(v) > 0.01: return str(int(v))
            return f'{v:.4f}' if abs(v) < 1 else f'{v:.2f}'
        return str(v)

    def _res(self, rule, status, sv, tv, msg):
        return {
            'rule_id': rule['id'], 'rule_desc': rule['d'], 'status': status,
            'source_sheet': SHEET_NAMES.get(rule['ss'], rule['ss']),
            'source_field': rule['sf'], 'source_value': sv,
            'target_sheet': SHEET_NAMES.get(rule['ts'], rule['ts']),
            'target_field': rule['tf'], 'target_value': tv, 'message': msg,
        }


# ====================================================================
#  FixSuggester -- 自動修正建議引擎
# ====================================================================

class FixSuggester:
    """針對校驗錯誤產生具體修正建議"""

    TEMPLATES = {
        'V-001': lambda r: f"請將 {r['target_sheet']} 的發票號碼統一修改為「{r['source_value']}」",
        'V-002': lambda r: f"請確認 {r['source_sheet']} 與 {r['target_sheet']} 的產品項數差異，確保所有產品都已列入",
        'V-003': lambda r: f"請核對數量合計，以 {r['source_sheet']} 的 {r['source_value']} 為基準統一",
        'V-004': lambda r: f"請逐項核對 {r['source_sheet']} 中各產品的金額(數量x單價)，確保合計等於總金額",
        'V-005': lambda r: f"請將 {r['target_sheet']} 的發票號碼修改為「{r['source_value']}」",
        'V-006': lambda r: f"請將 {r['target_sheet']} 的收貨人名稱統一修改為「{r['source_value']}」",
        'V-007': lambda r: f"請確認 {r['target_sheet']} 的FOB價值是否應等於 {r['source_sheet']} 總金額，注意CIF條款下FOB可能不同",
        'V-008': lambda r: f"請確認 {r['target_sheet']} 是否遺漏產品項目，應與 {r['source_sheet']} 保持一致（{r['source_value']}項）",
        'V-009': lambda r: f"請將 {r['target_sheet']} 的收貨人統一修改為「{r['source_value']}」",
        'V-010': lambda r: f"請確認 {r['target_sheet']} 的卸貨港應包含或等於「{r['source_value']}」",
        'V-011': lambda r: f"請核對 {r['target_sheet']} 的數量合計，應與 {r['source_sheet']} 一致（{r['source_value']}）",
        'V-012': lambda r: f"請核對 {r['target_sheet']} 的淨重合計，以 {r['source_sheet']} 的 {r['source_value']} 為基準統一",
        'V-013': lambda r: f"請核對 {r['target_sheet']} 的毛重合計，以 {r['source_sheet']} 的 {r['source_value']} 為基準統一",
        'V-014': lambda r: f"請核對 {r['target_sheet']} 的CBM合計，以 {r['source_sheet']} 的 {r['source_value']} 為基準統一",
        'V-015': lambda r: f"請確認 {r['target_sheet']} 與 {r['source_sheet']} 的產品項數一致",
        'V-016': lambda r: f"請逐項比對 {r['source_sheet']} 與 {r['target_sheet']} 的HS編碼，確保每個產品完全一致",
        'V-017': lambda r: f"請將 {r['target_sheet']} 的訂單編號統一為「{r['source_value']}」",
        'V-018': lambda r: f"請將 {r['target_sheet']} 的訂單編號統一為「{r['source_value']}」",
        'V-019': lambda r: f"請將 {r['target_sheet']} 的日期統一為「{r['source_value']}」",
        'V-020': lambda r: f"請將 {r['target_sheet']} 的日期統一為「{r['source_value']}」",
    }

    def suggest(self, results):
        for r in results:
            if r['status'] in ('錯誤', '警告'):
                fn = self.TEMPLATES.get(r['rule_id'])
                r['fix_suggestion'] = fn(r) if fn else f"請以 {r['source_sheet']} 的值「{r['source_value']}」為基準修正 {r['target_sheet']}"
            else:
                r['fix_suggestion'] = ''
        return results


# ====================================================================
#  ReportGenerator -- 校驗報告生成器
# ====================================================================

class ReportGenerator:
    """將校驗結果輸出為格式化 Excel 報告"""

    GREEN = 'C6EFCE'; YELLOW = 'FFEB9C'; RED = 'FFC7CE'
    HDR = '002060'; HDR_TXT = 'FFFFFF'; GRAY = 'F2F2F2'

    def generate(self, results, output_path, doc_path=''):
        wb = openpyxl.Workbook()
        pc = sum(1 for r in results if r['status']=='通過')
        wc = sum(1 for r in results if r['status']=='警告')
        ec = sum(1 for r in results if r['status']=='錯誤')
        sc = sum(1 for r in results if r['status']=='跳過')
        tot = len(results)
        conclusion = '需修正' if ec else ('建議確認' if wc else '通過')

        self._summary(wb, tot, pc, wc, ec, sc, conclusion, doc_path)
        self._detail(wb, results)
        os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
        wb.save(output_path)
        return output_path

    def _summary(self, wb, tot, pc, wc, ec, sc, conclusion, doc_path):
        ws = wb.active; ws.title = '校驗總覽'
        tf = Font(name='Microsoft JhengHei', size=16, bold=True, color=self.HDR)
        hf = Font(name='Microsoft JhengHei', size=11, bold=True, color=self.HDR_TXT)
        hfill = PatternFill('solid', fgColor=self.HDR)
        nf = Font(name='Microsoft JhengHei', size=11)
        bf = Font(name='Microsoft JhengHei', size=11, bold=True)

        ws.merge_cells('A1:F1')
        ws['A1'] = '報關單據一致性校驗報告'; ws['A1'].font = tf
        ws['A1'].alignment = Alignment(horizontal='center')

        for i, (l, v) in enumerate([
            ('校驗時間', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ('單據路徑', doc_path),
            ('校驗工具版本', '1.0.0'),
        ], 3):
            ws.cell(i, 1, l).font = bf; ws.cell(i, 2, v).font = nf

        sr = 7
        for j, h in enumerate(['統計項目','數量','百分比'], 1):
            c = ws.cell(sr, j, h); c.font = hf; c.fill = hfill
            c.alignment = Alignment(horizontal='center')

        cmap = {'通過': self.GREEN, '警告': self.YELLOW, '錯誤': self.RED}
        for i, (l, cnt, pct) in enumerate([
            ('總校驗項目', tot, '100%'),
            ('通過', pc, f'{pc/tot*100:.1f}%' if tot else '0%'),
            ('警告', wc, f'{wc/tot*100:.1f}%' if tot else '0%'),
            ('錯誤', ec, f'{ec/tot*100:.1f}%' if tot else '0%'),
            ('跳過', sc, f'{sc/tot*100:.1f}%' if tot else '0%'),
        ], sr+1):
            ws.cell(i, 1, l).font = bf
            ws.cell(i, 2, cnt).font = nf; ws.cell(i,2).alignment = Alignment(horizontal='center')
            ws.cell(i, 3, pct).font = nf; ws.cell(i,3).alignment = Alignment(horizontal='center')
            if l in cmap:
                fl = PatternFill('solid', fgColor=cmap[l])
                for j in range(1,4): ws.cell(i, j).fill = fl

        cr = sr + 7
        ws.cell(cr, 1, '校驗結論').font = bf
        cc = ws.cell(cr, 2, conclusion)
        if conclusion == '通過':
            cc.font = Font(name='Microsoft JhengHei', size=14, bold=True, color='006100')
            cc.fill = PatternFill('solid', fgColor=self.GREEN)
        elif conclusion == '需修正':
            cc.font = Font(name='Microsoft JhengHei', size=14, bold=True, color='9C0006')
            cc.fill = PatternFill('solid', fgColor=self.RED)
        else:
            cc.font = Font(name='Microsoft JhengHei', size=14, bold=True, color='9C6500')
            cc.fill = PatternFill('solid', fgColor=self.YELLOW)

        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 45
        ws.column_dimensions['C'].width = 14

    def _detail(self, wb, results):
        ws = wb.create_sheet('詳細結果')
        hdrs = ['校驗編號','校驗項目','來源單據','來源值','目標單據','目標值','結果','修正建議']
        hf = Font(name='Microsoft JhengHei', size=10, bold=True, color=self.HDR_TXT)
        hfill = PatternFill('solid', fgColor=self.HDR)
        for j, h in enumerate(hdrs, 1):
            c = ws.cell(1, j, h); c.font = hf; c.fill = hfill
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        nf = Font(name='Microsoft JhengHei', size=10)
        scolors = {'通過':self.GREEN, '警告':self.YELLOW, '錯誤':self.RED, '跳過':self.GRAY}

        for i, r in enumerate(results, 2):
            vals = [r.get('rule_id',''), r.get('rule_desc',''),
                    r.get('source_sheet',''), r.get('source_value',''),
                    r.get('target_sheet',''), r.get('target_value',''),
                    r.get('status',''), r.get('fix_suggestion','')]
            for j, v in enumerate(vals, 1):
                c = ws.cell(i, j, v); c.font = nf
                c.alignment = Alignment(vertical='center', wrap_text=True)
            st = r.get('status','')
            if st in scolors:
                ws.cell(i, 7).fill = PatternFill('solid', fgColor=scolors[st])
            if i % 2 == 0 and st not in ('錯誤','警告'):
                af = PatternFill('solid', fgColor=self.GRAY)
                for j in range(1, 9):
                    if j != 7: ws.cell(i, j).fill = af

        for j, w in enumerate([10,40,12,28,12,28,8,55], 1):
            ws.column_dimensions[get_column_letter(j)].width = w
        ws.freeze_panes = 'A2'


# ====================================================================
#  SimplifiedDocGenerator -- 簡化報關單據生成器 (用於 --full 模式)
# ====================================================================

class SimplifiedDocGenerator:
    """根據填充後的訂單數據生成含6個Sheet的報關單據"""

    CO_NAME = 'SHANGHAI YUDA GROUP'
    SELLER = '上海御大國際貿易有限公司'
    S_ADDR = '上海市浦東新區張衡路1888號A座1208室，郵編201203'

    def generate(self, order_data, output_path):
        wb = openpyxl.Workbook()
        m = self._meta(order_data)
        self._ci(wb, m); self._pl(wb, m); self._bl(wb, m)
        self._co(wb, m); self._de(wb, m); self._cl(wb, m)
        os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
        wb.save(output_path)
        print(f"[INFO] 報關單據已生成: {output_path}")

    def _meta(self, od):
        items = od.get('items', [])
        tq = sum(_parse_num(it.get('quantity',0)) for it in items)
        ta = sum(_parse_num(it.get('amount',0)) for it in items)
        tnw = sum(_parse_num(it.get('net_weight',0)) for it in items)
        tgw = sum(_parse_num(it.get('gross_weight',0)) for it in items)
        tcbm = 0; tpkg = 0
        for it in items:
            l,w,h = _parse_num(it.get('length',0)),_parse_num(it.get('width',0)),_parse_num(it.get('height',0))
            pk = it.get('packages',1) or 1
            if l and w and h: tcbm += (l/100)*(w/100)*(h/100)*pk
            tpkg += int(pk)
        ono = od.get('order_no','')
        return {
            'order_no':ono, 'invoice_no':f'CI-{ono}' if ono else 'CI-X',
            'date':od.get('order_date', datetime.now().strftime('%Y-%m-%d')),
            'customer':od.get('customer',''), 'address':od.get('address',''),
            'contract_no':od.get('contract_no',''), 'dest_port':od.get('dest_port',''),
            'trade_terms':od.get('trade_terms',''), 'payment':od.get('payment',''),
            'from_port':'Shanghai, China', 'items':items,
            'total_qty':tq, 'total_amount':ta, 'total_nw':tnw, 'total_gw':tgw,
            'total_cbm':round(tcbm,4), 'total_pkg':tpkg,
        }

    def _ci(self, wb, m):
        ws = wb.active; ws.title = SHEET_EXCEL['commercial_invoice']
        ws.cell(1,1,self.CO_NAME).font = Font(bold=True, size=14)
        ws.cell(2,1,'商業發票 / COMMERCIAL INVOICE').font = Font(bold=True, size=12)
        ws.cell(3,1,f"Invoice No. / 發票號碼: {m['invoice_no']}")
        ws.cell(5,1,'Invoice No. / 發票號碼'); ws.cell(5,3,m['invoice_no'])
        ws.cell(5,6,'Date / 日期'); ws.cell(5,8,m['date'])
        ws.cell(6,1,'Seller / 賣方'); ws.cell(6,3,self.SELLER)
        ws.cell(6,6,'Buyer / 買方'); ws.cell(6,8,m['customer'])
        ws.cell(7,1,'Address / 地址'); ws.cell(7,3,self.S_ADDR)
        ws.cell(7,6,'Buyer Address / 買方地址'); ws.cell(7,8,m.get('address',''))
        ws.cell(8,1,'Contract No. / 合同編號'); ws.cell(8,3,m['contract_no'])
        ws.cell(8,6,'Trade Terms / 貿易條款'); ws.cell(8,8,m['trade_terms'])
        ws.cell(10,1,'From / 起運港'); ws.cell(10,3,m['from_port'])
        ws.cell(10,6,'To / 目的港'); ws.cell(10,8,m['dest_port'])
        cs = m['customer'][:20] if m['customer'] else ''
        ws.cell(11,1,'Shipping Marks / 嘜頭'); ws.cell(11,3,f'N/M - {cs}')
        ws.cell(11,6,'Payment / 付款方式'); ws.cell(11,8,m['payment'])
        for j,h in enumerate(['Item No.\n項號','Description of Goods (English)\n貨物描述（英文）',
            'Model\n型號','HS Code\nHS編碼','Quantity\n數量','Unit\n單位',
            'Unit Price\n(USD)\n單價','Amount\n(USD)\n金額',
            'Country of\nOrigin\n原產地','Remarks\n備註'], 1):
            ws.cell(13,j,h).font = Font(bold=True, size=9)
        for i, it in enumerate(m['items']):
            r = 14+i
            ws.cell(r,1,i+1); ws.cell(r,2,it.get('product_name_en',''))
            ws.cell(r,3,it.get('model','')); ws.cell(r,4,str(it.get('hs_code','')))
            ws.cell(r,5,it.get('quantity',0)); ws.cell(r,6,it.get('unit',''))
            ws.cell(r,7,it.get('unit_price',0)); ws.cell(r,8,it.get('amount',0))
            ws.cell(r,9,it.get('origin','CN')); ws.cell(r,10,it.get('remarks',''))
        tr = 14+len(m['items'])
        ws.cell(tr,2,'TOTAL / 合計'); ws.cell(tr,5,m['total_qty']); ws.cell(tr,8,m['total_amount'])

    def _pl(self, wb, m):
        ws = wb.create_sheet(SHEET_EXCEL['packing_list'])
        ws.cell(1,1,self.CO_NAME).font = Font(bold=True, size=14)
        ws.cell(2,1,'裝箱單 / PACKING LIST').font = Font(bold=True, size=12)
        ws.cell(4,1,'Invoice No. / 發票號碼'); ws.cell(4,3,m['invoice_no'])
        ws.cell(4,6,'Date / 日期'); ws.cell(4,8,m['date'])
        ws.cell(5,1,'Seller / 賣方'); ws.cell(5,3,self.SELLER)
        ws.cell(5,6,'Buyer / 買方'); ws.cell(5,8,m['customer'])
        ws.cell(6,1,'From / 起運港'); ws.cell(6,3,m['from_port'])
        ws.cell(6,6,'To / 目的港'); ws.cell(6,8,m['dest_port'])
        cs = m['customer'][:20] if m['customer'] else ''
        ws.cell(7,1,'Shipping Marks / 嘜頭'); ws.cell(7,3,f'N/M - {cs}')
        ws.cell(7,6,'Contract No. / 合同編號'); ws.cell(7,8,m['contract_no'])
        for j,h in enumerate(['Item No.\n項號','Description\n貨物描述','Model\n型號',
            'Quantity\n數量','Packages\n件數','Net Weight\n(kg)\n淨重',
            'Gross Weight\n(kg)\n毛重','Measurement\n(cm)\n尺寸',
            'CBM\n(m³)\n體積','Remarks\n備註'], 1):
            ws.cell(9,j,h).font = Font(bold=True, size=9)
        for i, it in enumerate(m['items']):
            r = 10+i
            ws.cell(r,1,i+1); ws.cell(r,2,it.get('product_name_en',''))
            ws.cell(r,3,it.get('model','')); ws.cell(r,4,it.get('quantity',0))
            ws.cell(r,5,it.get('packages',1)); ws.cell(r,6,it.get('net_weight',0))
            ws.cell(r,7,it.get('gross_weight',0))
            l,w,h = it.get('length',0),it.get('width',0),it.get('height',0)
            ws.cell(r,8,f'{l}x{w}x{h}' if l and w and h else '')
            pk = it.get('packages',1) or 1
            cbm = round((l/100)*(w/100)*(h/100)*pk, 4) if l and w and h else 0
            ws.cell(r,9,cbm); ws.cell(r,10,it.get('remarks',''))
        tr = 10+len(m['items'])
        ws.cell(tr,2,'TOTAL / 合計'); ws.cell(tr,4,m['total_qty'])
        ws.cell(tr,5,m['total_pkg']); ws.cell(tr,6,m['total_nw'])
        ws.cell(tr,7,m['total_gw']); ws.cell(tr,9,m['total_cbm'])

    def _bl(self, wb, m):
        ws = wb.create_sheet(SHEET_EXCEL['bl_info'])
        ws.cell(1,1,self.CO_NAME).font = Font(bold=True, size=14)
        ws.cell(2,1,'提單資訊 / BILL OF LADING INFORMATION').font = Font(bold=True, size=12)
        ws.cell(4,1,'Shipper / 託運人'); ws.cell(4,3,self.SELLER)
        ws.cell(4,5,'Consignee / 收貨人'); ws.cell(4,7,m['customer'])
        ws.cell(5,1,'Address / 地址'); ws.cell(5,3,self.S_ADDR)
        ws.cell(5,5,'Address / 地址'); ws.cell(5,7,m.get('address',''))
        ws.cell(6,1,'Notify Party / 通知方'); ws.cell(6,3,m['customer'])
        ws.cell(6,5,'Notify Address / 通知地址'); ws.cell(6,7,m.get('address',''))
        ws.cell(8,1,'Vessel / 船名'); ws.cell(8,3,'TO BE NOMINATED / 待定')
        ws.cell(8,5,'Voyage / 航次'); ws.cell(8,7,'TO BE NOMINATED / 待定')
        ws.cell(9,1,'Port of Loading / 裝貨港'); ws.cell(9,3,m['from_port'])
        ws.cell(9,5,'Port of Discharge / 卸貨港'); ws.cell(9,7,m['dest_port'])
        ws.cell(10,1,'Place of Receipt / 收貨地'); ws.cell(10,3,m['from_port'])
        ws.cell(10,5,'Place of Delivery / 交貨地'); ws.cell(10,7,m['dest_port'])
        ws.cell(11,1,'B/L No. / 提單號'); ws.cell(11,3,'TO BE ASSIGNED / 待分配')
        ws.cell(11,5,'Freight / 運費'); ws.cell(11,7,'PREPAID')
        for j,h in enumerate(['Marks & Numbers\n嘜頭','Description of Goods\n貨物描述',
            'Quantity\n數量','Unit\n單位','Net Weight (kg)\n淨重',
            'Gross Weight (kg)\n毛重','Measurement (cm)\n尺寸','CBM (m³)\n體積'], 1):
            ws.cell(13,j,h).font = Font(bold=True, size=9)
        for i, it in enumerate(m['items']):
            r = 14+i
            if i == 0: ws.cell(r,1,f"N/M\n{m['invoice_no']}\n{m['dest_port']}")
            ws.cell(r,2,it.get('product_name_en',''))
            ws.cell(r,3,it.get('quantity',0)); ws.cell(r,4,it.get('unit',''))
            ws.cell(r,5,it.get('net_weight',0)); ws.cell(r,6,it.get('gross_weight',0))
            l,w,h = it.get('length',0),it.get('width',0),it.get('height',0)
            ws.cell(r,7,f'{l}x{w}x{h}' if l and w and h else '')
            pk = it.get('packages',1) or 1
            cbm = round((l/100)*(w/100)*(h/100)*pk, 4) if l and w and h else 0
            ws.cell(r,8,cbm)
        tr = 14+len(m['items'])
        ws.cell(tr,2,'TOTAL'); ws.cell(tr,3,m['total_qty'])
        ws.cell(tr,5,m['total_nw']); ws.cell(tr,6,m['total_gw']); ws.cell(tr,8,m['total_cbm'])

    def _co(self, wb, m):
        ws = wb.create_sheet(SHEET_EXCEL['co_application'])
        ws.cell(1,1,self.CO_NAME).font = Font(bold=True, size=14)
        ws.cell(2,1,'原產地證申請 / CERTIFICATE OF ORIGIN APPLICATION').font = Font(bold=True, size=12)
        ws.cell(3,1,'FORM E / 中國-東盟自由貿易區原產地證書')
        ws.cell(5,1,'Exporter / 出口商'); ws.cell(5,3,self.SELLER)
        ws.cell(5,5,'Consignee / 收貨人'); ws.cell(5,7,m['customer'])
        ws.cell(6,1,'Address / 地址'); ws.cell(6,3,self.S_ADDR)
        ws.cell(6,5,'Address / 地址'); ws.cell(6,7,m.get('address',''))
        ws.cell(7,1,'Transport Details / 運輸詳情')
        ws.cell(7,3,f"FROM {m['from_port'].upper()} TO {m['dest_port'].upper()}")
        ws.cell(7,5,'Invoice No. / 發票號碼'); ws.cell(7,7,m['invoice_no'])
        ws.cell(8,1,'Date / 日期'); ws.cell(8,3,m['date'])
        ws.cell(8,5,'Country of Origin / 原產國'); ws.cell(8,7,"PEOPLE'S REPUBLIC OF CHINA")
        for j,h in enumerate(['Item No.\n項號','Description of Goods\n貨物描述',
            'HS Code\nHS編碼','Origin Criterion\n原產地標準','Quantity\n數量',
            'Unit\n單位','FOB Value (USD)\nFOB價值','Country of Origin\n原產國'], 1):
            ws.cell(10,j,h).font = Font(bold=True, size=9)
        for i, it in enumerate(m['items']):
            r = 11+i
            ws.cell(r,1,i+1); ws.cell(r,2,it.get('product_name_en',''))
            ws.cell(r,3,str(it.get('hs_code',''))); ws.cell(r,4,'PS')
            ws.cell(r,5,it.get('quantity',0)); ws.cell(r,6,it.get('unit',''))
            ws.cell(r,7,it.get('amount',0)); ws.cell(r,8,'China')
        tr = 11+len(m['items'])
        ws.cell(tr,2,'TOTAL'); ws.cell(tr,7,m['total_amount'])

    def _de(self, wb, m):
        ws = wb.create_sheet(SHEET_EXCEL['declaration_elements'])
        ws.cell(1,1,self.CO_NAME).font = Font(bold=True, size=14)
        ws.cell(2,1,'海關申報要素表 / CUSTOMS DECLARATION ELEMENTS').font = Font(bold=True, size=12)
        ws.cell(3,1,f"Invoice No. / 發票號碼: {m['invoice_no']}  |  Date / 日期: {m['date']}")
        ws.cell(5,1,'訂單編號'); ws.cell(5,3,m['order_no'])
        ws.cell(5,5,'合同編號'); ws.cell(5,7,m['contract_no'])
        ws.cell(6,1,'客戶名稱'); ws.cell(6,3,m['customer'])
        ws.cell(6,5,'目的港'); ws.cell(6,7,m['dest_port'])
        cr = 8
        for i, it in enumerate(m['items']):
            cn = it.get('product_name_cn',''); en = it.get('product_name_en','')
            ws.cell(cr,1,f'Product #{i+1}: {cn} / {en}'); cr += 1
            for j,h in enumerate(['產品編碼','HS編碼','型號規格','品牌','單位','數量','單價USD','原產地'], 1):
                ws.cell(cr,j,h).font = Font(bold=True, size=9)
            cr += 1
            ws.cell(cr,1,it.get('product_code','')); ws.cell(cr,2,str(it.get('hs_code','')))
            ws.cell(cr,3,it.get('model','')); ws.cell(cr,4,it.get('brand',''))
            ws.cell(cr,5,it.get('unit','')); ws.cell(cr,6,it.get('quantity',0))
            ws.cell(cr,7,it.get('unit_price',0)); ws.cell(cr,8,it.get('origin','CN'))
            cr += 1
            ws.cell(cr,1,'申報要素項目'); ws.cell(cr,4,'中文內容'); ws.cell(cr,7,'英文內容'); cr += 1
            ws.cell(cr,1,'Product'); ws.cell(cr,7,en); cr += 1
            if it.get('brand'): ws.cell(cr,1,'Brand'); ws.cell(cr,7,it['brand']); cr += 1
            if it.get('model'): ws.cell(cr,1,'Model'); ws.cell(cr,7,it['model']); cr += 1
            cr += 1  # blank separator

    def _cl(self, wb, m):
        ws = wb.create_sheet(SHEET_EXCEL['checklist'])
        ws.cell(1,1,self.CO_NAME).font = Font(bold=True, size=14)
        ws.cell(2,1,'報關資料核對清單 / DOCUMENT CHECKLIST').font = Font(bold=True, size=12)
        ws.cell(3,1,f"訂單編號: {m['order_no']}")
        ws.cell(5,1,'訂單編號'); ws.cell(5,3,m['order_no'])
        ws.cell(5,4,'訂單日期'); ws.cell(5,6,m['date'])
        ws.cell(6,1,'合同編號'); ws.cell(6,3,m['contract_no'])
        ws.cell(6,4,'客戶名稱'); ws.cell(6,6,m['customer'])
        ws.cell(7,1,'目的港'); ws.cell(7,3,m['dest_port'])
        ws.cell(7,4,'貿易條款'); ws.cell(7,6,m['trade_terms'])
        for j,h in enumerate(['序號\nNo.','核對項目\nCheck Item','狀態\nStatus',
            '備註\nRemarks','核對人\nChecker','日期\nDate'], 1):
            ws.cell(9,j,h).font = Font(bold=True, size=9)
        checks = ['商業發票已製作','裝箱單已製作','提單資訊已確認','原產地證已申請',
                  '申報要素已核對','HS編碼已確認','品牌授權已確認','產品照片已準備',
                  '嘜頭已製作','合同副本已準備','報關委託書已簽署','其他']
        for i, c in enumerate(checks):
            ws.cell(10+i, 1, i+1); ws.cell(10+i, 2, c); ws.cell(10+i, 3, '\u2610')


# ====================================================================
#  注入測試用不一致數據
# ====================================================================

def inject_errors(data):
    """注入4項不一致，預期觸發 V-003, V-007, V-010, V-012"""
    pl = data.get('packing_list', {})
    if 'total_qty' in pl: pl['total_qty'] = 1750
    co = data.get('co_application', {})
    if 'total_fob' in co: co['total_fob'] = 25000
    bl = data.get('bl_info', {})
    if 'port_disch' in bl: bl['port_disch'] = 'Bangkok Port, Thailand'
    if 'total_nw' in bl: bl['total_nw'] = 2300
    print("[TEST] 已注入 4 項測試用不一致 (V-003, V-007, V-010, V-012)")
    return data


# ====================================================================
#  主程式
# ====================================================================

def run_validate(args):
    doc = os.path.abspath(args.document)
    if not os.path.exists(doc):
        print(f"[ERROR] 找不到檔案: {doc}"); sys.exit(1)
    print(f"\n{'='*60}\n  報關單據一致性校驗\n  檔案: {doc}\n{'='*60}\n")

    parser = DeclarationParser()
    print("[1/4] 解析報關單據...")
    data = parser.parse(doc)
    for k in SHEET_NAMES:
        d = data.get(k, {})
        if d.get('_missing'):
            print(f"  {SHEET_NAMES[k]}: 缺少")
        else:
            cnt = d.get('_count', len(d.get('items',[])))
            print(f"  {SHEET_NAMES[k]}: {cnt} 項產品")

    if args.inject_errors:
        print("\n[TEST] 注入測試用不一致數據...")
        data = inject_errors(data)

    print("\n[2/4] 執行 20 項校驗規則...")
    validator = CrossValidator()
    results = validator.validate(data)

    print("[3/4] 生成修正建議...")
    suggester = FixSuggester()
    results = suggester.suggest(results)

    print("[4/4] 產出校驗報告...")
    if args.output:
        rpt = os.path.abspath(args.output)
    else:
        doc_dir = os.path.dirname(doc) or '.'
        if os.path.basename(doc_dir) == '輸出':
            odir = doc_dir
        else:
            odir = os.path.join(doc_dir, '輸出')
        os.makedirs(odir, exist_ok=True)
        ono = 'UNKNOWN'
        ci = data.get('commercial_invoice', {})
        inv = _to_str(ci.get('invoice_no',''))
        if inv: ono = inv.replace('CI-','')
        rpt = os.path.join(odir, f"校驗報告_{ono}_{datetime.now().strftime('%Y%m%d')}.xlsx")

    ReportGenerator().generate(results, rpt, doc)

    pc = sum(1 for r in results if r['status']=='通過')
    wc = sum(1 for r in results if r['status']=='警告')
    ec = sum(1 for r in results if r['status']=='錯誤')
    sc = sum(1 for r in results if r['status']=='跳過')

    print(f"\n{'='*60}\n  校驗結果\n{'='*60}")
    print(f"  總校驗項目: {len(results)}")
    print(f"  通過: {pc}  |  警告: {wc}  |  錯誤: {ec}  |  跳過: {sc}")
    print(f"{'='*60}")
    if ec:
        print("\n  [!!] 校驗結論: 需修正")
        for r in results:
            if r['status'] == '錯誤':
                print(f"    - {r['rule_id']}: {r['rule_desc']}")
                print(f"      來源 ({r['source_sheet']}): {r['source_value']}")
                print(f"      目標 ({r['target_sheet']}): {r['target_value']}")
                if r.get('fix_suggestion'):
                    print(f"      建議: {r['fix_suggestion']}")
    elif wc:
        print("\n  [!] 校驗結論: 建議確認")
    else:
        print("\n  [OK] 校驗結論: 通過")
    print(f"\n  報告: {rpt}\n{'='*60}\n")


def run_fill(args):
    op = os.path.abspath(args.order)
    db = os.path.abspath(args.db)
    for p in [op, db]:
        if not os.path.exists(p): print(f"[ERROR] 找不到: {p}"); sys.exit(1)

    print(f"\n{'='*60}\n  智能填充模式\n  訂單: {op}\n  數據庫: {db}\n{'='*60}\n")

    filler = SmartFiller()
    print("[1/2] 偵測欄位 + 填充...")
    wb, hr, fm, items, meta = filler.fill_from_product_master(op, db)
    print(f"  表頭行: {hr}  |  欄位: {list(fm.keys())}  |  品項: {len(items)}")
    for i, it in enumerate(items):
        print(f"    #{i+1}: {it.get('product_code','?')} - {it.get('product_name_en','?')} | HS: {it.get('hs_code','?')}")

    out = os.path.abspath(args.output) if args.output else os.path.join('輸出',
        f"{os.path.splitext(os.path.basename(op))[0]}_filled.xlsx")
    print(f"\n[2/2] 儲存...")
    filler.save(wb, hr, fm, items, meta, out)
    print(f"\n{'='*60}\n  填充完成: {out}\n{'='*60}\n")


def run_full(args):
    op = os.path.abspath(args.order)
    db = os.path.abspath(args.db)
    for p in [op, db]:
        if not os.path.exists(p): print(f"[ERROR] 找不到: {p}"); sys.exit(1)

    print(f"\n{'='*60}\n  完整模式: 填充 + 生成 + 校驗\n{'='*60}\n")

    print("[STEP 1/3] 智能填充...")
    filler = SmartFiller()
    wb, hr, fm, items, meta = filler.fill_from_product_master(op, db)

    print(f"\n[STEP 2/3] 生成報關單據...")
    od = dict(meta); od['items'] = items
    odir = '輸出'; os.makedirs(odir, exist_ok=True)
    ds = datetime.now().strftime('%Y%m%d')
    base = os.path.splitext(os.path.basename(op))[0]
    doc_p = os.path.join(odir, f"報關單據_{base}_{ds}.xlsx")
    SimplifiedDocGenerator().generate(od, doc_p)

    print(f"\n[STEP 3/3] 校驗...")
    parser = DeclarationParser()
    data = parser.parse(doc_p)
    validator = CrossValidator()
    results = validator.validate(data)
    FixSuggester().suggest(results)

    rpt = os.path.join(odir, f"校驗報告_{base}_{ds}.xlsx")
    ReportGenerator().generate(results, rpt, doc_p)

    pc = sum(1 for r in results if r['status']=='通過')
    ec = sum(1 for r in results if r['status']=='錯誤')
    print(f"\n{'='*60}\n  完成!\n  報關單據: {doc_p}\n  校驗報告: {rpt}")
    print(f"  通過: {pc} | 錯誤: {ec} | 結論: {'通過' if not ec else '需修正'}\n{'='*60}\n")
    if ec: sys.exit(1)


def main():
    ap = argparse.ArgumentParser(
        description='報關模板智能填充與一致性校驗工具 v1.0',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""範例:
  python3 報關單據校驗助手.py --validate 輸出/報關單據_PO-2026-TH-0047_20260616.xlsx
  python3 報關單據校驗助手.py --fill 範例採購訂單_Sample_PO.xlsx --db 產品主數據庫_Product_Master.xlsx
  python3 報關單據校驗助手.py --full 範例採購訂單_Sample_PO.xlsx --db 產品主數據庫_Product_Master.xlsx
""")
    g = ap.add_mutually_exclusive_group(required=True)
    g.add_argument('--validate', dest='document', help='校驗模式：報關單據路徑')
    g.add_argument('--fill', dest='fill_order', help='智能填充：訂單路徑')
    g.add_argument('--full', dest='full_order', help='完整模式：訂單路徑')
    ap.add_argument('--db', help='產品主數據庫路徑')
    ap.add_argument('--output', '-o', help='輸出路徑')
    ap.add_argument('--inject-errors', action='store_true', help='注入測試用不一致數據')

    # Alias: allow 'order' as shortcut for fill_order/full_order
    args = ap.parse_args()
    if args.document:
        run_validate(args)
    elif args.fill_order:
        args.order = args.fill_order
        if not args.db: print("[ERROR] --fill 需要 --db"); sys.exit(1)
        run_fill(args)
    elif args.full_order:
        args.order = args.full_order
        if not args.db: print("[ERROR] --full 需要 --db"); sys.exit(1)
        run_full(args)


if __name__ == '__main__':
    main()
