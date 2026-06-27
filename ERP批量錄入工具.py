#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ERP批量自動錄入工具 (ERP Batch Auto-Import Tool)
===============================================
適用於泰國/印尼進出口貿易之ERP系統批量資料錄入。
支援採購訂單、供應商發票、料號主數據、合同資訊、銷售訂單、出入庫單據六類模板。
"""

import argparse
import csv
import datetime
import json
import os
import re
import sys
import textwrap
from copy import copy
from io import StringIO

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("錯誤: 需要安裝 openpyxl 套件。請執行: pip install openpyxl")
    sys.exit(1)

# ---------------------------------------------------------------------------
# 樣式常數
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Microsoft JhengHei", bold=True, color="FFFFFF", size=11)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_ALIGNMENT = Alignment(vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
ERROR_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
WARNING_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
OK_FILL = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
ERROR_FONT = Font(color="CC0000", bold=True)
WARNING_FONT = Font(color="996600", bold=True)


# ============================================================================
# ERPTemplateManager
# ============================================================================
class ERPTemplateManager:
    """管理6種標準化Excel資料蒐集模板。"""

    TEMPLATES = {
        "採購訂單": {
            "filename": "採購訂單PO模板.xlsx",
            "sheet": "採購訂單",
            "columns": [
                ("PO號碼", 18),
                ("供應商代碼", 14),
                ("供應商名稱", 28),
                ("料號", 12),
                ("料號描述(中)", 24),
                ("料號描述(英)", 30),
                ("數量", 10),
                ("單位", 8),
                ("單價", 12),
                ("幣別", 8),
                ("交貨日期", 14),
                ("目的倉庫", 16),
                ("備註", 24),
            ],
            "required": ["PO號碼", "供應商代碼", "供應商名稱", "料號", "數量", "單位", "單價", "幣別", "交貨日期"],
            "key_fields": ["PO號碼"],
            "dropdowns": {
                "幣別": ["THB", "IDR", "USD", "TWD", "CNY", "EUR", "JPY"],
                "單位": ["PCS", "KG", "SET", "M", "ROLL", "BOX", "LOT"],
            },
        },
        "供應商發票": {
            "filename": "供應商發票模板.xlsx",
            "sheet": "供應商發票",
            "columns": [
                ("發票號", 20),
                ("PO號碼", 18),
                ("供應商代碼", 14),
                ("發票日期", 14),
                ("幣別", 8),
                ("含稅總額", 16),
                ("稅額", 14),
                ("淨額", 16),
                ("到期日", 14),
            ],
            "required": ["發票號", "PO號碼", "供應商代碼", "發票日期", "幣別", "含稅總額", "淨額"],
            "key_fields": ["發票號"],
            "dropdowns": {
                "幣別": ["THB", "IDR", "USD", "TWD", "CNY", "EUR", "JPY"],
            },
        },
        "料號主數據": {
            "filename": "料號主數據模板.xlsx",
            "sheet": "料號主數據",
            "columns": [
                ("料號", 12),
                ("描述(中)", 28),
                ("描述(英)", 34),
                ("HS編碼", 16),
                ("單位", 8),
                ("淨重(kg)", 12),
                ("毛重(kg)", 12),
                ("產地", 10),
                ("類別", 12),
            ],
            "required": ["料號", "描述(中)", "描述(英)", "HS編碼", "單位"],
            "key_fields": ["料號"],
            "dropdowns": {
                "單位": ["PCS", "KG", "SET", "M", "ROLL", "BOX", "LOT"],
                "產地": ["TH", "ID", "TW", "CN", "VN", "MY", "JP", "KR", "US"],
                "類別": ["電子零件", "塑膠原料", "金屬材料", "包裝材料", "輔助材料", "成品"],
            },
        },
        "合同資訊": {
            "filename": "合同資訊模板.xlsx",
            "sheet": "合同資訊",
            "columns": [
                ("合同號", 20),
                ("合同名稱", 30),
                ("供應商", 28),
                ("簽署日期", 14),
                ("到期日", 14),
                ("金額", 18),
                ("幣別", 8),
                ("付款條件", 16),
            ],
            "required": ["合同號", "合同名稱", "供應商", "簽署日期", "到期日", "金額", "幣別"],
            "key_fields": ["合同號"],
            "dropdowns": {
                "幣別": ["THB", "IDR", "USD", "TWD", "CNY", "EUR", "JPY"],
                "付款條件": ["Net 30", "Net 60", "Net 90", "T/T Advance", "L/C at Sight", "COD"],
            },
        },
        "銷售訂單": {
            "filename": "銷售訂單SO模板.xlsx",
            "sheet": "銷售訂單",
            "columns": [
                ("SO號碼", 18),
                ("客戶代碼", 14),
                ("客戶名稱", 28),
                ("料號", 12),
                ("數量", 10),
                ("單價", 12),
                ("幣別", 8),
                ("交貨日期", 14),
                ("目的港", 18),
            ],
            "required": ["SO號碼", "客戶代碼", "客戶名稱", "料號", "數量", "單價", "幣別", "交貨日期"],
            "key_fields": ["SO號碼"],
            "dropdowns": {
                "幣別": ["THB", "IDR", "USD", "TWD", "CNY", "EUR", "JPY"],
                "目的港": ["BKK曼谷", "JKT雅加達", "KUL吉隆坡", "SGP新加坡", "HKG香港", "TYO東京", "SHA上海"],
            },
        },
        "出入庫單據": {
            "filename": "出入庫單據模板.xlsx",
            "sheet": "出入庫單據",
            "columns": [
                ("單據號", 18),
                ("類型(入庫/出庫)", 16),
                ("日期", 14),
                ("料號", 12),
                ("數量", 10),
                ("倉庫", 16),
                ("PO/SO關聯號", 18),
                ("操作人", 12),
            ],
            "required": ["單據號", "類型(入庫/出庫)", "日期", "料號", "數量", "倉庫"],
            "key_fields": ["單據號"],
            "dropdowns": {
                "類型(入庫/出庫)": ["入庫", "出庫"],
                "倉庫": ["曼谷主倉", "雅加達主倉", "春武里保稅倉", "泗水轉運倉", "台北總倉"],
            },
        },
    }

    # 範例資料
    SAMPLE_DATA = {
        "採購訂單": [
            ["PO-2026-TH-0050", "SUP-TH-001", "曼谷精密電子有限公司", "EL-001", "高精度電容器", "High-Precision Capacitor 100uF", 5000, "PCS", 12.50, "THB", "2026-07-15", "曼谷主倉", "急單-產線優先"],
            ["PO-2026-TH-0051", "SUP-TH-003", "清邁塑膠工業股份", "PM-002", "ABS塑膠原料", "ABS Plastic Resin Grade A", 2000, "KG", 85.00, "THB", "2026-07-20", "曼谷主倉", ""],
            ["PO-2026-TH-0052", "SUP-ID-002", "雅加達鋼鐵集團", "MT-003", "不鏽鋼板304", "Stainless Steel Sheet 304 2B", 500, "KG", 210.00, "USD", "2026-08-01", "雅加達主倉", "印尼進口-需報關"],
            ["PO-2026-TH-0053", "SUP-ID-005", "泗水包裝材料公司", "PK-004", "防潮包裝袋", "Moisture-Proof Packaging Bag L", 10000, "PCS", 3.80, "IDR", "2026-07-25", "泗水轉運倉", ""],
            ["PO-2026-TH-0054", "SUP-TH-001", "曼谷精密電子有限公司", "EL-006", "SMD電阻器", "SMD Resistor 0805 10K", 20000, "PCS", 0.15, "USD", "2026-08-10", "春武里保稅倉", "保稅區入庫"],
        ],
        "供應商發票": [
            ["INV-TH-2026-001", "PO-2026-TH-0050", "SUP-TH-001", "2026-06-10", "THB", 66375.00, 3875.00, 62500.00, "2026-07-10"],
            ["INV-TH-2026-002", "PO-2026-TH-0051", "SUP-TH-003", "2026-06-12", "THB", 181900.00, 11900.00, 170000.00, "2026-07-12"],
            ["INV-ID-2026-001", "PO-2026-TH-0052", "SUP-ID-002", "2026-06-15", "USD", 112350.00, 7350.00, 105000.00, "2026-08-15"],
            ["INV-ID-2026-002", "PO-2026-TH-0053", "SUP-ID-005", "2026-06-18", "IDR", 41040000.00, 2640000.00, 38400000.00, "2026-07-18"],
        ],
        "料號主數據": [
            ["EL-001", "高精度電容器", "High-Precision Capacitor 100uF", "8532.22.00", "PCS", 0.002, 0.003, "TH", "電子零件"],
            ["PM-002", "ABS塑膠原料", "ABS Plastic Resin Grade A", "3903.30.00", "KG", 1.000, 1.000, "TH", "塑膠原料"],
            ["MT-003", "不鏽鋼板304", "Stainless Steel Sheet 304 2B", "7219.33.00", "KG", 1.000, 1.000, "ID", "金屬材料"],
            ["PK-004", "防潮包裝袋", "Moisture-Proof Packaging Bag L", "3923.21.00", "PCS", 0.005, 0.008, "ID", "包裝材料"],
            ["EL-005", "連接器外殼", "Connector Housing 12-Pin", "8536.90.00", "PCS", 0.010, 0.015, "TW", "電子零件"],
            ["EL-006", "SMD電阻器", "SMD Resistor 0805 10K", "8533.21.00", "PCS", 0.001, 0.001, "TH", "電子零件"],
            ["AX-007", "潤滑油脂", "Industrial Lubricant Grease", "2710.19.00", "KG", 1.000, 1.050, "ID", "輔助材料"],
            ["FP-008", "組裝完成品A", "Assembly Unit Type-A", "8471.30.00", "SET", 2.500, 3.200, "TH", "成品"],
        ],
        "合同資訊": [
            ["CT-2026-TH-001", "2026年泰國電子零件採購框架合同", "曼谷精密電子有限公司", "2026-01-15", "2026-12-31", 5000000.00, "THB", "Net 60"],
            ["CT-2026-ID-001", "印尼金屬原料年度供應協議", "雅加達鋼鐵集團", "2026-02-01", "2027-01-31", 850000.00, "USD", "L/C at Sight"],
            ["CT-2026-TH-002", "塑膠原料季度供貨合同", "清邁塑膠工業股份", "2026-04-01", "2026-09-30", 2400000.00, "THB", "Net 30"],
        ],
        "銷售訂單": [
            ["SO-2026-TH-001", "CUS-TH-101", "曼谷智能科技股份", "EL-001", 2000, 18.50, "THB", "2026-07-30", "BKK曼谷"],
            ["SO-2026-ID-001", "CUS-ID-201", "雅加達自動化設備公司", "FP-008", 100, 4500.00, "USD", "2026-08-15", "JKT雅加達"],
            ["SO-2026-TH-002", "CUS-TH-103", "春武里汽車電子廠", "EL-006", 50000, 0.25, "USD", "2026-08-20", "BKK曼谷"],
            ["SO-2026-SG-001", "CUS-SG-301", "新加坡精密儀器貿易", "MT-003", 200, 280.00, "USD", "2026-09-01", "SGP新加坡"],
        ],
        "出入庫單據": [
            ["WH-2026-001", "入庫", "2026-06-10", "EL-001", 5000, "曼谷主倉", "PO-2026-TH-0050", "張偉明"],
            ["WH-2026-002", "入庫", "2026-06-12", "PM-002", 2000, "曼谷主倉", "PO-2026-TH-0051", "張偉明"],
            ["WH-2026-003", "出庫", "2026-06-14", "EL-001", 2000, "曼谷主倉", "SO-2026-TH-001", "李俊賢"],
            ["WH-2026-004", "入庫", "2026-06-15", "MT-003", 500, "雅加達主倉", "PO-2026-TH-0052", "Ahmad"],
            ["WH-2026-005", "出庫", "2026-06-16", "PK-004", 3000, "泗水轉運倉", "SO-2026-ID-001", "Budi"],
            ["WH-2026-006", "入庫", "2026-06-18", "EL-006", 20000, "春武里保稅倉", "PO-2026-TH-0054", "李俊賢"],
        ],
    }

    def create_template(self, template_name, output_dir="."):
        """建立單一模板檔案並填入範例資料。"""
        if template_name not in self.TEMPLATES:
            print(f"  錯誤: 未知模板類型 '{template_name}'")
            return None
        cfg = self.TEMPLATES[template_name]
        wb = Workbook()
        ws = wb.active
        ws.title = cfg["sheet"]
        headers = [c[0] for c in cfg["columns"]]
        widths = [c[1] for c in cfg["columns"]]

        # 寫入標題列
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

        # 設定欄寬
        for col_idx, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w

        # 凍結窗格
        ws.freeze_panes = "A2"
        # 自動篩選
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        # 下拉式選單 (Data Validation)
        for field, options in cfg.get("dropdowns", {}).items():
            if field in headers:
                col_idx = headers.index(field) + 1
                formula = '"' + ",".join(options) + '"'
                dv = DataValidation(type="list", formula1=formula, allow_blank=True)
                dv.error = f"請從下拉清單中選擇有效的{field}"
                dv.errorTitle = "無效輸入"
                dv.prompt = f"請選擇{field}"
                dv.promptTitle = field
                ws.add_data_validation(dv)
                for row in range(2, 102):  # 預設100行
                    dv.add(ws.cell(row=row, column=col_idx))

        # 寫入範例資料
        samples = self.SAMPLE_DATA.get(template_name, [])
        for row_idx, record in enumerate(samples, 2):
            for col_idx, val in enumerate(record, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = DATA_ALIGNMENT
                cell.border = THIN_BORDER

        filepath = os.path.join(output_dir, cfg["filename"])
        wb.save(filepath)
        return filepath

    def create_all_templates(self, output_dir="."):
        """建立所有6種模板。"""
        paths = []
        for name in self.TEMPLATES:
            p = self.create_template(name, output_dir)
            if p:
                paths.append(p)
        return paths

    def detect_template_type(self, file_path):
        """依據檔案標題列自動偵測模板類型。"""
        try:
            wb = load_workbook(file_path, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            wb.close()
        except Exception:
            return None

        best_match = None
        best_score = 0
        for tname, cfg in self.TEMPLATES.items():
            expected = [c[0] for c in cfg["columns"]]
            common = sum(1 for h in headers if h in expected)
            score = common / max(len(expected), 1)
            if score > best_score:
                best_score = score
                best_match = tname
        return best_match if best_score >= 0.5 else None

    def get_config(self, template_name):
        return self.TEMPLATES.get(template_name)


# ============================================================================
# DataPreprocessor
# ============================================================================
class DataPreprocessor:
    """資料驗證與清洗。"""

    def __init__(self, template_name, manager=None):
        self.template_name = template_name
        self.manager = manager or ERPTemplateManager()
        self.cfg = self.manager.get_config(template_name)
        self.errors = []
        self.warnings = []

    def _reset(self):
        self.errors = []
        self.warnings = []

    # ------------------------------------------------------------------
    def validate_required_fields(self, records):
        """檢查每筆記錄是否具備必填欄位。"""
        if not self.cfg:
            self.errors.append({"row": "-", "field": "-", "message": "未知模板類型", "severity": "錯誤"})
            return
        required = self.cfg["required"]
        headers = [c[0] for c in self.cfg["columns"]]
        for idx, rec in enumerate(records, 2):
            for field in required:
                if field in headers:
                    col_idx = headers.index(field)
                    val = rec[col_idx] if col_idx < len(rec) else None
                    if val is None or (isinstance(val, str) and val.strip() == ""):
                        self.errors.append({
                            "row": idx, "field": field,
                            "message": f"第{idx}列: 必填欄位「{field}」為空",
                            "severity": "錯誤",
                        })

    def normalize_dates(self, records):
        """統一日期格式為 YYYY-MM-DD。"""
        if not self.cfg:
            return records
        headers = [c[0] for c in self.cfg["columns"]]
        date_fields = [h for h in headers if "日期" in h or "到期" in h or "日" == h[-1:]]
        date_patterns = [
            (r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", "%Y-%m-%d"),
            (r"(\d{1,2})[-/](\d{1,2})[-/](\d{4})", "%m-%d-%Y"),
            (r"(\d{4})(\d{2})(\d{2})", "%Y%m%d"),
        ]
        for idx, rec in enumerate(records, 2):
            for field in date_fields:
                if field in headers:
                    col_idx = headers.index(field)
                    if col_idx < len(rec):
                        val = rec[col_idx]
                        if val is None or str(val).strip() == "":
                            continue
                        if isinstance(val, datetime.datetime):
                            rec[col_idx] = val.strftime("%Y-%m-%d")
                            continue
                        s = str(val).strip()
                        matched = False
                        for pat, fmt in date_patterns:
                            m = re.match(pat, s)
                            if m:
                                try:
                                    dt = datetime.datetime.strptime(
                                        m.group(0), fmt if "%" not in pat else pat.replace("(", "").replace(")", "")
                                    )
                                    rec[col_idx] = dt.strftime("%Y-%m-%d")
                                    matched = True
                                    break
                                except ValueError:
                                    try:
                                        groups = m.groups()
                                        if len(groups[0]) == 4:
                                            dt = datetime.datetime(int(groups[0]), int(groups[1]), int(groups[2]))
                                        else:
                                            dt = datetime.datetime(int(groups[2]), int(groups[0]), int(groups[1]))
                                        rec[col_idx] = dt.strftime("%Y-%m-%d")
                                        matched = True
                                        break
                                    except (ValueError, IndexError):
                                        pass
                        if not matched:
                            self.errors.append({
                                "row": idx, "field": field,
                                "message": f"第{idx}列: 日期格式無法識別「{s}」",
                                "severity": "錯誤",
                            })
        return records

    def normalize_numbers(self, records):
        """清理數值欄位（去逗號、轉數字）。"""
        if not self.cfg:
            return records
        headers = [c[0] for c in self.cfg["columns"]]
        num_keywords = ["數量", "單價", "金額", "總額", "稅額", "淨額", "淨重", "毛重"]
        num_fields = [h for h in headers if any(kw in h for kw in num_keywords)]
        for idx, rec in enumerate(records, 2):
            for field in num_fields:
                if field in headers:
                    col_idx = headers.index(field)
                    if col_idx < len(rec):
                        val = rec[col_idx]
                        if val is None or (isinstance(val, str) and val.strip() == ""):
                            continue
                        if isinstance(val, (int, float)):
                            continue
                        s = str(val).strip().replace(",", "").replace("，", "")
                        try:
                            rec[col_idx] = float(s)
                        except ValueError:
                            self.errors.append({
                                "row": idx, "field": field,
                                "message": f"第{idx}列: 數值無法轉換「{val}」",
                                "severity": "錯誤",
                            })
        return records

    def check_duplicates(self, records):
        """依據主鍵欄位偵測重複。"""
        if not self.cfg:
            return
        headers = [c[0] for c in self.cfg["columns"]]
        key_fields = self.cfg.get("key_fields", [])
        for kf in key_fields:
            if kf not in headers:
                continue
            col_idx = headers.index(kf)
            seen = {}
            for idx, rec in enumerate(records, 2):
                if col_idx < len(rec):
                    val = rec[col_idx]
                    if val in seen:
                        self.errors.append({
                            "row": idx, "field": kf,
                            "message": f"第{idx}列: {kf}「{val}」與第{seen[val]}列重複",
                            "severity": "錯誤",
                        })
                    else:
                        seen[val] = idx

    def validate_references(self, records, master_data=None):
        """交叉驗證參照完整性（PO號碼、供應商代碼等）。"""
        if not self.cfg or not master_data:
            return
        headers = [c[0] for c in self.cfg["columns"]]
        if self.template_name == "供應商發票" and "PO號碼" in headers:
            col_idx = headers.index("PO號碼")
            po_set = set()
            for rec in master_data.get("採購訂單", []):
                po_headers = [c[0] for c in self.manager.TEMPLATES["採購訂單"]["columns"]]
                if "PO號碼" in po_headers:
                    po_col = po_headers.index("PO號碼")
                    if po_col < len(rec):
                        po_set.add(rec[po_col])
            for idx, rec in enumerate(records, 2):
                if col_idx < len(rec):
                    val = rec[col_idx]
                    if val and po_set and val not in po_set:
                        self.warnings.append({
                            "row": idx, "field": "PO號碼",
                            "message": f"第{idx}列: PO號碼「{val}」未見於採購訂單主數據",
                            "severity": "警告",
                        })
        if "供應商代碼" in headers:
            col_idx = headers.index("供應商代碼")
            valid_prefixes = ["SUP-TH-", "SUP-ID-", "SUP-TW-", "SUP-CN-", "SUP-"]
            for idx, rec in enumerate(records, 2):
                if col_idx < len(rec):
                    val = rec[col_idx]
                    if val and not any(str(val).startswith(p) for p in valid_prefixes):
                        self.warnings.append({
                            "row": idx, "field": "供應商代碼",
                            "message": f"第{idx}列: 供應商代碼「{val}」格式可能不正確",
                            "severity": "警告",
                        })

    def generate_error_report(self):
        """產生驗證報告摘要。"""
        report_lines = []
        report_lines.append("=" * 60)
        report_lines.append("  資料驗證報告")
        report_lines.append("=" * 60)
        report_lines.append(f"  模板類型: {self.template_name}")
        report_lines.append(f"  錯誤數量: {len(self.errors)}")
        report_lines.append(f"  警告數量: {len(self.warnings)}")
        report_lines.append("-" * 60)
        if self.errors:
            report_lines.append("  【錯誤】")
            for e in self.errors:
                report_lines.append(f"    ✗ {e['message']}")
        if self.warnings:
            report_lines.append("  【警告】")
            for w in self.warnings:
                report_lines.append(f"    ⚠ {w['message']}")
        if not self.errors and not self.warnings:
            report_lines.append("  ✓ 所有資料通過驗證")
        report_lines.append("=" * 60)
        return "\n".join(report_lines)

    def run_full_validation(self, records, master_data=None):
        """執行完整驗證管線。"""
        self._reset()
        self.validate_required_fields(records)
        records = self.normalize_dates(records)
        records = self.normalize_numbers(records)
        self.check_duplicates(records)
        self.validate_references(records, master_data)
        return records


# ============================================================================
# RPAScriptGenerator
# ============================================================================
class RPAScriptGenerator:
    """產生RPA腳本與操作指南。"""

    @staticmethod
    def generate_pyautogui_script(records, erp_type="generic", template_name=""):
        """產生 pyautogui RPA Python腳本。"""
        lines = [
            "#!/usr/bin/env python3",
            "# -*- coding: utf-8 -*-",
            f'"""',
            f"ERP自動錄入RPA腳本",
            f"模板類型: {template_name}",
            f"記錄筆數: {len(records)}",
            f"產生時間: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f'"""',
            "",
            "import pyautogui",
            "import time",
            "import sys",
            "",
            "# 安全設定",
            "pyautogui.FAILSAFE = True",
            "pyautogui.PAUSE = 0.5",
            "",
            "# 等待使用者切換到ERP視窗",
            "print('=== ERP自動錄入RPA腳本 ===')",
            f"print('即將錄入 {len(records)} 筆「{template_name}」資料')",
            "print('請在 5 秒內切換到 ERP 系統視窗...')",
            "time.sleep(5)",
            "",
            "# 資料集",
            f"records = {repr(records)}",
            "",
            "def enter_record(idx, record):",
            '    """輸入單筆記錄。"""',
            "    print(f'正在輸入第 {{idx+1}} 筆記錄...')",
            "    for field_idx, value in enumerate(record):",
            "        if value is None or str(value).strip() == '':",
            "            pyautogui.press('tab')",
            "            continue",
            "        pyautogui.typewrite(str(value), interval=0.05)",
            "        pyautogui.press('tab')",
            "    # 儲存 (根據ERP系統調整快捷鍵)",
            "    pyautogui.hotkey('ctrl', 's')",
            "    time.sleep(1)",
            "    print(f'第 {{idx+1}} 筆記錄已儲存')",
            "",
            "def main():",
            "    try:",
            "        for idx, record in enumerate(records):",
            "            enter_record(idx, record)",
            "            time.sleep(0.5)",
            "        print(f'\\n全部完成! 共錄入 {{len(records)}} 筆記錄。')",
            "    except KeyboardInterrupt:",
            "        print('\\n使用者中斷操作。')",
            "        sys.exit(0)",
            "",
            "if __name__ == '__main__':",
            "    main()",
        ]
        return "\n".join(lines)

    @staticmethod
    def generate_power_automate_guide(records, template_name=""):
        """產生Power Automate操作指南。"""
        lines = [
            f"# Power Automate Desktop 自動錄入指南",
            f"## 模板: {template_name}  |  記錄數: {len(records)}",
            f"",
            f"## 前置準備",
            f"1. 安裝 **Power Automate Desktop** (Windows 10/11 免費版)",
            f"2. 確認 ERP 系統已開啟且可操作",
            f"3. 準備好匯入檔案（CSV 或 Excel）",
            f"",
            f"## 建立流程步驟",
            f"",
            f"### 步驟1: 讀取資料",
            f"- 使用「讀取 Excel 工作表」或「讀取 CSV 檔案」動作",
            f"- 將資料存入變數 `%DataTable%`",
            f"",
            f"### 步驟2: 迴圈處理每筆記錄",
            f"- 新增「For each」迴圈，迭代 `%DataTable%`",
            f"- 在迴圈內逐一填入欄位",
            f"",
            f"### 步驟3: 填入欄位（以通用ERP為例）",
            f"- 使用「填入視窗控制項中的文字」或「傳送按鍵」動作",
            f"- 每個欄位後按 Tab 切換到下一個輸入框",
            f"- 欄位對應:",
        ]
        if records:
            for i, col in enumerate(range(min(len(records[0]), 8))):
                lines.append(f"  - 欄位{i+1}: `%CurrentItem[{i}]%`")
        lines += [
            f"",
            f"### 步驟4: 儲存記錄",
            f"- 使用「傳送按鍵」動作: Ctrl+S 或點擊「儲存」按鈕",
            f"- 加入「等待」動作（1-2秒）確保儲存完成",
            f"",
            f"### 步驟5: 錯誤處理",
            f"- 加入「On block error」區塊",
            f"- 記錄失敗的列號以便後續重試",
            f"- 建議加入截圖動作以記錄錯誤狀態",
            f"",
            f"## 注意事項",
            f"- 首次執行建議先用 1-2 筆資料測試",
            f"- 確認 ERP 畫面解析度與定位一致",
            f"- 保持滑鼠/鍵盤不被其他程式干擾",
        ]
        return "\n".join(lines)

    @staticmethod
    def generate_uipath_guide(records, template_name=""):
        """產生UiPath Community Edition操作指南。"""
        lines = [
            f"# UiPath Community Edition 自動錄入指南",
            f"## 模板: {template_name}  |  記錄數: {len(records)}",
            f"",
            f"## 環境建置",
            f"1. 下載 UiPath Community Edition (免費)",
            f"2. 安裝 UiPath Studio",
            f"3. 建立新的 Process 專案",
            f"",
            f"## 工作流程設計",
            f"",
            f"### 1. 初始化 (Sequence)",
            f"- `Read Range` 讀取 Excel 資料 → `DataTable` 變數",
            f"- 設定 counter 變數 `rowIndex = 0`",
            f"",
            f"### 2. 主迴圈 (For Each Row in DataTable)",
            f"- `Attach Window` 鎖定 ERP 視窗",
            f"- 依序填入各欄位:",
            f"  ```",
            f"  Type Into → 欄位1: CurrentRow(0).ToString",
            f"  Type Into → 欄位2: CurrentRow(1).ToString",
            f"  ... (依此類推)",
            f"  ```",
            f"- `Click` 儲存按鈕 或 `Send Hotkey` Ctrl+S",
            f"- `Delay` 1-2秒",
            f"- `rowIndex = rowIndex + 1`",
            f"",
            f"### 3. 異常處理 (Try Catch)",
            f"- Try: 填入與儲存邏輯",
            f"- Catch: 記錄錯誤到 Log, 截圖保存",
            f"- Finally: 更新進度",
            f"",
            f"### 4. 結果輸出",
            f"- 使用 `Write Line` 輸出統計",
            f"- 將失敗記錄寫入 Excel",
            f"",
            f"## 部署建議",
            f"- 使用 UiPath Orchestrator 排程執行",
            f"- 或在 UiPath Assistant 手動觸發",
            f"- 建議在非工作時段（夜間/週末）批次執行",
        ]
        return "\n".join(lines)


# ============================================================================
# ERPBatchImporter
# ============================================================================
class ERPBatchImporter:
    """批量匯入處理器。"""

    def __init__(self):
        self.manager = ERPTemplateManager()
        self.preprocessor = None
        self.template_name = None
        self.records = []
        self.headers = []
        self.cleaned_records = []
        self.failed_records = []

    def load_template(self, file_path):
        """讀取Excel模板並自動偵測類型。"""
        self.template_name = self.manager.detect_template_type(file_path)
        if not self.template_name:
            print(f"  錯誤: 無法辨識檔案「{file_path}」的模板類型。")
            return False

        cfg = self.manager.get_config(self.template_name)
        self.preprocessor = DataPreprocessor(self.template_name, self.manager)

        try:
            wb = load_workbook(file_path, read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
        except Exception as e:
            print(f"  錯誤: 無法讀取檔案「{file_path}」- {e}")
            return False

        if not rows:
            print("  錯誤: 檔案為空。")
            return False

        self.headers = list(rows[0])
        self.records = [list(r) for r in rows[1:] if any(v is not None for v in r)]
        print(f"  已載入模板: {self.template_name}")
        print(f"  記錄筆數: {len(self.records)}")
        return True

    def preprocess(self, file_path=None):
        """執行 DataPreprocessor 管線。"""
        if not self.preprocessor:
            print("  錯誤: 請先載入模板 (load_template)。")
            return False
        master_data = {}
        for tname in self.manager.TEMPLATES:
            samples = self.manager.SAMPLE_DATA.get(tname, [])
            if samples:
                master_data[tname] = samples

        self.cleaned_records = self.preprocessor.run_full_validation(self.records, master_data)
        error_count = len(self.preprocessor.errors)
        warn_count = len(self.preprocessor.warnings)
        print(f"  驗證完成: {error_count} 個錯誤, {warn_count} 個警告")
        return error_count == 0

    def generate_import_file(self, records=None, fmt="csv", output_dir=".", filename_prefix=""):
        """產生ERP就緒的匯入檔案。"""
        recs = records if records is not None else self.cleaned_records
        if not recs:
            recs = self.records
        tname = self.template_name or "資料"
        prefix = filename_prefix or f"ERP匯入_{tname}"
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

        if fmt == "csv":
            fpath = os.path.join(output_dir, f"{prefix}_{ts}.csv")
            with open(fpath, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                if self.headers:
                    writer.writerow(self.headers)
                for rec in recs:
                    writer.writerow(rec)
        else:
            fpath = os.path.join(output_dir, f"{prefix}_{ts}.xlsx")
            wb = Workbook()
            ws = wb.active
            ws.title = tname
            cfg = self.manager.get_config(self.template_name) if self.template_name else None
            if self.headers:
                for col_idx, h in enumerate(self.headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=h)
                    cell.font = HEADER_FONT
                    cell.fill = HEADER_FILL
                    cell.alignment = HEADER_ALIGNMENT
                    cell.border = THIN_BORDER
                if cfg:
                    for col_idx, (_, w) in enumerate(cfg["columns"], 1):
                        ws.column_dimensions[get_column_letter(col_idx)].width = w
                ws.freeze_panes = "A2"
                ws.auto_filter.ref = f"A1:{get_column_letter(len(self.headers))}1"
            for row_idx, rec in enumerate(recs, 2):
                for col_idx, val in enumerate(rec, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.alignment = DATA_ALIGNMENT
                    cell.border = THIN_BORDER
            wb.save(fpath)

        print(f"  已產生匯入檔案: {fpath}")
        return fpath

    def generate_import_summary(self):
        """產生匯入統計摘要。"""
        total = len(self.records)
        errors = len(self.preprocessor.errors) if self.preprocessor else 0
        warnings = len(self.preprocessor.warnings) if self.preprocessor else 0
        passed = total - len(set(e["row"] for e in (self.preprocessor.errors if self.preprocessor else [])))
        passed = max(0, passed)
        failed = total - passed

        lines = []
        lines.append("")
        lines.append("=" * 60)
        lines.append("  ERP批量錄入 - 匯入摘要報告")
        lines.append("=" * 60)
        lines.append(f"  模板類型:     {self.template_name or '未載入'}")
        lines.append(f"  總記錄數:     {total}")
        lines.append(f"  通過驗證:     {passed}")
        lines.append(f"  未通過驗證:   {failed}")
        lines.append(f"  警告數:       {warnings}")
        lines.append(f"  通過率:       {passed/total*100:.1f}%" if total > 0 else "  通過率:       N/A")
        lines.append("-" * 60)
        if self.preprocessor and self.preprocessor.errors:
            lines.append("  錯誤明細:")
            for e in self.preprocessor.errors:
                lines.append(f"    [{e['severity']}] {e['message']}")
        if self.preprocessor and self.preprocessor.warnings:
            lines.append("  警告明細:")
            for w in self.preprocessor.warnings:
                lines.append(f"    [{w['severity']}] {w['message']}")
        lines.append("=" * 60)
        return "\n".join(lines)

    def create_rpa_script(self, records=None, output_dir="."):
        """產生RPA腳本。"""
        recs = records if records is not None else (self.cleaned_records if self.cleaned_records else self.records)
        rpa = RPAScriptGenerator()
        tname = self.template_name or "通用"

        # pyautogui
        script = rpa.generate_pyautogui_script(recs, template_name=tname)
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        script_path = os.path.join(output_dir, f"RPA腳本_{tname}_{ts}.py")
        with open(script_path, "w", encoding="utf-8") as f:
            f.write(script)
        print(f"  已產生 pyautogui RPA腳本: {script_path}")

        # Power Automate guide
        pa_guide = rpa.generate_power_automate_guide(recs, template_name=tname)
        pa_path = os.path.join(output_dir, f"Power_Automate指南_{tname}_{ts}.md")
        with open(pa_path, "w", encoding="utf-8") as f:
            f.write(pa_guide)
        print(f"  已產生 Power Automate 指南: {pa_path}")

        # UiPath guide
        uipath_guide = rpa.generate_uipath_guide(recs, template_name=tname)
        uipath_path = os.path.join(output_dir, f"UiPath指南_{tname}_{ts}.md")
        with open(uipath_path, "w", encoding="utf-8") as f:
            f.write(uipath_guide)
        print(f"  已產生 UiPath 指南: {uipath_path}")

        return script_path, pa_path, uipath_path

    def export_validation_report(self, output_dir="."):
        """產生色彩標記的Excel驗證報告。"""
        if not self.preprocessor:
            print("  錯誤: 請先執行驗證。")
            return None

        tname = self.template_name or "資料"
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        fpath = os.path.join(output_dir, f"驗證報告_{tname}_{ts}.xlsx")

        wb = Workbook()

        # 摘要頁
        ws_summary = wb.active
        ws_summary.title = "驗證摘要"
        summary_headers = ["項目", "數值"]
        for col_idx, h in enumerate(summary_headers, 1):
            cell = ws_summary.cell(row=1, column=col_idx, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

        total = len(self.records)
        err_count = len(self.preprocessor.errors)
        warn_count = len(self.preprocessor.warnings)
        passed = total - len(set(e["row"] for e in self.preprocessor.errors))
        passed = max(0, passed)

        summary_data = [
            ("模板類型", tname),
            ("總記錄數", total),
            ("通過驗證", passed),
            ("錯誤數", err_count),
            ("警告數", warn_count),
            ("通過率", f"{passed/total*100:.1f}%" if total > 0 else "N/A"),
            ("驗證時間", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ]
        for row_idx, (item, val) in enumerate(summary_data, 2):
            ws_summary.cell(row=row_idx, column=1, value=item).border = THIN_BORDER
            cell = ws_summary.cell(row=row_idx, column=2, value=val)
            cell.border = THIN_BORDER
            if item == "錯誤數" and err_count > 0:
                cell.fill = ERROR_FILL
                cell.font = ERROR_FONT
            elif item == "警告數" and warn_count > 0:
                cell.fill = WARNING_FILL
                cell.font = WARNING_FONT
            elif item == "通過驗證":
                cell.fill = OK_FILL

        ws_summary.column_dimensions["A"].width = 16
        ws_summary.column_dimensions["B"].width = 24
        ws_summary.freeze_panes = "A2"

        # 錯誤明細頁
        if self.preprocessor.errors or self.preprocessor.warnings:
            ws_detail = wb.create_sheet("問題明細")
            detail_headers = ["嚴重度", "列號", "欄位", "說明"]
            for col_idx, h in enumerate(detail_headers, 1):
                cell = ws_detail.cell(row=1, column=col_idx, value=h)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = HEADER_ALIGNMENT
                cell.border = THIN_BORDER

            row_idx = 2
            all_issues = self.preprocessor.errors + self.preprocessor.warnings
            for issue in all_issues:
                ws_detail.cell(row=row_idx, column=1, value=issue["severity"]).border = THIN_BORDER
                ws_detail.cell(row=row_idx, column=2, value=issue["row"]).border = THIN_BORDER
                ws_detail.cell(row=row_idx, column=3, value=issue["field"]).border = THIN_BORDER
                cell = ws_detail.cell(row=row_idx, column=4, value=issue["message"])
                cell.border = THIN_BORDER
                if issue["severity"] == "錯誤":
                    for c in range(1, 5):
                        ws_detail.cell(row=row_idx, column=c).fill = ERROR_FILL
                else:
                    for c in range(1, 5):
                        ws_detail.cell(row=row_idx, column=c).fill = WARNING_FILL
                row_idx += 1

            ws_detail.column_dimensions["A"].width = 10
            ws_detail.column_dimensions["B"].width = 8
            ws_detail.column_dimensions["C"].width = 18
            ws_detail.column_dimensions["D"].width = 60
            ws_detail.freeze_panes = "A2"

        # 原始資料頁（標記問題列）
        ws_data = wb.create_sheet("原始資料")
        error_rows = set()
        for e in self.preprocessor.errors:
            if isinstance(e["row"], int):
                error_rows.add(e["row"])
        warn_rows = set()
        for w in self.preprocessor.warnings:
            if isinstance(w["row"], int):
                warn_rows.add(w["row"])

        if self.headers:
            for col_idx, h in enumerate(self.headers, 1):
                cell = ws_data.cell(row=1, column=col_idx, value=h)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = HEADER_ALIGNMENT
                cell.border = THIN_BORDER

        for row_idx, rec in enumerate(self.records, 2):
            for col_idx, val in enumerate(rec, 1):
                cell = ws_data.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = DATA_ALIGNMENT
                cell.border = THIN_BORDER
                if row_idx in error_rows:
                    cell.fill = ERROR_FILL
                elif row_idx in warn_rows:
                    cell.fill = WARNING_FILL

        if self.headers:
            for col_idx in range(1, len(self.headers) + 1):
                ws_data.column_dimensions[get_column_letter(col_idx)].width = 18
            ws_data.freeze_panes = "A2"
            ws_data.auto_filter.ref = f"A1:{get_column_letter(len(self.headers))}1"

        wb.save(fpath)
        print(f"  已產生驗證報告: {fpath}")
        return fpath


# ============================================================================
# ERPImportTool - 主程式 CLI
# ============================================================================
class ERPImportTool:
    """ERP批量錄入工具 - CLI 主程式。"""

    def __init__(self):
        self.manager = ERPTemplateManager()
        self.importer = ERPBatchImporter()
        self.output_dir = "."

    def cmd_create_templates(self, output_dir="."):
        """建立所有模板。"""
        self.output_dir = output_dir
        print("\n" + "=" * 60)
        print("  建立ERP批量錄入模板")
        print("=" * 60)
        paths = self.manager.create_all_templates(output_dir)
        print(f"\n  已建立 {len(paths)} 個模板檔案:")
        for p in paths:
            print(f"    - {p}")
        print()
        return paths

    def cmd_validate(self, file_path):
        """驗證資料檔案。"""
        print("\n" + "=" * 60)
        print("  資料驗證")
        print("=" * 60)
        if not self.importer.load_template(file_path):
            return False
        self.importer.preprocess()
        report = self.importer.preprocessor.generate_error_report()
        print(report)
        self.importer.export_validation_report(self.output_dir)
        return True

    def cmd_process(self, file_path, output_dir="."):
        """完整管線: 驗證 → 清洗 → 產生匯入檔。"""
        self.output_dir = output_dir
        print("\n" + "=" * 60)
        print("  ERP批量錄入 - 完整處理管線")
        print("=" * 60)
        print("\n[1/4] 載入模板...")
        if not self.importer.load_template(file_path):
            return False
        print("\n[2/4] 資料驗證與清洗...")
        clean = self.importer.preprocess()
        print("\n[3/4] 產生匯入檔案...")
        self.importer.generate_import_file(fmt="csv", output_dir=output_dir)
        self.importer.generate_import_file(fmt="xlsx", output_dir=output_dir)
        print("\n[4/4] 產生驗證報告...")
        self.importer.export_validation_report(output_dir)
        summary = self.importer.generate_import_summary()
        print(summary)
        return True

    def cmd_rpa_script(self, file_path, output_dir="."):
        """產生RPA腳本。"""
        self.output_dir = output_dir
        print("\n" + "=" * 60)
        print("  產生RPA自動錄入腳本")
        print("=" * 60)
        if not self.importer.load_template(file_path):
            return False
        self.importer.preprocess()
        self.importer.create_rpa_script(output_dir=output_dir)
        return True

    def cmd_summary(self, file_path):
        """顯示匯入摘要。"""
        print("\n" + "=" * 60)
        print("  匯入摘要統計")
        print("=" * 60)
        if not self.importer.load_template(file_path):
            return False
        self.importer.preprocess()
        summary = self.importer.generate_import_summary()
        print(summary)
        return True

    def cmd_demo(self, output_dir="."):
        """執行完整Demo。"""
        self.output_dir = output_dir
        print("")
        print("*" * 64)
        print("*" + " " * 62 + "*")
        print("*" + "  ERP批量自動錄入工具 - 完整Demo".ljust(58) + "    *")
        print("*" + "  適用範圍: 泰國/印尼 進出口貿易".ljust(56) + "      *")
        print("*" + " " * 62 + "*")
        print("*" * 64)

        # ---------------------------------------------------------------
        # 步驟1: 建立所有模板
        # ---------------------------------------------------------------
        print("\n" + "=" * 64)
        print("  步驟1: 建立所有6種資料蒐集模板")
        print("=" * 64)
        paths = self.manager.create_all_templates(output_dir)
        print(f"\n  已建立 {len(paths)} 個模板檔案:")
        for p in paths:
            print(f"    - {os.path.basename(p)}")

        # ---------------------------------------------------------------
        # 步驟2: 載入採購訂單模板
        # ---------------------------------------------------------------
        print("\n" + "=" * 64)
        print("  步驟2: 載入「採購訂單」模板")
        print("=" * 64)
        po_path = os.path.join(output_dir, "採購訂單PO模板.xlsx")
        importer = ERPBatchImporter()
        if not importer.load_template(po_path):
            print("  Demo中斷: 無法載入採購訂單模板。")
            return

        # ---------------------------------------------------------------
        # 步驟3: 引入故意錯誤
        # ---------------------------------------------------------------
        print("\n" + "=" * 64)
        print("  步驟3: 模擬人為輸入錯誤（測試驗證能力）")
        print("=" * 64)
        records = importer.records

        # 錯誤1: 清空一筆的PO號碼（必填缺漏）
        if len(records) > 1:
            records[1][0] = ""  # 第2列PO號碼清空
            print("  >> 已引入錯誤1: 第3列 PO號碼設為空白（必填缺漏）")

        # 錯誤2: 新增一筆重複PO號碼
        dup_record = list(records[0])  # 複製第1筆
        records.append(dup_record)
        print(f"  >> 已引入錯誤2: 新增重複PO號碼「{dup_record[0]}」")

        # 錯誤3: 無效日期
        if len(records) > 2:
            records[2][10] = "2026/13/45"  # 無效月份與日期
            print("  >> 已引入錯誤3: 第4列 交貨日期設為「2026/13/45」（無效日期）")

        print(f"\n  目前總記錄數: {len(records)}")

        # ---------------------------------------------------------------
        # 步驟4: 執行驗證
        # ---------------------------------------------------------------
        print("\n" + "=" * 64)
        print("  步驟4: 執行資料驗證")
        print("=" * 64)
        importer.records = records
        importer.preprocess()
        report = importer.preprocessor.generate_error_report()
        print(report)

        # ---------------------------------------------------------------
        # 步驟5: 修正錯誤並重新驗證
        # ---------------------------------------------------------------
        print("\n" + "=" * 64)
        print("  步驟5: 修正錯誤並產生乾淨的匯入檔案")
        print("=" * 64)

        # 修正: 移除有問題的記錄, 重新載入乾淨資料
        print("  >> 移除錯誤記錄, 重新載入乾淨的採購訂單資料...")
        clean_importer = ERPBatchImporter()
        clean_importer.load_template(po_path)
        clean_importer.preprocess()

        print("\n  產生乾淨的匯入檔案:")
        csv_path = clean_importer.generate_import_file(fmt="csv", output_dir=output_dir, filename_prefix="乾淨匯入_採購訂單")
        xlsx_path = clean_importer.generate_import_file(fmt="xlsx", output_dir=output_dir, filename_prefix="乾淨匯入_採購訂單")

        # ---------------------------------------------------------------
        # 步驟6: 產生RPA指南
        # ---------------------------------------------------------------
        print("\n" + "=" * 64)
        print("  步驟6: 產生RPA自動錄入指南文件")
        print("=" * 64)
        rpa_paths = clean_importer.create_rpa_script(output_dir=output_dir)

        # ---------------------------------------------------------------
        # 步驟7: 驗證報告與摘要
        # ---------------------------------------------------------------
        print("\n" + "=" * 64)
        print("  步驟7: 產生驗證報告與統計摘要")
        print("=" * 64)

        # 含錯誤的報告
        print("\n  [含錯誤版本的驗證報告]")
        importer.export_validation_report(output_dir)

        # 乾淨版的摘要
        print("\n  [乾淨版本的統計摘要]")
        summary = clean_importer.generate_import_summary()
        print(summary)

        # ---------------------------------------------------------------
        # 完成
        # ---------------------------------------------------------------
        print("\n" + "=" * 64)
        print("  Demo完成! 輸出檔案總覽:")
        print("=" * 64)
        output_files = []
        for f in os.listdir(output_dir):
            if f.endswith((".xlsx", ".csv", ".py", ".md")) and not f.startswith("."):
                fpath = os.path.join(output_dir, f)
                size = os.path.getsize(fpath)
                output_files.append((f, size))

        output_files.sort()
        for fname, size in output_files:
            size_str = f"{size:,} bytes" if size < 1024 else f"{size/1024:.1f} KB"
            print(f"    {fname:<48s} {size_str:>12s}")

        print(f"\n  共產生 {len(output_files)} 個檔案。")
        print("\n" + "*" * 64)
        print("  感謝使用 ERP批量自動錄入工具!")
        print("*" * 64 + "\n")


# ============================================================================
# CLI argparse
# ============================================================================
def main():
    parser = argparse.ArgumentParser(
        description="ERP批量自動錄入工具 - 泰國/印尼進出口貿易",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=textwrap.dedent("""\
            使用範例:
              python3 ERP批量錄入工具.py --create-templates
              python3 ERP批量錄入工具.py --validate 採購訂單PO模板.xlsx
              python3 ERP批量錄入工具.py --process 採購訂單PO模板.xlsx
              python3 ERP批量錄入工具.py --rpa-script 採購訂單PO模板.xlsx
              python3 ERP批量錄入工具.py --summary 採購訂單PO模板.xlsx
              python3 ERP批量錄入工具.py --demo
        """),
    )
    parser.add_argument("--create-templates", action="store_true", help="建立所有6種Excel模板（含範例資料）")
    parser.add_argument("--validate", metavar="FILE", help="驗證資料檔案")
    parser.add_argument("--process", metavar="FILE", help="完整處理: 驗證→清洗→產生匯入檔")
    parser.add_argument("--rpa-script", metavar="FILE", help="產生RPA自動錄入腳本")
    parser.add_argument("--summary", metavar="FILE", help="顯示匯入摘要統計")
    parser.add_argument("--demo", action="store_true", help="執行完整Demo流程")
    parser.add_argument("--output-dir", default=".", help="輸出目錄 (預設: 目前目錄)")

    args = parser.parse_args()
    tool = ERPImportTool()
    output_dir = args.output_dir

    if args.create_templates:
        tool.cmd_create_templates(output_dir)
    elif args.validate:
        tool.output_dir = output_dir
        tool.cmd_validate(args.validate)
    elif args.process:
        tool.cmd_process(args.process, output_dir)
    elif args.rpa_script:
        tool.cmd_rpa_script(args.rpa_script, output_dir)
    elif args.summary:
        tool.cmd_summary(args.summary)
    elif args.demo:
        tool.cmd_demo(output_dir)
    else:
        parser.print_help()
        print("\n提示: 使用 --demo 可查看完整功能演示。")


if __name__ == "__main__":
    main()
