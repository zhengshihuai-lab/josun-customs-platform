#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
訂艙比價自動化工具 (Freight Booking & Price Comparison Automation)
報關部門專用 — 海運費比價、貨代管理、詢價自動生成
"""

import argparse
import json
import os
import sys
from datetime import datetime, timedelta
from statistics import mean, stdev

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
        numbers,
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print("錯誤：缺少 openpyxl 套件，請執行 pip install openpyxl")
    sys.exit(1)

# ── 常數 ──────────────────────────────────────────────────────────────────────
DEEP_BLUE = "1F4E79"
WHITE = "FFFFFF"
LIGHT_BLUE = "D6E4F0"
LIGHT_GRAY = "F2F2F2"
ACCENT_GREEN = "70AD47"
ACCENT_RED = "FF4444"
ACCENT_ORANGE = "FFA500"

HEADER_FILL = PatternFill(start_color=DEEP_BLUE, end_color=DEEP_BLUE, fill_type="solid")
HEADER_FONT = Font(name="Microsoft JhengHei", bold=True, color=WHITE, size=11)
ALT_FILL = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(SCRIPT_DIR, "貨代報價數據庫.xlsx")

# ── 輔助函式 ──────────────────────────────────────────────────────────────────

def fmt_usd(val):
    """格式化美元金額：$1,234"""
    if val is None:
        return "N/A"
    return f"${val:,.0f}"


def style_sheet(ws, header_count=None):
    """為工作表套用深藍色標題樣式、凍結窗格、自動篩選"""
    if ws.max_row is None or ws.max_row < 1:
        return
    cols = ws.max_column
    if header_count is None:
        header_count = cols

    # 標題列樣式
    for col_idx in range(1, cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # 資料列樣式（交替底色）
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = THIN_BORDER
            cell.alignment = CENTER
            if row_idx % 2 == 0:
                cell.fill = ALT_FILL

    # 自動欄寬
    for col_idx in range(1, cols + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                # 粗略估計：中文字元算 2 寬度
                length = sum(2 if ord(c) > 127 else 1 for c in str(val))
                max_len = max(max_len, length)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # 凍結首行
    ws.freeze_panes = "A2"

    # 自動篩選
    ws.auto_filter.ref = f"A1:{get_column_letter(cols)}{ws.max_row}"


# ══════════════════════════════════════════════════════════════════════════════
# 1. FreightForwarderDB — 貨代資料庫
# ══════════════════════════════════════════════════════════════════════════════

SAMPLE_FORWARDERS = [
    {
        "貨代名稱": "上海遠洋國際物流",
        "聯絡人": "張偉",
        "WeCom": "zhangwei_logistics",
        "電話": "+86-21-5888-1001",
        "郵箱": "zhangwei@shyy-logistics.com",
        "優勢航線": "Shanghai→Laem Chabang, Shanghai→Bangkok",
        "服務評分": 4.5,
        "合作年資": 8,
        "備註": "泰國線老牌貨代，艙位穩定",
    },
    {
        "貨代名稱": "寧波海通貨運代理",
        "聯絡人": "李娜",
        "WeCom": "lina_haitong",
        "電話": "+86-574-8765-2200",
        "郵箱": "lina@nb-haitong.com",
        "優勢航線": "Ningbo→Bangkok, Ningbo→Ho Chi Minh",
        "服務評分": 4.2,
        "合作年資": 5,
        "備註": "寧波口岸操作效率高",
    },
    {
        "貨代名稱": "深圳鵬程國際貨代",
        "聯絡人": "王磊",
        "WeCom": "wanglei_pengcheng",
        "電話": "+86-755-2600-3300",
        "郵箱": "wanglei@sz-pengcheng.com",
        "優勢航線": "Shenzhen→Jakarta, Shenzhen→Surabaya",
        "服務評分": 4.7,
        "合作年資": 10,
        "備註": "印尼線首選，價格競爭力強",
    },
    {
        "貨代名稱": "環太平洋航運有限公司",
        "聯絡人": "陳曉明",
        "WeCom": "chenxm_pacific",
        "電話": "+86-21-3456-7890",
        "郵箱": "cxm@transpacific-shipping.com",
        "優勢航線": "Shanghai→Jakarta, Shanghai→Surabaya",
        "服務評分": 4.0,
        "合作年資": 6,
        "備註": "東南亞航線覆蓋廣",
    },
    {
        "貨代名稱": "東盟物流集團",
        "聯絡人": "黃建華",
        "WeCom": "huangjh_asean",
        "電話": "+66-2-678-9012",
        "郵箱": "huangjh@asean-logistics.com",
        "優勢航線": "Shanghai→Laem Chabang, Ningbo→Bangkok",
        "服務評分": 4.8,
        "合作年資": 12,
        "備註": "泰國本地清關能力強，評分最高",
    },
    {
        "貨代名稱": "華南國際船務代理",
        "聯絡人": "劉芳",
        "WeCom": "liufang_huanan",
        "電話": "+86-755-8888-4400",
        "郵箱": "liufang@hn-shipagent.com",
        "優勢航線": "Shenzhen→Jakarta, Shanghai→Jakarta",
        "服務評分": 3.8,
        "合作年資": 3,
        "備註": "新合作貨代，價格有優勢",
    },
    {
        "貨代名稱": "中遠海運物流(東南亞)",
        "聯絡人": "趙剛",
        "WeCom": "zhaogang_cosco",
        "電話": "+86-21-6596-1122",
        "郵箱": "zhaogang@cosco-sea.com",
        "優勢航線": "Shanghai→Laem Chabang, Shanghai→Jakarta, Ningbo→Bangkok",
        "服務評分": 4.3,
        "合作年資": 15,
        "備註": "大型央企背景，信譽佳",
    },
    {
        "貨代名稱": "嘉里國際貨運",
        "聯絡人": "林雅琪",
        "WeCom": "linyq_kerry",
        "電話": "+852-2888-6600",
        "郵箱": "linyq@kerry-freight.com",
        "優勢航線": "Shanghai→Surabaya, Shenzhen→Jakarta, Shanghai→Jakarta",
        "服務評分": 4.6,
        "合作年資": 9,
        "備註": "跨國集團，服務網絡完善",
    },
]


class FreightForwarderDB:
    """貨代資料庫管理"""

    def __init__(self):
        self.forwarders = list(SAMPLE_FORWARDERS)

    # ── 新增貨代 ──
    def add_forwarder(self, forwarder_dict):
        required = ["貨代名稱", "聯絡人", "優勢航線", "服務評分"]
        for key in required:
            if key not in forwarder_dict:
                raise ValueError(f"缺少必填欄位：{key}")
        forwarder_dict.setdefault("WeCom", "")
        forwarder_dict.setdefault("電話", "")
        forwarder_dict.setdefault("郵箱", "")
        forwarder_dict.setdefault("合作年資", 0)
        forwarder_dict.setdefault("備註", "")
        self.forwarders.append(forwarder_dict)
        return forwarder_dict["貨代名稱"]

    # ── 依航線查詢 ──
    def get_forwarders_by_route(self, origin, destination):
        route_key = f"{origin}→{destination}"
        result = []
        for f in self.forwarders:
            routes = f.get("優勢航線", "")
            if route_key in routes or f"{origin}→{destination}" in routes:
                result.append(f)
        # 也做模糊比對
        if not result:
            for f in self.forwarders:
                routes = f.get("優勢航線", "")
                if origin.lower() in routes.lower() and destination.lower() in routes.lower():
                    result.append(f)
        return sorted(result, key=lambda x: x.get("服務評分", 0), reverse=True)

    # ── 評分 ──
    def rate_forwarder(self, name, new_rating):
        if not 1 <= new_rating <= 5:
            raise ValueError("評分須介於 1-5 之間")
        for f in self.forwarders:
            if f["貨代名稱"] == name:
                f["服務評分"] = round(new_rating, 1)
                return True
        return False

    # ── 評分排行 ──
    def get_top_rated(self, n=5):
        return sorted(self.forwarders, key=lambda x: x.get("服務評分", 0), reverse=True)[:n]

    # ── 顯示表格 ──
    def display_table(self, forwarder_list=None):
        if forwarder_list is None:
            forwarder_list = self.forwarders
        if not forwarder_list:
            print("  （無資料）")
            return
        headers = ["貨代名稱", "聯絡人", "WeCom", "電話", "優勢航線", "服務評分", "合作年資", "備註"]
        col_widths = [22, 8, 20, 18, 42, 8, 8, 28]
        sep = "┼".join("─" * w for w in col_widths)
        header_line = "│".join(h.center(w) for h, w in zip(headers, col_widths))
        print(f"  {'│'.join('─' * w for w in col_widths)}")
        print(f"  │{header_line}│")
        print(f"  {sep}")
        for f in forwarder_list:
            row = "│".join(
                str(f.get(h, "")).center(w) for h, w in zip(headers, col_widths)
            )
            print(f"  │{row}│")
        print(f"  {'│'.join('─' * w for w in col_widths)}")


# ══════════════════════════════════════════════════════════════════════════════
# 2. PriceHistoryDB — 歷史報價資料庫
# ══════════════════════════════════════════════════════════════════════════════

def _build_sample_quotes():
    """產生 30+ 筆歷史報價資料"""
    routes = [
        ("Shanghai", "Laem Chabang", {"20GP": (700, 850), "40GP": (900, 1100), "40HQ": (950, 1200)}),
        ("Shanghai", "Jakarta", {"20GP": (600, 780), "40GP": (800, 1050), "40HQ": (850, 1100)}),
        ("Shanghai", "Surabaya", {"20GP": (650, 820), "40GP": (850, 1080), "40HQ": (900, 1150)}),
        ("Ningbo", "Bangkok", {"20GP": (680, 830), "40GP": (880, 1070), "40HQ": (920, 1130)}),
        ("Shenzhen", "Jakarta", {"20GP": (580, 750), "40GP": (780, 1000), "40HQ": (820, 1060)}),
    ]
    forwarders_names = [
        "上海遠洋國際物流", "寧波海通貨運代理", "深圳鵬程國際貨運",
        "環太平洋航運有限公司", "東盟物流集團", "華南國際船務代理",
        "中遠海運物流(東南亞)", "嘉里國際貨運",
    ]
    months = [
        ("2026-01", 15), ("2026-02", 12), ("2026-03", 18),
        ("2026-04", 14), ("2026-05", 20), ("2026-06", 10),
    ]
    import random
    random.seed(42)  # 可重現

    quotes = []
    quote_id = 1
    for (origin, dest, price_ranges) in routes:
        for (month, _) in months:
            # 每個航線每月 1~2 筆
            count = random.randint(1, 2)
            for _ in range(count):
                container = random.choice(list(price_ranges.keys()))
                low, high = price_ranges[container]
                ocean = random.randint(low, high)
                thc = random.randint(95, 145)
                doc_fee = random.randint(35, 65)
                other = random.randint(0, 40)
                total = ocean + thc + doc_fee + other
                transit = random.randint(5, 12)
                day = random.randint(1, 28)
                date_str = f"{month}-{day:02d}"
                valid_days = random.choice([7, 14, 30])
                fwd = random.choice(forwarders_names)
                quotes.append({
                    "報價編號": f"Q{quote_id:04d}",
                    "日期": date_str,
                    "起運港": origin,
                    "目的港": dest,
                    "櫃型": container,
                    "貨代": fwd,
                    "海運費(USD)": ocean,
                    "THC": thc,
                    "文件費": doc_fee,
                    "其他費用": other,
                    "總費用": total,
                    "Transit Time(天)": transit,
                    "有效期": f"{valid_days}天",
                    "備註": random.choice(["", "旺季附加費", "含ISPS", "限時優惠", "含AMS申報", ""]),
                })
                quote_id += 1
    return quotes


SAMPLE_QUOTES = _build_sample_quotes()


class PriceHistoryDB:
    """歷史報價資料庫"""

    def __init__(self):
        self.quotes = list(SAMPLE_QUOTES)

    def add_quote(self, quote_dict):
        required = ["日期", "起運港", "目的港", "櫃型", "貨代", "海運費(USD)", "總費用"]
        for key in required:
            if key not in quote_dict:
                raise ValueError(f"缺少必填欄位：{key}")
        quote_dict.setdefault("報價編號", f"Q{len(self.quotes)+1:04d}")
        quote_dict.setdefault("THC", 0)
        quote_dict.setdefault("文件費", 0)
        quote_dict.setdefault("其他費用", 0)
        quote_dict.setdefault("Transit Time(天)", 0)
        quote_dict.setdefault("有效期", "14天")
        quote_dict.setdefault("備註", "")
        self.quotes.append(quote_dict)
        return quote_dict["報價編號"]

    def get_quotes_by_route(self, origin, destination, container_type=None):
        results = []
        for q in self.quotes:
            if q["起運港"].lower() == origin.lower() and q["目的港"].lower() == destination.lower():
                if container_type is None or q["櫃型"] == container_type:
                    results.append(q)
        return sorted(results, key=lambda x: x["日期"])

    def get_price_trend(self, origin, destination, container_type="40GP"):
        quotes = self.get_quotes_by_route(origin, destination, container_type)
        if not quotes:
            return []
        # 按月聚合
        monthly = {}
        for q in quotes:
            month = q["日期"][:7]
            monthly.setdefault(month, []).append(q["總費用"])
        trend = []
        for month in sorted(monthly.keys()):
            prices = monthly[month]
            trend.append({
                "月份": month,
                "平均費用": round(mean(prices), 0),
                "最低費用": min(prices),
                "最高費用": max(prices),
                "報價數量": len(prices),
            })
        return trend

    def get_average_price(self, origin, destination, container_type="40GP"):
        quotes = self.get_quotes_by_route(origin, destination, container_type)
        if not quotes:
            return None
        prices = [q["總費用"] for q in quotes]
        return {
            "平均": round(mean(prices), 2),
            "最低": min(prices),
            "最高": max(prices),
            "標準差": round(stdev(prices), 2) if len(prices) > 1 else 0,
            "樣本數": len(prices),
        }

    def get_best_price(self, origin, destination, container_type="40GP"):
        quotes = self.get_quotes_by_route(origin, destination, container_type)
        if not quotes:
            return None
        return min(quotes, key=lambda x: x["總費用"])


# ══════════════════════════════════════════════════════════════════════════════
# 3. InquiryGenerator — 標準化詢價訊息
# ══════════════════════════════════════════════════════════════════════════════

class InquiryGenerator:
    """標準化詢價訊息產生器"""

    _INQUIRY_COUNTER = 0

    @classmethod
    def _next_id(cls):
        cls._INQUIRY_COUNTER += 1
        return f"INQ-{datetime.now().strftime('%Y%m%d')}-{cls._INQUIRY_COUNTER:03d}"

    # ── FCL 詢價 ──
    def generate_inquiry(self, origin, destination, cargo_info, container_type="40GP",
                         ship_date="", inquiry_type="FCL"):
        inquiry_id = self._next_id()
        if inquiry_type == "FCL":
            return self._fcl_template(inquiry_id, origin, destination, cargo_info,
                                      container_type, ship_date)
        elif inquiry_type == "LCL":
            return self._lcl_template(inquiry_id, origin, destination, cargo_info, ship_date)
        elif inquiry_type == "SPECIAL":
            return self._special_template(inquiry_id, origin, destination, cargo_info,
                                          container_type, ship_date)
        else:
            raise ValueError(f"不支援的詢價類型：{inquiry_type}")

    def _fcl_template(self, iid, origin, destination, cargo, container, ship_date):
        return (
            f"══════════════════════════════════════════\n"
            f"  【整櫃(FCL)詢價單】 編號：{iid}\n"
            f"══════════════════════════════════════════\n"
            f"  日期：{datetime.now().strftime('%Y-%m-%d')}\n"
            f"  起運港：{origin}\n"
            f"  目的港：{destination}\n"
            f"  貨物描述：{cargo}\n"
            f"  櫃型：{container}\n"
            f"  預計出貨日：{ship_date}\n"
            f"──────────────────────────────────────────\n"
            f"  請提供以下費用明細：\n"
            f"    1. 海運費 (Ocean Freight)\n"
            f"    2. THC (碼頭操作費)\n"
            f"    3. 文件費 (Documentation Fee)\n"
            f"    4. 其他附加費用（請列明）\n"
            f"    5. Transit Time（航程天數）\n"
            f"    6. 報價有效期\n"
            f"    7. 最近可訂艙船期\n"
            f"──────────────────────────────────────────\n"
            f"  備註：請於 3 個工作天內回覆，謝謝。\n"
            f"══════════════════════════════════════════"
        )

    def _lcl_template(self, iid, origin, destination, cargo, ship_date):
        return (
            f"══════════════════════════════════════════\n"
            f"  【拼箱(LCL)詢價單】 編號：{iid}\n"
            f"══════════════════════════════════════════\n"
            f"  日期：{datetime.now().strftime('%Y-%m-%d')}\n"
            f"  起運港：{origin}\n"
            f"  目的港：{destination}\n"
            f"  貨物描述：{cargo}\n"
            f"  預計出貨日：{ship_date}\n"
            f"──────────────────────────────────────────\n"
            f"  請提供以下費用明細：\n"
            f"    1. 拼箱海運費 (per CBM / per W/M)\n"
            f"    2. CFS 費用（倉庫操作費）\n"
            f"    3. 文件費\n"
            f"    4. 其他附加費用\n"
            f"    5. Transit Time\n"
            f"    6. 報價有效期\n"
            f"══════════════════════════════════════════"
        )

    def _special_template(self, iid, origin, destination, cargo, container, ship_date):
        return (
            f"══════════════════════════════════════════\n"
            f"  【特種貨物詢價單】 編號：{iid}\n"
            f"══════════════════════════════════════════\n"
            f"  日期：{datetime.now().strftime('%Y-%m-%d')}\n"
            f"  起運港：{origin}\n"
            f"  目的港：{destination}\n"
            f"  貨物描述：{cargo}\n"
            f"  櫃型需求：{container}\n"
            f"  預計出貨日：{ship_date}\n"
            f"──────────────────────────────────────────\n"
            f"  特殊要求：\n"
            f"    □ 危險品 (DG Cargo)  □ 冷凍櫃 (Reefer)\n"
            f"    □ 開頂櫃 (Open Top)  □ 框架櫃 (Flat Rack)\n"
            f"    □ 超重貨物          □ 其他：_________\n"
            f"──────────────────────────────────────────\n"
            f"  請提供：\n"
            f"    1. 海運費 + 特殊附加費\n"
            f"    2. 特殊操作費用明細\n"
            f"    3. 所需文件（DG 申報 / 溫度記錄等）\n"
            f"    4. Transit Time & 報價有效期\n"
            f"══════════════════════════════════════════"
        )

    # ── 批量詢價 ──
    def generate_batch_inquiry(self, inquiry_params, forwarder_list):
        results = []
        inquiry_msg = self.generate_inquiry(**inquiry_params)
        for fwd in forwarder_list:
            results.append({
                "貨代": fwd["貨代名稱"],
                "聯絡人": fwd["聯絡人"],
                "WeCom": fwd.get("WeCom", ""),
                "詢價訊息": inquiry_msg,
            })
        return results

    # ── 追蹤催詢 ──
    def generate_follow_up(self, inquiry_id, days_pending):
        urgency = "普通"
        if days_pending >= 5:
            urgency = "緊急"
        elif days_pending >= 3:
            urgency = "一般催促"
        return (
            f"══════════════════════════════════════════\n"
            f"  【詢價追蹤 — {urgency}】\n"
            f"══════════════════════════════════════════\n"
            f"  詢價編號：{inquiry_id}\n"
            f"  已等待天數：{days_pending} 天\n"
            f"   urgency：{urgency}\n"
            f"──────────────────────────────────────────\n"
            f"  您好，\n"
            f"  關於上述詢價單（{inquiry_id}），\n"
            f"  目前尚未收到貴司回覆。\n"
            f"  因出貨日期臨近，懇請儘速提供報價。\n"
            f"  如有任何問題請隨時聯繫，謝謝。\n"
            f"══════════════════════════════════════════"
        )


# ══════════════════════════════════════════════════════════════════════════════
# 4. PriceComparator — 比價分析引擎
# ══════════════════════════════════════════════════════════════════════════════

class PriceComparator:
    """比價與分析"""

    def __init__(self, price_db=None):
        self.price_db = price_db or PriceHistoryDB()

    # ── 解析報價 ──
    def parse_quote(self, quote_input):
        if isinstance(quote_input, dict):
            q = dict(quote_input)
            # 確保總費用有值
            if "總費用" not in q:
                q["總費用"] = (
                    q.get("海運費(USD)", 0)
                    + q.get("THC", 0)
                    + q.get("文件費", 0)
                    + q.get("其他費用", 0)
                )
            return q
        # 嘗試 JSON 字串
        try:
            data = json.loads(quote_input)
            return self.parse_quote(data)
        except (json.JSONDecodeError, TypeError):
            pass
        return None

    # ── 並排比較表 ──
    def compare_quotes(self, quotes):
        parsed = [self.parse_quote(q) for q in quotes]
        parsed = [p for p in parsed if p is not None]
        if not parsed:
            print("  （無可比較的報價）")
            return parsed

        headers = ["項目"]
        for i, q in enumerate(parsed):
            name = q.get("貨代", f"報價 {i+1}")
            headers.append(name[:14])

        rows_data = [
            ("海運費(USD)", "海運費(USD)"),
            ("THC", "THC"),
            ("文件費", "文件費"),
            ("其他費用", "其他費用"),
            ("總費用", "總費用"),
            ("Transit Time(天)", "Transit Time(天)"),
        ]

        col_w = [16] + [16] * len(parsed)
        sep = "┼".join("─" * w for w in col_w)

        def _fmt(label, val):
            if label == "總費用":
                return fmt_usd(val)
            if label == "Transit Time(天)":
                return f"{val} 天"
            return fmt_usd(val) if isinstance(val, (int, float)) else str(val)

        print(f"  {'│'.join('─' * w for w in col_w)}")
        print(f"  │{'│'.join(h.center(w) for h, w in zip(headers, col_w))}│")
        print(f"  {sep}")
        for label, key in rows_data:
            cells = [label.center(col_w[0])]
            for q in parsed:
                val = q.get(key, "N/A")
                cells.append(_fmt(label, val).center(col_w[1]))
            print(f"  │{'│'.join(cells)}│")
        print(f"  {'│'.join('─' * w for w in col_w)}")

        # 找出最低
        min_total = min(parsed, key=lambda x: x.get("總費用", float("inf")))
        min_transit = min(parsed, key=lambda x: x.get("Transit Time(天)", 999))
        print(f"\n  >>> 最低總費用：{min_total.get('貨代', '未知')}（{fmt_usd(min_total.get('總費用', 0))}）")
        print(f"  >>> 最短航程  ：{min_transit.get('貨代', '未知')}（{min_transit.get('Transit Time(天)', '?')} 天）")
        return parsed

    # ── 價格合理性分析 ──
    def analyze_fairness(self, quote, origin, destination):
        q = self.parse_quote(quote)
        if not q:
            return None
        container = q.get("櫃型", "40GP")
        stats = self.price_db.get_average_price(origin, destination, container)
        if not stats:
            return {"判定": "無歷史資料可比對", "報價": q.get("總費用", 0)}

        price = q.get("總費用", 0)
        avg = stats["平均"]
        diff_pct = ((price - avg) / avg) * 100 if avg else 0

        if diff_pct <= -10:
            verdict = "偏低 ⚠ 請確認是否隱藏附加費"
        elif diff_pct <= -5:
            verdict = "低於平均（優惠）"
        elif diff_pct <= 5:
            verdict = "合理範圍"
        elif diff_pct <= 10:
            verdict = "略高於平均"
        else:
            verdict = "偏高 ⚠ 建議重新議價"

        return {
            "報價總費用": price,
            "歷史平均": round(avg, 0),
            "歷史最低": stats["最低"],
            "歷史最高": stats["最高"],
            "偏離幅度": f"{diff_pct:+.1f}%",
            "合理性判定": verdict,
            "樣本數": stats["樣本數"],
        }

    # ── 加權推薦 ──
    def recommend_best(self, quotes, weights=None):
        if weights is None:
            weights = {"price": 0.4, "transit": 0.3, "service": 0.3}
        parsed = [self.parse_quote(q) for q in quotes]
        parsed = [p for p in parsed if p is not None]
        if not parsed:
            return []

        totals = [p.get("總費用", 0) for p in parsed]
        transits = [p.get("Transit Time(天)", 0) for p in parsed]
        services = [p.get("服務評分", 4.0) for p in parsed]

        min_t, max_t = min(totals), max(totals)
        min_tr, max_tr = min(transits), max(transits)
        min_s, max_s = min(services), max(services)

        def normalize(val, lo, hi):
            if lo == hi:
                return 1.0
            return (hi - val) / (hi - lo)  # 越低越好 → 越高得分

        def norm_service(val, lo, hi):
            if lo == hi:
                return 1.0
            return (val - lo) / (hi - lo)  # 越高越好

        scored = []
        for i, p in enumerate(parsed):
            price_score = normalize(totals[i], min_t, max_t)
            transit_score = normalize(transits[i], min_tr, max_tr)
            service_score = norm_service(services[i], min_s, max_s)
            total_score = (
                weights["price"] * price_score
                + weights["transit"] * transit_score
                + weights["service"] * service_score
            )
            scored.append({
                "排名": 0,
                "貨代": p.get("貨代", f"報價 {i+1}"),
                "總費用": totals[i],
                "Transit Time": transits[i],
                "服務評分": services[i],
                "價格得分": round(price_score * 100, 1),
                "時效得分": round(transit_score * 100, 1),
                "服務得分": round(service_score * 100, 1),
                "加權總分": round(total_score * 100, 1),
            })

        scored.sort(key=lambda x: x["加權總分"], reverse=True)
        for rank, s in enumerate(scored, 1):
            s["排名"] = rank
        return scored

    # ── 異常偵測 ──
    def detect_anomalies(self, quote, origin=None, destination=None):
        q = self.parse_quote(quote)
        if not q:
            return []
        anomalies = []
        origin = origin or q.get("起運港", "")
        destination = destination or q.get("目的港", "")
        container = q.get("櫃型", "40GP")
        stats = self.price_db.get_average_price(origin, destination, container)
        if stats and stats["樣本數"] >= 3:
            avg = stats["平均"]
            sd = stats["標準差"]
            price = q.get("總費用", 0)
            if sd > 0:
                z = (price - avg) / sd
                if z > 2:
                    anomalies.append(f"總費用 {fmt_usd(price)} 顯著偏高（Z-score: {z:.2f}）")
                elif z < -2:
                    anomalies.append(f"總費用 {fmt_usd(price)} 顯著偏低（Z-score: {z:.2f}）— 請留意隱藏費用")

        ocean = q.get("海運費(USD)", 0)
        thc = q.get("THC", 0)
        if thc > 0 and ocean > 0 and thc / ocean > 0.25:
            anomalies.append(f"THC 佔海運費比例過高（{thc/ocean*100:.0f}%）")

        transit = q.get("Transit Time(天)", 0)
        if transit > 15:
            anomalies.append(f"航程天數偏長（{transit} 天），可能影響交期")
        return anomalies

    # ── 產生比價報告 Excel ──
    def generate_comparison_report(self, quotes, route_info, output_path=None):
        if output_path is None:
            output_path = os.path.join(
                SCRIPT_DIR,
                f"比價報告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            )

        wb = Workbook()
        # Sheet 1: 報價比較
        ws1 = wb.active
        ws1.title = "報價比較"
        parsed = [self.parse_quote(q) for q in quotes]
        parsed = [p for p in parsed if p is not None]

        headers = ["貨代", "海運費(USD)", "THC", "文件費", "其他費用", "總費用",
                    "Transit Time(天)", "服務評分"]
        ws1.append(headers)
        for p in parsed:
            ws1.append([
                p.get("貨代", ""),
                p.get("海運費(USD)", 0),
                p.get("THC", 0),
                p.get("文件費", 0),
                p.get("其他費用", 0),
                p.get("總費用", 0),
                p.get("Transit Time(天)", 0),
                p.get("服務評分", 0),
            ])
        style_sheet(ws1)

        # Sheet 2: 推薦排名
        ws2 = wb.create_sheet("推薦排名")
        recommendations = self.recommend_best(quotes)
        rec_headers = ["排名", "貨代", "總費用", "Transit Time", "服務評分",
                       "價格得分", "時效得分", "服務得分", "加權總分"]
        ws2.append(rec_headers)
        for r in recommendations:
            ws2.append([
                r["排名"], r["貨代"], r["總費用"], r["Transit Time"],
                r["服務評分"], r["價格得分"], r["時效得分"],
                r["服務得分"], r["加權總分"],
            ])
        style_sheet(ws2)

        # Sheet 3: 歷史參考
        ws3 = wb.create_sheet("歷史價格參考")
        origin = route_info.get("origin", "")
        dest = route_info.get("destination", "")
        container = route_info.get("container", "40GP")
        hist_quotes = self.price_db.get_quotes_by_route(origin, dest, container)
        if hist_quotes:
            hist_headers = list(hist_quotes[0].keys())
            ws3.append(hist_headers)
            for hq in hist_quotes:
                ws3.append([hq.get(h, "") for h in hist_headers])
            style_sheet(ws3)

        wb.save(output_path)
        return output_path


# ══════════════════════════════════════════════════════════════════════════════
# 5. BookingAssistant — 訂艙工作流管理
# ══════════════════════════════════════════════════════════════════════════════

class BookingAssistant:
    """訂艙工作流管理"""

    def __init__(self):
        self._counter = 0
        self.bookings = {}

    def _next_id(self):
        self._counter += 1
        return f"BK-{datetime.now().strftime('%Y%m%d')}-{self._counter:03d}"

    def create_booking_request(self, forwarder, route, cargo, dates):
        booking_id = self._next_id()
        booking = {
            "訂艙編號": booking_id,
            "貨代": forwarder if isinstance(forwarder, str) else forwarder.get("貨代名稱", ""),
            "起運港": route.get("origin", ""),
            "目的港": route.get("destination", ""),
            "貨物描述": cargo,
            "櫃型": route.get("container", "40GP"),
            "預計出貨日": dates.get("ship_date", ""),
            "截止訂艙日": dates.get("cutoff_date", ""),
            "狀態": "待確認",
            "建立時間": datetime.now().strftime("%Y-%m-%d %H:%M"),
        }
        self.bookings[booking_id] = booking
        return booking

    def track_booking(self, booking_id, status=None):
        if booking_id not in self.bookings:
            return f"找不到訂艙編號：{booking_id}"
        if status:
            self.bookings[booking_id]["狀態"] = status
        return self.bookings[booking_id]

    def generate_booking_summary(self):
        if not self.bookings:
            return "目前無任何訂艙記錄。"
        lines = [
            "══════════════════════════════════════════",
            "  【訂艙總覽】",
            f"  統計時間：{datetime.now().strftime('%Y-%m-%d %H:%M')}",
            "══════════════════════════════════════════",
        ]
        status_counts = {}
        for bk in self.bookings.values():
            s = bk["狀態"]
            status_counts[s] = status_counts.get(s, 0) + 1
        lines.append(f"  總訂艙數：{len(self.bookings)}")
        for s, c in status_counts.items():
            lines.append(f"    {s}：{c} 筆")
        lines.append("──────────────────────────────────────────")
        for bk in self.bookings.values():
            lines.append(
                f"  [{bk['訂艙編號']}] {bk['貨代']}  "
                f"{bk['起運港']}→{bk['目的港']}  "
                f"{bk['櫃型']}  狀態：{bk['狀態']}"
            )
        lines.append("══════════════════════════════════════════")
        return "\n".join(lines)

    def display_booking_form(self, booking):
        lines = [
            "══════════════════════════════════════════",
            "  【訂艙委託單】",
            "══════════════════════════════════════════",
        ]
        for k, v in booking.items():
            lines.append(f"  {k}：{v}")
        lines.append("══════════════════════════════════════════")
        return "\n".join(lines)


# ══════════════════════════════════════════════════════════════════════════════
# 6. Excel 資料庫產生器
# ══════════════════════════════════════════════════════════════════════════════

def generate_excel_db(path=None):
    """產生 貨代報價數據庫.xlsx"""
    if path is None:
        path = DB_PATH
    wb = Workbook()

    # ── Sheet 1: 貨代資料 ──
    ws1 = wb.active
    ws1.title = "貨代資料"
    fwd_headers = ["貨代名稱", "聯絡人", "WeCom", "電話", "郵箱",
                   "優勢航線", "服務評分", "合作年資", "備註"]
    ws1.append(fwd_headers)
    for f in SAMPLE_FORWARDERS:
        ws1.append([f.get(h, "") for h in fwd_headers])
    style_sheet(ws1)

    # ── Sheet 2: 歷史報價 ──
    ws2 = wb.create_sheet("歷史報價")
    q_headers = ["報價編號", "日期", "起運港", "目的港", "櫃型", "貨代",
                 "海運費(USD)", "THC", "文件費", "其他費用", "總費用",
                 "Transit Time(天)", "有效期", "備註"]
    ws2.append(q_headers)
    for q in SAMPLE_QUOTES:
        ws2.append([q.get(h, "") for h in q_headers])
    style_sheet(ws2)

    # ── Sheet 3: 航線參考 ──
    ws3 = wb.create_sheet("航線參考")
    route_headers = ["航線", "起運港代碼", "目的港代碼", "起運港", "目的港",
                     "標準航程(天)", "20GP價格區間(USD)", "40GP價格區間(USD)",
                     "40HQ價格區間(USD)", "主要船公司", "備註"]
    ws3.append(route_headers)
    route_data = [
        ["Shanghai→Laem Chabang", "CNSHA", "THLCH", "Shanghai", "Laem Chabang",
         "6-9", "$700-$850", "$900-$1,100", "$950-$1,200",
         "SITC, TS Lines, KMTC", "泰國最大港，轉關便利"],
        ["Shanghai→Jakarta", "CNSHA", "IDJKT", "Shanghai", "Jakarta",
         "7-10", "$600-$780", "$800-$1,050", "$850-$1,100",
         "ONE, Evergreen, PIL", "印尼主港，需注意清關時效"],
        ["Shanghai→Surabaya", "CNSHA", "IDSUB", "Shanghai", "Surabaya",
         "8-12", "$650-$820", "$850-$1,080", "$900-$1,150",
         "Maersk, CMA CGM, Hapag", "印尼第二大港"],
        ["Ningbo→Bangkok", "CNNGB", "THBKK", "Ningbo", "Bangkok",
         "5-8", "$680-$830", "$880-$1,070", "$920-$1,130",
         "SITC, TS Lines, WanHai", "寧波出發曼谷快船"],
        ["Shenzhen→Jakarta", "CNSZX", "IDJKT", "Shenzhen", "Jakarta",
         "6-9", "$580-$750", "$780-$1,000", "$820-$1,060",
         "PIL, WanHai, OOCL", "華南出發印尼線"],
    ]
    for row in route_data:
        ws3.append(row)
    style_sheet(ws3)

    wb.save(path)
    return path


# ══════════════════════════════════════════════════════════════════════════════
# 7. 文字圖表：價格趨勢
# ══════════════════════════════════════════════════════════════════════════════

def print_trend_chart(trend_data, title=""):
    """終端機文字長條圖"""
    if not trend_data:
        print("  （無趨勢資料）")
        return
    if title:
        print(f"\n  【{title}】")
    max_val = max(t["平均費用"] for t in trend_data)
    bar_max = 40
    print()
    for t in trend_data:
        avg = t["平均費用"]
        bar_len = int((avg / max_val) * bar_max) if max_val else 0
        bar = "█" * bar_len
        print(f"  {t['月份']}  {bar} {fmt_usd(avg)}  "
              f"(最低:{fmt_usd(t['最低費用'])} 最高:{fmt_usd(t['最高費用'])}  "
              f"樣本:{t['報價數量']})")
    print()


# ══════════════════════════════════════════════════════════════════════════════
# 8. CLI 主程式
# ══════════════════════════════════════════════════════════════════════════════

def run_demo():
    """完整 Demo 演示"""
    sep = "=" * 66
    print()
    print(sep)
    print("   訂艙比價自動化工具 — 完整演示 (Demo)")
    print(f"   執行時間：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(sep)

    # ── 初始化 ──
    fwd_db = FreightForwarderDB()
    price_db = PriceHistoryDB()
    inquiry_gen = InquiryGenerator()
    comparator = PriceComparator(price_db)
    booking_asst = BookingAssistant()

    # ── 1. 顯示 8 家貨代 ──
    print(f"\n{sep}")
    print("  ▶ 情境 1：顯示所有貨代（共 {0} 家）及評分".format(len(fwd_db.forwarders)))
    print(sep)
    fwd_db.display_table()
    print(f"\n  ★ 評分排行 Top 3：")
    for i, f in enumerate(fwd_db.get_top_rated(3), 1):
        print(f"    {i}. {f['貨代名稱']} — 評分 {f['服務評分']}，合作 {f['合作年資']} 年")

    # ── 2. 詢價單 ──
    print(f"\n{sep}")
    print("  ▶ 情境 2：產生詢價單 — Shanghai→Jakarta, 40HQ, 15CBM, 2026-06-25")
    print(sep)
    inquiry = inquiry_gen.generate_inquiry(
        origin="Shanghai",
        destination="Jakarta",
        cargo_info="電子零件，15CBM，毛重 8,000 KGS",
        container_type="40HQ",
        ship_date="2026-06-25",
        inquiry_type="FCL",
    )
    print(inquiry)

    # ── 3. 比價 ──
    print(f"\n{sep}")
    print("  ▶ 情境 3：比較 4 筆報價 — Shanghai→Laem Chabang, 40GP")
    print(sep)
    demo_quotes = [
        {
            "貨代": "上海遠洋國際物流",
            "起運港": "Shanghai", "目的港": "Laem Chabang",
            "櫃型": "40GP",
            "海運費(USD)": 850, "THC": 120, "文件費": 50, "其他費用": 0,
            "總費用": 1020, "Transit Time(天)": 7, "服務評分": 4.5,
        },
        {
            "貨代": "東盟物流集團",
            "起運港": "Shanghai", "目的港": "Laem Chabang",
            "櫃型": "40GP",
            "海運費(USD)": 780, "THC": 135, "文件費": 55, "其他費用": 0,
            "總費用": 970, "Transit Time(天)": 9, "服務評分": 4.8,
        },
        {
            "貨代": "中遠海運物流(東南亞)",
            "起運港": "Shanghai", "目的港": "Laem Chabang",
            "櫃型": "40GP",
            "海運費(USD)": 920, "THC": 110, "文件費": 45, "其他費用": 0,
            "總費用": 1075, "Transit Time(天)": 6, "服務評分": 4.3,
        },
        {
            "貨代": "嘉里國際貨運",
            "起運港": "Shanghai", "目的港": "Laem Chabang",
            "櫃型": "40GP",
            "海運費(USD)": 810, "THC": 125, "文件費": 50, "其他費用": 0,
            "總費用": 985, "Transit Time(天)": 8, "服務評分": 4.6,
        },
    ]
    comparator.compare_quotes(demo_quotes)

    # ── 4. 合理性分析 ──
    print(f"\n{sep}")
    print("  ▶ 情境 4：合理性分析 — 各報價 vs 歷史平均")
    print(sep)
    for dq in demo_quotes:
        fairness = comparator.analyze_fairness(dq, "Shanghai", "Laem Chabang")
        if fairness:
            print(f"\n  【{dq['貨代']}】")
            for k, v in fairness.items():
                print(f"    {k}：{v}")

    # ── 5. 推薦排名 ──
    print(f"\n{sep}")
    print("  ▶ 情境 5：加權推薦排名（價格 40% / 時效 30% / 服務 30%）")
    print(sep)
    recommendations = comparator.recommend_best(demo_quotes)
    rec_headers = ["排名", "貨代", "總費用", "Transit", "服務", "價格分", "時效分", "服務分", "加權總分"]
    rec_widths = [4, 22, 10, 8, 6, 8, 8, 8, 10]
    print(f"  {'│'.join(h.center(w) for h, w in zip(rec_headers, rec_widths))}")
    print(f"  {'┼'.join('─' * w for w in rec_widths)}")
    for r in recommendations:
        row_vals = [
            str(r["排名"]),
            r["貨代"],
            fmt_usd(r["總費用"]),
            f'{r["Transit Time"]}天',
            str(r["服務評分"]),
            str(r["價格得分"]),
            str(r["時效得分"]),
            str(r["服務得分"]),
            str(r["加權總分"]),
        ]
        print(f"  {'│'.join(v.center(w) for v, w in zip(row_vals, rec_widths))}")

    if recommendations:
        best = recommendations[0]
        print(f"\n  ★★★ 最佳推薦：{best['貨代']}（加權總分 {best['加權總分']}）")
        print(f"      總費用 {fmt_usd(best['總費用'])}，航程 {best['Transit Time']} 天，"
              f"服務評分 {best['服務評分']}")

    # ── 6. 價格趨勢 ──
    print(f"\n{sep}")
    print("  ▶ 情境 6：價格趨勢 — Shanghai→Jakarta 近 6 個月")
    print(sep)
    for ctype in ["20GP", "40GP", "40HQ"]:
        trend = price_db.get_price_trend("Shanghai", "Jakarta", ctype)
        print_trend_chart(trend, f"Shanghai→Jakarta {ctype}")

    # ── 7. 產生比價報告 Excel ──
    print(f"\n{sep}")
    print("  ▶ 情境 7：產生比價報告 Excel")
    print(sep)
    report_path = comparator.generate_comparison_report(
        demo_quotes,
        {"origin": "Shanghai", "destination": "Laem Chabang", "container": "40GP"},
    )
    print(f"  報告已儲存：{report_path}")

    # ── 8. 訂艙 ──
    print(f"\n{sep}")
    print("  ▶ 附加：訂艙工作流演示")
    print(sep)
    bk1 = booking_asst.create_booking_request(
        "東盟物流集團",
        {"origin": "Shanghai", "destination": "Laem Chabang", "container": "40GP"},
        "電子零件，28CBM，毛重 12,000 KGS",
        {"ship_date": "2026-06-25", "cutoff_date": "2026-06-22"},
    )
    print(booking_asst.display_booking_form(bk1))

    bk2 = booking_asst.create_booking_request(
        "深圳鵬程國際貨代",
        {"origin": "Shenzhen", "destination": "Jakarta", "container": "40HQ"},
        "紡織品，55CBM，毛重 18,000 KGS",
        {"ship_date": "2026-07-01", "cutoff_date": "2026-06-28"},
    )
    booking_asst.track_booking(bk2["訂艙編號"], "已確認")
    print(booking_asst.generate_booking_summary())

    # ── 產生資料庫 Excel ──
    print(f"\n{sep}")
    print("  ▶ 產生貨代報價資料庫 Excel")
    print(sep)
    db_path = generate_excel_db()
    print(f"  資料庫已儲存：{db_path}")

    # ── 異常偵測 ──
    print(f"\n{sep}")
    print("  ▶ 附加：異常偵測演示")
    print(sep)
    anomaly_quote = {
        "貨代": "測試貨代",
        "起運港": "Shanghai", "目的港": "Laem Chabang",
        "櫃型": "40GP",
        "海運費(USD)": 1800, "THC": 500, "文件費": 80, "其他費用": 100,
        "總費用": 2480, "Transit Time(天)": 18, "服務評分": 3.0,
    }
    anomalies = comparator.detect_anomalies(anomaly_quote)
    print(f"  測試報價：總費用 {fmt_usd(2480)}，航程 18 天")
    if anomalies:
        for a in anomalies:
            print(f"  ⚠ {a}")
    else:
        print("  無異常")

    print(f"\n{sep}")
    print("  Demo 完成！所有功能正常運作。")
    print(sep)
    print()


# ══════════════════════════════════════════════════════════════════════════════
# CLI
# ══════════════════════════════════════════════════════════════════════════════

def cmd_forwarders(args):
    db = FreightForwarderDB()
    if args.route:
        parts = args.route.split("→")
        if len(parts) == 2:
            results = db.get_forwarders_by_route(parts[0].strip(), parts[1].strip())
            print(f"\n  航線 {args.route} 的推薦貨代（共 {len(results)} 家）：")
            db.display_table(results)
        else:
            print("  格式錯誤，請使用：起運港→目的港")
    else:
        print(f"\n  所有貨代（共 {len(db.forwarders)} 家）：")
        db.display_table()


def cmd_add_quote(args):
    price_db = PriceHistoryDB()
    quote = {
        "日期": args.date or datetime.now().strftime("%Y-%m-%d"),
        "起運港": args.origin,
        "目的港": args.destination,
        "櫃型": args.container,
        "貨代": args.forwarder,
        "海運費(USD)": args.ocean,
        "THC": args.thc,
        "文件費": args.doc_fee,
        "其他費用": args.other_fee,
        "總費用": args.ocean + args.thc + args.doc_fee + args.other_fee,
        "Transit Time(天)": args.transit,
    }
    qid = price_db.add_quote(quote)
    print(f"\n  報價已新增：{qid}")
    print(f"  航線：{args.origin}→{args.destination}  櫃型：{args.container}")
    print(f"  總費用：{fmt_usd(quote['總費用'])}  航程：{args.transit} 天")


def cmd_compare(args):
    comparator = PriceComparator()
    if args.FILE and os.path.exists(args.FILE):
        wb = load_workbook(args.FILE)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        quotes = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            q = dict(zip(headers, row))
            quotes.append(q)
        print(f"\n  從 {args.FILE} 讀取 {len(quotes)} 筆報價：")
        comparator.compare_quotes(quotes)
    else:
        print(f"  檔案不存在：{args.FILE}")


def cmd_inquiry(args):
    gen = InquiryGenerator()
    inquiry_type = args.type.upper() if args.type else "FCL"
    msg = gen.generate_inquiry(
        origin=args.origin,
        destination=args.destination,
        cargo_info=args.cargo or "一般貨物",
        container_type=args.container,
        ship_date=args.date,
        inquiry_type=inquiry_type,
    )
    print(f"\n{msg}")


def cmd_recommend(args):
    comparator = PriceComparator()
    if args.FILE and os.path.exists(args.FILE):
        wb = load_workbook(args.FILE)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        quotes = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            q = dict(zip(headers, row))
            quotes.append(q)
        weights = {"price": args.w_price, "transit": args.w_transit, "service": args.w_service}
        print(f"\n  加權推薦（價格:{args.w_price} / 時效:{args.w_transit} / 服務:{args.w_service}）：\n")
        recs = comparator.recommend_best(quotes, weights)
        for r in recs:
            print(f"    #{r['排名']} {r['貨代']}  "
                  f"總分:{r['加權總分']}  "
                  f"費用:{fmt_usd(r['總費用'])}  "
                  f"航程:{r['Transit Time']}天  "
                  f"服務:{r['服務評分']}")
    else:
        print(f"  檔案不存在：{args.FILE}")


def cmd_history(args):
    price_db = PriceHistoryDB()
    quotes = price_db.get_quotes_by_route(args.origin, args.destination, args.container)
    if not quotes:
        print(f"\n  無 {args.origin}→{args.destination} ({args.container}) 的歷史報價")
        return
    print(f"\n  歷史報價：{args.origin}→{args.destination} ({args.container})  共 {len(quotes)} 筆\n")
    headers = ["日期", "貨代", "海運費(USD)", "THC", "文件費", "總費用", "Transit Time(天)"]
    widths = [12, 22, 12, 8, 8, 12, 12]
    print(f"  {'│'.join(h.center(w) for h, w in zip(headers, widths))}")
    print(f"  {'┼'.join('─' * w for w in widths)}")
    for q in quotes:
        vals = [
            q.get("日期", ""),
            q.get("貨代", "")[:14],
            fmt_usd(q.get("海運費(USD)", 0)),
            fmt_usd(q.get("THC", 0)),
            fmt_usd(q.get("文件費", 0)),
            fmt_usd(q.get("總費用", 0)),
            f'{q.get("Transit Time(天)", "")}天',
        ]
        print(f"  {'│'.join(str(v).center(w) for v, w in zip(vals, widths))}")

    stats = price_db.get_average_price(args.origin, args.destination, args.container)
    if stats:
        print(f"\n  統計：平均 {fmt_usd(stats['平均'])}  "
              f"最低 {fmt_usd(stats['最低'])}  "
              f"最高 {fmt_usd(stats['最高'])}  "
              f"標準差 {fmt_usd(stats['標準差'])}")


def cmd_trend(args):
    price_db = PriceHistoryDB()
    origin = args.origin or "Shanghai"
    dest = args.destination or "Jakarta"
    container = args.container or "40GP"
    trend = price_db.get_price_trend(origin, dest, container)
    print_trend_chart(trend, f"{origin}→{dest} {container} 價格趨勢")


def cmd_report(args):
    comparator = PriceComparator()
    price_db = PriceHistoryDB()
    origin = args.origin or "Shanghai"
    dest = args.destination or "Laem Chabang"
    container = args.container or "40GP"
    quotes = price_db.get_quotes_by_route(origin, dest, container)
    if len(quotes) < 2:
        print("  歷史報價不足 2 筆，無法產生報告。")
        return
    # 取最近 6 筆
    recent = quotes[-6:]
    path = comparator.generate_comparison_report(
        recent,
        {"origin": origin, "destination": dest, "container": container},
    )
    print(f"\n  比價報告已產生：{path}")


def main():
    parser = argparse.ArgumentParser(
        description="訂艙比價自動化工具 — 報關部門海運費比價、貨代管理、詢價自動生成",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "範例：\n"
            "  python3 訂艙比價工具.py --demo\n"
            "  python3 訂艙比價工具.py --forwarders\n"
            "  python3 訂艙比價工具.py --forwarders --route 'Shanghai→Jakarta'\n"
            "  python3 訂艙比價工具.py --inquiry --origin Shanghai --destination Jakarta --container 40HQ --date 2026-06-25\n"
            "  python3 訂艙比價工具.py --history --origin Shanghai --destination Jakarta --container 40GP\n"
            "  python3 訂艙比價工具.py --trend --origin Shanghai --destination Jakarta\n"
            "  python3 訂艙比價工具.py --compare quotes.xlsx\n"
            "  python3 訂艙比價工具.py --recommend quotes.xlsx\n"
            "  python3 訂艙比價工具.py --report --origin Shanghai --destination 'Laem Chabang'\n"
        ),
    )
    parser.add_argument("--demo", action="store_true", help="執行完整演示")
    parser.add_argument("--forwarders", action="store_true", help="列出所有貨代")
    parser.add_argument("--route", type=str, default=None, help="依航線篩選貨代（格式：起運港→目的港）")

    # 新增報價
    parser.add_argument("--add-quote", action="store_true", help="新增一筆報價")
    parser.add_argument("--origin", type=str, default="Shanghai", help="起運港")
    parser.add_argument("--destination", type=str, default="Jakarta", help="目的港")
    parser.add_argument("--container", type=str, default="40GP", help="櫃型 (20GP/40GP/40HQ)")
    parser.add_argument("--forwarder", type=str, default="", help="貨代名稱")
    parser.add_argument("--ocean", type=float, default=0, help="海運費 USD")
    parser.add_argument("--thc", type=float, default=0, help="THC")
    parser.add_argument("--doc-fee", type=float, default=0, help="文件費")
    parser.add_argument("--other-fee", type=float, default=0, help="其他費用")
    parser.add_argument("--transit", type=int, default=7, help="航程天數")
    parser.add_argument("--date", type=str, default=None, help="報價日期")

    # 比價 & 推薦
    parser.add_argument("--compare", metavar="FILE", help="從 Excel 比較報價")
    parser.add_argument("--recommend", metavar="FILE", help="從 Excel 推薦最佳報價")
    parser.add_argument("--w-price", type=float, default=0.4, help="價格權重（預設 0.4）")
    parser.add_argument("--w-transit", type=float, default=0.3, help="時效權重（預設 0.3）")
    parser.add_argument("--w-service", type=float, default=0.3, help="服務權重（預設 0.3）")

    # 詢價
    parser.add_argument("--inquiry", action="store_true", help="產生詢價訊息")
    parser.add_argument("--cargo", type=str, default=None, help="貨物描述")
    parser.add_argument("--type", type=str, default="FCL", help="詢價類型 (FCL/LCL/SPECIAL)")

    # 歷史 & 趨勢
    parser.add_argument("--history", action="store_true", help="顯示歷史報價")
    parser.add_argument("--trend", action="store_true", help="顯示價格趨勢")

    # 報告
    parser.add_argument("--report", action="store_true", help="產生比價報告 Excel")

    args = parser.parse_args()

    if args.demo:
        run_demo()
    elif args.forwarders:
        cmd_forwarders(args)
    elif args.add_quote:
        cmd_add_quote(args)
    elif args.compare:
        cmd_compare(args)
    elif args.inquiry:
        cmd_inquiry(args)
    elif args.recommend:
        cmd_recommend(args)
    elif args.history:
        cmd_history(args)
    elif args.trend:
        cmd_trend(args)
    elif args.report:
        cmd_report(args)
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
